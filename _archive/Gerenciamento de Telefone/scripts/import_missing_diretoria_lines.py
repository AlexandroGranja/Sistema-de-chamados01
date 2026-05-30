"""Importa linhas da aba Diretoria que faltam no banco atual."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.db.repository import load_linhas, save_linhas


ROOT = Path(__file__).resolve().parent.parent
PLANILHAS_DIR = ROOT / "data" / "planilhas"
ATIVAS_PATH = PLANILHAS_DIR / "Telefones11.25_SomenteAtivas.xlsx"
RELACAO_PATH = PLANILHAS_DIR / "Relação Linhas Prosper270226.xlsx"


def digits_only(value: object) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def _load_relacao_set() -> set[str]:
    wb = load_workbook(RELACAO_PATH, data_only=True)
    ws = wb["ListaAtual"] if "ListaAtual" in wb.sheetnames else wb[wb.sheetnames[0]]
    linhas: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        linha = digits_only(ws.cell(row_idx, 4).value)
        if 10 <= len(linha) <= 13:
            linhas.add(linha)
    return linhas


def _load_diretoria_rows() -> list[dict[str, object]]:
    wb = load_workbook(ATIVAS_PATH, data_only=True)
    if "Diretoria" not in wb.sheetnames:
        return []

    ws = wb["Diretoria"]
    headers = [str(ws.cell(1, col_idx).value or "").strip() for col_idx in range(1, ws.max_column + 1)]
    relacao_set = _load_relacao_set()
    rows: list[dict[str, object]] = []
    for row_idx in range(2, ws.max_row + 1):
        linha = digits_only(ws.cell(row_idx, 2).value)
        if linha not in relacao_set:
            continue
        row = {headers[col_idx - 1] if headers[col_idx - 1] else f"col_{col_idx}": ws.cell(row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)}
        row["LinhaNorm"] = linha
        rows.append(row)
    return rows


def _build_new_rows(rows: list[dict[str, object]], existing_lines: set[str]) -> pd.DataFrame:
    payload: list[dict[str, object]] = []
    for row in rows:
        linha = str(row.get("LinhaNorm") or "").strip()
        if not linha or linha in existing_lines:
            continue

        cargo = str(row.get("Cargo") or "").strip()
        aparelho = str(row.get("Aparelho") or "").strip()
        modelo = str(row.get("Modelo") or "").strip()
        setor = str(row.get("Setor") or "").strip()
        perfil = str(row.get("Perfil") or "").strip()
        imei_a = digits_only(row.get("IMEI"))
        imei_b = digits_only(row.get("IMEI2"))
        chip = digits_only(row.get("CHIP"))

        payload.append(
            {
                "Codigo": "",
                "Nome": str(row.get("Nomes") or "").strip(),
                "Equipe": "Diretoria",
                "EquipePadrao": "Diretoria",
                "GrupoEquipe": "Equipe Interna",
                "TipoEquipe": "Interna",
                "Localidade": "Diretoria",
                "Gestor": "",
                "Supervisor": "",
                "Segmento": "Internos",
                "Papel": cargo.title() if cargo else "",
                "Linha": linha,
                "E-mail": "",
                "Gerenciamento": "",
                "Data da Troca": "",
                "Data Retorno": "",
                "Data Ocorrência": "",
                "Data Solicitação TBS": "",
                "Motivo": "",
                "Observação": "",
                "IMEI A": imei_a,
                "IMEI B": imei_b,
                "Marca": aparelho,
                "CHIP": chip,
                "Aparelho": aparelho,
                "Modelo": modelo,
                "Setor": setor,
                "Cargo": cargo,
                "Desconto": "",
                "Perfil": perfil,
                "Empresa": "",
                "Ativo": "",
                "Numero de Serie": "",
                "Patrimonio": "",
                "Operadora": "",
                "Aba": "Diretoria",
            }
        )
    return pd.DataFrame(payload)


def main() -> int:
    atuais = load_linhas("ativas")
    existing_lines = set(atuais["Linha"].fillna("").astype(str).str.strip()) if "Linha" in atuais.columns else set()
    diretoria_rows = _load_diretoria_rows()
    novos = _build_new_rows(diretoria_rows, existing_lines)

    if novos.empty:
        print("Nenhuma nova linha da Diretoria para importar.")
        return 0

    atualizados = pd.concat([atuais, novos], ignore_index=True)
    atualizados = atualizados.drop_duplicates(subset=["Linha"], keep="first")
    total = save_linhas(atualizados, modo="ativas")

    print(f"Linhas da Diretoria importadas: {len(novos)}")
    print(f"Total de linhas ativas gravadas: {total}")
    print("Linhas adicionadas:")
    print(novos[["Linha", "Nome", "EquipePadrao", "Segmento", "Aparelho", "Modelo", "Cargo"]].to_string(index=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
