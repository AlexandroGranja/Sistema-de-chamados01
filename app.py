from __future__ import annotations

import io
import json
import os
import secrets
import base64
import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote

import pandas as pd
import streamlit as st

# Cookie manager para login persistente (inicializado em main, após set_page_config)
HAS_COOKIES = False
_cookies = None

def _init_cookies():
    """Inicializa o gerenciador de cookies. Chamar após set_page_config."""
    global _cookies, HAS_COOKIES
    if _cookies is not None:
        return
    try:
        from streamlit_cookies_manager import EncryptedCookieManager
        _cookies = EncryptedCookieManager(
            prefix="planilhas_tel/",
            password=os.environ.get("COOKIES_PASSWORD", "planilhas_telefones_secret_2024"),
        )
        HAS_COOKIES = True
    except ImportError:
        HAS_COOKIES = False
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Suporte a banco de dados (opcional)
try:
    from src.db.repository import (
        load_linhas, has_data, get_db_path, save_linhas,
        init_db, verificar_login, criar_usuario, listar_usuarios, excluir_usuario, atualizar_senha_usuario, tem_usuarios,
        listar_usuarios_com_status_chamados, contar_usuarios_sem_chamados,
        obter_usuario_por_username, obter_usuario_app_id, criar_sso_code, criar_sessao, validar_sessao, encerrar_sessao,
        registrar_auditoria, listar_auditoria,
        resolver_referencia_chamado, preparar_referencia_chamado,
        garantir_chamado_stub, vincular_chamado_linha, registrar_chamado_evento, registrar_movimentacao_linha,
        criar_chamado, listar_chamados, obter_chamado, atualizar_status_chamado,
    )
    HAS_DB = True
except ImportError:
    HAS_DB = False
    init_db = verificar_login = criar_usuario = listar_usuarios = excluir_usuario = atualizar_senha_usuario = tem_usuarios = None
    listar_usuarios_com_status_chamados = contar_usuarios_sem_chamados = None
    obter_usuario_por_username = criar_sessao = validar_sessao = encerrar_sessao = None
    registrar_auditoria = listar_auditoria = None
    resolver_referencia_chamado = preparar_referencia_chamado = None
    garantir_chamado_stub = vincular_chamado_linha = registrar_chamado_evento = registrar_movimentacao_linha = None
    criar_chamado = listar_chamados = obter_chamado = atualizar_status_chamado = None


from src.core.config import (
    RULES_FILE, DOC_DIR, is_postgres_configured, get_chamados_app_url,
)
from src.core.streamlit_access import resolve_streamlit_access
ABAS_ALIMENTO = ["Centro Alimentos"]
ABAS_MEDICAMENTO = ["Distrito Norte", "Distrito Sul"]
ABAS_FOCO = ["Distrito Norte", "Distrito Sul", "Centro Alimentos", "Promotores", "Internos", "Troca de Aparelho", "Devolução Manutenção", "Roubo-Perda"]
ABAS_PROMOTORES = ["Promotores"]
ABAS_INTERNOS = ["Internos"]
ABAS_MANUTENCAO = ["Troca de Aparelho", "Devolução Manutenção", "Devolucao Manutencao"]
ABAS_ROUBO_PERDA = ["Roubo-Perda", "Roubo e Perda"]
EQUIPES_PROMOTORES = ["Promotores"]
EQUIPES_INTERNOS = ["Internos"]
EQUIPES_MANUTENCAO = ["Manutenção"]
EQUIPES_ROUBO_PERDA = ["Roubo e Perda"]
EQUIPES_ALIMENTO = [
    "Gerentes do Alimento",
    "Consumo Baixada",
    "Consumo Oeste",
    "Consumo Zona Norte",
    "Consumo Niteroi",
    "Equipe Especial",
    "Gerente Senior",
]
EQUIPES_MEDICAMENTO = [
    "Gerentes do Medicamento",
    "Distrito Norte",
    "Distrito Sul",
]
GESTORES_MEDICAMENTO = {
    "Distrito Norte": "Gestora Norte Demo",
    "Distrito Sul": "Gestor Sul Demo",
}
DEFAULT_COLUMNS = [
    "Codigo",
    "Nome",
    "Nome de Guerra",
    "Equipe",
    "Linha",
    "E-mail",
    "Gerenciamento",
    "Data da Troca",
    "Data Retorno",
    "Data Ocorrência",
    "Data Solicitação TBS",
    "Motivo",
    "Observação",
    "IMEI A",
    "IMEI B",
    "Marca",
    "CHIP",
    "Aparelho",
    "Modelo",
    "Setor",
    "Cargo",
    "Desconto",
    "Perfil",
    "Empresa",
    "Ativo",
    "Numero de Serie",
    "Patrimonio",
    "Operadora",
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn"
    )
    return text.lower()


def title_case_safe(value: str) -> str:
    tokens = [t for t in value.replace("/", " / ").split(" ") if t != ""]
    if not tokens:
        return "Sem Equipe"
    out: list[str] = []
    for token in tokens:
        if token == "/":
            out.append(token)
        elif token.isupper() and len(token) <= 4:
            out.append(token)
        else:
            out.append(token.capitalize())
    return " ".join(out).replace(" / ", "/")


def normalize_team_key(value: Any) -> str:
    text = normalize_text(value)
    text = " ".join(text.split())
    return text


def digits_only(value: Any) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value) if ch.isdigit())


def is_valid_phone(value: Any) -> bool:
    digits = digits_only(value)
    return 10 <= len(digits) <= 13


def build_vaga_mask(df: pd.DataFrame) -> pd.Series:
    """Linha vaga = tem linha e nao tem usuario cadastrado ou esta marcada como VAGO."""
    if df.empty:
        return pd.Series(dtype=bool)
    linhas_preenchidas = (
        df["Linha"].fillna("").astype(str).str.strip() != ""
        if "Linha" in df.columns
        else pd.Series(False, index=df.index)
    )
    nomes_norm = (
        df["Nome"].fillna("").astype(str).map(normalize_text)
        if "Nome" in df.columns
        else pd.Series("", index=df.index)
    )
    sem_usuario = nomes_norm.isin(["", "vago"])
    return linhas_preenchidas & sem_usuario


def is_valid_email(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", text))


def is_valid_imei(value: Any) -> bool:
    digits = digits_only(value)
    if not digits:
        return True
    return len(digits) == 15


def is_temporary_line_identifier(value: Any) -> bool:
    text = str(value or "").strip().upper()
    return text.startswith("NOVA-") or text.startswith("MANUT-")


def collect_editor_validation_errors(changed_rows_df: pd.DataFrame, full_df: pd.DataFrame) -> list[dict[str, str]]:
    """Valida campos criticos apenas nas linhas alteradas antes de salvar."""
    if changed_rows_df.empty:
        return []

    errors: list[dict[str, str]] = []
    duplicate_counts: dict[str, int] = {}

    if "Linha" in full_df.columns:
        full_lines = full_df["Linha"].fillna("").astype(str).str.strip()
        valid_lines = [
            line
            for line in full_lines.tolist()
            if line and not is_temporary_line_identifier(line)
        ]
        duplicate_counts = pd.Series(valid_lines).value_counts().to_dict() if valid_lines else {}

    for _, row in changed_rows_df.iterrows():
        linha_raw = str(row.get("Linha") or "").strip()
        linha_label = linha_raw or "(sem linha)"

        if not linha_raw:
            errors.append(
                {
                    "Linha": linha_label,
                    "Campo": "Linha",
                    "Problema": "Preencha a linha antes de salvar.",
                }
            )
        elif not is_temporary_line_identifier(linha_raw) and not is_valid_phone(linha_raw):
            errors.append(
                {
                    "Linha": linha_label,
                    "Campo": "Linha",
                    "Problema": "Informe uma linha com 10 a 13 digitos.",
                }
            )
        elif duplicate_counts.get(linha_raw, 0) > 1:
            errors.append(
                {
                    "Linha": linha_label,
                    "Campo": "Linha",
                    "Problema": "Ja existe outra linha com esse mesmo numero neste modo.",
                }
            )

        email_value = str(row.get("E-mail") or "").strip()
        if email_value and not is_valid_email(email_value):
            errors.append(
                {
                    "Linha": linha_label,
                    "Campo": "E-mail",
                    "Problema": "Informe um e-mail valido.",
                }
            )

        for imei_field in ("IMEI A", "IMEI B"):
            imei_value = str(row.get(imei_field) or "").strip()
            if imei_value and not is_valid_imei(imei_value):
                errors.append(
                    {
                        "Linha": linha_label,
                        "Campo": imei_field,
                        "Problema": "O IMEI deve ter 15 digitos.",
                    }
                )

    return errors


def collect_changed_editor_rows(base_views: list[pd.DataFrame], edited_views: list[pd.DataFrame]) -> pd.DataFrame:
    """Retorna apenas as linhas realmente alteradas no editor."""
    if not base_views or not edited_views:
        return pd.DataFrame()

    base_df = pd.concat(base_views, ignore_index=True)
    edited_df = pd.concat(edited_views, ignore_index=True)
    if "__row_key" not in base_df.columns or "__row_key" not in edited_df.columns:
        return pd.DataFrame()

    base_idx = base_df.set_index("__row_key", drop=False)
    edit_idx = edited_df.set_index("__row_key", drop=False)
    cols_compare = [
        c
        for c in edited_df.columns
        if c in base_df.columns and c not in ("__row_key", "Conflito", "Validacao", "Selecionar", "⬆", "⬇", "Mover equipe", "✏", "↕ Ordem")
    ]

    changed_keys: list[str] = []
    for rk in edited_df["__row_key"].dropna().astype(str).unique():
        if rk not in base_idx.index or rk not in edit_idx.index:
            continue
        base_row = base_idx.loc[rk]
        edit_row = edit_idx.loc[rk]
        if any(str(base_row.get(col, "") or "").strip() != str(edit_row.get(col, "") or "").strip() for col in cols_compare):
            changed_keys.append(rk)

    if not changed_keys:
        return pd.DataFrame(columns=edited_df.columns)

    return edited_df[edited_df["__row_key"].fillna("").astype(str).isin(changed_keys)].copy()


def build_candidate_full_df(full_df: pd.DataFrame, changed_rows_df: pd.DataFrame) -> pd.DataFrame:
    """Monta a versao candidata para validacao antes do salvamento."""
    if changed_rows_df.empty or "__row_key" not in full_df.columns or "__row_key" not in changed_rows_df.columns:
        return full_df.copy()

    candidate_df = full_df.copy()
    mask_changed = candidate_df["__row_key"].fillna("").astype(str).isin(
        changed_rows_df["__row_key"].fillna("").astype(str)
    )
    unchanged_df = candidate_df[~mask_changed].copy()
    return pd.concat([unchanged_df, changed_rows_df.copy()], ignore_index=True)


def find_header_row(ws) -> tuple[int | None, int | None]:
    max_r = min(80, ws.max_row)
    max_c = min(80, ws.max_column)
    for row in range(1, max_r + 1):
        for col in range(1, max_c + 1):
            if normalize_text(ws.cell(row, col).value) == "linha":
                return row, col
    return None, None


def collect_headers(ws, header_row: int, start_col: int, limit: int = 40) -> list[str]:
    headers: list[str] = []
    empty_streak = 0
    max_c = min(ws.max_column, start_col + limit)
    for col in range(1, max_c + 1):
        value = ws.cell(header_row, col).value
        text = str(value).strip() if value is not None else ""
        if text == "":
            headers.append(f"col_{col}")
            empty_streak += 1
        else:
            headers.append(text)
            empty_streak = 0
        if col >= start_col and empty_streak >= 8:
            break
    return headers


def map_known_column(column_name: str) -> str:
    c = normalize_text(column_name)
    mapping = {
        "codigo": "Codigo",
        "codigos": "Codigo",
        "asset": "Codigo",
        "asset tag": "Codigo",
        "nome": "Nome",
        "nomes": "Nome",
        "nome colaborador": "Nome",
        "nome de guerra": "Nome de Guerra",
        "equipe": "Equipe",
        "local": "Localidade",
        "localidade": "Localidade",
        "data da troca": "Data da Troca",
        "data troca": "Data da Troca",
        "data ocorrencia": "Data Ocorrência",
        "data ocorrência": "Data Ocorrência",
        "data retorno": "Data Retorno",
        "data de retorno": "Data Retorno",
        "data solicitacao tbs": "Data Solicitação TBS",
        "data solicitação tbs": "Data Solicitação TBS",
        "linha": "Linha",
        "email": "E-mail",
        "e-mail": "E-mail",
        "gerenciamento": "Gerenciamento",
        "bloqueio": "Gerenciamento",
        "imei": "IMEI A",
        "imei a": "IMEI A",
        "imei2": "IMEI B",
        "imei b": "IMEI B",
        "chip": "CHIP",
        "marca": "Marca",
        "aparelho": "Aparelho",
        "modelo": "Modelo",
        "setor": "Setor",
        "cargo": "Cargo",
        "desconto": "Desconto",
        "perfil": "Perfil",
        "empresa": "Empresa",
        "ativo": "Ativo",
        "ativos": "Ativo",
        "numero de serie": "Numero de Serie",
        "patrimonio": "Patrimonio",
        "motivo": "Motivo",
        "obs": "Observação",
        "ns": "Numero de Serie",
        "operadora": "Operadora",
        "s/n": "Numero de Serie",
        "s n": "Numero de Serie",
        "nº serie": "Numero de Serie",
        # Prosper Norte usa nome do gerente como cabeçalho da coluna código
        "fernando goncalves de mello": "Codigo",
        "goncalves d": "Codigo",
        "goncalves": "Codigo",
    }
    return mapping.get(c, column_name)


def detect_tipo_equipe(equipe: str, aba: str) -> str:
    e = normalize_text(equipe)
    a = normalize_text(aba)
    if "intern" in e or "intern" in a:
        return "Interna"
    return "Externa"


def detect_localidade(equipe: str) -> str:
    e = str(equipe).strip()
    if e == "":
        return ""
    markers = ["gerencia", "gerente", "supervisor", "promotor", "diretoria", "internos"]
    low = normalize_text(e)
    if any(marker in low for marker in markers):
        return ""
    return title_case_safe(e)


def _build_alimento_map(rules_path: Path) -> dict[str, dict[str, str]]:
    """Monta mapeamento localidade -> equipe, gestor, supervisor para Alimento (equipes_alimento.csv)."""
    ali_path = rules_path.parent / "equipes_alimento.csv"
    out: dict[str, dict[str, str]] = {}
    team_defaults: dict[str, dict[str, str]] = {}

    if not ali_path.exists():
        return out

    try:
        ref = pd.read_csv(ali_path, dtype=str).fillna("")
        for _, r in ref.iterrows():
            er = str(r.get("equipe_real", "")).strip()
            loc = str(r.get("localidade", "")).strip()
            gerente = str(r.get("gerente", "")).strip()
            supervisor = str(r.get("supervisor", "")).strip()
            if loc:
                key = normalize_team_key(loc)
                out[key] = {"equipe": er, "gestor": gerente, "supervisor": supervisor}
            else:
                eq_key = normalize_team_key(er)
                team_defaults[eq_key] = {"equipe": er, "gestor": gerente, "supervisor": supervisor}
        for eq in ["consumo baixada", "consumo oeste", "consumo zona norte", "consumo niteroi"]:
            out[eq] = team_defaults.get(eq, {"equipe": title_case_safe(eq), "gestor": "Marcelo Neves", "supervisor": ""})
        out["especial consumo"] = {"equipe": "Equipe Especial", "gestor": "Marco Antonio Neves Suzart", "supervisor": "Ricardo Cascao"}
        for loc, eq, sup in [
            ("alcantara", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
            ("cabucu", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
            ("piabetá", "Consumo Baixada", ""),
            ("piabeta", "Consumo Baixada", ""),
            ("realengo", "Consumo Oeste", "Marcelo Martins Da Costa"),
            ("barra da tijuca", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("recreio", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("botafogo", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("taquara", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("anchieta", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("lote xv", "Consumo Baixada", ""),
            ("niteroi i", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
            ("niterói i", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
            ("conveniencia farma 12", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("conveniencia farma 20", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("key account", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("supervisora senior", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("gerencia varejo ii", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("consumo zona sul", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("itaborai", "Consumo Baixada", ""),
        ]:
            out[normalize_team_key(loc)] = {"equipe": eq, "gestor": "Marcelo Neves", "supervisor": sup}
        for k in ["rede especial 1", "rede especial 2", "rede especial 3", "rede especial 4", "rota especial 1", "rota especial 2",
                  "rota especial 3", "rota especial 4", "impulso 01", "impulso 02", "impulso 03", "impulso 04",
                  "impulso 05", "auto servico 01", "auto servico 02", "auto servico 03", "auto servico 04",
                  "auto servico 05", "auto servico 06", "a s impulso 2", "a s impulso 3", "a s impulso 4",
                  "a s impulso 5"]:
            out[k] = {"equipe": "Equipe Especial", "gestor": "Marco Antonio Neves Suzart", "supervisor": "Ricardo Cascao"}
    except Exception:
        pass
    return out


def detect_grupo_equipe(equipe: str, tipo: str) -> str:
    e = normalize_text(equipe)
    special_tokens = ["especial", "impulso", "auto servico", "auto servico", "rota especial"]
    if any(token in e for token in special_tokens):
        return "Equipe Especial"
    if normalize_text(tipo) == "interna":
        return "Equipe Interna"
    return "Outras Equipes Externas"


def ensure_rules_file(df: pd.DataFrame, rules_path: Path) -> pd.DataFrame:
    if rules_path.exists():
        rules = pd.read_csv(rules_path, dtype=str).fillna("")
        # Backward-compatible migration of rule columns.
        if "equipe_key" not in rules.columns:
            rules["equipe_key"] = rules.get("equipe_origem", "").map(normalize_team_key)
        if "equipe_origem" not in rules.columns:
            rules["equipe_origem"] = ""
        if "equipe_padrao" not in rules.columns:
            rules["equipe_padrao"] = rules["equipe_origem"]
        if "tipo_equipe" not in rules.columns:
            rules["tipo_equipe"] = "Externa"
        if "localidade" not in rules.columns:
            rules["localidade"] = ""
        if "grupo_equipe" not in rules.columns:
            rules["grupo_equipe"] = rules.apply(
                lambda r: detect_grupo_equipe(str(r.get("equipe_padrao", "")), str(r.get("tipo_equipe", ""))),
                axis=1,
            )
        if "gestor" not in rules.columns:
            rules["gestor"] = ""
        if "supervisor" not in rules.columns:
            rules["supervisor"] = ""
        if "segmento" not in rules.columns:
            rules["segmento"] = "Alimento"
        if "eh_equipe" not in rules.columns:
            rules["eh_equipe"] = "True"
        if "equipe_pai" not in rules.columns:
            rules["equipe_pai"] = ""
        rules = rules[
            [
                "equipe_key",
                "equipe_origem",
                "equipe_padrao",
                "grupo_equipe",
                "tipo_equipe",
                "localidade",
                "gestor",
                "supervisor",
                "segmento",
                "eh_equipe",
                "equipe_pai",
            ]
        ]
        rules.to_csv(rules_path, index=False, encoding="utf-8-sig")
        return rules

    base = (
        df[["Equipe", "Aba"]]
        .drop_duplicates()
        .assign(equipe_key=lambda x: x["Equipe"].map(normalize_team_key))
        .sort_values("Equipe")
    )
    base["equipe_origem"] = base["Equipe"]
    base["equipe_padrao"] = base["Equipe"].map(lambda v: title_case_safe(str(v)))
    base["tipo_equipe"] = base.apply(lambda r: detect_tipo_equipe(str(r["Equipe"]), str(r["Aba"])), axis=1)
    base["localidade"] = base["Equipe"].map(detect_localidade)
    base["grupo_equipe"] = base.apply(
        lambda r: detect_grupo_equipe(str(r["equipe_padrao"]), str(r["tipo_equipe"])),
        axis=1,
    )
    base["gestor"] = ""
    base["supervisor"] = ""
    base["segmento"] = "Alimento"
    base["eh_equipe"] = "True"
    base["equipe_pai"] = ""
    rules = base[
        [
            "equipe_key",
            "equipe_origem",
            "equipe_padrao",
            "grupo_equipe",
            "tipo_equipe",
            "localidade",
            "gestor",
            "supervisor",
            "segmento",
            "eh_equipe",
            "equipe_pai",
        ]
    ]
    rules.to_csv(rules_path, index=False, encoding="utf-8-sig")
    return rules


def apply_team_standardization(df: pd.DataFrame, rules_path: Path) -> pd.DataFrame:
    rules = ensure_rules_file(df, rules_path)
    merged = df.copy()
    merged["equipe_key"] = merged["Equipe"].map(normalize_team_key)

    dedup_rules = rules.drop_duplicates(subset=["equipe_key"], keep="first")
    merged = merged.merge(
        dedup_rules[
            [
                "equipe_key",
                "equipe_padrao",
                "grupo_equipe",
                "tipo_equipe",
                "localidade",
                "gestor",
                "supervisor",
                "segmento",
            ]
        ],
        on="equipe_key",
        how="left",
    )

    merged["EquipePadrao"] = merged["equipe_padrao"].fillna("").astype(str).str.strip()
    merged.loc[merged["EquipePadrao"] == "", "EquipePadrao"] = merged["Equipe"].map(
        lambda v: title_case_safe(str(v))
    )

    merged["TipoEquipe"] = merged["tipo_equipe"].fillna("").astype(str).str.strip()
    merged.loc[merged["TipoEquipe"] == "", "TipoEquipe"] = merged.apply(
        lambda r: detect_tipo_equipe(str(r["Equipe"]), str(r["Aba"])),
        axis=1,
    )

    merged["GrupoEquipe"] = merged["grupo_equipe"].fillna("").astype(str).str.strip()
    merged.loc[merged["GrupoEquipe"] == "", "GrupoEquipe"] = merged.apply(
        lambda r: detect_grupo_equipe(str(r["EquipePadrao"]), str(r["TipoEquipe"])),
        axis=1,
    )

    merged["Localidade"] = merged["localidade"].fillna("").astype(str).str.strip()
    merged["Gestor"] = merged["gestor"].fillna("").astype(str).str.strip()
    merged["Supervisor"] = merged["supervisor"].fillna("").astype(str).str.strip()
    merged["Segmento"] = merged["segmento"].fillna("Alimento").astype(str).str.strip()
    merged.loc[merged["Segmento"] == "", "Segmento"] = "Alimento"
    merged.drop(
        columns=[
            "equipe_key",
            "equipe_padrao",
            "grupo_equipe",
            "tipo_equipe",
            "localidade",
            "gestor",
            "supervisor",
            "segmento",
        ],
        inplace=True,
    )

    # Alimento (Nova Prosper): aplicar mesmo padrão de equipes/gerentes que linhas ativas
    mask_ali = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_ALIMENTO]
    )
    if mask_ali.any():
        ali_map = _build_alimento_map(rules_path)
        for idx in merged.index[mask_ali]:
            key = normalize_team_key(merged.at[idx, "Equipe"])
            if key in ali_map:
                merged.at[idx, "EquipePadrao"] = ali_map[key]["equipe"]
                merged.at[idx, "Gestor"] = ali_map[key]["gestor"]
                merged.at[idx, "Supervisor"] = ali_map[key]["supervisor"]
                merged.at[idx, "Localidade"] = merged.at[idx, "Equipe"]
            else:
                # Fallback: Equipe Especial ou Consumo conforme padrão do nome
                low = key
                if any(t in low for t in ["especial", "impulso", "rota especial", "auto servico", "a s impulso"]):
                    merged.at[idx, "EquipePadrao"] = "Equipe Especial"
                    merged.at[idx, "Gestor"] = "Marco Antonio Neves Suzart"
                    merged.at[idx, "Supervisor"] = "Ricardo Cascao"
                else:
                    merged.at[idx, "EquipePadrao"] = "Consumo Zona Norte"
                    merged.at[idx, "Gestor"] = "Marcelo Neves"
                    merged.at[idx, "Supervisor"] = "Paulo Roberto Ferreira Chaves"
                merged.at[idx, "Localidade"] = merged.at[idx, "Equipe"]

    # Medicamento: Prosper Norte/Sul — EquipePadrao = Aba, Gestor por aba. Sem supervisor (não usar regras de Alimento).
    mask_med = merged["Aba"].fillna("").astype(str).str.strip().isin(ABAS_MEDICAMENTO)
    if mask_med.any():
        merged.loc[mask_med, "Segmento"] = "Medicamento"
        merged.loc[mask_med, "EquipePadrao"] = merged.loc[mask_med, "Aba"]
        merged.loc[mask_med, "Localidade"] = merged.loc[mask_med, "Equipe"]
        merged.loc[mask_med, "Supervisor"] = ""
        for aba, gestor in GESTORES_MEDICAMENTO.items():
            merged.loc[mask_med & (merged["Aba"] == aba), "Gestor"] = gestor

    # Promotores (aba) — Segmento Promotores
    mask_prom = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_PROMOTORES]
    )
    if mask_prom.any():
        merged.loc[mask_prom, "Segmento"] = "Promotores"
        merged.loc[mask_prom, "EquipePadrao"] = "Promotores"
        merged.loc[mask_prom, "Localidade"] = merged.loc[mask_prom, "Equipe"]

    # Internos (aba) — Segmento Internos
    mask_int = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_INTERNOS]
    )
    if mask_int.any():
        merged.loc[mask_int, "Segmento"] = "Internos"
        merged.loc[mask_int, "EquipePadrao"] = "Internos"
        mask_loc_vazio = mask_int & (merged["Localidade"].fillna("").astype(str).str.strip() == "")
        merged.loc[mask_loc_vazio, "Localidade"] = merged.loc[mask_loc_vazio, "Equipe"]

    # Manutenção (abas de troca e devolução)
    mask_manut = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_MANUTENCAO]
    )
    if mask_manut.any():
        merged.loc[mask_manut, "Segmento"] = "Manutenção"
        merged.loc[mask_manut, "EquipePadrao"] = "Manutenção"
        mask_loc_vazio = mask_manut & (merged["Localidade"].fillna("").astype(str).str.strip() == "")
        merged.loc[mask_loc_vazio, "Localidade"] = merged.loc[mask_loc_vazio, "Equipe"]

    # Roubo e Perda
    mask_roubo = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_ROUBO_PERDA]
    )
    if mask_roubo.any():
        merged.loc[mask_roubo, "Segmento"] = "Roubo e Perda"
        merged.loc[mask_roubo, "EquipePadrao"] = "Roubo e Perda"
        mask_loc_vazio = mask_roubo & (merged["Localidade"].fillna("").astype(str).str.strip() == "")
        merged.loc[mask_loc_vazio, "Localidade"] = merged.loc[mask_loc_vazio, "Equipe"]

    return merged


def apply_filters(
    df: pd.DataFrame,
    grupos: list[str],
    segmentos: list[str],
    teams: list[str],
    tipos: list[str],
    localidades: list[str],
    query: str,
) -> pd.DataFrame:
    result = df.copy()
    if grupos:
        result = result[result["GrupoEquipe"].isin(grupos)]
    if segmentos:
        result = result[result["Segmento"].isin(segmentos)]
    if teams:
        result = result[result["EquipePadrao"].isin(teams)]
    if tipos:
        result = result[result["TipoEquipe"].isin(tipos)]
    if localidades:
        result = result[result["Localidade"].isin(localidades)]

    query = query.strip().lower()
    if query:
        columns = [
            "Linha",
            "Codigo",
            "Nome",
            "Equipe",
            "GrupoEquipe",
            "Segmento",
            "EquipePadrao",
            "TipoEquipe",
            "Localidade",
            "Gestor",
            "Supervisor",
            "Aparelho",
            "Modelo",
            "Aba",
        ]
        mask = pd.Series(False, index=result.index)
        for col in columns:
            mask = mask | result[col].astype(str).str.lower().str.contains(query, na=False)
        result = result[mask]

    return result.sort_values(by=["Segmento", "GrupoEquipe", "EquipePadrao", "Nome", "Linha"], kind="stable")


def non_empty_or_default(value: Any, default: str) -> str:
    text = str(value).strip() if value is not None else ""
    return text if text else default


def normalize_name(value: Any) -> str:
    return normalize_text(value).replace(" ", "")


def classify_papel(row: pd.Series) -> str:
    nome = normalize_name(row.get("Nome", ""))
    gestor = normalize_name(row.get("Gestor", ""))
    supervisor = normalize_name(row.get("Supervisor", ""))
    if nome != "" and gestor != "" and nome == gestor:
        return "Gerente"
    if nome != "" and supervisor != "" and nome == supervisor:
        return "Supervisor"
    if "vago" in normalize_text(row.get("Nome", "")):
        return "Vago"
    return "Vendedor"


def build_full_table(df: pd.DataFrame) -> pd.DataFrame:
    table = df.copy()
    table["Papel"] = table.apply(classify_papel, axis=1)
    cols = [
        "Segmento",
        "Data da Troca",
        "Data Retorno",
        "Data Ocorrência",
        "Data Solicitação TBS",
        "GrupoEquipe",
        "EquipePadrao",
        "TipoEquipe",
        "Localidade",
        "Gestor",
        "Supervisor",
        "Papel",
        "Nome de Guerra",
        "Codigo",
        "Nome",
        "Linha",
        "E-mail",
        "Gerenciamento",
        "Motivo",
        "Observação",
        "Marca",
        "Aparelho",
        "Modelo",
        "CHIP",
        "IMEI A",
        "IMEI B",
        "Setor",
        "Cargo",
        "Desconto",
        "Perfil",
        "Empresa",
        "Ativo",
        "Numero de Serie",
        "Patrimonio",
        "Operadora",
        "Aba",
        "Equipe",
    ]
    existing = [c for c in cols if c in table.columns]
    return table[existing].sort_values(
        ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
        kind="stable",
    )


def _get_gerentes_medicamento(rules_path: Path) -> set[str]:
    """Retorna o conjunto de nomes normalizados dos gerentes do Medicamento."""
    out: set[str] = set()
    for nome in GESTORES_MEDICAMENTO.values():
        if nome:
            out.add(normalize_text(nome))
    med_path = rules_path.parent / "equipes_medicamento.csv"
    if med_path.exists():
        try:
            ref = pd.read_csv(med_path, dtype=str).fillna("")
            for _, r in ref.iterrows():
                g = str(r.get("gerente", "")).strip()
                if g:
                    out.add(normalize_text(g))
        except Exception:
            pass
    return out


def _get_gerentes_alimento(rules_path: Path) -> set[str]:
    """Retorna o conjunto de nomes normalizados dos gerentes do Alimento (equipes_alimento.csv)."""
    ali_path = rules_path.parent / "equipes_alimento.csv"
    out: set[str] = set()
    if not ali_path.exists():
        return out
    try:
        ref = pd.read_csv(ali_path, dtype=str).fillna("")
        for _, r in ref.iterrows():
            g = str(r.get("gerente", "")).strip()
            if g:
                out.add(normalize_text(g))
        if "marco antonio neves suzart" not in out:
            out.add("marco antonio neves suzart")
        if "marcelo neves" not in out:
            out.add("marcelo neves")
    except Exception:
        pass
    return out


def _get_supervisor_from_rules(equipe: str, rules_path: Path) -> str:
    """Obtém supervisor da equipe. Medicamento usa equipes_medicamento (geralmente vago)."""
    eq_key = normalize_team_key(equipe)
    if eq_key in ("prosper norte", "prosper sul"):
        med_path = rules_path.parent / "equipes_medicamento.csv"
        if med_path.exists():
            try:
                ref = pd.read_csv(med_path, dtype=str).fillna("")
                for _, r in ref.iterrows():
                    if normalize_team_key(str(r.get("equipe_real", ""))) == eq_key:
                        sup = str(r.get("supervisor", "")).strip()
                        return sup if sup else ""
            except Exception:
                pass
        return ""
    ali_path = rules_path.parent / "equipes_alimento.csv"
    if ali_path.exists():
        try:
            ref = pd.read_csv(ali_path, dtype=str).fillna("")
            fallback_sup = ""
            for _, r in ref.iterrows():
                er = str(r.get("equipe_real", "")).strip()
                loc = str(r.get("localidade", "")).strip()
                if normalize_team_key(er) != eq_key:
                    continue
                sup = str(r.get("supervisor", "")).strip()
                if not sup:
                    continue
                if not loc:
                    return sup
                fallback_sup = sup
            if fallback_sup:
                return fallback_sup
        except Exception:
            pass
    if rules_path.exists():
        try:
            rules = pd.read_csv(rules_path, dtype=str).fillna("")
            for _, r in rules.iterrows():
                if normalize_team_key(str(r.get("equipe_padrao", ""))) == eq_key:
                    sup = str(r.get("supervisor", "")).strip()
                    if sup:
                        return sup
        except Exception:
            pass
    return ""


def _supervisor_display(dfe: pd.DataFrame, eq: str, rules_path: Path) -> str:
    """Nome do supervisor para exibição no cabeçalho. Vago se vazio, senão nome. Fallback em regras."""
    sup_col = dfe["Supervisor"].fillna("").astype(str).str.strip()
    sup = ""
    for v in sup_col:
        if v and v not in ("—", "-", "vago", "Sem Supervisor"):
            sup = v
            break
    if sup:
        return sup
    if eq == "Promotores":
        for _, row in dfe.iterrows():
            loc = normalize_text(str(row.get("Localidade", "") or row.get("Equipe", "")).strip())
            if "supervisor" in loc:
                return non_empty_or_default(row.get("Nome", ""), "—")
    sup_from_rules = _get_supervisor_from_rules(eq, rules_path)
    if sup_from_rules:
        return sup_from_rules
    return "Vago"


def _codigo_supervisor(dfe: pd.DataFrame) -> str:
    """Obtem o codigo do supervisor (linha onde Nome == Supervisor)."""
    sup = dfe["Supervisor"].fillna("").astype(str).str.strip()
    nomes = dfe["Nome"].fillna("").astype(str).str.strip()
    cods = dfe["Codigo"].fillna("").astype(str).str.strip()
    for i, s in enumerate(sup):
        if s and s != "Sem Supervisor":
            nm = nomes.iloc[i]
            if nm and normalize_text(nm) == normalize_text(s):
                return cods.iloc[i] or "—"
    return "—"


def _celula_vago(v: Any) -> str:
    txt = str(v).strip() if v is not None else ""
    if normalize_text(txt) == "vago":
        return '<span style="color:#c00;font-weight:600">VAGO</span>'
    return _escape_html(txt) if txt else ""


def _escape_html(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _render_equipe_tabela(dfe: pd.DataFrame, eq: str) -> str:
    """Tabela com supervisor como 1ª linha e demais vendedores. Cabeçalho com rótulos claros."""
    cod_sup = _codigo_supervisor(dfe)
    nome_sup = non_empty_or_default(dfe["Supervisor"].iloc[0], "—")
    linha_contexto = str(((st.session_state.get("chamado_context") or {}).get("linha") or "")).strip()

    cols_full = [
        "Codigo", "Nome", "Equipe", "Linha", "E-mail", "Gerenciamento",
        "Aparelho", "Modelo", "CHIP", "IMEI A", "IMEI B",
        "Setor", "Cargo", "Desconto", "Perfil", "Empresa", "Ativo",
        "Numero de Serie", "Operadora",
    ]
    labels_full = [
        "Código", "Nome", "Localidade", "Linha", "E-mail", "Gerenciamento",
        "Aparelho", "Modelo", "CHIP", "IMEI A", "IMEI B",
        "Setor", "Cargo", "Desconto", "Perfil", "Empresa", "Ativo",
        "Nº Série", "Operadora",
    ]
    cols = [c for c in cols_full if c in dfe.columns]
    labels = [labels_full[cols_full.index(c)] for c in cols]

    sup_norm = normalize_text(nome_sup)
    row_supervisor = None
    rows_data = []

    def _is_supervisor_row(row: pd.Series) -> bool:
        if sup_norm and normalize_text(str(row.get("Nome", "")).strip()) == sup_norm:
            return True
        if eq == "Promotores":
            loc = normalize_text(str(row.get("Localidade", "") or row.get("Equipe", "")).strip())
            if "supervisor" in loc:
                return True
        return False

    for _, row in dfe.iterrows():
        if _is_supervisor_row(row):
            row_supervisor = row
            continue
        rows_data.append(row)

    html = ['<table class="tbl-equipe">']
    html.append("<thead><tr>")
    for L in labels:
        html.append(f"<th>{L}</th>")
    html.append("</tr></thead><tbody>")

    if row_supervisor is not None:
        display_nome_sup = nome_sup if (nome_sup and nome_sup != "—") else non_empty_or_default(row_supervisor.get("Nome", ""), "—")
        row_classes = ["row-supervisor"]
        linha_sup = str(row_supervisor.get("Linha", "") or "").strip()
        if linha_contexto and linha_sup == linha_contexto:
            row_classes.append("row-target")
        html.append(f'<tr class="{" ".join(row_classes)}">')
        for c in cols:
            v = row_supervisor.get(c, "")
            s = str(v).strip() if v is not None else ""
            if not s and c == "Codigo":
                s = cod_sup or non_empty_or_default(row_supervisor.get("Codigo", ""), "—")
            elif not s and c == "Nome":
                s = display_nome_sup
            elif not s and c == "Equipe":
                s = eq
            html.append(f'<td><strong>{_escape_html(s)}</strong></td>')
        html.append("</tr>")

    for row in rows_data:
        linha_row = str(row.get("Linha", "") or "").strip()
        row_class = ' class="row-target"' if linha_contexto and linha_row == linha_contexto else ""
        html.append(f"<tr{row_class}>")
        for c in cols:
            v = row.get(c, "")
            s = str(v).strip() if v is not None else ""
            if c in ("Codigo", "Nome") and normalize_text(s) == "vago":
                html.append(f'<td>{_celula_vago(s)}</td>')
            else:
                html.append(f'<td>{_escape_html(s)}</td>')
        html.append("</tr>")
    html.append("</tbody></table>")
    return '<div class="tbl-wrapper">' + "\n".join(html) + '</div>'


def _salvar_sessao_cookie(user: dict) -> None:
    """Cria sessão e salva token no cookie para login persistente."""
    token = secrets.token_hex(32)
    criar_sessao(token, user["username"])
    st.session_state.auth_token = token
    if HAS_COOKIES and _cookies.ready():
        _cookies["auth_token"] = token
        _cookies.save()


def _restaurar_sessao_cookie() -> bool:
    """Tenta restaurar sessão a partir do cookie. Retorna True se restaurou."""
    if not (HAS_COOKIES and HAS_DB and _cookies.ready()):
        return False
    token = _cookies.get("auth_token")
    if not token:
        return False
    user = validar_sessao(token)
    if user:
        st.session_state.authenticated = True
        st.session_state.user = user
        st.session_state.auth_token = token
        return True
    # Token inválido ou expirado
    try:
        del _cookies["auth_token"]
        _cookies.save()
    except Exception:
        pass
    return False


def _sync_usuarios_chamados(*, dry_run: bool = False) -> tuple[bool, str]:
    try:
        from scripts.sync_usuarios_chamados import executar_sync

        return executar_sync(dry_run=dry_run)
    except Exception as exc:
        return False, str(exc)


def _bind_chamados_api_token(username: str, password: str = "") -> None:
    """Obtém JWT da API Chamados após login Streamlit (Fase C1)."""
    try:
        from src.core.config import use_telefones_api
        from src.services.telefones_api_client import is_enabled, login_chamados_api, set_api_token

        if not (is_enabled() and use_telefones_api()):
            return
        pwd = (password or "").strip()
        if not pwd:
            return
        token = login_chamados_api(username, pwd)
        set_api_token(token)
        if token:
            st.session_state["chamados_api_token"] = token
    except Exception:
        pass


def _render_login_or_first_user() -> bool:
    """
    Renderiza login ou formulário do primeiro usuário.
    Retorna True se o usuário está autenticado (mostrar app principal).
    """
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if "user" not in st.session_state:
        st.session_state.user = None
    if "auth_token" not in st.session_state:
        st.session_state.auth_token = None

    if st.session_state.authenticated:
        return True

    if not HAS_DB:
        st.warning("Login requer banco de dados. Execute o sync e ative o banco em Config.")
        return True

    init_db()
    if not tem_usuarios():
        st.markdown("### Criar primeiro usuário (administrador)")
        st.caption("Nenhum usuário existe. Crie o administrador inicial.")
        with st.form("primeiro_usuario"):
            u = st.text_input("Usuário", key="fu_user")
            p = st.text_input("Senha", type="password", key="fu_pass")
            if st.form_submit_button("Criar"):
                if u.strip() and len(p) >= 4:
                    if criar_usuario(u.strip(), p, is_admin=True):
                        if is_postgres_configured():
                            _sync_usuarios_chamados()
                        _bind_chamados_api_token(u.strip(), p)
                        user = {"username": u.strip().lower(), "is_admin": True}
                        st.session_state.authenticated = True
                        st.session_state.user = user
                        _salvar_sessao_cookie(user)
                        st.success("Usuário criado! Redirecionando...")
                        st.rerun()
                    else:
                        st.error("Erro ao criar. Usuário já existe?")
                else:
                    st.error("Usuário e senha (mín. 4 caracteres) obrigatórios.")
        return False

    # Verificar sessão salva no cookie (login persistente)
    if _restaurar_sessao_cookie():
        st.rerun()

    st.markdown("### Login")
    with st.form("login"):
        u = st.text_input("Usuário", key="login_user")
        p = st.text_input("Senha", type="password", key="login_pass")
        if st.form_submit_button("Entrar"):
            user = verificar_login(u, p)
            if user:
                st.session_state.authenticated = True
                st.session_state.user = user
                _bind_chamados_api_token(u, p)
                _salvar_sessao_cookie(user)
                st.rerun()
            else:
                st.error("Usuário ou senha incorretos.")
    return False


def _current_user_is_admin() -> bool:
    return bool((st.session_state.get("user") or {}).get("is_admin"))


def _streamlit_access_mode(current_page: str, chamado_context: dict | None = None) -> str:
    ctx = chamado_context or st.session_state.get("chamado_context") or {}
    return resolve_streamlit_access(
        is_admin=_current_user_is_admin(),
        page=str(current_page or "painel"),
        chamado_id=str(ctx.get("chamado_id") or ""),
        linha=str(ctx.get("linha") or ""),
        return_url=str(ctx.get("return_url") or ""),
    )


def _render_operador_sem_contexto(chamados_app_url: str) -> None:
    st.markdown("### Gerenciamento de Telefones")
    st.info(
        "O painel completo de linhas e restrito a **administradores** (Fase C3). "
        "Operadores devem trabalhar no **Sistema de Chamados** (React)."
    )
    url = (chamados_app_url or "").strip() or "http://localhost:3000"
    st.link_button("Abrir Sistema de Chamados", url=url.rstrip("/"), use_container_width=False)
    st.caption(
        "Para editar uma linha a partir de um ticket, use o botao no Chamados que abre este app "
        "com contexto de chamado (integracao B2)."
    )
    st.caption("Documentacao: doc/UI_PAPEL_APPS_C3.md")


def _audit(
    acao: str,
    entidade: str,
    chave: str = "",
    chamado_id: str = "",
    antes: dict | None = None,
    depois: dict | None = None,
    detalhes: str = "",
) -> None:
    """Wrapper seguro de auditoria (não quebra fluxo principal)."""
    if not (HAS_DB and registrar_auditoria):
        return
    try:
        user = st.session_state.get("user", {}) or {}
        chamado_ref = str(chamado_id or st.session_state.get("chamado_id") or "").strip()
        registrar_auditoria(
            acao=acao,
            entidade=entidade,
            chave_registro=chave,
            chamado_id=chamado_ref,
            antes=antes,
            depois=depois,
            detalhes=detalhes,
            user_id=str(user.get("id", "")),
            username=str(user.get("username", "")),
            origem="app",
        )
        # Unificacao progressiva: também registra no `chamado_eventos` quando houver contexto de chamado.
        if registrar_chamado_evento and chamado_ref and str(chamado_ref).isdigit():
            try:
                registrar_chamado_evento(
                    chamado_id=chamado_ref,
                    tipo_evento=str(acao or "").strip(),
                    descricao=str(detalhes or "").strip(),
                    antes=antes,
                    depois=depois,
                    user_id=str(user.get("id", "")),
                )
            except Exception:
                pass
    except Exception:
        pass


def _set_post_chamado_banner(default_text: str = "Alteração concluída com sucesso.") -> None:
    """Prepara um banner de retorno ao chamado após ações relevantes."""
    chamado_context = st.session_state.get("chamado_context") or {}
    chamado_id = str(chamado_context.get("chamado_id") or st.session_state.get("chamado_id") or "").strip()
    return_url = str(chamado_context.get("return_url") or "").strip()
    if not chamado_id and not return_url:
        return
    st.session_state["_post_chamado_return_banner"] = {
        "text": default_text,
        "chamado_id": chamado_id,
        "return_url": return_url,
    }


def _auto_mark_context_line_vago(df_current: pd.DataFrame, modo_db: str, dados_do_banco: bool) -> pd.DataFrame:
    """Marca Nome/Codigo como VAGO ao abrir um chamado com linha em contexto."""
    chamado_context = st.session_state.get("chamado_context") or {}
    chamado_id = str(chamado_context.get("chamado_id") or st.session_state.get("chamado_id") or "").strip()
    linha_ctx = str(chamado_context.get("linha") or "").strip()
    auto_signature = f"{chamado_id}|{linha_ctx}|{modo_db}"
    if not chamado_id or not linha_ctx:
        return df_current
    if st.session_state.get("_auto_vago_signature") == auto_signature:
        return df_current
    if not (HAS_DB and dados_do_banco and save_linhas and load_linhas):
        st.session_state["_auto_vago_signature"] = auto_signature
        st.session_state["_post_conflict_feedback"] = {
            "type": "warning",
            "text": "Nao foi possivel marcar a linha como VAGO automaticamente porque o banco nao esta ativo.",
        }
        return df_current

    modos_tentativa = [modo_db] + [m for m in ["ativas", "desativadas"] if m != modo_db]
    for modo_target in modos_tentativa:
        if not has_data(modo_target):
            continue
        df_mode = load_linhas(modo=modo_target)
        if "Linha" not in df_mode.columns:
            continue
        mask_linha = df_mode["Linha"].fillna("").astype(str).str.strip() == linha_ctx
        if not mask_linha.any():
            continue
        if int(mask_linha.sum()) > 1:
            st.session_state["_auto_vago_signature"] = auto_signature
            st.session_state["_post_conflict_feedback"] = {
                "type": "warning",
                "text": f"A linha `{linha_ctx}` apareceu mais de uma vez e nao foi marcada como VAGO automaticamente.",
            }
            return df_current

        row_idx = df_mode[mask_linha].index[0]
        before_row = df_mode.loc[row_idx].to_dict()
        nome_norm = normalize_text(before_row.get("Nome", ""))
        codigo_norm = normalize_text(before_row.get("Codigo", ""))
        if nome_norm == "vago" and codigo_norm == "vago":
            st.session_state["_auto_vago_signature"] = auto_signature
            return df_mode if modo_target == modo_db else df_current

        df_mode.loc[row_idx, "Nome"] = "VAGO"
        if "Codigo" in df_mode.columns:
            df_mode.loc[row_idx, "Codigo"] = "VAGO"
        save_linhas(df_mode, modo=modo_target)
        _audit(
            acao="marcar_linha_vaga_chamado",
            entidade="linhas",
            chave=linha_ctx,
            chamado_id=chamado_id,
            antes=before_row,
            depois={"Nome": "VAGO", "Codigo": "VAGO", "modo": modo_target},
            detalhes="Nome e codigo marcados como VAGO automaticamente ao abrir chamado.",
        )
        st.session_state["_auto_vago_signature"] = auto_signature
        st.session_state["_post_conflict_feedback"] = {
            "type": "success",
            "text": f"Linha `{linha_ctx}` marcada como VAGO automaticamente para o chamado `{chamado_id}`.",
        }
        return df_mode if modo_target == modo_db else df_current

    st.session_state["_auto_vago_signature"] = auto_signature
    st.session_state["_post_conflict_feedback"] = {
        "type": "warning",
        "text": f"A linha `{linha_ctx}` nao foi encontrada para marcar como VAGO automaticamente.",
    }
    return df_current


def main() -> None:
    st.set_page_config(
        page_title="Gerenciamento de Telefones",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    _init_cookies()

    if not _render_login_or_first_user():
        return

    cwd = Path.cwd()
    rules_path = DOC_DIR / RULES_FILE if (DOC_DIR / RULES_FILE).exists() else cwd / RULES_FILE

    st.markdown(
        """
        <style>
        #MainMenu, header, footer { visibility: hidden; }
        header[data-testid="stHeader"] { display: none; }
        div[data-testid="stToolbar"] { display: none !important; }
        section[data-testid="stSidebar"] > div:first-child { display: none; }
        .block-container { padding-top: 0 !important; padding-bottom: 1rem !important; }
        [data-testid="stHorizontalBlock"]:first-of-type {
            background: linear-gradient(90deg, #f3c300 0%, #ffcf22 100%);
            padding: 10px 20px;
            border-radius: 8px;
            margin-bottom: 1rem;
            flex-wrap: nowrap !important;
            gap: 30px !important;
            border: 1px solid #b78900;
        }
        [data-testid="stHorizontalBlock"]:first-of-type label,
        [data-testid="stHorizontalBlock"]:first-of-type p,
        [data-testid="stHorizontalBlock"]:first-of-type span,
        [data-testid="stHorizontalBlock"]:first-of-type strong,
        [data-testid="stHorizontalBlock"]:first-of-type div[data-testid="stMarkdown"],
        [data-testid="stHorizontalBlock"]:first-of-type button {
            color: #171717 !important;
            opacity: 1 !important;
        }
        [data-testid="stHorizontalBlock"]:first-of-type > div {
            overflow: visible !important;
        }
        /* Segmento em duplas (2 colunas) na navbar */
        [data-testid="stHorizontalBlock"]:first-of-type > div:nth-child(3) [role="radiogroup"] {
            display: grid !important;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            column-gap: 14px;
            row-gap: 2px;
        }
        [data-testid="stHorizontalBlock"]:first-of-type > div:nth-child(3) [role="radiogroup"] label {
            white-space: nowrap !important;
        }
        [data-testid="stHorizontalBlock"]:first-of-type > div:nth-child(3) [role="radiogroup"] p {
            white-space: nowrap !important;
            word-break: keep-all !important;
        }
        [data-testid="stHorizontalBlock"]:first-of-type button {
            background-color: transparent !important;
            border-color: rgba(23,23,23,0.75) !important;
            white-space: nowrap !important;
            min-width: fit-content !important;
        }
        .main-title { font-size: 1.8rem; font-weight: 600; color: #1f1f1f !important; margin-bottom: 0; opacity: 1 !important; }
        .main-subtitle { font-size: 0.95rem; color: #5a6c7d !important; margin-bottom: 1rem; opacity: 1 !important; }
        .team-header {
            background: linear-gradient(90deg, #2d5a24 0%, #3d7a34 100%);
            color: #ffffff !important;
            padding: 12px 16px;
            border-radius: 8px;
            margin: 24px 0 4px 0;
            font-weight: 600;
            font-size: 1.1rem;
            opacity: 1 !important;
        }
        .team-header-inactive {
            background: linear-gradient(90deg, #8b1e1e 0%, #b83434 100%);
            color: #ffffff !important;
            padding: 12px 16px;
            border-radius: 8px;
            margin: 24px 0 4px 0;
            font-weight: 600;
            font-size: 1.1rem;
            opacity: 1 !important;
        }
        .tbl-equipe {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 24px;
            font-size: 0.9rem;
            box-shadow: 0 1px 3px rgba(0,0,0,0.08);
            border-radius: 8px;
            overflow: hidden;
        }
        .tbl-equipe thead tr { background: #191919; color: #ffd64a !important; font-weight: 600; opacity: 1 !important; }
        .tbl-equipe thead th { color: #ffffff !important; opacity: 1 !important; }
        .tbl-equipe th, .tbl-equipe td {
            padding: 10px 12px;
            text-align: left;
            border-bottom: 1px solid #e8ecf0;
            white-space: nowrap;
            opacity: 1 !important;
        }
        .tbl-equipe tbody tr:nth-child(even) { background: #f8faf8; }
        .tbl-equipe tbody tr:hover { background: #eef5ee; }
        .tbl-equipe tr.row-supervisor { background: #e8f0e8 !important; }
        .tbl-equipe tr.row-target {
            background: #fff3bf !important;
            box-shadow: inset 4px 0 0 #d97706;
        }
        .tbl-equipe tr.row-target td {
            font-weight: 600;
        }
        .tbl-wrapper { overflow-x: auto; margin-bottom: 24px; }
        .navbar-top {
            background: linear-gradient(90deg, #1e3a5f 0%, #2d5a7b 100%);
            color: white;
            padding: 14px 20px;
            border-radius: 8px;
            margin-bottom: 20px;
            display: flex;
            align-items: center;
            gap: 16px;
            flex-wrap: wrap;
        }
        .navbar-top .brand { font-weight: 700; font-size: 1.15rem; white-space: nowrap; }
        .navbar-top span { white-space: nowrap; }
        div[data-testid="stSidebar"], section[data-testid="stSidebar"] { display: none; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # Navbar no topo da página
    col_brand, col_modo, col_seg, col_cfg, col_chamados, col_logout = st.columns([1.55, 1.35, 2.25, 1.0, 0.95, 0.9])
    query_params = st.query_params

    def _qp_value(*keys: str) -> str:
        for qp_key in keys:
            raw_qp = query_params.get(qp_key)
            if isinstance(raw_qp, list):
                raw_qp = raw_qp[0] if raw_qp else ""
            raw_qp = str(raw_qp or "").strip()
            if raw_qp:
                return raw_qp
        return ""

    chamado_qp = _qp_value("ticket_id", "chamado_id", "id_chamado", "chamado")
    linha_ctx_qp = _qp_value("linha", "linha_telefone", "numero_linha", "phone_line")
    segmento_ctx_explicit = _qp_value("segmento_chamado", "ctx_segmento")
    equipe_ctx_explicit = _qp_value("equipe_chamado", "ctx_equipe")
    busca_ctx_explicit = _qp_value("busca_chamado", "ctx_busca")
    return_url_qp = _qp_value("return_url", "return_to", "back_url", "chamados_url")
    contexto_externo_base = bool(chamado_qp or linha_ctx_qp or segmento_ctx_explicit or equipe_ctx_explicit or busca_ctx_explicit or return_url_qp)
    segmento_ctx_qp = segmento_ctx_explicit or (_qp_value("segmento") if contexto_externo_base else "")
    equipe_ctx_qp = equipe_ctx_explicit or (_qp_value("equipe") if contexto_externo_base else "")
    busca_ctx_qp = busca_ctx_explicit or (_qp_value("busca") if contexto_externo_base else "")

    if chamado_qp:
        st.session_state["chamado_id"] = chamado_qp
    elif "chamado_id" not in st.session_state:
        st.session_state["chamado_id"] = ""

    chamado_context = {
        "chamado_id": chamado_qp,
        "linha": linha_ctx_qp,
        "segmento": segmento_ctx_qp,
        "equipe": equipe_ctx_qp,
        "busca": busca_ctx_qp,
        "return_url": return_url_qp,
        "chamado_source": "",
        "chamado_label": "",
        "chamado_numero": "",
        "chamado_aviso": "",
        "chamado_legado": False,
    }
    if chamado_qp and resolver_referencia_chamado:
        chamado_ref = resolver_referencia_chamado(chamado_qp)
        if chamado_ref.get("valido") and chamado_ref.get("resolved_id") is not None:
            resolved_id = str(chamado_ref["resolved_id"])
            st.session_state["chamado_id"] = resolved_id
            chamado_context["chamado_id"] = resolved_id
        chamado_context["chamado_source"] = str(chamado_ref.get("source") or "")
        chamado_context["chamado_label"] = str(chamado_ref.get("label") or "")
        chamado_context["chamado_numero"] = str(chamado_ref.get("numero") or "")
        chamado_context["chamado_aviso"] = str(chamado_ref.get("aviso") or "")
        chamado_context["chamado_legado"] = bool(chamado_ref.get("legado"))
    has_chamado_context = any(bool(v) for v in chamado_context.values())
    ctx_signature = "|".join(chamado_context.get(k, "") for k in ["chamado_id", "linha", "segmento", "equipe", "busca", "return_url"])
    segmentos_validos_ctx = ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"]
    if has_chamado_context and st.session_state.get("_chamado_context_signature") != ctx_signature:
        st.session_state["chamado_context"] = chamado_context
        st.session_state["_chamado_context_signature"] = ctx_signature
        if chamado_context["segmento"] in segmentos_validos_ctx:
            st.session_state["nav_segmento"] = chamado_context["segmento"]
        if chamado_context["equipe"]:
            st.session_state["filtro_equipe"] = chamado_context["equipe"]
        if chamado_context["linha"]:
            st.session_state["filtro_busca"] = chamado_context["linha"]
        elif chamado_context["busca"]:
            st.session_state["filtro_busca"] = chamado_context["busca"]
    elif not has_chamado_context:
        st.session_state["chamado_context"] = {}
        st.session_state["_chamado_context_signature"] = ""
    if "_qp_hydrated" not in st.session_state:
        modo_qp = query_params.get("modo")
        seg_qp = query_params.get("segmento")
        equipe_qp = query_params.get("equipe")
        busca_qp = query_params.get("busca")
        busca_tipo_qp = query_params.get("busca_tipo")
        somente_vagas_qp = query_params.get("somente_vagas")
        if modo_qp in ["Linhas ativas", "Linhas desativadas"]:
            st.session_state["nav_modo"] = modo_qp
        if seg_qp in ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"]:
            st.session_state["nav_segmento"] = seg_qp
        if equipe_qp:
            st.session_state["filtro_equipe"] = equipe_qp
        if busca_qp is not None:
            st.session_state["filtro_busca"] = busca_qp
        if busca_tipo_qp in ["Geral", "Linha", "IMEI", "Aparelho", "Motivo"]:
            st.session_state["filtro_busca_tipo"] = busca_tipo_qp
        if somente_vagas_qp in ["0", "1"]:
            st.session_state["filtro_somente_vagas"] = somente_vagas_qp == "1"
        st.session_state["_qp_hydrated"] = True
    pagina_qp = query_params.get("pagina", "painel")
    current_page = pagina_qp if pagina_qp in ["painel", "config"] else "painel"
    chamados_app_url = get_chamados_app_url()
    with col_brand:
        logo_path = Path("assets/logo.png")
        if logo_path.exists():
            logo_b64 = base64.b64encode(logo_path.read_bytes()).decode("utf-8")
            st.markdown(
                f"""
                <div style="display:flex; flex-direction:column; align-items:center; justify-content:center; text-align:center; margin-top:2px;">
                    <img src="data:image/png;base64,{logo_b64}" style="width:64px; height:auto; object-fit:contain; margin-bottom:4px;" />
                    <p style="font-weight:700; color:#1a1a1a; margin:0; white-space:nowrap;">Gerenciamento de Telefones</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        else:
            st.markdown('<p style="font-weight:700; color:#1a1a1a; margin:0; white-space:nowrap;">Gerenciamento de Telefones</p>', unsafe_allow_html=True)
    with col_modo:
        modos_opts = ["Linhas ativas", "Linhas desativadas"]
        if st.session_state.get("nav_modo") not in modos_opts:
            st.session_state["nav_modo"] = modos_opts[0]
        modo_sel = st.radio("Modo", options=modos_opts, horizontal=True, key="nav_modo")
    with col_seg:
        segmentos_opts = ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"]
        segmento_pref = st.session_state.get("pending_segmento") or st.session_state.get("nav_segmento")
        # Se existir segmento pendente (linha recém-criada), aplica antes de instanciar o widget.
        if segmento_pref in segmentos_opts and st.session_state.get("nav_segmento") != segmento_pref:
            st.session_state["nav_segmento"] = segmento_pref
        if st.session_state.get("nav_segmento") not in segmentos_opts:
            st.session_state["nav_segmento"] = segmentos_opts[0]
        segmento_sel = st.radio("Segmento", options=segmentos_opts, horizontal=True, key="nav_segmento")
    modo_db = "ativas" if modo_sel == "Linhas ativas" else "desativadas"
    segmento_sem_filtro_modo = segmento_sel in ("Manutenção", "Roubo e Perda")
    st.query_params["modo"] = modo_sel
    st.query_params["segmento"] = segmento_sel
    def _render_config_content() -> None:
        if st.session_state.get("user", {}).get("is_admin"):
            with st.expander("Gerenciar usuários", expanded=False):
                nu = st.text_input("Novo usuário", key="new_user")
                np = st.text_input("Senha", type="password", key="new_pass")
                is_adm = st.checkbox("Administrador", key="new_admin")
                if st.button("Criar usuário", key="btn_new_user"):
                    if nu.strip() and len(np) >= 4:
                        if criar_usuario(nu.strip(), np, is_admin=is_adm):
                            if is_postgres_configured():
                                sync_ok, sync_msg = _sync_usuarios_chamados()
                                if sync_ok:
                                    st.success("Usuário criado e sincronizado com Chamados.")
                                else:
                                    st.success("Usuário criado no Gerenciamento.")
                                    st.warning(f"Sincronização com Chamados falhou: {sync_msg}")
                            else:
                                st.success("Usuário criado!")
                            st.rerun()
                        else:
                            st.error("Usuário já existe.")
                    else:
                        st.error("Usuário e senha (mín. 4 caracteres) obrigatórios.")
                users = listar_usuarios()
                if users:
                    st.caption("Usuários cadastrados")

                    modal_target = str(st.session_state.get("_pwd_modal_user", "") or "").strip().lower()
                    logged_user = str((st.session_state.get("user", {}) or {}).get("username", "")).strip().lower()

                    if modal_target and hasattr(st, "dialog"):
                        @st.dialog("Alterar senha de usuário")
                        def _show_password_modal() -> None:
                            st.markdown(f"**Usuário:** `{modal_target}`")
                            senha_atual_modal = ""
                            if modal_target == logged_user:
                                senha_atual_modal = st.text_input(
                                    "Senha atual",
                                    type="password",
                                    key=f"pwd_current_{modal_target}",
                                )
                            senha_nova_modal = st.text_input(
                                "Nova senha",
                                type="password",
                                key=f"pwd_new_{modal_target}",
                            )
                            senha_conf_modal = st.text_input(
                                "Confirmar nova senha",
                                type="password",
                                key=f"pwd_confirm_{modal_target}",
                            )
                            c_modal1, c_modal2 = st.columns(2)
                            with c_modal1:
                                salvar_modal = st.button("Salvar senha", key=f"pwd_save_{modal_target}")
                            with c_modal2:
                                cancelar_modal = st.button("Cancelar", key=f"pwd_cancel_{modal_target}")

                            if cancelar_modal:
                                st.session_state["_pwd_modal_user"] = ""
                                st.rerun()

                            if not salvar_modal:
                                return

                            if len(senha_nova_modal) < 4:
                                st.error("A nova senha deve ter no mínimo 4 caracteres.")
                                return
                            if senha_nova_modal != senha_conf_modal:
                                st.error("A confirmação da nova senha não confere.")
                                return
                            if modal_target == logged_user and not verificar_login(logged_user, senha_atual_modal):
                                st.error("Senha atual inválida.")
                                return

                            if atualizar_senha_usuario and atualizar_senha_usuario(modal_target, senha_nova_modal):
                                detalhe_audit = (
                                    f"Senha alterada pelo próprio usuário: {modal_target}"
                                    if modal_target == logged_user
                                    else f"Senha alterada por administrador para usuário: {modal_target}"
                                )
                                _audit(
                                    acao="alterar_senha_usuario",
                                    entidade="usuarios",
                                    chave=modal_target,
                                    detalhes=detalhe_audit,
                                )
                                st.session_state["_pwd_modal_user"] = ""
                                st.success("Senha atualizada com sucesso.")
                                st.rerun()
                            else:
                                st.error("Não foi possível atualizar a senha do usuário.")

                        _show_password_modal()

                    elif modal_target and not hasattr(st, "dialog"):
                        st.warning("Sua versão do Streamlit não suporta modal (`st.dialog`).")
                        if st.button("Fechar aviso", key="btn_close_pwd_warning"):
                            st.session_state["_pwd_modal_user"] = ""
                            st.rerun()

                    for usr in users:
                        adm = " (admin)" if usr["is_admin"] else ""
                        c1, c2, c3 = st.columns([2, 1, 1])
                        with c1:
                            st.text(f"{usr['username']}{adm}")
                        with c2:
                            if st.button("🔒 Alterar senha", key=f"pwd_{usr['username']}", help=f"Alterar senha de {usr['username']}"):
                                st.session_state["_pwd_modal_user"] = str(usr["username"]).strip().lower()
                                st.rerun()
                        with c3:
                            if st.button("🗑️ Excluir", key=f"del_{usr['username']}", help=f"Remover {usr['username']}"):
                                excluir_usuario(usr["username"])
                                if usr["username"] == st.session_state.get("user", {}).get("username"):
                                    token = st.session_state.get("auth_token")
                                    if token:
                                        encerrar_sessao(token)
                                    if HAS_COOKIES and _cookies.ready():
                                        try:
                                            del _cookies["auth_token"]
                                            _cookies.save()
                                        except Exception:
                                            pass
                                    st.session_state.authenticated = False
                                    st.session_state.user = None
                                    st.session_state.auth_token = None
                                st.rerun()
            st.divider()
            if is_postgres_configured() and listar_usuarios_com_status_chamados:
                with st.expander("Sincronização com Chamados (users)", expanded=False):
                    st.caption(
                        "Espelha `usuarios_app` na tabela `users` do React. "
                        "Necessário para SSO e FKs de chamados. "
                        "Senhas: ver doc/POLITICA_SENHAS.md."
                    )
                    status_rows = listar_usuarios_com_status_chamados()
                    pendentes = [u for u in status_rows if u.get("ativo", True) and not u.get("tem_chamados")]
                    if pendentes:
                        st.warning(
                            f"{len(pendentes)} usuário(s) no Gerenciamento **sem** espelho em Chamados: "
                            + ", ".join(f"`{u['username']}`" for u in pendentes[:8])
                            + ("…" if len(pendentes) > 8 else "")
                        )
                    elif status_rows:
                        st.success("Todos os usuários ativos estão espelhados em Chamados.")
                    if status_rows:
                        st.dataframe(
                            [
                                {
                                    "Usuário": u["username"],
                                    "Admin": "Sim" if u.get("is_admin") else "Não",
                                    "Chamados (users)": "OK" if u.get("tem_chamados") else "Pendente",
                                }
                                for u in status_rows
                            ],
                            hide_index=True,
                            use_container_width=True,
                        )
                    c_sync1, c_sync2 = st.columns(2)
                    with c_sync1:
                        if st.button("Sincronizar agora", key="btn_sync_chamados_users"):
                            sync_ok, sync_msg = _sync_usuarios_chamados()
                            if sync_ok:
                                st.success(sync_msg)
                            else:
                                st.error(sync_msg)
                            st.rerun()
                    with c_sync2:
                        if st.button("Simular sync (dry-run)", key="btn_sync_chamados_dry"):
                            sync_ok, sync_msg = _sync_usuarios_chamados(dry_run=True)
                            if sync_ok:
                                st.info(sync_msg)
                            else:
                                st.error(sync_msg)
            st.divider()
        st.markdown("**Configurações gerais**")
        if HAS_DB:
            if is_postgres_configured():
                st.success("PostgreSQL: conexão ok.")
            else:
                st.warning("PostgreSQL: não configurado.")
            try:
                from src.core.config import use_telefones_api
                from src.services.telefones_api_client import get_api_token, is_enabled

                if use_telefones_api() and is_enabled():
                    if get_api_token():
                        st.info(
                            "Modo C1: leitura e gravação de linhas via API Chamados (token ativo). "
                            "Auditoria local do Streamlit continua ativa."
                        )
                    else:
                        st.warning(
                            "USE_TELEFONES_API=true, mas sem token JWT — grid usa SQL direto "
                            "(faça login com senha ou desative a flag)."
                        )
            except Exception:
                pass
        if st.session_state.get("user", {}).get("is_admin"):
            st.divider()
            with st.expander("Histórico de alterações (auditoria)", expanded=False):
                chamado_context_audit = str(((st.session_state.get("chamado_context") or {}).get("chamado_id") or "")).strip()
                c_audit_1, c_audit_2 = st.columns([2.2, 1])
                with c_audit_1:
                    audit_chamado = st.text_input(
                        "Filtrar por chamado",
                        value=st.session_state.get("audit_chamado_filter", ""),
                        placeholder="Ex.: 123",
                        key="audit_chamado_filter",
                    ).strip()
                with c_audit_2:
                    st.caption(" ")
                    if chamado_context_audit and st.button("Usar chamado atual", key="btn_use_current_chamado_filter"):
                        st.session_state["audit_chamado_filter"] = chamado_context_audit
                        st.rerun()
                audit_limit = st.number_input(
                    "Quantidade de registros",
                    min_value=20,
                    max_value=2000,
                    value=200,
                    step=20,
                    key="audit_limit",
                )
                logs = listar_auditoria(
                    limit=int(audit_limit),
                    chamado_id=audit_chamado,
                ) if listar_auditoria else []
                if not logs:
                    st.caption("Sem registros de auditoria.")
                else:
                    def _parse_json_safe(value: Any) -> Any:
                        if value is None:
                            return None
                        if isinstance(value, float) and pd.isna(value):
                            return None
                        if isinstance(value, (dict, list, int, float, bool)):
                            return value
                        txt = str(value).strip()
                        if not txt or txt.lower() in {"none", "nan"}:
                            return None
                        try:
                            return json.loads(txt)
                        except Exception:
                            return txt

                    def _friendly_action_name(action: str) -> str:
                        action_map = {
                            "salvar_edicoes": "Salvar edições",
                            "criar_linha": "Criar linha",
                            "excluir_linha": "Excluir linha",
                            "mover_equipe": "Mover equipe",
                            "mover_setor": "Mover setor",
                            "mudar_modo_linha": "Ativar/Desativar linha",
                            "enviar_manutencao": "Enviar para manutenção",
                            "sincronizar_banco": "Sincronizar banco",
                            "login": "Login",
                            "logout": "Logout",
                            "onboarding_linha": "Onboarding de linha",
                            "manutencao_aparelho": "Manutenção de aparelho",
                            "roubo_perda_linha": "Roubo/perda de linha",
                            "transferencia_equipe": "Transferência de equipe",
                            "desligamento_linha": "Desligamento de linha",
                        }
                        action_norm = str(action or "").strip()
                        return action_map.get(action_norm, action_norm.replace("_", " ").title())

                    def _fmt_audit_val(v: Any) -> str:
                        txt = "—" if v is None else str(v).strip()
                        if not txt:
                            txt = "—"
                        return txt if len(txt) <= 36 else (txt[:33] + "...")

                    def _render_alteracoes_summary(
                        antes_ev: Any,
                        depois_ev: Any,
                        *,
                        fallback_segmento: str = "—",
                        fallback_modo: str = "—",
                    ) -> str | None:
                        alteracoes: list[dict[str, Any]] = []
                        total_alteracoes = 0
                        if isinstance(antes_ev, dict):
                            raw_alter = antes_ev.get("alteracoes") or []
                            if isinstance(raw_alter, list):
                                alteracoes = raw_alter
                            total_alteracoes = int(antes_ev.get("alteracoes_total") or 0)
                        if not alteracoes and isinstance(depois_ev, dict):
                            raw_alter = depois_ev.get("alteracoes") or []
                            if isinstance(raw_alter, list):
                                alteracoes = raw_alter
                            total_alteracoes = int(depois_ev.get("alteracoes_total") or total_alteracoes or 0)

                        if alteracoes:
                            partes: list[str] = []
                            for ch in alteracoes[:2]:
                                if not isinstance(ch, dict):
                                    continue
                                ln = str(ch.get("linha") or "—").strip() or "—"
                                campo = str(ch.get("campo") or "—").strip() or "—"
                                antes_v = _fmt_audit_val(ch.get("antes"))
                                depois_v = _fmt_audit_val(ch.get("depois"))
                                partes.append(f"Linha {ln} | {campo}: {antes_v} -> {depois_v}")
                            if partes:
                                total_base = total_alteracoes if total_alteracoes > 0 else len(alteracoes)
                                extra = max(0, total_base - len(partes))
                                sufixo = f" (+{extra} alteração(ões))" if extra > 0 else ""
                                return " ; ".join(partes) + sufixo

                        segmento = fallback_segmento
                        modo_ev = fallback_modo
                        if isinstance(depois_ev, dict):
                            segmento = str(depois_ev.get("segmento", "") or "").strip() or segmento
                            modo_ev = str(depois_ev.get("modo", "") or "").strip() or modo_ev
                        linhas_editadas: list[str] = []
                        campos_alterados: list[str] = []
                        if isinstance(antes_ev, dict):
                            linhas_editadas = antes_ev.get("linhas_editadas") or []
                            campos_alterados = antes_ev.get("campos_alterados") or []
                        if not campos_alterados and isinstance(depois_ev, dict):
                            campos_alterados = depois_ev.get("campos_alterados") or []
                        qtd_linhas = len(linhas_editadas) if isinstance(linhas_editadas, list) else 0
                        campos_validos = [str(c).strip() for c in campos_alterados if str(c).strip()]
                        if not campos_validos and qtd_linhas == 0:
                            return None
                        campos_preview = ", ".join(campos_validos[:3])
                        if len(campos_validos) > 3:
                            campos_preview += f" +{len(campos_validos) - 3}"
                        if qtd_linhas == 1 and isinstance(linhas_editadas, list) and linhas_editadas:
                            alvo = f"linha {str(linhas_editadas[0]).strip()}"
                        else:
                            alvo = f"{qtd_linhas} linha(s)"
                        if campos_preview:
                            return f"Alterou {alvo} em {segmento} ({modo_ev}) | Campos: {campos_preview}"
                        return f"Alterou {alvo} em {segmento} ({modo_ev})"

                    def _build_what_edited(row: pd.Series) -> str:
                        acao_ev = str(row.get("acao", "") or "").strip()
                        antes_ev = _parse_json_safe(row.get("antes_json"))
                        depois_ev = _parse_json_safe(row.get("depois_json"))
                        detalhes_ev = str(row.get("detalhes", "") or "").strip()
                        chave_ev = str(row.get("chave_registro", "") or row.get("entidade_id", "") or "").strip()

                        alteracoes_resumo = _render_alteracoes_summary(antes_ev, depois_ev)
                        if alteracoes_resumo and acao_ev in {
                            "salvar_edicoes",
                            "onboarding_linha",
                            "manutencao_aparelho",
                            "roubo_perda_linha",
                            "transferencia_equipe",
                            "desligamento_linha",
                        }:
                            return alteracoes_resumo

                        if acao_ev == "salvar_edicoes" and isinstance(depois_ev, dict):
                            resumo = _render_alteracoes_summary(
                                antes_ev,
                                depois_ev,
                                fallback_segmento=str(depois_ev.get("segmento", "") or "").strip() or "—",
                                fallback_modo=str(depois_ev.get("modo", "") or "").strip() or "—",
                            )
                            if resumo:
                                return resumo
                        if acao_ev == "criar_linha":
                            return f"Criou a linha {chave_ev or '—'}"
                        if acao_ev == "excluir_linha":
                            return f"Excluiu a linha {chave_ev or '—'}"
                        if acao_ev == "mover_equipe":
                            eq_dest = ""
                            if isinstance(depois_ev, dict):
                                eq_dest = str(depois_ev.get('EquipePadrao') or depois_ev.get('Equipe') or "").strip()
                            return f"Moveu a linha {chave_ev or '—'} para equipe {eq_dest or '—'}"
                        if acao_ev == "mover_setor":
                            seg_dest = ""
                            if isinstance(depois_ev, dict):
                                seg_dest = str(depois_ev.get('Segmento') or "").strip()
                            return f"Moveu a linha {chave_ev or '—'} para o setor {seg_dest or '—'}"
                        if acao_ev == "mudar_modo_linha":
                            modo_dest = ""
                            if isinstance(depois_ev, dict):
                                modo_dest = str(depois_ev.get('modo') or "").strip()
                            return f"Alterou a linha {chave_ev or '—'} para {modo_dest or '—'}"
                        if acao_ev == "enviar_manutencao":
                            return f"Enviou para manutenção: {chave_ev or '—'}"
                        if acao_ev == "onboarding_linha":
                            nome_dest = ""
                            if isinstance(depois_ev, dict):
                                nome_dest = str(depois_ev.get("nome") or depois_ev.get("nome_usuario_snapshot") or "").strip()
                            return f"Atribuiu linha {chave_ev or '—'}" + (f" a {nome_dest}" if nome_dest else "")
                        if acao_ev == "manutencao_aparelho":
                            return f"Atualizou aparelho da linha {chave_ev or '—'}"
                        if acao_ev == "roubo_perda_linha":
                            return f"Registrou roubo/perda na linha {chave_ev or '—'}"
                        if acao_ev == "transferencia_equipe":
                            eq_dest = ""
                            if isinstance(depois_ev, dict):
                                eq_dest = str(depois_ev.get("equipe") or depois_ev.get("equipe_padrao") or "").strip()
                            return f"Transferiu linha {chave_ev or '—'}" + (f" para equipe {eq_dest}" if eq_dest else "")
                        if acao_ev == "desligamento_linha":
                            return f"Desligamento — linha {chave_ev or '—'}"
                        if acao_ev in ("sincronizar_banco", "login", "logout"):
                            return f"{_friendly_action_name(acao_ev)}"

                        if detalhes_ev:
                            return detalhes_ev
                        return _friendly_action_name(acao_ev) or "Edição"

                    def _extract_chamado_id(row: pd.Series) -> str:
                        raw = row.get("chamado_id", "")
                        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                            return "—"
                        txt = str(raw).strip()
                        if not txt or txt.lower() == "nan":
                            return "—"
                        return txt

                    def _to_local_datetime(series: pd.Series) -> pd.Series:
                        # `criado_em` no SQLite vem em UTC; converter para horário local melhora a leitura.
                        dt_utc = pd.to_datetime(series, errors="coerce", utc=True)
                        tz_name = os.environ.get("APP_TIMEZONE", "America/Sao_Paulo")
                        try:
                            return dt_utc.dt.tz_convert(tz_name)
                        except Exception:
                            return dt_utc

                    df_logs = pd.DataFrame(logs)
                    acoes_edicao = {
                        "salvar_edicoes",
                        "criar_linha",
                        "excluir_linha",
                        "mover_equipe",
                        "mover_setor",
                        "mudar_modo_linha",
                        "enviar_manutencao",
                        "onboarding_linha",
                        "manutencao_aparelho",
                        "roubo_perda_linha",
                        "transferencia_equipe",
                        "desligamento_linha",
                    }
                    if "acao" in df_logs.columns:
                        df_logs = df_logs[df_logs["acao"].isin(acoes_edicao)]
                    if "criado_em" in df_logs.columns:
                        df_logs["__ord_dt"] = _to_local_datetime(df_logs["criado_em"])
                        sort_cols = ["__ord_dt"] + (["id"] if "id" in df_logs.columns else [])
                        sort_asc = [False] * len(sort_cols)
                        df_logs = df_logs.sort_values(sort_cols, ascending=sort_asc, kind="stable")
                    elif "id" in df_logs.columns:
                        df_logs = df_logs.sort_values(["id"], ascending=[False], kind="stable")
                    df_logs = df_logs.reset_index(drop=True)
                    if df_logs.empty:
                        st.caption("Sem alterações de edição registradas.")
                    else:
                        df_view = pd.DataFrame()
                        if "username" in df_logs.columns:
                            df_view["Quem editou"] = df_logs["username"].fillna("").astype(str).str.strip().replace("", "—")
                        else:
                            df_view["Quem editou"] = "—"
                        df_view["Chamado"] = df_logs.apply(_extract_chamado_id, axis=1)
                        df_view["O que editou"] = df_logs.apply(_build_what_edited, axis=1)
                        dt_fmt = df_logs["__ord_dt"] if "__ord_dt" in df_logs.columns else _to_local_datetime(df_logs.get("criado_em"))
                        df_view["Quando editou"] = dt_fmt.dt.strftime("%d/%m/%Y %H:%M:%S").fillna("—")
                        st.dataframe(df_view[["Quem editou", "Chamado", "O que editou", "Quando editou"]], width="stretch", hide_index=True)

    def _render_chamados_content() -> None:
        """Fase C2: UI legada descontinuada — use o Sistema de Chamados (React)."""
        st.markdown("### 📌 Sistema de Chamados")
        st.warning(
            "Esta página interna do Streamlit foi **descontinuada** (Fase C2). "
            "Abra, acompanhe e feche chamados no **Sistema de Chamados** (React)."
        )
        chamados_url = get_chamados_app_url().rstrip("/") or "http://localhost:3000"
        st.link_button("Abrir Chamados (React)", url=f"{chamados_url}/tickets", use_container_width=False)
        st.caption("Use o botão **📌 Chamados** na barra superior para login automático (SSO).")
        if HAS_DB and is_postgres_configured():
            try:
                from src.db.repository import resumir_chamados_legado

                resumo = resumir_chamados_legado()
                st.markdown("#### Resumo do banco")
                c1, c2, c3 = st.columns(3)
                c1.metric("Tickets (React)", resumo.get("total_tickets", 0))
                c2.metric("Chamados legado", resumo.get("total_chamados", 0))
                c3.metric("Auditoria só legado", resumo.get("auditoria_so_legado", 0))
                if int(resumo.get("auditoria_orfa") or 0) > 0:
                    st.error(
                        f"{resumo.get('auditoria_orfa')} registro(s) de auditoria com chamado_id órfão. "
                        "Execute: python -m scripts.verificar_chamados_legado"
                    )
            except Exception as exc:
                st.caption(f"Não foi possível carregar resumo: {exc}")
        st.caption("Documentação: doc/CHAMADOS_TICKETS_UNIFICACAO.md")

    with col_cfg:
        if current_page == "config":
            if st.button("⬅ Painel", key="btn_go_panel"):
                st.query_params["pagina"] = "painel"
                st.rerun()
        elif _current_user_is_admin():
            with st.popover("➕ Adicionar"):
                if HAS_DB:
                    st.caption("Passo 1: escolha Segmento e Equipe. Passo 2: complete os dados na tabela.")
                    seg_add = st.selectbox("Segmento", ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"], key="add_seg")
                    eq_ref = (
                        EQUIPES_ALIMENTO
                        if seg_add == "Alimento"
                        else (
                            EQUIPES_MEDICAMENTO
                            if seg_add == "Medicamento"
                            else (
                                EQUIPES_PROMOTORES
                                if seg_add == "Promotores"
                                else (
                                    EQUIPES_INTERNOS
                                    if seg_add == "Internos"
                                    else (EQUIPES_MANUTENCAO if seg_add == "Manutenção" else EQUIPES_ROUBO_PERDA)
                                )
                            )
                        )
                    )
                    eq_add = st.selectbox("Equipe", eq_ref, key="add_equipe")

                    if st.button("Criar linha na equipe", key="add_criar_linha"):
                        try:
                            df_atual = load_linhas(modo=modo_db)
                            cols = df_atual.columns.tolist() or [
                                "Codigo", "Nome", "Equipe", "EquipePadrao", "GrupoEquipe", "TipoEquipe",
                                "Localidade", "Gestor", "Supervisor", "Segmento", "Papel", "Linha",
                                "E-mail", "Gerenciamento", "IMEI A", "IMEI B", "CHIP", "Aparelho", "Modelo",
                                "Setor", "Cargo", "Desconto", "Perfil", "Empresa", "Ativo",
                                "Numero de Serie", "Operadora", "Aba",
                            ]
                            linha_temp = f"NOVA-{secrets.token_hex(3).upper()}"
                            new_row = {c: "" for c in cols}
                            new_row.update({
                                "Segmento": seg_add,
                                "EquipePadrao": eq_add,
                                "Equipe": eq_add,
                                "GrupoEquipe": seg_add,
                                "TipoEquipe": "",
                                "Linha": linha_temp,
                            })
                            df_novo = pd.concat([df_atual, pd.DataFrame([new_row])], ignore_index=True)
                            df_novo = df_novo.sort_values(
                                ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
                                kind="stable",
                            )
                            save_linhas(df_novo, modo=modo_db)
                            _audit(
                                acao="criar_linha",
                                entidade="linhas",
                                chave=linha_temp,
                                depois={
                                    "segmento": seg_add,
                                    "equipe": eq_add,
                                    "modo": modo_db,
                                    "linha": linha_temp,
                                },
                                detalhes="Linha criada via botão Adicionar",
                            )
                            _set_post_chamado_banner("Linha criada com sucesso.")
                            st.session_state.pending_segmento = seg_add
                            st.session_state.pending_equipe = eq_add
                            st.session_state.pending_linha = linha_temp
                            st.session_state.filtro_busca = ""
                            st.success("Linha criada! Complete os dados na tabela abaixo.")
                            st.rerun()
                        except Exception as exc:
                            st.error(f"Erro ao criar linha: {exc}")
                else:
                    st.caption("Banco indisponivel no momento para adicionar linhas.")
            if st.button("⚙ Configuração", key="btn_go_config"):
                st.query_params["pagina"] = "config"
                st.rerun()
    with col_chamados:
        if not chamados_app_url:
            st.caption("Sem URL")
        else:
            authenticated = bool(st.session_state.get("authenticated", False))
            if not (authenticated and HAS_DB):
                st.caption("Faça login para usar o sistema de chamados.")
            else:
                user = st.session_state.get("user", {}) or {}
                username = str(user.get("username") or "").strip().lower()
                usuario_app_id = obter_usuario_app_id(username) if username else None

                # IMPORTANTE:
                # `st.link_button` pode reutilizar o `sso_code` que já estava gerado no último render.
                # Por isso, geramos o `sso_code` SOMENTE no clique, e redirecionamos imediatamente.
                if st.button("📌 Chamados", key="btn_chamados_redirect"):
                    if usuario_app_id:
                        expira_em = datetime.utcnow() + timedelta(minutes=5)
                        sso_code = criar_sso_code(usuario_app_id, expira_em)
                        code_preview = f"{sso_code[:8]}...{sso_code[-4:]}" if sso_code and len(sso_code) > 12 else sso_code
                        print(f"[GER] Gerado sso_code no clique: {code_preview}")
                        url = (
                            f"{chamados_app_url}?sso_code={sso_code}"
                            if sso_code
                            else chamados_app_url
                        )
                        chamado_id_sso = str(
                            st.session_state.get("chamado_id")
                            or (st.session_state.get("chamado_context") or {}).get("chamado_id")
                            or ""
                        ).strip()
                        if sso_code and chamado_id_sso:
                            redirect_path = quote(f"/tickets?ticket_id={chamado_id_sso}", safe="")
                            url = f"{url}&redirect={redirect_path}"
                        st.session_state["_sso_redirect_url"] = url
                        st.rerun()
                    else:
                        st.caption("Sem usuário cadastrado no banco para SSO.")

                redirect_url = st.session_state.get("_sso_redirect_url")
                if redirect_url:
                    st.session_state["_sso_redirect_url"] = None
                    # Abre o sistema React de Chamados em nova aba (evita trocar a aba atual).
                    # Obs.: o `sso_code` está na query string e é consumido no backend ao abrir a nova aba.
                    safe_url = str(redirect_url).replace("\\", "\\\\").replace('"', '\\"')
                    st.components.v1.html(
                        f"""
                        <script>
                          window.open("{safe_url}", "_blank", "noopener,noreferrer");
                        </script>
                        """,
                        height=0,
                    )
                    st.caption("Abrindo Chamados em uma nova aba...")
    with col_logout:
        user_display = st.session_state.get("user", {}).get("username", "")
        st.markdown(
            f'<span style="color: #1a1a1a; font-size: 0.9rem; white-space: nowrap;">{user_display}</span>&nbsp;&nbsp;',
            unsafe_allow_html=True,
        )
        if st.button("Sair", key="btn_logout"):
            token = st.session_state.get("auth_token")
            if token and HAS_DB:
                encerrar_sessao(token)
            if HAS_COOKIES and _cookies.ready():
                try:
                    del _cookies["auth_token"]
                    _cookies.save()
                except Exception:
                    pass
            st.session_state.authenticated = False
            st.session_state.user = None
            st.session_state.auth_token = None
            st.rerun()

    if current_page == "config":
        if _streamlit_access_mode("config") == "blocked_config":
            st.markdown("### Configuracao")
            st.warning("Area restrita a administradores. Use o Sistema de Chamados para operacao diaria.")
            if chamados_app_url:
                st.link_button(
                    "Abrir Sistema de Chamados",
                    url=chamados_app_url.rstrip("/"),
                    use_container_width=False,
                )
            if st.button("Voltar ao painel", key="btn_blocked_config_panel"):
                st.query_params["pagina"] = "painel"
                st.rerun()
            return
        st.markdown("### ⚙ Configurações")
        _render_config_content()
        return

    active_chamado_context = st.session_state.get("chamado_context", {}) or {}
    panel_access = _streamlit_access_mode("painel", active_chamado_context)
    if panel_access == "blocked_panel":
        _render_operador_sem_contexto(chamados_app_url)
        return
    if panel_access == "operador_chamado":
        st.caption(
            "Modo operador: edicao vinculada ao chamado. Painel completo disponivel apenas para administradores."
        )

    if any(bool(v) for k, v in active_chamado_context.items() if k not in {"chamado_source", "chamado_label", "chamado_numero", "chamado_aviso", "chamado_legado"}):
        contexto_partes: list[str] = []
        chamado_id_ctx = str(active_chamado_context.get("chamado_id") or "").strip()
        chamado_source_ctx = str(active_chamado_context.get("chamado_source") or "").strip()
        chamado_label_ctx = str(active_chamado_context.get("chamado_label") or "").strip()
        if chamado_id_ctx:
            if chamado_source_ctx == "tickets":
                rotulo = chamado_label_ctx or f"Ticket #{chamado_id_ctx}"
                contexto_partes.append(f"**Ticket (Chamados):** {rotulo} (id `{chamado_id_ctx}`)")
            else:
                rotulo = chamado_label_ctx or f"Chamado #{chamado_id_ctx}"
                contexto_partes.append(f"**Chamado:** {rotulo} (id `{chamado_id_ctx}`)")
        if active_chamado_context.get("linha"):
            contexto_partes.append(f"**Linha alvo:** {active_chamado_context['linha']}")
        if active_chamado_context.get("segmento"):
            contexto_partes.append(f"**Segmento:** {active_chamado_context['segmento']}")
        if active_chamado_context.get("equipe"):
            contexto_partes.append(f"**Equipe:** {active_chamado_context['equipe']}")
        contexto_partes.append(f"**Usuário:** {st.session_state.get('user', {}).get('username', '—')}")
        st.info("Contexto de chamado ativo. Filtros aplicados automaticamente.  \n" + " | ".join(contexto_partes))
        chamado_aviso_ctx = str(active_chamado_context.get("chamado_aviso") or "").strip()
        if chamado_aviso_ctx:
            st.warning(chamado_aviso_ctx)
        link_cols = st.columns([1, 1])
        with link_cols[0]:
            if active_chamado_context.get("return_url"):
                st.link_button("Voltar para o chamado", url=str(active_chamado_context["return_url"]), use_container_width=False)
        with link_cols[1]:
            if chamado_source_ctx == "tickets" and chamados_app_url and chamado_id_ctx:
                ticket_url = f"{chamados_app_url.rstrip('/')}/tickets?ticket_id={chamado_id_ctx}"
                st.link_button("Abrir ticket no Chamados", url=ticket_url, use_container_width=False)
            elif chamados_app_url and chamado_id_ctx:
                ticket_url = f"{chamados_app_url.rstrip('/')}/tickets?ticket_id={chamado_id_ctx}"
                st.link_button("Abrir chamado no Chamados", url=ticket_url, use_container_width=False)

    post_conflict_feedback = st.session_state.get("_post_conflict_feedback") or {}
    if post_conflict_feedback:
        feedback_type = str(post_conflict_feedback.get("type") or "").strip().lower()
        feedback_text = str(post_conflict_feedback.get("text") or "").strip()
        if feedback_text:
            if feedback_type == "success":
                st.success(feedback_text)
            else:
                st.warning(feedback_text)
        st.session_state["_post_conflict_feedback"] = {}

    post_chamado_banner = st.session_state.get("_post_chamado_return_banner") or {}
    if post_chamado_banner:
        banner_text = str(post_chamado_banner.get("text") or "").strip() or "Alteração concluída com sucesso."
        chamado_banner = str(post_chamado_banner.get("chamado_id") or "").strip()
        return_url_banner = str(post_chamado_banner.get("return_url") or "").strip()
        if chamado_banner:
            banner_text = f"{banner_text} Chamado: {chamado_banner}."
        st.success(banner_text)
        if return_url_banner:
            st.link_button("Retornar ao chamado", url=return_url_banner, use_container_width=False)
        st.session_state["_post_chamado_return_banner"] = {}

    # Carregar dados: somente banco de dados
    if not HAS_DB:
        st.error("O backend de banco de dados nao esta disponivel.")
        st.stop()
    dados_do_banco = False
    if has_data(modo_db) or (segmento_sem_filtro_modo and (has_data("ativas") or has_data("desativadas"))):
        try:
            if segmento_sem_filtro_modo:
                df_ativas = load_linhas(modo="ativas") if has_data("ativas") else pd.DataFrame()
                df_des = load_linhas(modo="desativadas") if has_data("desativadas") else pd.DataFrame()
                df = pd.concat([df_ativas, df_des], ignore_index=True)
            else:
                df = load_linhas(modo=modo_db)
            dados_do_banco = True
        except Exception as exc:
            st.error(f"Erro ao ler o banco: {exc}")
            st.stop()
    if not dados_do_banco:
        st.error("O banco de dados esta vazio ou nao possui dados para este modo. O sistema nao usa mais planilhas como fonte.")
        st.stop()

    # Fase B1: resolve tickets.id (React) ou chamados legado; stub só quando necessário.
    if dados_do_banco and active_chamado_context.get("chamado_id") and preparar_referencia_chamado:
        try:
            chamado_ref_prep = preparar_referencia_chamado(str(active_chamado_context.get("chamado_id")))
            if chamado_ref_prep.get("valido"):
                active_chamado_context["chamado_source"] = str(chamado_ref_prep.get("source") or "")
                active_chamado_context["chamado_label"] = str(chamado_ref_prep.get("label") or "")
                active_chamado_context["chamado_aviso"] = str(chamado_ref_prep.get("aviso") or "")
                if chamado_ref_prep.get("source") == "chamados" and active_chamado_context.get("linha"):
                    vincular_chamado_linha(active_chamado_context.get("chamado_id"), active_chamado_context.get("linha"))
        except Exception:
            pass

    if dados_do_banco and not segmento_sem_filtro_modo:
        df = _auto_mark_context_line_vago(df, modo_db, dados_do_banco)

    prefixo = "Linhas" if segmento_sem_filtro_modo else ("Linhas Ativas" if modo_sel == "Linhas ativas" else "Linhas Desativadas")
    df_full_mode = df.copy()
    # ID técnico por linha para merge seguro no data_editor (evita depender da coluna Linha).
    df_full_mode = df_full_mode.reset_index(drop=True)
    df_full_mode["__row_key"] = [f"{modo_db}:{i}" for i in range(len(df_full_mode))]
    df_full_mode["__base_linha"] = df_full_mode["Linha"].fillna("").astype(str).str.strip() if "Linha" in df_full_mode.columns else ""

    df = df.reset_index(drop=True).copy()
    df["__row_key"] = df_full_mode["__row_key"].values
    df["__base_linha"] = df_full_mode["__base_linha"].values
    if segmento_sel == "Alimento":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "alimento"]
        if (not dados_do_banco) and "Aba" in df.columns:
            aba_norm = df["Aba"].fillna("").astype(str).str.strip()
            df = df[aba_norm.str.lower().isin([a.lower() for a in ABAS_ALIMENTO])]
        equipes_ref = EQUIPES_ALIMENTO
        titulo = f"{prefixo} — Alimento"
    elif segmento_sel == "Medicamento":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "medicamento"]
        equipes_ref = EQUIPES_MEDICAMENTO
        titulo = f"{prefixo} — Medicamento"
    elif segmento_sel == "Promotores":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "promotores"]
        equipes_ref = EQUIPES_PROMOTORES
        titulo = f"{prefixo} — Promotores"
    elif segmento_sel == "Internos":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "internos"]
        equipes_ref = EQUIPES_INTERNOS
        titulo = f"{prefixo} — Internos"
    elif segmento_sel == "Manutenção":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "manutenção"]
        equipes_ref = EQUIPES_MANUTENCAO
        titulo = f"{prefixo} — Manutenção"
    else:
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "roubo e perda"]
        equipes_ref = EQUIPES_ROUBO_PERDA
        titulo = f"{prefixo} — Roubo e Perda"

    if df.empty:
        st.warning(f"Nenhum dado de {segmento_sel} encontrado.")
        st.stop()

    linhas_preenchidas_segmento = (
        df["Linha"].fillna("").astype(str).str.strip() != ""
        if "Linha" in df.columns
        else pd.Series(False, index=df.index)
    )
    linhas_vagas_segmento = build_vaga_mask(df)
    linhas_em_uso_segmento = linhas_preenchidas_segmento & ~linhas_vagas_segmento
    escopo_modo = "ativas" if modo_sel == "Linhas ativas" else "desativadas"
    total_segmento = int(len(df))
    painel_contexto = segmento_sel if segmento_sem_filtro_modo else f"{segmento_sel} | {modo_sel}"
    total_label = f"Total {segmento_sel}" if segmento_sem_filtro_modo else f"Total {escopo_modo}"
    st.caption(f"Painel operacional: {painel_contexto}")
    d1, d2, d3 = st.columns(3)
    d1.metric(f"Em uso ({segmento_sel})", int(linhas_em_uso_segmento.sum()))
    d2.metric(f"Vagas ({segmento_sel})", int(linhas_vagas_segmento.sum()))
    d3.metric(total_label, total_segmento)

    if dados_do_banco:
        tbl = df.copy()
        tbl = tbl.sort_values(
            ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
            kind="stable",
        )
    else:
        tbl = build_full_table(df)
    # Normaliza valores nulos para não exibir "None" no editor/tabelas.
    tbl = tbl.replace({None: ""}).fillna("")
    if segmento_sel == "Roubo e Perda":
        cols_front = [
            "Nome de Guerra",
            "Nome",
            "Equipe",
            "Data Ocorrência",
            "Linha",
            "Marca",
            "Modelo",
            "IMEI A",
            "IMEI B",
            "CHIP",
            "Patrimonio",
            "Numero de Serie",
            "Motivo",
            "Data Solicitação TBS",
            "Observação",
        ]
    elif segmento_sel == "Manutenção":
        cols_front = [
            "Data da Troca",
            "Data Retorno",
            "Codigo",
            "Nome",
            "Equipe",
            "Linha",
            "E-mail",
            "IMEI A",
            "IMEI B",
            "CHIP",
            "Aparelho",
            "Modelo",
            "Setor",
            "Cargo",
            "Patrimonio",
            "Numero de Serie",
            "Motivo",
            "Observação",
        ]
    elif segmento_sel == "Internos":
        cols_front = [
            "Nome",
            "Localidade",
            "Linha",
            "E-mail",
            "Gerenciamento",
            "IMEI A",
            "IMEI B",
            "CHIP",
            "Aparelho",
            "Modelo",
            "Setor",
            "Cargo",
            "Desconto",
            "Perfil",
            "Empresa",
            "Ativo",
            "Numero de Serie",
        ]
    else:
        cols_front = [
            "Codigo",
            "Nome",
            "Equipe",
            "Linha",
            "E-mail",
            "Gerenciamento",
            "IMEI A",
            "IMEI B",
            "CHIP",
            "Aparelho",
            "Modelo",
            "Setor",
            "Cargo",
            "Desconto",
            "Perfil",
            "Empresa",
            "Ativo",
            "Numero de Serie",
            "Operadora",
        ]
    existing_front = [c for c in cols_front if c in tbl.columns]
    tail = [c for c in tbl.columns if c not in existing_front]
    tbl = tbl[existing_front + tail]
    if not dados_do_banco and segmento_sel == "Alimento":
        gerentes_set = _get_gerentes_alimento(rules_path)
        mask_papel = tbl["Papel"].fillna("").astype(str).str.strip().str.lower() == "gerente"
        mask_nome = tbl["Nome"].fillna("").astype(str).apply(
            lambda x: normalize_text(x.strip()) in gerentes_set
        )
        mask_gerente = mask_papel | mask_nome
        if mask_gerente.any():
            tbl.loc[mask_gerente, "EquipePadrao"] = "Gerentes do Alimento"
            tbl = tbl.sort_values(
                ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
                kind="stable",
            )
    if not dados_do_banco and segmento_sel == "Medicamento":
        gerentes_set = _get_gerentes_medicamento(rules_path)
        mask_papel = tbl["Papel"].fillna("").astype(str).str.strip().str.lower() == "gerente"
        mask_nome = tbl["Nome"].fillna("").astype(str).apply(
            lambda x: normalize_text(x.strip()) in gerentes_set
        )
        mask_gerente = mask_papel | mask_nome
        if mask_gerente.any():
            tbl.loc[mask_gerente, "EquipePadrao"] = "Gerentes do Medicamento"
            tbl = tbl.sort_values(
                ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
                kind="stable",
            )
    equipes_validas = sorted(set(tbl["EquipePadrao"].dropna().astype(str).str.strip().unique()))
    equipes = [e for e in equipes_ref if e in equipes_validas]
    equipes += [e for e in equipes_validas if e not in equipes]
    # No segmento Manutenção, "Mover equipe" = mover entre as duas seções (não entre equipes externas)
    equipes_mover_manut = ["Manutenção", "Manutenções antigas"]
    tbl_segment = tbl.copy()

    st.markdown(f'<p class="main-title">{titulo}</p>', unsafe_allow_html=True)
    st.markdown('<p class="main-subtitle">Controle de linhas telefônicas por equipe. Selecione uma equipe para ver os detalhes.</p>', unsafe_allow_html=True)

    col_filtro, col_busca_tipo, col_busca, col_vagas = st.columns([1, 1.1, 2.0, 0.9])
    with col_filtro:
        equipe_options = ["Todas as equipes"] + equipes
        pending_equipe = st.session_state.get("pending_equipe")
        if pending_equipe in equipe_options:
            st.session_state.filtro_equipe = pending_equipe
        current_equipe_filter = st.session_state.get("filtro_equipe")
        if current_equipe_filter not in equipe_options:
            st.session_state["filtro_equipe"] = "Todas as equipes"
        equipe_sel = st.selectbox(
            "Equipe",
            options=equipe_options,
            format_func=lambda x: "Todas as equipes" if x == "Todas as equipes" else x,
            label_visibility="collapsed",
            key="filtro_equipe",
        )
    with col_busca_tipo:
        busca_tipo_options = ["Geral", "Linha", "IMEI", "Aparelho", "Motivo"]
        if st.session_state.get("filtro_busca_tipo") not in busca_tipo_options:
            st.session_state["filtro_busca_tipo"] = "Geral"
        busca_tipo = st.selectbox(
            "Tipo de busca",
            options=busca_tipo_options,
            label_visibility="collapsed",
            key="filtro_busca_tipo",
        )
    with col_busca:
        placeholder_map = {
            "Geral": "Buscar por nome, codigo, linha, aparelho, IMEI...",
            "Linha": "Buscar numero da linha...",
            "IMEI": "Buscar IMEI A ou IMEI B...",
            "Aparelho": "Buscar aparelho ou modelo...",
            "Motivo": "Buscar motivo ou observacao...",
        }
        busca = st.text_input(
            "Buscar",
            placeholder=placeholder_map.get(busca_tipo, "Buscar..."),
            label_visibility="collapsed",
            key="filtro_busca",
        )
    with col_vagas:
        if "filtro_somente_vagas" not in st.session_state:
            st.session_state["filtro_somente_vagas"] = False
        somente_vagas = st.checkbox("Somente vagas", key="filtro_somente_vagas")
    st.query_params["equipe"] = equipe_sel
    st.query_params["busca"] = busca
    st.query_params["busca_tipo"] = busca_tipo
    st.query_params["somente_vagas"] = "1" if somente_vagas else "0"
    editor_context_key = f"{modo_db}|{segmento_sel}|{equipe_sel}|{busca_tipo}|{busca}|{int(somente_vagas)}"

    query = busca.strip().lower()
    if query:
        mask = pd.Series(False, index=tbl.index)
        if busca_tipo == "Linha":
            columns_to_search = ["Linha"]
        elif busca_tipo == "IMEI":
            columns_to_search = ["IMEI A", "IMEI B"]
        elif busca_tipo == "Aparelho":
            columns_to_search = ["Aparelho", "Modelo", "Marca"]
        elif busca_tipo == "Motivo":
            columns_to_search = ["Motivo", "Observação", "Gerenciamento"]
        else:
            columns_to_search = [
                "Nome",
                "Codigo",
                "Linha",
                "Aparelho",
                "Modelo",
                "Marca",
                "EquipePadrao",
                "Supervisor",
                "Gestor",
                "Motivo",
                "Observação",
                "IMEI A",
                "IMEI B",
                "E-mail",
            ]
        for c in columns_to_search:
            if c in tbl.columns:
                mask = mask | tbl[c].astype(str).str.lower().str.contains(query, na=False)
        tbl = tbl[mask]

    if equipe_sel != "Todas as equipes":
        tbl = tbl[tbl["EquipePadrao"] == equipe_sel]

    if somente_vagas:
        tbl = tbl[build_vaga_mask(tbl)]

    for qp_key in ["com_motivo", "sem_aparelho", "imei_pendente", "sem_email"]:
        if qp_key in st.query_params:
            del st.query_params[qp_key]

    pending_linha = st.session_state.get("pending_linha")
    if pending_linha and not tbl.empty and "Linha" in tbl.columns:
        mask_pending = tbl["Linha"].fillna("").astype(str) == pending_linha
        if mask_pending.any():
            st.info(f"Nova linha criada na equipe **{equipe_sel}**. Procure por `{pending_linha}` e edite os dados.")
            st.session_state.pending_segmento = None
            st.session_state.pending_equipe = None
            st.session_state.pending_linha = None

    linha_contexto_ativa = str(((st.session_state.get("chamado_context") or {}).get("linha") or "")).strip()
    if linha_contexto_ativa and not tbl.empty and "Linha" in tbl.columns:
        mask_contexto = tbl["Linha"].fillna("").astype(str).str.strip() == linha_contexto_ativa
        if mask_contexto.any():
            st.success(f"Linha do chamado localizada: `{linha_contexto_ativa}`")

    conflict_payload = st.session_state.get("_editor_conflict_payload") or {}

    def _build_conflict_payload_from_pending(
        pending_changes: list[dict[str, str]],
        context_key: str,
    ) -> dict[str, Any]:
        if not pending_changes:
            return {}
        resumo: list[dict[str, str]] = []
        detalhes: list[dict[str, str]] = []
        grouped: dict[str, list[dict[str, str]]] = {}
        for ch in pending_changes:
            base_linha = str(ch.get("base_linha") or "").strip()
            if not base_linha:
                continue
            grouped.setdefault(base_linha, []).append(ch)
        for base_linha, items in grouped.items():
            campos = [str(it.get("campo") or "").strip() for it in items if str(it.get("campo") or "").strip()]
            resumo.append(
                {
                    "Linha": base_linha or "—",
                    "Motivo": "Alguns campos continuam em conflito.",
                    "Campos com diferenca": ", ".join(campos[:6]) + (f" (+{len(campos) - 6})" if len(campos) > 6 else ""),
                }
            )
            for it in items:
                detalhes.append(
                    {
                        "Linha": base_linha or "—",
                        "Campo": str(it.get("campo") or "—"),
                        "Valor original": str(it.get("original") or "—"),
                        "Minha edição": str(it.get("mine") or "—"),
                        "Valor atual no banco": str(it.get("current") or "—"),
                    }
                )
        return {
            "context": context_key,
            "resumo": resumo,
            "detalhes": detalhes,
            "pending_changes": pending_changes,
        }

    if conflict_payload and conflict_payload.get("context") == editor_context_key:
        st.warning("Existe um conflito de edição pendente. Revise as diferenças antes de tentar salvar novamente.")
        resumo_conf = conflict_payload.get("resumo") or []
        detalhes_conf = conflict_payload.get("detalhes") or []
        if resumo_conf:
            st.dataframe(pd.DataFrame(resumo_conf), width="stretch", hide_index=True)
        if detalhes_conf:
            with st.expander("Ver diferenças do conflito", expanded=False):
                st.dataframe(pd.DataFrame(detalhes_conf), width="stretch", hide_index=True)
            linhas_conflito_opts = sorted(
                {
                    str(item.get("Linha") or "").strip()
                    for item in detalhes_conf
                    if str(item.get("Linha") or "").strip()
                }
            )
            if linhas_conflito_opts:
                linha_revisao = st.selectbox(
                    "Revisar conflito por linha",
                    options=linhas_conflito_opts,
                    key="conflict_line_review",
                )
                detalhes_linha = [
                    item for item in detalhes_conf
                    if str(item.get("Linha") or "").strip() == linha_revisao
                ]
                if detalhes_linha:
                    st.caption(f"Comparativo da linha `{linha_revisao}`")
                    st.dataframe(pd.DataFrame(detalhes_linha), width="stretch", hide_index=True)
                    campos_linha_opts = [
                        str(item.get("Campo") or "").strip()
                        for item in detalhes_linha
                        if str(item.get("Campo") or "").strip()
                    ]
                    if campos_linha_opts:
                        campo_revisao = st.selectbox(
                            "Resolver campo",
                            options=campos_linha_opts,
                            key="conflict_field_review",
                        )
                        detalhe_campo = next(
                            (
                                item for item in detalhes_linha
                                if str(item.get("Campo") or "").strip() == campo_revisao
                            ),
                            None,
                        )
                        if detalhe_campo:
                            a_res1, a_res2 = st.columns([1, 1.2])
                            with a_res1:
                                if st.button("Aceitar valor atual do banco", key="btn_accept_db_value_conflict"):
                                    pending_changes = conflict_payload.get("pending_changes") or []
                                    pending_changes = [
                                        ch
                                        for ch in pending_changes
                                        if not (
                                            str(ch.get("base_linha") or "").strip() == linha_revisao
                                            and str(ch.get("campo") or "").strip() == campo_revisao
                                        )
                                    ]
                                    st.session_state["_editor_conflict_payload"] = _build_conflict_payload_from_pending(
                                        pending_changes,
                                        editor_context_key,
                                    )
                                    st.session_state["_post_conflict_feedback"] = {
                                        "type": "success",
                                        "text": f"Campo `{campo_revisao}` da linha `{linha_revisao}` resolvido com o valor atual do banco.",
                                    }
                                    st.rerun()
                            with a_res2:
                                if st.button("Sobrescrever com minha edição", key="btn_force_my_value_conflict"):
                                    pending_changes = conflict_payload.get("pending_changes") or []
                                    pending_change = next(
                                        (
                                            ch for ch in pending_changes
                                            if str(ch.get("base_linha") or "").strip() == linha_revisao
                                            and str(ch.get("campo") or "").strip() == campo_revisao
                                        ),
                                        None,
                                    )
                                    if pending_change is None:
                                        st.warning("Nao foi possivel localizar o conflito selecionado.")
                                    else:
                                        df_current_mode = load_linhas(modo=modo_db)
                                        seg_lower_force = segmento_sel.strip().lower()
                                        mask_seg_force = df_current_mode["Segmento"].fillna("").astype(str).str.strip().str.lower() == seg_lower_force
                                        base_linha_force = str(pending_change.get("base_linha") or "").strip()
                                        campo_force = str(pending_change.get("campo") or "").strip()
                                        novo_valor_raw = str(pending_change.get("mine_raw") or "")
                                        current_matches = df_current_mode[
                                            mask_seg_force & (df_current_mode["Linha"].fillna("").astype(str).str.strip() == base_linha_force)
                                        ]
                                        if current_matches.empty:
                                            st.warning("A linha nao foi encontrada no banco para sobrescrever.")
                                        elif len(current_matches) > 1:
                                            st.warning("Existe mais de uma linha com esse identificador. Resolva manualmente para evitar sobrescrita indevida.")
                                        elif not campo_force or campo_force not in df_current_mode.columns:
                                            st.warning("Campo invalido para sobrescrita.")
                                        else:
                                            row_idx_force = current_matches.index[0]
                                            before_force = df_current_mode.loc[row_idx_force].to_dict()
                                            df_current_mode.loc[row_idx_force, campo_force] = novo_valor_raw
                                            save_linhas(df_current_mode, modo=modo_db)
                                            _audit(
                                                acao="resolver_conflito_edicao",
                                                entidade="linhas",
                                                chave=base_linha_force,
                                                antes=before_force,
                                                depois={campo_force: novo_valor_raw, "modo": modo_db},
                                                detalhes=f"Conflito resolvido manualmente mantendo a edicao do usuario no campo {campo_force}.",
                                            )
                                            _set_post_chamado_banner("Conflito resolvido e alteração aplicada com sucesso.")
                                            pending_changes = [
                                                ch
                                                for ch in pending_changes
                                                if not (
                                                    str(ch.get("base_linha") or "").strip() == linha_revisao
                                                    and str(ch.get("campo") or "").strip() == campo_revisao
                                                )
                                            ]
                                            st.session_state["_editor_conflict_payload"] = _build_conflict_payload_from_pending(
                                                pending_changes,
                                                editor_context_key,
                                            )
                                            st.session_state["_editor_base_context"] = ""
                                            st.session_state["_editor_base_snapshot"] = []
                                            st.session_state["_post_conflict_feedback"] = {
                                                "type": "success",
                                                "text": f"Campo `{campo_revisao}` da linha `{linha_revisao}` sobrescrito com a sua edicao.",
                                            }
                                            st.rerun()
        c_conf1, c_conf2, c_conf3 = st.columns([1, 1, 1.2])
        with c_conf1:
            if st.button("Atualizar dados", key="btn_refresh_after_conflict"):
                st.session_state["_editor_base_context"] = ""
                st.session_state["_editor_base_snapshot"] = []
                st.session_state["_editor_conflict_payload"] = {}
                st.rerun()
        with c_conf2:
            if st.button("Fechar aviso", key="btn_close_conflict_warning"):
                st.session_state["_editor_conflict_payload"] = {}
                st.rerun()
        with c_conf3:
            if st.button("Reaplicar alterações seguras", key="btn_reapply_safe_changes"):
                st.session_state["_reapply_conflict_request"] = editor_context_key

    conflict_map_by_line: dict[str, list[str]] = {}
    if conflict_payload and conflict_payload.get("context") == editor_context_key:
        pending_conflict_changes = conflict_payload.get("pending_changes") or []
        for ch in pending_conflict_changes:
            base_linha = str(ch.get("base_linha") or "").strip()
            campo = str(ch.get("campo") or "").strip()
            if not base_linha or not campo:
                continue
            conflict_map_by_line.setdefault(base_linha, [])
            if campo not in conflict_map_by_line[base_linha]:
                conflict_map_by_line[base_linha].append(campo)

    def _decorate_editor_df(df_in: pd.DataFrame, manut_antigos: bool = False) -> pd.DataFrame:
        df_out = df_in.copy()
        df_out = df_out.drop(columns=["__manut_antigo", "Histórico"], errors="ignore")
        if segmento_sel == "Manutenção" and manut_antigos:
            df_out = df_out.drop(columns=["↕ Ordem"], errors="ignore")
        elif segmento_sel == "Manutenção" and not manut_antigos:
            if "↕ Ordem" not in df_out.columns:
                df_out["↕ Ordem"] = list(range(1, len(df_out) + 1))
            else:
                df_out["↕ Ordem"] = pd.to_numeric(df_out["↕ Ordem"], errors="coerce")
        else:
            if "↕ Ordem" not in df_out.columns:
                df_out["↕ Ordem"] = list(range(1, len(df_out) + 1))
            else:
                ordem_num = pd.to_numeric(df_out["↕ Ordem"], errors="coerce")
                ordem_seq = pd.Series(range(1, len(df_out) + 1), index=df_out.index, dtype="int64")
                df_out["↕ Ordem"] = ordem_num.where(ordem_num.notna(), ordem_seq).astype(int)
        if "Mover equipe" not in df_out.columns:
            df_out["Mover equipe"] = ""
        if conflict_map_by_line:
            base_linhas = df_out["__base_linha"].fillna("").astype(str).str.strip() if "__base_linha" in df_out.columns else pd.Series([""] * len(df_out), index=df_out.index)
            conflito_textos = []
            conflito_ord = []
            for ln in base_linhas:
                campos = conflict_map_by_line.get(ln, [])
                if campos:
                    txt = ", ".join(campos[:3])
                    if len(campos) > 3:
                        txt += f" +{len(campos) - 3}"
                    conflito_textos.append(f"Conflito: {txt}")
                    conflito_ord.append(0)
                else:
                    conflito_textos.append("")
                    conflito_ord.append(1)
            df_out["Conflito"] = conflito_textos
            df_out["__conflict_ord"] = conflito_ord
            sort_cols_conf = ["__conflict_ord"] + [c for c in ["Papel", "Nome", "Linha"] if c in df_out.columns]
            ascending_conf = [True] + [True] * (len(sort_cols_conf) - 1)
            df_out = df_out.sort_values(sort_cols_conf, ascending=ascending_conf, kind="stable")
            df_out = df_out.drop(columns=["__conflict_ord"], errors="ignore")
        cols = df_out.columns.tolist()
        controls_front = [c for c in ["↕ Ordem", "Mover equipe"] if c in cols]
        cols = controls_front + [c for c in cols if c not in controls_front]
        if "Conflito" in cols:
            cols = controls_front + ["Conflito"] + [c for c in cols if c not in (set(controls_front) | {"Conflito"})]
        df_out = df_out[cols]
        return df_out

    n_linhas = tbl["Linha"].nunique()
    n_pessoas = len(tbl)
    lbl_linhas = "linhas" if segmento_sem_filtro_modo else ("linhas desativadas" if modo_sel == "Linhas desativadas" else "linhas ativas")
    st.caption(f"**{n_pessoas}** registros | **{n_linhas}** {lbl_linhas}")

    if tbl.empty:
        st.info("Nenhum resultado. Tente outro filtro ou termo de busca.")
        st.stop()

    if segmento_sel == "Medicamento":
        sort_cols = ["Localidade", "Papel", "Nome", "Linha"]
        sort_ascending = [True, True, True, True]
    elif segmento_sel == "Manutenção":
        sort_cols = ["Data da Troca", "Data Retorno", "Nome", "Linha"]
        sort_ascending = [False, False, True, True]
    elif segmento_sel == "Roubo e Perda":
        sort_cols = ["Data Ocorrência", "Data da Troca", "Nome", "Linha"]
        sort_ascending = [False, False, True, True]
    else:
        sort_cols = ["Papel", "Nome", "Linha"]
        sort_ascending = [True, True, True]
    header_cls = "team-header-inactive" if modo_sel == "Linhas desativadas" else "team-header"
    usar_editor = HAS_DB and dados_do_banco
    save_clicked_top = False
    if usar_editor:
        st.caption("✏️ Edite direto na linha. Em '↕ Ordem' informe a posição no bloco e em 'Mover equipe' troque a equipe da linha.")
        col_save_btn, _ = st.columns([0.38, 0.62])
        with col_save_btn:
            save_clicked_top = st.button(
                "Salvar alterações",
                key="btn_save_changes_top",
                type="primary",
                use_container_width=False,
            )

    def _sort_for_display(df_in: pd.DataFrame) -> pd.DataFrame:
        """Ordena visualização, priorizando datas em ordem decrescente quando aplicável."""
        df_sort = df_in.copy()
        if "↕ Ordem" in df_sort.columns:
            df_sort["↕ Ordem"] = pd.to_numeric(df_sort["↕ Ordem"], errors="coerce")
        # Regra específica de Manutenção:
        # o mais recente sempre deve ser Ordem 1, depois 2, 3...
        # (independente de ordem manual anterior).
        # Manutenção: lógica movida para _sort_manut_block (duas seções separadas)
        sort_eff: list[str] = []
        asc_eff: list[bool] = []
        date_cols = {"Data da Troca", "Data Retorno", "Data Ocorrência", "Data Solicitação TBS"}
        if "↕ Ordem" in df_sort.columns:
            sort_eff.append("↕ Ordem")
            asc_eff.append(True)
        for col, asc in zip(sort_cols, sort_ascending):
            if col not in df_sort.columns:
                continue
            if col in date_cols:
                aux = "__ord_" + normalize_text(col).replace(" ", "_")
                df_sort[aux] = pd.to_datetime(df_sort[col], errors="coerce")
                sort_eff.append(aux)
            else:
                sort_eff.append(col)
            asc_eff.append(asc)
        if sort_eff:
            df_sort = df_sort.sort_values(sort_eff, kind="stable", ascending=asc_eff)
        aux_cols = [c for c in df_sort.columns if c.startswith("__ord_")]
        if aux_cols:
            df_sort = df_sort.drop(columns=aux_cols, errors="ignore")
        return df_sort

    def _apply_manual_team_order(df_in: pd.DataFrame, equipe_nome: str) -> pd.DataFrame:
        """Aplica ordem manual persistida em sessão para uma equipe."""
        if segmento_sel == "Manutenção":
            # Em manutenção a ordem é automática por recência (mais novo = 1).
            return df_in
        if df_in.empty or "Linha" not in df_in.columns:
            return df_in
        order_map = st.session_state.get("_manual_order_map") or {}
        order_key = f"{modo_db}|{segmento_sel}|{str(equipe_nome).strip()}"
        ordered_lines = order_map.get(order_key) or []
        if not ordered_lines:
            return df_in
        rank = {str(ln).strip(): i for i, ln in enumerate(ordered_lines)}
        df_ord = df_in.copy()
        df_ord["__manual_rank"] = df_ord["Linha"].fillna("").astype(str).str.strip().map(lambda v: rank.get(v, 10**9))
        df_ord["__manual_seq"] = range(len(df_ord))
        df_ord = df_ord.sort_values(["__manual_rank", "__manual_seq"], kind="stable")
        return df_ord.drop(columns=["__manual_rank", "__manual_seq"], errors="ignore")

    def _split_manut_novos_antigos(tbl_manut: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
        """Separa linhas de manutenção em novas (hoje+) e antigas."""
        from datetime import date
        today_ref = date.today()
        if "Data da Troca" not in tbl_manut.columns or tbl_manut.empty:
            return pd.DataFrame(), tbl_manut.copy()
        df = tbl_manut.copy()
        # Aceita YYYY-MM-DD, DD/MM/YYYY e outros formatos
        df["__ord_dt"] = pd.to_datetime(df["Data da Troca"], errors="coerce", dayfirst=True)
        df["__ord_date"] = df["__ord_dt"].dt.date
        # data_troca >= hoje = novos; NaT/data inválida vai para antigos
        mask_novos = (df["__ord_date"].notna()) & (df["__ord_date"] >= today_ref)
        df = df.drop(columns=["__ord_dt", "__ord_date"], errors="ignore")
        return df[mask_novos].copy(), df[~mask_novos].copy()

    def _sort_manut_block(df_in: pd.DataFrame, is_novos: bool) -> pd.DataFrame:
        """Ordena bloco de manutenção. Novos: ordem 1,2,3. Antigos: só por data."""
        df_sort = df_in.copy()
        if df_sort.empty:
            return df_sort
        sort_cols: list[str] = []
        asc_list: list[bool] = []
        for col, asc in [("Data da Troca", False), ("Data Retorno", False), ("Nome", True), ("Linha", True)]:
            if col not in df_sort.columns:
                continue
            if col in {"Data da Troca", "Data Retorno"}:
                aux = "__ord_" + normalize_text(col).replace(" ", "_")
                df_sort[aux] = pd.to_datetime(df_sort[col], errors="coerce")
                sort_cols.append(aux)
            else:
                sort_cols.append(col)
            asc_list.append(asc)
        if sort_cols:
            df_sort = df_sort.sort_values(sort_cols, kind="stable", ascending=asc_list)
        aux_cols = [c for c in df_sort.columns if c.startswith("__ord_")]
        if aux_cols:
            df_sort = df_sort.drop(columns=aux_cols, errors="ignore")
        if is_novos and len(df_sort) > 0:
            df_sort["↕ Ordem"] = list(range(1, len(df_sort) + 1))
        return df_sort

    def _render_team_actions(eq_nome: str, df_equipe: pd.DataFrame, df_editor_view: pd.DataFrame | None = None) -> None:
        """Ações por bloco de equipe: excluir, mover e ativar/desativar."""
        linhas_team = sorted([str(x).strip() for x in df_equipe["Linha"].dropna().astype(str).tolist() if str(x).strip()])
        if not linhas_team:
            return
        key_base = normalize_text(eq_nome).replace(" ", "_")
        linhas_selecionadas_editor: list[str] = []
        if isinstance(df_editor_view, pd.DataFrame) and not df_editor_view.empty:
            if "Selecionar" in df_editor_view.columns and "Linha" in df_editor_view.columns:
                mask_sel = df_editor_view["Selecionar"].fillna(False).astype(bool)
                if mask_sel.any():
                    linhas_selecionadas_editor = sorted(
                        {
                            str(v).strip()
                            for v in df_editor_view.loc[mask_sel, "Linha"].tolist()
                            if str(v).strip()
                        }
                    )
        with st.expander(f"Ações da equipe: {eq_nome}", expanded=False):
            action_options = [
                "Excluir linha",
                "Mover para outra equipe",
                "Mover para outro setor",
            ]
            if segmento_sel != "Manutenção":
                action_options.append("Enviar aparelho para manutenção")
            action_options.append("Desativar linha" if modo_db == "ativas" else "Ativar linha")

            c1, c2, c3 = st.columns([1.3, 1.4, 1.3])
            with c1:
                linha_default = ""
                if len(linhas_selecionadas_editor) == 1:
                    linha_default = linhas_selecionadas_editor[0]
                elif len(linhas_selecionadas_editor) > 1:
                    st.warning("Mais de uma linha foi selecionada no editor. Para esta ação, mantenha somente uma selecionada.")
                linha_options = [""] + linhas_team
                idx_default = linha_options.index(linha_default) if linha_default in linha_options else 0
                linha_alvo = st.selectbox("Linha", options=linha_options, index=idx_default, key=f"acao_linha_{key_base}")
                if linha_default:
                    st.caption("Linha carregada automaticamente da seleção do editor.")
            with c2:
                acao_linha = st.selectbox(
                    "Ação",
                    options=action_options,
                    key=f"acao_tipo_{key_base}",
                )
            linha_selecionada = {}
            if linha_alvo:
                matches = df_equipe[df_equipe["Linha"].fillna("").astype(str).str.strip() == linha_alvo]
                if not matches.empty:
                    linha_selecionada = matches.iloc[0].to_dict()
            destino_setor = ""
            destino_equipe = ""
            with c3:
                if acao_linha == "Mover para outra equipe":
                    destino_equipe = st.selectbox("Equipe destino", options=equipes, key=f"acao_eq_dest_{key_base}")
                elif acao_linha == "Mover para outro setor":
                    destino_setor = st.selectbox("Setor destino", options=["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"], key=f"acao_seg_dest_{key_base}")
                    eq_dest_ref = (
                        EQUIPES_ALIMENTO
                        if destino_setor == "Alimento"
                        else (
                            EQUIPES_MEDICAMENTO
                            if destino_setor == "Medicamento"
                            else (
                                EQUIPES_PROMOTORES
                                if destino_setor == "Promotores"
                                else (
                                    EQUIPES_INTERNOS
                                    if destino_setor == "Internos"
                                    else (EQUIPES_MANUTENCAO if destino_setor == "Manutenção" else EQUIPES_ROUBO_PERDA)
                                )
                            )
                        )
                    )
                    destino_equipe = st.selectbox("Equipe destino", options=eq_dest_ref, key=f"acao_eq_dest_setor_{key_base}")

            if linha_selecionada:
                resumo_linha = []
                for campo in ["Nome", "Aparelho", "Modelo", "Segmento", "EquipePadrao"]:
                    valor = str(linha_selecionada.get(campo, "") or "").strip()
                    if valor:
                        resumo_linha.append(f"{campo}: {valor}")
                if resumo_linha:
                    st.caption("Linha selecionada: " + " | ".join(resumo_linha))

            action_guidance = {
                "Excluir linha": "Remove a linha deste modo. Use apenas quando o registro nao deve mais existir aqui.",
                "Mover para outra equipe": "Mantem o mesmo setor e apenas troca a equipe responsavel.",
                "Mover para outro setor": "Move a linha para outro setor e atualiza tambem a equipe destino.",
                "Enviar aparelho para manutenção": "Cria uma copia em Manutenção e mantem a linha original para voce atualizar o aparelho depois.",
                "Desativar linha": "Move a linha de ativas para desativadas sem perder o historico do registro.",
                "Ativar linha": "Move a linha de desativadas para ativas para voltar ao uso operacional.",
            }
            if acao_linha in action_guidance:
                guidance_text = action_guidance[acao_linha]
                if acao_linha == "Mover para outro setor" and destino_setor == "Roubo e Perda":
                    guidance_text += " Depois da movimentacao, revise Data Ocorrencia, Motivo e observacoes."
                elif acao_linha == "Mover para outro setor" and destino_setor == "Manutenção":
                    guidance_text += " Depois da movimentacao, revise datas e dados do aparelho em manutencao."
                elif acao_linha == "Mover para outra equipe" and destino_equipe:
                    guidance_text += f" Destino selecionado: {destino_equipe}."
                st.info(guidance_text)

            def _format_action_date(value: Any) -> str:
                if value in (None, ""):
                    return ""
                try:
                    return pd.Timestamp(value).strftime("%d/%m/%Y")
                except Exception:
                    return str(value).strip()

            motivo_operacao = ""
            observacao_operacao = ""
            data_ocorrencia_operacao = ""
            data_troca_operacao = ""
            data_retorno_operacao = ""

            if acao_linha == "Mover para outro setor" and destino_setor == "Roubo e Perda":
                f_roubo1, f_roubo2 = st.columns([1, 1.2])
                with f_roubo1:
                    data_ocorrencia_operacao = _format_action_date(
                        st.date_input(
                            "Data ocorrencia",
                            value=pd.Timestamp.today().date(),
                            key=f"acao_data_ocorrencia_{key_base}",
                        )
                    )
                    motivo_operacao = st.text_input(
                        "Motivo",
                        key=f"acao_motivo_roubo_{key_base}",
                        placeholder="Ex.: Roubo, perda ou furto",
                    ).strip()
                with f_roubo2:
                    observacao_operacao = st.text_area(
                        "Observacao",
                        key=f"acao_obs_roubo_{key_base}",
                        placeholder="Detalhes importantes para acompanhamento",
                        height=80,
                    ).strip()
            elif acao_linha in {"Enviar aparelho para manutenção"} or (
                acao_linha == "Mover para outro setor" and destino_setor == "Manutenção"
            ):
                f_manut1, f_manut2 = st.columns([1, 1.2])
                with f_manut1:
                    data_troca_operacao = _format_action_date(
                        st.date_input(
                            "Data da troca",
                            value=pd.Timestamp.today().date(),
                            key=f"acao_data_troca_{key_base}",
                        )
                    )
                    data_retorno_operacao = _format_action_date(
                        st.date_input(
                            "Data retorno prevista",
                            value=None,
                            key=f"acao_data_retorno_{key_base}",
                        )
                    )
                with f_manut2:
                    motivo_operacao = st.text_input(
                        "Motivo da manutencao",
                        key=f"acao_motivo_manut_{key_base}",
                        placeholder="Ex.: Defeito no aparelho, troca preventiva",
                    ).strip()
                    observacao_operacao = st.text_area(
                        "Observacao",
                        key=f"acao_obs_manut_{key_base}",
                        placeholder="Informacoes para a equipe de manutencao",
                        height=80,
                    ).strip()
            elif acao_linha in {"Desativar linha", "Ativar linha"}:
                f_status1, f_status2 = st.columns([1, 1.2])
                with f_status1:
                    motivo_operacao = st.text_input(
                        "Motivo operacional",
                        key=f"acao_motivo_status_{key_base}",
                        placeholder="Ex.: Desligamento, substituicao, reativacao",
                    ).strip()
                with f_status2:
                    observacao_operacao = st.text_area(
                        "Observacao",
                        key=f"acao_obs_status_{key_base}",
                        placeholder="Contexto opcional da alteracao de status",
                        height=80,
                    ).strip()

            requires_confirmation = (
                acao_linha in {"Excluir linha", "Enviar aparelho para manutenção", "Desativar linha", "Ativar linha"}
                or (acao_linha == "Mover para outro setor" and destino_setor in {"Manutenção", "Roubo e Perda"})
            )
            action_confirmed = True
            if requires_confirmation:
                confirmation_labels = {
                    "Excluir linha": "Confirmo a exclusao desta linha.",
                    "Enviar aparelho para manutenção": "Confirmo o envio para manutencao e vou revisar a linha original depois.",
                    "Desativar linha": "Confirmo que esta linha deve sair das ativas.",
                    "Ativar linha": "Confirmo que esta linha deve voltar para as ativas.",
                }
                confirm_label = confirmation_labels.get(
                    acao_linha,
                    "Confirmo a movimentacao desta linha para um fluxo operacional sensivel.",
                )
                action_confirmed = st.checkbox(
                    confirm_label,
                    key=f"acao_confirm_{key_base}_{normalize_text(acao_linha).replace(' ', '_')}",
                )

            if linhas_selecionadas_editor:
                st.caption(f"{len(linhas_selecionadas_editor)} linha(s) selecionada(s) no editor.")

            mover_lote_disponivel = acao_linha in {"Mover para outra equipe", "Mover para outro setor"}
            if mover_lote_disponivel and st.button("Mover linhas selecionadas", key=f"acao_move_selected_{key_base}"):
                if not linhas_selecionadas_editor:
                    st.warning("Selecione ao menos uma linha no checkbox 'Selecionar'.")
                    return
                if acao_linha == "Mover para outra equipe" and (not destino_equipe or destino_equipe == eq_nome):
                    st.warning("Selecione uma equipe de destino diferente da atual.")
                    return
                if acao_linha == "Mover para outro setor":
                    if not destino_setor or not destino_equipe:
                        st.warning("Selecione o setor e a equipe de destino.")
                        return
                    if destino_setor == "Roubo e Perda" and not motivo_operacao:
                        st.warning("Informe o motivo para mover a linha para Roubo e Perda.")
                        return
                if requires_confirmation and not action_confirmed:
                    st.warning("Confirme a operacao para continuar.")
                    return
                try:
                    modo_origem = modo_db
                    df_origem = load_linhas(modo=modo_origem)
                    mask_lote = df_origem["Linha"].fillna("").astype(str).isin(linhas_selecionadas_editor)
                    if not mask_lote.any():
                        st.error("Nenhuma linha selecionada foi encontrada no modo atual.")
                        return
                    before_rows = df_origem[mask_lote].copy()

                    if acao_linha == "Mover para outra equipe":
                        df_origem.loc[mask_lote, "EquipePadrao"] = destino_equipe
                        df_origem.loc[mask_lote, "Equipe"] = destino_equipe
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="mover_equipe_lote",
                            entidade="linhas",
                            chave=f"{eq_nome}:{len(before_rows)}",
                            antes={"linhas": before_rows["Linha"].fillna("").astype(str).tolist()},
                            depois={"EquipePadrao": destino_equipe, "Equipe": destino_equipe, "quantidade": int(len(before_rows)), "modo": modo_origem},
                            detalhes=f"Movimentacao em lote para equipe {destino_equipe}",
                        )
                        try:
                            if active_chamado_context.get("chamado_id"):
                                for _, row_mv in before_rows.iterrows():
                                    linha_mv = str(row_mv.get("Linha", "")).strip()
                                    if not linha_mv:
                                        continue
                                    registrar_movimentacao_linha(
                                        chamado_id=str(active_chamado_context.get("chamado_id")),
                                        linha_num=linha_mv,
                                        tipo_movimentacao="mover_equipe",
                                        antes=row_mv.to_dict(),
                                        depois={"EquipePadrao": destino_equipe, "Equipe": destino_equipe},
                                        motivo="",
                                        observacao=f"Origem equipe: {eq_nome} (lote)",
                                        user_id=str(st.session_state.get("user", {}).get("id", "")),
                                    )
                        except Exception:
                            pass
                        _set_post_chamado_banner("Linhas movidas com sucesso.")
                        st.session_state.pending_equipe = destino_equipe
                        st.success(f"{len(before_rows)} linha(s) movida(s) para equipe {destino_equipe}.")
                        st.rerun()

                    elif acao_linha == "Mover para outro setor":
                        df_origem.loc[mask_lote, "Segmento"] = destino_setor
                        df_origem.loc[mask_lote, "GrupoEquipe"] = destino_setor
                        df_origem.loc[mask_lote, "EquipePadrao"] = destino_equipe
                        df_origem.loc[mask_lote, "Equipe"] = destino_equipe
                        if destino_setor == "Roubo e Perda":
                            df_origem.loc[mask_lote, "Data Ocorrência"] = data_ocorrencia_operacao
                            df_origem.loc[mask_lote, "Motivo"] = motivo_operacao
                            if observacao_operacao:
                                df_origem.loc[mask_lote, "Observação"] = observacao_operacao
                        elif destino_setor == "Manutenção":
                            df_origem.loc[mask_lote, "Data da Troca"] = data_troca_operacao
                            if data_retorno_operacao:
                                df_origem.loc[mask_lote, "Data Retorno"] = data_retorno_operacao
                            if motivo_operacao:
                                df_origem.loc[mask_lote, "Motivo"] = motivo_operacao
                            if observacao_operacao:
                                df_origem.loc[mask_lote, "Observação"] = observacao_operacao
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="mover_setor_lote",
                            entidade="linhas",
                            chave=f"{eq_nome}:{len(before_rows)}",
                            antes={"linhas": before_rows["Linha"].fillna("").astype(str).tolist()},
                            depois={"Segmento": destino_setor, "EquipePadrao": destino_equipe, "quantidade": int(len(before_rows)), "modo": modo_origem},
                            detalhes=f"Movimentacao em lote para setor {destino_setor} / equipe {destino_equipe}",
                        )
                        try:
                            if active_chamado_context.get("chamado_id"):
                                for _, row_mv in before_rows.iterrows():
                                    linha_mv = str(row_mv.get("Linha", "")).strip()
                                    if not linha_mv:
                                        continue
                                    registrar_movimentacao_linha(
                                        chamado_id=str(active_chamado_context.get("chamado_id")),
                                        linha_num=linha_mv,
                                        tipo_movimentacao="mover_setor",
                                        antes=row_mv.to_dict(),
                                        depois={"Segmento": destino_setor, "EquipePadrao": destino_equipe, "Motivo": motivo_operacao, "Observação": observacao_operacao},
                                        motivo=motivo_operacao,
                                        observacao=observacao_operacao,
                                        user_id=str(st.session_state.get("user", {}).get("id", "")),
                                    )
                        except Exception:
                            pass
                        _set_post_chamado_banner("Linhas movidas de setor com sucesso.")
                        st.session_state.pending_segmento = destino_setor
                        st.session_state.pending_equipe = destino_equipe
                        st.success(f"{len(before_rows)} linha(s) movida(s) para setor {destino_setor} / equipe {destino_equipe}.")
                        st.rerun()
                except Exception as exc:
                    st.error(f"Erro ao mover linhas selecionadas: {exc}")

            if st.button("Executar ação", key=f"acao_exec_{key_base}"):
                if not linha_alvo:
                    st.error("Selecione uma linha.")
                    return
                if acao_linha == "Mover para outra equipe" and (not destino_equipe or destino_equipe == eq_nome):
                    st.warning("Selecione uma equipe de destino diferente da atual.")
                    return
                if acao_linha == "Mover para outro setor":
                    segmento_atual = str(linha_selecionada.get("Segmento", "") or "").strip()
                    equipe_atual = str(linha_selecionada.get("EquipePadrao", "") or "").strip()
                    if not destino_setor or not destino_equipe:
                        st.warning("Selecione o setor e a equipe de destino.")
                        return
                    if destino_setor == segmento_atual and destino_equipe == equipe_atual:
                        st.warning("Selecione um destino diferente do setor e equipe atuais.")
                        return
                    if destino_setor == "Roubo e Perda" and not motivo_operacao:
                        st.warning("Informe o motivo para mover a linha para Roubo e Perda.")
                        return
                if acao_linha == "Enviar aparelho para manutenção" and not motivo_operacao:
                    st.warning("Informe o motivo da manutencao antes de continuar.")
                    return
                if acao_linha == "Desativar linha" and not motivo_operacao:
                    st.warning("Informe o motivo operacional da desativacao.")
                    return
                if requires_confirmation and not action_confirmed:
                    st.warning("Confirme a operacao para continuar.")
                    return
                try:
                    modo_origem = modo_db
                    df_origem = load_linhas(modo=modo_origem)
                    mask_origem = df_origem["Linha"].fillna("").astype(str) == linha_alvo
                    if not mask_origem.any():
                        st.error("Linha não encontrada no modo atual.")
                        return
                    before_row = df_origem[mask_origem].iloc[0].to_dict()

                    if acao_linha == "Excluir linha":
                        try:
                            if active_chamado_context.get("chamado_id"):
                                # Registra movimentação antes de remover a linha do banco,
                                # para respeitar a FK de `movimentacoes_linha.linha_id`.
                                registrar_movimentacao_linha(
                                    chamado_id=str(active_chamado_context.get("chamado_id")),
                                    linha_num=linha_alvo,
                                    tipo_movimentacao="excluir_linha",
                                    antes=before_row,
                                    depois=None,
                                    motivo="",
                                    observacao=f"Exclusão na equipe {eq_nome}",
                                    user_id=str(st.session_state.get("user", {}).get("id", "")),
                                )
                        except Exception:
                            pass
                        df_origem = df_origem[~mask_origem].copy()
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="excluir_linha",
                            entidade="linhas",
                            chave=linha_alvo,
                            antes=before_row,
                            detalhes=f"Exclusão na equipe {eq_nome}",
                        )
                        _set_post_chamado_banner("Linha excluída com sucesso.")
                        st.success("Linha excluída com sucesso.")
                        st.rerun()

                    elif acao_linha == "Mover para outra equipe":
                        df_origem.loc[mask_origem, "EquipePadrao"] = destino_equipe
                        df_origem.loc[mask_origem, "Equipe"] = destino_equipe
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="mover_equipe",
                            entidade="linhas",
                            chave=linha_alvo,
                            antes=before_row,
                            depois={
                                "EquipePadrao": destino_equipe,
                                "Equipe": destino_equipe,
                                "modo": modo_origem,
                            },
                            detalhes=f"Origem equipe: {eq_nome}",
                        )
                        try:
                            if active_chamado_context.get("chamado_id"):
                                registrar_movimentacao_linha(
                                    chamado_id=str(active_chamado_context.get("chamado_id")),
                                    linha_num=linha_alvo,
                                    tipo_movimentacao="mover_equipe",
                                    antes=before_row,
                                    depois={
                                        "EquipePadrao": destino_equipe,
                                        "Equipe": destino_equipe,
                                    },
                                    motivo="",
                                    observacao=f"Origem equipe: {eq_nome}",
                                    user_id=str(st.session_state.get("user", {}).get("id", "")),
                                )
                        except Exception:
                            pass
                        _set_post_chamado_banner("Linha movida com sucesso.")
                        st.session_state.pending_equipe = destino_equipe
                        st.success(f"Linha movida para equipe {destino_equipe}.")
                        st.rerun()

                    elif acao_linha == "Mover para outro setor":
                        df_origem.loc[mask_origem, "Segmento"] = destino_setor
                        df_origem.loc[mask_origem, "GrupoEquipe"] = destino_setor
                        df_origem.loc[mask_origem, "EquipePadrao"] = destino_equipe
                        df_origem.loc[mask_origem, "Equipe"] = destino_equipe
                        if destino_setor == "Roubo e Perda":
                            df_origem.loc[mask_origem, "Data Ocorrência"] = data_ocorrencia_operacao
                            df_origem.loc[mask_origem, "Motivo"] = motivo_operacao
                            if observacao_operacao:
                                df_origem.loc[mask_origem, "Observação"] = observacao_operacao
                        elif destino_setor == "Manutenção":
                            df_origem.loc[mask_origem, "Data da Troca"] = data_troca_operacao
                            if data_retorno_operacao:
                                df_origem.loc[mask_origem, "Data Retorno"] = data_retorno_operacao
                            if motivo_operacao:
                                df_origem.loc[mask_origem, "Motivo"] = motivo_operacao
                            if observacao_operacao:
                                df_origem.loc[mask_origem, "Observação"] = observacao_operacao
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="mover_setor",
                            entidade="linhas",
                            chave=linha_alvo,
                            antes=before_row,
                            depois={
                                "Segmento": destino_setor,
                                "GrupoEquipe": destino_setor,
                                "EquipePadrao": destino_equipe,
                                "Equipe": destino_equipe,
                                "Data Ocorrência": data_ocorrencia_operacao if destino_setor == "Roubo e Perda" else "",
                                "Data da Troca": data_troca_operacao if destino_setor == "Manutenção" else "",
                                "Data Retorno": data_retorno_operacao if destino_setor == "Manutenção" else "",
                                "Motivo": motivo_operacao,
                                "Observação": observacao_operacao,
                                "modo": modo_origem,
                            },
                        )
                        try:
                            if active_chamado_context.get("chamado_id"):
                                registrar_movimentacao_linha(
                                    chamado_id=str(active_chamado_context.get("chamado_id")),
                                    linha_num=linha_alvo,
                                    tipo_movimentacao="mover_setor",
                                    antes=before_row,
                                    depois={
                                        "Segmento": destino_setor,
                                        "EquipePadrao": destino_equipe,
                                        "Motivo": motivo_operacao,
                                        "Observação": observacao_operacao,
                                    },
                                    motivo=motivo_operacao,
                                    observacao=observacao_operacao,
                                    user_id=str(st.session_state.get("user", {}).get("id", "")),
                                )
                        except Exception:
                            pass
                        _set_post_chamado_banner("Linha movida de setor com sucesso.")
                        st.session_state.pending_segmento = destino_setor
                        st.session_state.pending_equipe = destino_equipe
                        st.success(f"Linha movida para setor {destino_setor} / equipe {destino_equipe}.")
                        st.rerun()

                    elif acao_linha == "Enviar aparelho para manutenção":
                        linha_origem = df_origem[mask_origem].iloc[0].copy()
                        linha_manut = linha_origem.copy()
                        linha_manut["Segmento"] = "Manutenção"
                        linha_manut["GrupoEquipe"] = "Manutenção"
                        linha_manut["EquipePadrao"] = "Manutenção"
                        linha_manut["Equipe"] = "Manutenção"
                        linha_manut["Linha"] = f"MANUT-{secrets.token_hex(3).upper()}"
                        linha_manut["Nome"] = f"{str(linha_origem.get('Nome', '')).strip()} (Manutenção)".strip()
                        base_ger = str(linha_manut.get("Gerenciamento", "") or "").strip()
                        linha_manut["Gerenciamento"] = f"{base_ger} | Origem: {str(linha_origem.get('Linha', '')).strip()}".strip(" |")
                        linha_manut["Data da Troca"] = data_troca_operacao
                        linha_manut["Data Retorno"] = data_retorno_operacao
                        linha_manut["Motivo"] = motivo_operacao
                        if observacao_operacao:
                            linha_manut["Observação"] = observacao_operacao
                        df_origem = pd.concat([df_origem, pd.DataFrame([linha_manut])], ignore_index=True)
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="enviar_manutencao",
                            entidade="linhas",
                            chave=str(linha_manut["Linha"]),
                            antes=before_row,
                            depois=linha_manut.to_dict(),
                            detalhes="Criada cópia para manutenção",
                        )
                        try:
                            if active_chamado_context.get("chamado_id"):
                                registrar_movimentacao_linha(
                                    chamado_id=str(active_chamado_context.get("chamado_id")),
                                    linha_num=str(linha_manut["Linha"]),
                                    tipo_movimentacao="enviar_manutencao",
                                    antes=before_row,
                                    depois=linha_manut.to_dict(),
                                    motivo=motivo_operacao,
                                    observacao=observacao_operacao,
                                    user_id=str(st.session_state.get("user", {}).get("id", "")),
                                )
                        except Exception:
                            pass
                        _set_post_chamado_banner("Envio para manutenção registrado com sucesso.")
                        st.session_state.pending_segmento = "Manutenção"
                        st.session_state.pending_equipe = "Manutenção"
                        st.session_state.pending_linha = linha_manut["Linha"]
                        st.success("Aparelho enviado para Manutenção. Atualize a linha original com o novo aparelho.")
                        st.rerun()

                    elif acao_linha in ("Desativar linha", "Ativar linha"):
                        destino_modo = "desativadas" if acao_linha == "Desativar linha" else "ativas"
                        if destino_modo == modo_origem:
                            st.info("A linha já está neste modo.")
                            return
                        df_dest = load_linhas(modo=destino_modo)
                        movidas = df_origem[mask_origem].copy()
                        df_origem = df_origem[~mask_origem].copy()
                        linhas_movidas = set(movidas["Linha"].fillna("").astype(str))
                        if not df_dest.empty:
                            mask_rm = ~df_dest["Linha"].fillna("").astype(str).isin(linhas_movidas)
                            df_dest = df_dest[mask_rm].copy()
                        df_dest = pd.concat([df_dest, movidas], ignore_index=True)
                        if motivo_operacao:
                            df_dest.loc[df_dest["Linha"].fillna("").astype(str) == linha_alvo, "Motivo"] = motivo_operacao
                        if observacao_operacao:
                            df_dest.loc[df_dest["Linha"].fillna("").astype(str) == linha_alvo, "Observação"] = observacao_operacao
                        save_linhas(df_origem, modo=modo_origem)
                        save_linhas(df_dest, modo=destino_modo)
                        _audit(
                            acao="mudar_modo_linha",
                            entidade="linhas",
                            chave=linha_alvo,
                            antes=before_row,
                            depois={**before_row, "modo": destino_modo, "Motivo": motivo_operacao, "Observação": observacao_operacao},
                        )
                        try:
                            if active_chamado_context.get("chamado_id"):
                                registrar_movimentacao_linha(
                                    chamado_id=str(active_chamado_context.get("chamado_id")),
                                    linha_num=linha_alvo,
                                    tipo_movimentacao="mudar_modo_linha",
                                    antes=before_row,
                                    depois={"modo": destino_modo, "Motivo": motivo_operacao, "Observação": observacao_operacao},
                                    motivo=motivo_operacao,
                                    observacao=observacao_operacao,
                                    user_id=str(st.session_state.get("user", {}).get("id", "")),
                                )
                        except Exception:
                            pass
                        _set_post_chamado_banner("Alteração de status da linha concluída com sucesso.")
                        st.success(f"Linha movida para {destino_modo}.")
                        st.rerun()
                except Exception as exc:
                    st.error(f"Erro ao executar ação: {exc}")

    all_edited: list[pd.DataFrame] = []
    all_base_views: list[pd.DataFrame] = []

    if segmento_sel == "Manutenção":
        # Remove coluna Histórico – usamos seções separadas em vez disso
        tbl = tbl.drop(columns=["Histórico", "__manut_antigo"], errors="ignore")
        tbl_novos, tbl_antigos = _split_manut_novos_antigos(tbl)
        for block_name, block_df, is_novos in [
            ("Manutenção", tbl_novos, True),
            ("Manutenções antigas", tbl_antigos, False),
        ]:
            dfe = _sort_manut_block(block_df, is_novos)
            header_text = f'<div class="{header_cls}">📋 {block_name}</div>'
            st.markdown(header_text, unsafe_allow_html=True)
            if usar_editor:
                all_base_views.append(dfe.copy())
                dfe_editor = _decorate_editor_df(dfe, manut_antigos=not is_novos)
                ed_key = "ed_manut_novos" if is_novos else "ed_manut_antigos"
                col_config = {
                    "__row_key": None,
                    "__base_linha": None,
                    "↕ Ordem": st.column_config.NumberColumn("↕ Ordem", min_value=1, step=1, format="%d"),
                    "Mover equipe": st.column_config.SelectboxColumn("Mover equipe", options=[""] + equipes_mover_manut),
                }
                if not is_novos and "↕ Ordem" not in dfe_editor.columns:
                    col_config.pop("↕ Ordem", None)
                ed = st.data_editor(
                    dfe_editor,
                    width="stretch",
                    num_rows="fixed",
                    hide_index=True,
                    key=ed_key,
                    column_config=col_config,
                    disabled=["__row_key", "__base_linha", "Conflito"],
                )
                all_edited.append(ed)
            else:
                st.markdown(_render_equipe_tabela(dfe, block_name), unsafe_allow_html=True)
    elif equipe_sel == "Todas as equipes":
        eq_order = [e for e in equipes if e in tbl["EquipePadrao"].unique()]
        eq_order += [e for e in sorted(tbl["EquipePadrao"].unique()) if e not in eq_order]
        for eq in eq_order:
            dfe = _sort_for_display(tbl[tbl["EquipePadrao"] == eq])
            dfe = _apply_manual_team_order(dfe, eq)
            if segmento_sel in ("Internos", "Roubo e Perda"):
                header_text = f'<div class="{header_cls}">📋 {eq}</div>'
            elif eq in ("Gerentes do Alimento", "Gerentes do Medicamento"):
                gestor = "—"
                supervisor = "—"
                header_text = f'<div class="{header_cls}">📋 {eq} — <strong>Gerente:</strong> {gestor}  |  <strong>Supervisor:</strong> {supervisor}</div>'
            else:
                gestor = non_empty_or_default(dfe["Gestor"].iloc[0], "—")
                supervisor = _supervisor_display(dfe, eq, rules_path)
                header_text = f'<div class="{header_cls}">📋 {eq} — <strong>Gerente:</strong> {gestor}  |  <strong>Supervisor:</strong> {supervisor}</div>'
            st.markdown(header_text, unsafe_allow_html=True)
            if usar_editor:
                all_base_views.append(dfe.copy())
                dfe_editor = _decorate_editor_df(dfe)
                ed = st.data_editor(
                    dfe_editor,
                    width="stretch",
                    num_rows="fixed",
                    hide_index=True,
                    key=f"ed_{eq}",
                    column_config={
                        "__row_key": None,
                        "__base_linha": None,
                        "↕ Ordem": st.column_config.NumberColumn("↕ Ordem", min_value=1, step=1, format="%d"),
                        "Mover equipe": st.column_config.SelectboxColumn("Mover equipe", options=[""] + equipes),
                    },
                    disabled=["__row_key", "__base_linha", "Conflito"],
                )
                all_edited.append(ed)
            else:
                st.markdown(_render_equipe_tabela(dfe, eq), unsafe_allow_html=True)
    else:
        dfe = _sort_for_display(tbl)
        dfe = _apply_manual_team_order(dfe, equipe_sel)
        if segmento_sel in ("Internos", "Roubo e Perda"):
            header_text = f'<div class="{header_cls}">📋 {equipe_sel}</div>'
        elif equipe_sel in ("Gerentes do Alimento", "Gerentes do Medicamento"):
            gestor = "—"
            supervisor = "—"
            header_text = f'<div class="{header_cls}">📋 {equipe_sel} — <strong>Gerente:</strong> {gestor}  |  <strong>Supervisor:</strong> {supervisor}</div>'
        else:
            gestor = non_empty_or_default(dfe["Gestor"].iloc[0], "—")
            supervisor = _supervisor_display(dfe, equipe_sel, rules_path)
            header_text = f'<div class="{header_cls}">📋 {equipe_sel} — <strong>Gerente:</strong> {gestor}  |  <strong>Supervisor:</strong> {supervisor}</div>'
        st.markdown(header_text, unsafe_allow_html=True)
        if usar_editor:
            all_base_views.append(dfe.copy())
            dfe_editor = _decorate_editor_df(dfe)
            ed = st.data_editor(
                dfe_editor,
                width="stretch",
                num_rows="fixed",
                hide_index=True,
                key=f"ed_{equipe_sel}",
                column_config={
                    "__row_key": None,
                    "__base_linha": None,
                    "↕ Ordem": st.column_config.NumberColumn("↕ Ordem", min_value=1, step=1, format="%d"),
                    "Mover equipe": st.column_config.SelectboxColumn("Mover equipe", options=[""] + equipes),
                },
                disabled=["__row_key", "__base_linha", "Conflito"],
            )
            all_edited.append(ed)
        else:
            st.markdown(_render_equipe_tabela(dfe, equipe_sel), unsafe_allow_html=True)

    if usar_editor and all_base_views:
        stored_editor_context = st.session_state.get("_editor_base_context")
        if stored_editor_context != editor_context_key:
            base_snapshot_df = pd.concat(all_base_views, ignore_index=True)
            st.session_state["_editor_base_context"] = editor_context_key
            st.session_state["_editor_base_snapshot"] = base_snapshot_df.to_dict("records")
            st.session_state["_editor_conflict_payload"] = {}

        conflict_payload = st.session_state.get("_editor_conflict_payload") or {}
        reapply_request = st.session_state.get("_reapply_conflict_request")
        if conflict_payload and conflict_payload.get("context") == editor_context_key and reapply_request == editor_context_key:
            pending_changes = conflict_payload.get("pending_changes") or []
            st.session_state["_reapply_conflict_request"] = ""
            if pending_changes:
                seg_lower_reapply = segmento_sel.strip().lower()
                full_current = df_full_mode.copy()
                mask_seg_reapply = full_current["Segmento"].fillna("").astype(str).str.strip().str.lower() == seg_lower_reapply
                segment_current = full_current[mask_seg_reapply].copy()
                others_current = full_current[~mask_seg_reapply].copy()
                grouped_changes: dict[str, list[dict[str, str]]] = {}
                for ch in pending_changes:
                    base_linha = str(ch.get("base_linha") or "").strip()
                    if not base_linha:
                        continue
                    grouped_changes.setdefault(base_linha, []).append(ch)

                def _norm_cmp(v: Any) -> str:
                    if v is None:
                        return ""
                    if isinstance(v, float) and pd.isna(v):
                        return ""
                    return str(v).strip()

                applied_fields = 0
                applied_rows: set[str] = set()
                remaining_pending: list[dict[str, str]] = []
                remaining_resumo: list[dict[str, str]] = []
                remaining_detalhes: list[dict[str, str]] = []

                if "Linha" in segment_current.columns:
                    segment_current["__base_linha_norm"] = segment_current["Linha"].fillna("").astype(str).str.strip()

                for base_linha, changes in grouped_changes.items():
                    if "__base_linha_norm" not in segment_current.columns:
                        remaining_resumo.append({"Linha": base_linha or "—", "Motivo": "Nao foi possivel comparar a linha no banco."})
                        remaining_pending.extend(changes)
                        continue
                    current_matches = segment_current[segment_current["__base_linha_norm"] == base_linha]
                    if current_matches.empty:
                        remaining_resumo.append({"Linha": base_linha or "—", "Motivo": "A linha nao foi encontrada no banco na versao atual."})
                        remaining_pending.extend(changes)
                        continue
                    row_idx = current_matches.index[0]
                    row_remaining_fields: list[str] = []
                    for ch in changes:
                        campo = str(ch.get("campo") or "").strip()
                        valor_original = _norm_cmp(ch.get("original"))
                        minha_edicao = _norm_cmp(ch.get("mine"))
                        if not campo or campo not in segment_current.columns:
                            remaining_pending.append(ch)
                            row_remaining_fields.append(campo or "—")
                            continue
                        valor_atual = _norm_cmp(segment_current.at[row_idx, campo])
                        if valor_atual == valor_original or valor_atual == minha_edicao:
                            if valor_atual != minha_edicao:
                                segment_current.at[row_idx, campo] = minha_edicao
                            applied_fields += 1
                            applied_rows.add(base_linha)
                        else:
                            remaining_pending.append(ch)
                            row_remaining_fields.append(campo)
                            remaining_detalhes.append(
                                {
                                    "Linha": base_linha or "—",
                                    "Campo": campo,
                                    "Valor original": valor_original or "—",
                                    "Minha edição": minha_edicao or "—",
                                    "Valor atual no banco": valor_atual or "—",
                                }
                            )
                    if row_remaining_fields:
                        remaining_resumo.append(
                            {
                                "Linha": base_linha or "—",
                                "Motivo": "Alguns campos continuam em conflito.",
                                "Campos com diferenca": ", ".join(row_remaining_fields[:6]) + (f" (+{len(row_remaining_fields) - 6})" if len(row_remaining_fields) > 6 else ""),
                            }
                        )

                if applied_fields > 0:
                    segment_current = segment_current.drop(columns=["__base_linha_norm"], errors="ignore")
                    updated_full = pd.concat([others_current, segment_current], ignore_index=True)
                    updated_full = updated_full.drop(columns=["__row_key", "__base_linha"], errors="ignore")
                    sort_cols_save = [c for c in ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"] if c in updated_full.columns]
                    if sort_cols_save:
                        updated_full = updated_full.sort_values(sort_cols_save, kind="stable")
                    n_reapply = save_linhas(updated_full, modo=modo_db)
                    _audit(
                        acao="reaplicar_edicoes_seguras",
                        entidade="linhas",
                        chave=modo_db,
                        antes={
                            "linhas_reaplicadas": sorted(applied_rows),
                            "campos_reaplicados": int(applied_fields),
                        },
                        depois={
                            "registros_gravados_modo": int(n_reapply),
                            "segmento": segmento_sel,
                            "modo": modo_db,
                        },
                        detalhes=f"Reaplicacao segura executada em {len(applied_rows)} linha(s).",
                    )

                if remaining_pending:
                    st.session_state["_editor_conflict_payload"] = {
                        "context": editor_context_key,
                        "resumo": remaining_resumo,
                        "detalhes": remaining_detalhes,
                        "pending_changes": remaining_pending,
                    }
                else:
                    st.session_state["_editor_conflict_payload"] = {}

                st.session_state["_editor_base_context"] = ""
                st.session_state["_editor_base_snapshot"] = []

                if applied_fields > 0 and remaining_pending:
                    st.session_state["_post_conflict_feedback"] = {
                        "type": "warning",
                        "text": f"Algumas alteracoes seguras foram reaplicadas ({applied_fields} campo(s)), mas ainda existem conflitos pendentes.",
                    }
                elif applied_fields > 0:
                    st.session_state["_post_conflict_feedback"] = {
                        "type": "success",
                        "text": f"Reaplicacao segura concluida com sucesso. {applied_fields} campo(s) reaplicado(s).",
                    }
                else:
                    st.session_state["_post_conflict_feedback"] = {
                        "type": "warning",
                        "text": "Nenhuma alteracao segura pode ser reaplicada automaticamente.",
                    }
                st.rerun()

    live_changed_rows_df = pd.DataFrame()
    live_validation_errors: list[dict[str, str]] = []
    if usar_editor and all_base_views and all_edited:
        live_changed_rows_df = collect_changed_editor_rows(all_base_views, all_edited)
        if not live_changed_rows_df.empty:
            live_candidate_df = build_candidate_full_df(
                df_full_mode,
                live_changed_rows_df.drop(columns=["Conflito", "Validacao", "↕ Ordem", "Mover equipe"], errors="ignore"),
            )
            live_validation_errors = collect_editor_validation_errors(
                live_changed_rows_df.drop(columns=["Conflito", "Validacao", "__row_key", "__base_linha", "↕ Ordem", "Mover equipe"], errors="ignore"),
                live_candidate_df.drop(columns=["Conflito", "Validacao", "__base_linha", "↕ Ordem", "Mover equipe"], errors="ignore"),
            )
            if live_validation_errors:
                linhas_com_pendencia = sorted(
                    {
                        str(item.get("Linha") or "").strip()
                        for item in live_validation_errors
                        if str(item.get("Linha") or "").strip()
                    }
                )
                st.warning(
                    f"Existem pendencias de validacao em {len(linhas_com_pendencia)} linha(s) alterada(s). Corrija antes de salvar."
                )
                with st.expander("Ver pendencias de validacao", expanded=False):
                    st.dataframe(pd.DataFrame(live_validation_errors), width="stretch", hide_index=True)

    if usar_editor and all_edited:
        if save_clicked_top:
            seg_lower = segmento_sel.strip().lower()
            mask_seg = df_full_mode["Segmento"].fillna("").astype(str).str.strip().str.lower() == seg_lower
            outros = df_full_mode[~mask_seg].copy()
            segment_atual = df_full_mode[mask_seg].copy()
            edited_df = pd.concat(all_edited, ignore_index=True)
            edited_df = edited_df.drop(columns=["Conflito"], errors="ignore")

            has_reorder_action = "↕ Ordem" in edited_df.columns
            has_team_move_action = False
            if "Mover equipe" in edited_df.columns:
                has_team_move_action = bool(edited_df["Mover equipe"].fillna("").astype(str).str.strip().ne("").any())
            if "Mover equipe" in edited_df.columns:
                dest_series = edited_df["Mover equipe"].fillna("").astype(str).str.strip()
                mask_move_team = dest_series != ""
                if segmento_sel == "Manutenção":
                    # No segmento Manutenção, "Mover equipe" = mover entre seções (Data da Troca)
                    today_str = datetime.now().strftime("%Y-%m-%d")
                    yesterday_str = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
                    mask_para_antigos = mask_move_team & (dest_series == "Manutenções antigas")
                    mask_para_novos = mask_move_team & (dest_series == "Manutenção")
                    if "Data da Troca" in edited_df.columns:
                        edited_df.loc[mask_para_antigos, "Data da Troca"] = yesterday_str
                        edited_df.loc[mask_para_novos, "Data da Troca"] = today_str
                    # EquipePadrao permanece "Manutenção" (não alterar para equipes externas)
                else:
                    if "EquipePadrao" in edited_df.columns:
                        edited_df.loc[mask_move_team, "EquipePadrao"] = dest_series[mask_move_team]
                    if "Equipe" in edited_df.columns:
                        edited_df.loc[mask_move_team, "Equipe"] = dest_series[mask_move_team]
            if has_reorder_action and "EquipePadrao" in edited_df.columns:
                ordem_num = pd.to_numeric(edited_df["↕ Ordem"], errors="coerce")
                edited_df["__ordem_manual"] = ordem_num.fillna(10**9)
                edited_df["__ordem_idx"] = range(len(edited_df))
                edited_df = edited_df.sort_values(
                    ["EquipePadrao", "__ordem_manual", "__ordem_idx"],
                    kind="stable",
                ).drop(columns=["__ordem_manual", "__ordem_idx"], errors="ignore")
                # Renumeração automática da coluna de ordem por equipe (1..N).
                if "↕ Ordem" in edited_df.columns:
                    edited_df["↕ Ordem"] = (
                        edited_df.groupby("EquipePadrao", sort=False).cumcount() + 1
                    )
            elif "↕ Ordem" in edited_df.columns:
                # Mesmo sem mudança manual de ordem, garante sequência válida na equipe.
                edited_df["↕ Ordem"] = (
                    edited_df.groupby("EquipePadrao", sort=False).cumcount() + 1
                )
                # Persiste ordem manual em sessão para exibição consistente.
                if "Linha" in edited_df.columns:
                    manual_map = st.session_state.get("_manual_order_map") or {}
                    for eq_name, grp in edited_df.groupby("EquipePadrao", sort=False):
                        key = f"{modo_db}|{segmento_sel}|{str(eq_name).strip()}"
                        linhas_ord = [
                            str(v).strip()
                            for v in grp["Linha"].fillna("").astype(str).tolist()
                            if str(v).strip()
                        ]
                        manual_map[key] = linhas_ord
                    st.session_state["_manual_order_map"] = manual_map
            row_keys_editadas = set(edited_df["__row_key"].dropna().astype(str).unique()) if "__row_key" in edited_df.columns else set()
            base_snapshot_records = st.session_state.get("_editor_base_snapshot") or []
            base_snapshot_df = pd.DataFrame(base_snapshot_records) if base_snapshot_records else pd.DataFrame()

            def _norm_cmp(v: Any) -> str:
                if v is None:
                    return ""
                if isinstance(v, float) and pd.isna(v):
                    return ""
                return str(v).strip()

            row_keys_alteradas: set[str] = set()
            campos_alterados_set: set[str] = set()
            linhas_editadas_lista: list[str] = []
            celulas_alteradas = 0
            alteracoes_detalhadas: list[dict[str, str]] = []
            max_alteracoes_detalhadas = 200
            conflitos_detectados: list[dict[str, str]] = []
            conflitos_detalhados: list[dict[str, str]] = []
            pending_conflict_changes: list[dict[str, str]] = []

            if row_keys_editadas and "__row_key" in edited_df.columns and "__row_key" in base_snapshot_df.columns:
                base_idx = base_snapshot_df.set_index("__row_key", drop=False)
                edit_idx = edited_df.set_index("__row_key", drop=False)
                cols_compare = [
                    c for c in edited_df.columns
                    if c in base_snapshot_df.columns and c not in ("__row_key", "↕ Ordem", "Mover equipe")
                ]
                for rk in row_keys_editadas:
                    if rk not in base_idx.index or rk not in edit_idx.index:
                        continue
                    base_row = base_idx.loc[rk]
                    edit_row = edit_idx.loc[rk]
                    linha_ref = _norm_cmp(edit_row.get("Linha")) or _norm_cmp(base_row.get("Linha")) or str(rk)
                    houve_mudanca = False
                    for col in cols_compare:
                        base_norm = _norm_cmp(base_row.get(col))
                        edit_norm = _norm_cmp(edit_row.get(col))
                        if base_norm != edit_norm:
                            houve_mudanca = True
                            celulas_alteradas += 1
                            campos_alterados_set.add(str(col))
                            if len(alteracoes_detalhadas) < max_alteracoes_detalhadas:
                                alteracoes_detalhadas.append(
                                    {
                                        "linha": linha_ref,
                                        "campo": str(col),
                                        "antes": base_norm or "—",
                                        "depois": edit_norm or "—",
                                    }
                                )
                    if houve_mudanca:
                        row_keys_alteradas.add(str(rk))
                        linhas_editadas_lista.append(linha_ref)

            linhas_editadas_lista = sorted(set(linhas_editadas_lista))
            campos_alterados_lista = sorted(campos_alterados_set)

            if has_reorder_action and not row_keys_alteradas and "__row_key" in edited_df.columns:
                # Reordenação sem alteração de campo: força salvamento da nova ordem.
                row_keys_alteradas = set(edited_df["__row_key"].dropna().astype(str).unique())
                linhas_editadas_lista = sorted(
                    {
                        str(v).strip()
                        for v in edited_df["Linha"].fillna("").astype(str).tolist()
                        if str(v).strip()
                    }
                )
                if not campos_alterados_lista:
                    campos_alterados_lista = ["ordem_linha"]
            if has_team_move_action and "__row_key" in edited_df.columns:
                row_keys_alteradas.update(set(edited_df["__row_key"].dropna().astype(str).unique()))
                if "mover_equipe_inline" not in campos_alterados_lista:
                    campos_alterados_lista.append("mover_equipe_inline")

            if not row_keys_alteradas:
                st.info("Nenhuma alteração detectada para salvar.")
            else:
                latest_segment_idx = segment_atual.copy()
                latest_segment_idx["__base_linha_norm"] = latest_segment_idx["Linha"].fillna("").astype(str).str.strip() if "Linha" in latest_segment_idx.columns else ""
                base_snapshot_idx = base_snapshot_df.set_index("__row_key", drop=False) if "__row_key" in base_snapshot_df.columns else pd.DataFrame()
                edit_idx = edited_df.set_index("__row_key", drop=False) if "__row_key" in edited_df.columns else pd.DataFrame()
                cols_conf_compare = [c for c in base_snapshot_df.columns if c in segment_atual.columns and c not in ("__row_key", "__base_linha")]
                if segmento_sel == "Manutenção" and has_team_move_action:
                    cols_conf_compare = [c for c in cols_conf_compare if c != "Data da Troca"]
                for rk in sorted(row_keys_alteradas):
                    if base_snapshot_idx.empty or rk not in base_snapshot_idx.index:
                        continue
                    if not edit_idx.empty and rk in edit_idx.index:
                        edit_row = edit_idx.loc[rk]
                    else:
                        edit_row = base_snapshot_idx.loc[rk]
                    base_row = base_snapshot_idx.loc[rk]
                    base_linha = _norm_cmp(base_row.get("__base_linha") or base_row.get("Linha"))
                    current_matches = latest_segment_idx[latest_segment_idx["__base_linha_norm"] == base_linha] if base_linha else pd.DataFrame()
                    if current_matches.empty:
                        conflitos_detectados.append(
                            {
                                "Linha": base_linha or "—",
                                "Motivo": "A linha nao foi encontrada no banco na versao atual.",
                            }
                        )
                        continue
                    current_row = current_matches.iloc[0]
                    campos_conflito: list[str] = []
                    for col in cols_conf_compare:
                        if _norm_cmp(base_row.get(col)) != _norm_cmp(current_row.get(col)):
                            campos_conflito.append(str(col))
                            minha_edicao = _norm_cmp(edit_row.get(col))
                            valor_original = _norm_cmp(base_row.get(col)) or "—"
                            valor_banco = _norm_cmp(current_row.get(col)) or "—"
                            pending_conflict_changes.append(
                                {
                                    "base_linha": base_linha or "—",
                                    "campo": str(col),
                                    "original": valor_original,
                                    "mine": minha_edicao or "—",
                                    "current": valor_banco,
                                    "original_raw": _norm_cmp(base_row.get(col)),
                                    "mine_raw": _norm_cmp(edit_row.get(col)),
                                    "current_raw": _norm_cmp(current_row.get(col)),
                                }
                            )
                            conflitos_detalhados.append(
                                {
                                    "Linha": base_linha or "—",
                                    "Campo": str(col),
                                    "Valor original": valor_original,
                                    "Minha edição": minha_edicao or "—",
                                    "Valor atual no banco": valor_banco,
                                }
                            )
                    if campos_conflito:
                        conflitos_detectados.append(
                            {
                                "Linha": base_linha or "—",
                                "Motivo": "Outro usuario alterou esta linha apos a abertura da tela.",
                                "Campos com diferenca": ", ".join(campos_conflito[:6]) + (f" (+{len(campos_conflito) - 6})" if len(campos_conflito) > 6 else ""),
                            }
                        )

                if conflitos_detectados:
                    st.session_state["_editor_conflict_payload"] = {
                        "context": editor_context_key,
                        "resumo": conflitos_detectados,
                        "detalhes": conflitos_detalhados,
                        "pending_changes": pending_conflict_changes,
                    }
                    st.error("Conflito de edicao detectado. Atualize a tela antes de salvar para evitar sobrescrever alteracoes de outra pessoa.")
                    st.dataframe(pd.DataFrame(conflitos_detectados), width="stretch", hide_index=True)
                    if conflitos_detalhados:
                        with st.expander("Ver diferenças do conflito", expanded=False):
                            st.dataframe(pd.DataFrame(conflitos_detalhados), width="stretch", hide_index=True)
                    return

                mask_rows_alteradas = edited_df["__row_key"].fillna("").astype(str).isin(row_keys_alteradas)
                edited_changed_df = edited_df[mask_rows_alteradas].copy()
                if has_reorder_action:
                    novos_seg = edited_df.copy()
                else:
                    mask_manter = ~segment_atual["__row_key"].fillna("").astype(str).isin(row_keys_alteradas)
                    segment_nao_editado = segment_atual[mask_manter].copy()
                    novos_seg = pd.concat([segment_nao_editado, edited_changed_df], ignore_index=True)
                novos = pd.concat([outros, novos_seg], ignore_index=True)
                novos = novos.drop(columns=["__row_key"], errors="ignore")
                control_only_change = celulas_alteradas == 0 and (has_reorder_action or has_team_move_action)
                validation_errors: list[dict[str, str]] = []
                if not control_only_change:
                    validation_errors = collect_editor_validation_errors(
                        edited_changed_df.drop(columns=["__row_key", "__base_linha", "↕ Ordem", "Mover equipe"], errors="ignore"),
                        novos.drop(columns=["__base_linha", "↕ Ordem", "Mover equipe"], errors="ignore"),
                    )
                if validation_errors:
                    st.error("Nao foi possivel salvar porque existem campos invalidos nas linhas alteradas.")
                    st.dataframe(pd.DataFrame(validation_errors), width="stretch", hide_index=True)
                    return
                novos = novos.drop(columns=["Mover equipe"], errors="ignore")
                n = save_linhas(novos, modo=modo_db)
                campos_txt = ", ".join(campos_alterados_lista[:4]) if campos_alterados_lista else "campos não identificados"
                if campos_alterados_lista and len(campos_alterados_lista) > 4:
                    campos_txt += f" (+{len(campos_alterados_lista) - 4})"
                _audit(
                    acao="salvar_edicoes",
                    entidade="linhas",
                    chave=modo_db,
                    antes={
                        "registros_segmento_antes": int(len(segment_atual)),
                        "linhas_editadas": linhas_editadas_lista,
                        "campos_alterados": campos_alterados_lista,
                        "celulas_alteradas": int(celulas_alteradas),
                        "alteracoes_total": int(celulas_alteradas),
                        "alteracoes": alteracoes_detalhadas,
                    },
                    depois={
                        "registros_segmento_depois": int(len(novos_seg)),
                        "registros_gravados_modo": int(n),
                        "segmento": segmento_sel,
                        "modo": modo_db,
                        "campos_alterados": campos_alterados_lista,
                        "alteracoes_total": int(celulas_alteradas),
                        "alteracoes": alteracoes_detalhadas,
                    },
                    detalhes=f"Edição de {len(linhas_editadas_lista)} linha(s). Campos: {campos_txt}",
                )
                _set_post_chamado_banner("Alterações salvas com sucesso.")
                st.session_state["_editor_base_context"] = ""
                st.session_state["_editor_base_snapshot"] = []
                st.session_state["_editor_conflict_payload"] = {}
                st.success(f"Alterações salvas! {len(linhas_editadas_lista)} linha(s) alterada(s).")
                st.rerun()

    st.markdown("---")
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for segmento in ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"]:
            seg_lower = segmento.strip().lower()
            mask = df_full_mode["Segmento"].fillna("").astype(str).str.strip().str.lower() == seg_lower
            df_seg = df_full_mode[mask].drop(columns=["__row_key", "__base_linha"], errors="ignore")
            if not df_seg.empty:
                df_seg.to_excel(writer, sheet_name=segmento, index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    header_fill = PatternFill(start_color="2D5A24", end_color="2D5A24", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    gerente_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    supervisor_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for ws in wb.worksheets:
        header_map = {str(cell.value or "").strip(): col_idx for col_idx, cell in enumerate(ws[1], 1)}
        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center")
            row_num = row[0].row
            papel_col = header_map.get("Papel")
            equipe_col = header_map.get("EquipePadrao")
            is_gerente = False
            is_supervisor = False
            if papel_col:
                papel_val = str(ws.cell(row=row_num, column=papel_col).value or "").strip().lower()
                if papel_val == "gerente":
                    is_gerente = True
                elif papel_val == "supervisor":
                    is_supervisor = True
            if equipe_col:
                eq_val = str(ws.cell(row=row_num, column=equipe_col).value or "").strip()
                if eq_val in ("Gerentes do Alimento", "Gerentes do Medicamento"):
                    is_gerente = True
            row_fill = gerente_fill if is_gerente else (supervisor_fill if is_supervisor else None)
            if row_fill:
                for c in row:
                    c.fill = row_fill
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 14
            for cell in ws[col_letter]:
                try:
                    max_len = max(max_len, min(len(str(cell.value or "")), 50))
                except TypeError:
                    pass
            ws.column_dimensions[col_letter].width = max_len
        ws.freeze_panes = "A2"
    buf_out = io.BytesIO()
    wb.save(buf_out)
    buf_out.seek(0)
    st.download_button(
        "Exportar tudo para Excel",
        data=buf_out,
        file_name="linhas_telefones.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
