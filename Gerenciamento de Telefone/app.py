from __future__ import annotations

import io
import json
import os
import secrets
import base64
import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st

# Keycloak OIDC — único provedor de identidade
_KEYCLOAK_URL = os.environ.get("KEYCLOAK_URL", "http://localhost:8080")
_KEYCLOAK_REALM = os.environ.get("KEYCLOAK_REALM", "prosper")
_KEYCLOAK_CLIENT_ID = os.environ.get("KEYCLOAK_CLIENT_ID_STREAMLIT", "streamlit-app")
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Suporte a banco de dados (opcional)
try:
    from src.db.repository import (
        load_linhas, has_data, get_db_path, save_linhas,
        init_db, criar_usuario, listar_usuarios, excluir_usuario,
        obter_usuario_por_username,
        registrar_auditoria, listar_auditoria,
        garantir_chamado_stub, vincular_chamado_linha, registrar_chamado_evento, registrar_movimentacao_linha,
        criar_chamado, listar_chamados, obter_chamado, atualizar_status_chamado,
    )
    HAS_DB = True
except ImportError:
    HAS_DB = False
    init_db = criar_usuario = listar_usuarios = excluir_usuario = None
    obter_usuario_por_username = None
    registrar_auditoria = listar_auditoria = None
    garantir_chamado_stub = vincular_chamado_linha = registrar_chamado_evento = registrar_movimentacao_linha = None
    criar_chamado = listar_chamados = obter_chamado = atualizar_status_chamado = None


from src.core.config import (
    RULES_FILE, DOC_DIR, is_postgres_configured, get_chamados_app_url,
)
from src.components.sidebar import render_sidebar
ABAS_ALIMENTO = ["Nova Prosper"]
ABAS_MEDICAMENTO = ["Prosper Norte", "Prosper Sul"]
ABAS_FOCO = ["Prosper Norte", "Prosper Sul", "Nova Prosper", "Promotores", "Internos", "Troca de Aparelho", "Devolução Manutenção", "Roubo-Perda"]
ABAS_PROMOTORES = ["Promotores"]
ABAS_INTERNOS = ["Internos"]
ABAS_MANUTENCAO = ["Troca de Aparelho", "Devolução Manutenção", "Devolucao Manutencao"]
ABAS_ROUBO_PERDA = ["Roubo-Perda", "Roubo e Perda"]
EQUIPES_PROMOTORES = ["Promotores"]
EQUIPES_INTERNOS = ["Internos"]
EQUIPES_MANUTENCAO = ["Manutenção"]
EQUIPES_ROUBO_PERDA = ["Roubo e Perda"]
EQUIPES_ALIMENTO = [
    "Gerentes do Alimento",
    "Consumo Baixada",
    "Consumo Oeste",
    "Consumo Zona Norte",
    "Consumo Niteroi",
    "Equipe Especial",
    "Gerente Senior",
]
EQUIPES_MEDICAMENTO = [
    "Gerentes do Medicamento",
    "Prosper Norte",
    "Prosper Sul",
]
GESTORES_MEDICAMENTO = {
    "Prosper Norte": "Priscila Rangel Manhães",
    "Prosper Sul": "Gustavo Luis Dias De Armada",
}
from src.core.config import NEW_COLUMNS_ORDER, HIDDEN_COLUMNS
DEFAULT_COLUMNS = NEW_COLUMNS_ORDER


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn"
    )
    return text.lower()


def title_case_safe(value: str) -> str:
    tokens = [t for t in value.replace("/", " / ").split(" ") if t != ""]
    if not tokens:
        return "Sem Equipe"
    out: list[str] = []
    for token in tokens:
        if token == "/":
            out.append(token)
        elif token.isupper() and len(token) <= 4:
            out.append(token)
        else:
            out.append(token.capitalize())
    return " ".join(out).replace(" / ", "/")


def normalize_team_key(value: Any) -> str:
    text = normalize_text(value)
    text = " ".join(text.split())
    return text


def digits_only(value: Any) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value) if ch.isdigit())


def is_valid_phone(value: Any) -> bool:
    digits = digits_only(value)
    return 10 <= len(digits) <= 13


def build_vaga_mask(df: pd.DataFrame) -> pd.Series:
    """Linha vaga = tem linha e nao tem usuario cadastrado ou esta marcada como VAGO."""
    if df.empty:
        return pd.Series(dtype=bool)
    linhas_preenchidas = (
        df["Linha"].fillna("").astype(str).str.strip() != ""
        if "Linha" in df.columns
        else pd.Series(False, index=df.index)
    )
    nomes_norm = (
        df["Nome"].fillna("").astype(str).map(normalize_text)
        if "Nome" in df.columns
        else pd.Series("", index=df.index)
    )
    sem_usuario = nomes_norm.isin(["", "vago"])
    return linhas_preenchidas & sem_usuario


def is_valid_email(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", text))


def is_valid_imei(value: Any) -> bool:
    digits = digits_only(value)
    if not digits:
        return True
    return len(digits) == 15


def is_temporary_line_identifier(value: Any) -> bool:
    text = str(value or "").strip().upper()
    return text.startswith("NOVA-") or text.startswith("MANUT-")


def collect_editor_validation_errors(changed_rows_df: pd.DataFrame, full_df: pd.DataFrame) -> list[dict[str, str]]:
    """Valida campos criticos apenas nas linhas alteradas antes de salvar."""
    if changed_rows_df.empty:
        return []

    errors: list[dict[str, str]] = []
    duplicate_counts: dict[str, int] = {}

    if "Linha" in full_df.columns:
        full_lines = full_df["Linha"].fillna("").astype(str).str.strip()
        valid_lines = [
            line
            for line in full_lines.tolist()
            if line and not is_temporary_line_identifier(line)
        ]
        duplicate_counts = pd.Series(valid_lines).value_counts().to_dict() if valid_lines else {}

    for _, row in changed_rows_df.iterrows():
        linha_raw = str(row.get("Linha") or "").strip()
        linha_label = linha_raw or "(sem linha)"

        if not linha_raw:
            errors.append(
                {
                    "Linha": linha_label,
                    "Campo": "Linha",
                    "Problema": "Preencha a linha antes de salvar.",
                }
            )
        elif not is_temporary_line_identifier(linha_raw) and not is_valid_phone(linha_raw):
            errors.append(
                {
                    "Linha": linha_label,
                    "Campo": "Linha",
                    "Problema": "Informe uma linha com 10 a 13 digitos.",
                }
            )
        elif duplicate_counts.get(linha_raw, 0) > 1:
            errors.append(
                {
                    "Linha": linha_label,
                    "Campo": "Linha",
                    "Problema": "Ja existe outra linha com esse mesmo numero neste modo.",
                }
            )

        email_value = str(row.get("E-mail") or "").strip()
        if email_value and not is_valid_email(email_value):
            errors.append(
                {
                    "Linha": linha_label,
                    "Campo": "E-mail",
                    "Problema": "Informe um e-mail valido.",
                }
            )

        for imei_field in ("IMEI A", "IMEI B"):
            imei_value = str(row.get(imei_field) or "").strip()
            if imei_value and not is_valid_imei(imei_value):
                errors.append(
                    {
                        "Linha": linha_label,
                        "Campo": imei_field,
                        "Problema": "O IMEI deve ter 15 digitos.",
                    }
                )

    return errors


def collect_changed_editor_rows(base_views: list[pd.DataFrame], edited_views: list[pd.DataFrame]) -> pd.DataFrame:
    """Retorna apenas as linhas realmente alteradas no editor."""
    if not base_views or not edited_views:
        return pd.DataFrame()

    base_df = pd.concat(base_views, ignore_index=True)
    edited_df = pd.concat(edited_views, ignore_index=True)
    if "__row_key" not in base_df.columns or "__row_key" not in edited_df.columns:
        return pd.DataFrame()

    base_idx = base_df.set_index("__row_key", drop=False)
    edit_idx = edited_df.set_index("__row_key", drop=False)
    cols_compare = [
        c
        for c in edited_df.columns
        if c in base_df.columns and c not in ("__row_key", "Conflito", "Validacao", "Selecionar", "⬆", "⬇", "Mover equipe", "✏", "↕ Ordem")
    ]

    changed_keys: list[str] = []
    for rk in edited_df["__row_key"].dropna().astype(str).unique():
        if rk not in base_idx.index or rk not in edit_idx.index:
            continue
        base_row = base_idx.loc[rk]
        edit_row = edit_idx.loc[rk]
        if any(str(base_row.get(col, "") or "").strip() != str(edit_row.get(col, "") or "").strip() for col in cols_compare):
            changed_keys.append(rk)

    if not changed_keys:
        return pd.DataFrame(columns=edited_df.columns)

    return edited_df[edited_df["__row_key"].fillna("").astype(str).isin(changed_keys)].copy()


def build_candidate_full_df(full_df: pd.DataFrame, changed_rows_df: pd.DataFrame) -> pd.DataFrame:
    """Monta a versao candidata para validacao antes do salvamento."""
    if changed_rows_df.empty or "__row_key" not in full_df.columns or "__row_key" not in changed_rows_df.columns:
        return full_df.copy()

    candidate_df = full_df.copy()
    mask_changed = candidate_df["__row_key"].fillna("").astype(str).isin(
        changed_rows_df["__row_key"].fillna("").astype(str)
    )
    unchanged_df = candidate_df[~mask_changed].copy()
    return pd.concat([unchanged_df, changed_rows_df.copy()], ignore_index=True)


def find_header_row(ws) -> tuple[int | None, int | None]:
    max_r = min(80, ws.max_row)
    max_c = min(80, ws.max_column)
    for row in range(1, max_r + 1):
        for col in range(1, max_c + 1):
            if normalize_text(ws.cell(row, col).value) == "linha":
                return row, col
    return None, None


def collect_headers(ws, header_row: int, start_col: int, limit: int = 40) -> list[str]:
    headers: list[str] = []
    empty_streak = 0
    max_c = min(ws.max_column, start_col + limit)
    for col in range(1, max_c + 1):
        value = ws.cell(header_row, col).value
        text = str(value).strip() if value is not None else ""
        if text == "":
            headers.append(f"col_{col}")
            empty_streak += 1
        else:
            headers.append(text)
            empty_streak = 0
        if col >= start_col and empty_streak >= 8:
            break
    return headers


def map_known_column(column_name: str) -> str:
    c = normalize_text(column_name)
    mapping = {
        "codigo": "Codigo",
        "codigos": "Codigo",
        "asset": "Codigo",
        "asset tag": "Codigo",
        "nome": "Nome",
        "nomes": "Nome",
        "nome colaborador": "Nome",
        "nome de guerra": "Nome de Guerra",
        "equipe": "Equipe",
        "local": "Localidade",
        "localidade": "Localidade",
        "data da troca": "Data da Troca",
        "data troca": "Data da Troca",
        "data ocorrencia": "Data Ocorrência",
        "data ocorrência": "Data Ocorrência",
        "data retorno": "Data Retorno",
        "data de retorno": "Data Retorno",
        "data solicitacao tbs": "Data Solicitação TBS",
        "data solicitação tbs": "Data Solicitação TBS",
        "linha": "Linha",
        "email": "E-mail",
        "e-mail": "E-mail",
        "gerenciamento": "Gerenciamento",
        "bloqueio": "Gerenciamento",
        "imei": "IMEI A",
        "imei a": "IMEI A",
        "imei2": "IMEI B",
        "imei b": "IMEI B",
        "chip": "CHIP",
        "marca": "Marca",
        "aparelho": "Aparelho",
        "modelo": "Modelo",
        "setor": "Setor",
        "cargo": "Cargo",
        "desconto": "Desconto",
        "perfil": "Perfil",
        "empresa": "Empresa",
        "ativo": "Ativo",
        "ativos": "Ativo",
        "numero de serie": "Numero de Serie",
        "patrimonio": "Patrimonio",
        "motivo": "Motivo",
        "obs": "Observação",
        "ns": "Numero de Serie",
        "operadora": "Operadora",
        "s/n": "Numero de Serie",
        "s n": "Numero de Serie",
        "nº serie": "Numero de Serie",
        # Prosper Norte usa nome do gerente como cabeçalho da coluna código
        "fernando goncalves de mello": "Codigo",
        "goncalves d": "Codigo",
        "goncalves": "Codigo",
    }
    return mapping.get(c, column_name)


def detect_tipo_equipe(equipe: str, aba: str) -> str:
    e = normalize_text(equipe)
    a = normalize_text(aba)
    if "intern" in e or "intern" in a:
        return "Interna"
    return "Externa"


def detect_localidade(equipe: str) -> str:
    e = str(equipe).strip()
    if e == "":
        return ""
    markers = ["gerencia", "gerente", "supervisor", "promotor", "diretoria", "internos"]
    low = normalize_text(e)
    if any(marker in low for marker in markers):
        return ""
    return title_case_safe(e)


def _build_alimento_map(rules_path: Path) -> dict[str, dict[str, str]]:
    """Monta mapeamento localidade -> equipe, gestor, supervisor para Alimento (equipes_alimento.csv)."""
    ali_path = rules_path.parent / "equipes_alimento.csv"
    out: dict[str, dict[str, str]] = {}
    team_defaults: dict[str, dict[str, str]] = {}

    if not ali_path.exists():
        return out

    try:
        ref = pd.read_csv(ali_path, dtype=str).fillna("")
        for _, r in ref.iterrows():
            er = str(r.get("equipe_real", "")).strip()
            loc = str(r.get("localidade", "")).strip()
            gerente = str(r.get("gerente", "")).strip()
            supervisor = str(r.get("supervisor", "")).strip()
            if loc:
                key = normalize_team_key(loc)
                out[key] = {"equipe": er, "gestor": gerente, "supervisor": supervisor}
            else:
                eq_key = normalize_team_key(er)
                team_defaults[eq_key] = {"equipe": er, "gestor": gerente, "supervisor": supervisor}
        for eq in ["consumo baixada", "consumo oeste", "consumo zona norte", "consumo niteroi"]:
            out[eq] = team_defaults.get(eq, {"equipe": title_case_safe(eq), "gestor": "Marcelo Neves", "supervisor": ""})
        out["especial consumo"] = {"equipe": "Equipe Especial", "gestor": "Marco Antonio Neves Suzart", "supervisor": "Ricardo Cascao"}
        for loc, eq, sup in [
            ("alcantara", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
            ("cabucu", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
            ("piabetá", "Consumo Baixada", ""),
            ("piabeta", "Consumo Baixada", ""),
            ("realengo", "Consumo Oeste", "Marcelo Martins Da Costa"),
            ("barra da tijuca", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("recreio", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("botafogo", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("taquara", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("anchieta", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("lote xv", "Consumo Baixada", ""),
            ("niteroi i", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
            ("niterói i", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
            ("conveniencia farma 12", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("conveniencia farma 20", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("key account", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("supervisora senior", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("gerencia varejo ii", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("consumo zona sul", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("itaborai", "Consumo Baixada", ""),
        ]:
            out[normalize_team_key(loc)] = {"equipe": eq, "gestor": "Marcelo Neves", "supervisor": sup}
        for k in ["rede especial 1", "rede especial 2", "rede especial 3", "rede especial 4", "rota especial 1", "rota especial 2",
                  "rota especial 3", "rota especial 4", "impulso 01", "impulso 02", "impulso 03", "impulso 04",
                  "impulso 05", "auto servico 01", "auto servico 02", "auto servico 03", "auto servico 04",
                  "auto servico 05", "auto servico 06", "a s impulso 2", "a s impulso 3", "a s impulso 4",
                  "a s impulso 5"]:
            out[k] = {"equipe": "Equipe Especial", "gestor": "Marco Antonio Neves Suzart", "supervisor": "Ricardo Cascao"}
    except Exception:
        pass
    return out


def detect_grupo_equipe(equipe: str, tipo: str) -> str:
    e = normalize_text(equipe)
    special_tokens = ["especial", "impulso", "auto servico", "auto servico", "rota especial"]
    if any(token in e for token in special_tokens):
        return "Equipe Especial"
    if normalize_text(tipo) == "interna":
        return "Equipe Interna"
    return "Outras Equipes Externas"


def ensure_rules_file(df: pd.DataFrame, rules_path: Path) -> pd.DataFrame:
    if rules_path.exists():
        rules = pd.read_csv(rules_path, dtype=str).fillna("")
        # Backward-compatible migration of rule columns.
        if "equipe_key" not in rules.columns:
            rules["equipe_key"] = rules.get("equipe_origem", "").map(normalize_team_key)
        if "equipe_origem" not in rules.columns:
            rules["equipe_origem"] = ""
        if "equipe_padrao" not in rules.columns:
            rules["equipe_padrao"] = rules["equipe_origem"]
        if "tipo_equipe" not in rules.columns:
            rules["tipo_equipe"] = "Externa"
        if "localidade" not in rules.columns:
            rules["localidade"] = ""
        if "grupo_equipe" not in rules.columns:
            rules["grupo_equipe"] = rules.apply(
                lambda r: detect_grupo_equipe(str(r.get("equipe_padrao", "")), str(r.get("tipo_equipe", ""))),
                axis=1,
            )
        if "gestor" not in rules.columns:
            rules["gestor"] = ""
        if "supervisor" not in rules.columns:
            rules["supervisor"] = ""
        if "segmento" not in rules.columns:
            rules["segmento"] = "Alimento"
        if "eh_equipe" not in rules.columns:
            rules["eh_equipe"] = "True"
        if "equipe_pai" not in rules.columns:
            rules["equipe_pai"] = ""
        rules = rules[
            [
                "equipe_key",
                "equipe_origem",
                "equipe_padrao",
                "grupo_equipe",
                "tipo_equipe",
                "localidade",
                "gestor",
                "supervisor",
                "segmento",
                "eh_equipe",
                "equipe_pai",
            ]
        ]
        rules.to_csv(rules_path, index=False, encoding="utf-8-sig")
        return rules

    base = (
        df[["Equipe", "Aba"]]
        .drop_duplicates()
        .assign(equipe_key=lambda x: x["Equipe"].map(normalize_team_key))
        .sort_values("Equipe")
    )
    base["equipe_origem"] = base["Equipe"]
    base["equipe_padrao"] = base["Equipe"].map(lambda v: title_case_safe(str(v)))
    base["tipo_equipe"] = base.apply(lambda r: detect_tipo_equipe(str(r["Equipe"]), str(r["Aba"])), axis=1)
    base["localidade"] = base["Equipe"].map(detect_localidade)
    base["grupo_equipe"] = base.apply(
        lambda r: detect_grupo_equipe(str(r["equipe_padrao"]), str(r["tipo_equipe"])),
        axis=1,
    )
    base["gestor"] = ""
    base["supervisor"] = ""
    base["segmento"] = "Alimento"
    base["eh_equipe"] = "True"
    base["equipe_pai"] = ""
    rules = base[
        [
            "equipe_key",
            "equipe_origem",
            "equipe_padrao",
            "grupo_equipe",
            "tipo_equipe",
            "localidade",
            "gestor",
            "supervisor",
            "segmento",
            "eh_equipe",
            "equipe_pai",
        ]
    ]
    rules.to_csv(rules_path, index=False, encoding="utf-8-sig")
    return rules


def apply_team_standardization(df: pd.DataFrame, rules_path: Path) -> pd.DataFrame:
    rules = ensure_rules_file(df, rules_path)
    merged = df.copy()
    merged["equipe_key"] = merged["Equipe"].map(normalize_team_key)

    dedup_rules = rules.drop_duplicates(subset=["equipe_key"], keep="first")
    merged = merged.merge(
        dedup_rules[
            [
                "equipe_key",
                "equipe_padrao",
                "grupo_equipe",
                "tipo_equipe",
                "localidade",
                "gestor",
                "supervisor",
                "segmento",
            ]
        ],
        on="equipe_key",
        how="left",
    )

    merged["EquipePadrao"] = merged["equipe_padrao"].fillna("").astype(str).str.strip()
    merged.loc[merged["EquipePadrao"] == "", "EquipePadrao"] = merged["Equipe"].map(
        lambda v: title_case_safe(str(v))
    )

    merged["TipoEquipe"] = merged["tipo_equipe"].fillna("").astype(str).str.strip()
    merged.loc[merged["TipoEquipe"] == "", "TipoEquipe"] = merged.apply(
        lambda r: detect_tipo_equipe(str(r["Equipe"]), str(r["Aba"])),
        axis=1,
    )

    merged["GrupoEquipe"] = merged["grupo_equipe"].fillna("").astype(str).str.strip()
    merged.loc[merged["GrupoEquipe"] == "", "GrupoEquipe"] = merged.apply(
        lambda r: detect_grupo_equipe(str(r["EquipePadrao"]), str(r["TipoEquipe"])),
        axis=1,
    )

    merged["Localidade"] = merged["localidade"].fillna("").astype(str).str.strip()
    merged["Gestor"] = merged["gestor"].fillna("").astype(str).str.strip()
    merged["Supervisor"] = merged["supervisor"].fillna("").astype(str).str.strip()
    merged["Segmento"] = merged["segmento"].fillna("Alimento").astype(str).str.strip()
    merged.loc[merged["Segmento"] == "", "Segmento"] = "Alimento"
    merged.drop(
        columns=[
            "equipe_key",
            "equipe_padrao",
            "grupo_equipe",
            "tipo_equipe",
            "localidade",
            "gestor",
            "supervisor",
            "segmento",
        ],
        inplace=True,
    )

    # Alimento (Nova Prosper): aplicar mesmo padrão de equipes/gerentes que linhas ativas
    mask_ali = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_ALIMENTO]
    )
    if mask_ali.any():
        ali_map = _build_alimento_map(rules_path)
        for idx in merged.index[mask_ali]:
            key = normalize_team_key(merged.at[idx, "Equipe"])
            if key in ali_map:
                merged.at[idx, "EquipePadrao"] = ali_map[key]["equipe"]
                merged.at[idx, "Gestor"] = ali_map[key]["gestor"]
                merged.at[idx, "Supervisor"] = ali_map[key]["supervisor"]
                merged.at[idx, "Localidade"] = merged.at[idx, "Equipe"]
            else:
                # Fallback: Equipe Especial ou Consumo conforme padrão do nome
                low = key
                if any(t in low for t in ["especial", "impulso", "rota especial", "auto servico", "a s impulso"]):
                    merged.at[idx, "EquipePadrao"] = "Equipe Especial"
                    merged.at[idx, "Gestor"] = "Marco Antonio Neves Suzart"
                    merged.at[idx, "Supervisor"] = "Ricardo Cascao"
                else:
                    merged.at[idx, "EquipePadrao"] = "Consumo Zona Norte"
                    merged.at[idx, "Gestor"] = "Marcelo Neves"
                    merged.at[idx, "Supervisor"] = "Paulo Roberto Ferreira Chaves"
                merged.at[idx, "Localidade"] = merged.at[idx, "Equipe"]

    # Medicamento: Prosper Norte/Sul — EquipePadrao = Aba, Gestor por aba. Sem supervisor (não usar regras de Alimento).
    mask_med = merged["Aba"].fillna("").astype(str).str.strip().isin(ABAS_MEDICAMENTO)
    if mask_med.any():
        merged.loc[mask_med, "Segmento"] = "Medicamento"
        merged.loc[mask_med, "EquipePadrao"] = merged.loc[mask_med, "Aba"]
        merged.loc[mask_med, "Localidade"] = merged.loc[mask_med, "Equipe"]
        merged.loc[mask_med, "Supervisor"] = ""
        for aba, gestor in GESTORES_MEDICAMENTO.items():
            merged.loc[mask_med & (merged["Aba"] == aba), "Gestor"] = gestor

    # Promotores (aba) — Segmento Promotores
    mask_prom = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_PROMOTORES]
    )
    if mask_prom.any():
        merged.loc[mask_prom, "Segmento"] = "Promotores"
        merged.loc[mask_prom, "EquipePadrao"] = "Promotores"
        merged.loc[mask_prom, "Localidade"] = merged.loc[mask_prom, "Equipe"]

    # Internos (aba) — Segmento Internos
    mask_int = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_INTERNOS]
    )
    if mask_int.any():
        merged.loc[mask_int, "Segmento"] = "Internos"
        merged.loc[mask_int, "EquipePadrao"] = "Internos"
        mask_loc_vazio = mask_int & (merged["Localidade"].fillna("").astype(str).str.strip() == "")
        merged.loc[mask_loc_vazio, "Localidade"] = merged.loc[mask_loc_vazio, "Equipe"]

    # Manutenção (abas de troca e devolução)
    mask_manut = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_MANUTENCAO]
    )
    if mask_manut.any():
        merged.loc[mask_manut, "Segmento"] = "Manutenção"
        merged.loc[mask_manut, "EquipePadrao"] = "Manutenção"
        mask_loc_vazio = mask_manut & (merged["Localidade"].fillna("").astype(str).str.strip() == "")
        merged.loc[mask_loc_vazio, "Localidade"] = merged.loc[mask_loc_vazio, "Equipe"]

    # Roubo e Perda
    mask_roubo = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_ROUBO_PERDA]
    )
    if mask_roubo.any():
        merged.loc[mask_roubo, "Segmento"] = "Roubo e Perda"
        merged.loc[mask_roubo, "EquipePadrao"] = "Roubo e Perda"
        mask_loc_vazio = mask_roubo & (merged["Localidade"].fillna("").astype(str).str.strip() == "")
        merged.loc[mask_loc_vazio, "Localidade"] = merged.loc[mask_loc_vazio, "Equipe"]

    return merged


def apply_filters(
    df: pd.DataFrame,
    grupos: list[str],
    segmentos: list[str],
    teams: list[str],
    tipos: list[str],
    localidades: list[str],
    query: str,
) -> pd.DataFrame:
    result = df.copy()
    if grupos:
        result = result[result["GrupoEquipe"].isin(grupos)]
    if segmentos:
        result = result[result["Segmento"].isin(segmentos)]
    if teams:
        result = result[result["EquipePadrao"].isin(teams)]
    if tipos:
        result = result[result["TipoEquipe"].isin(tipos)]
    if localidades:
        result = result[result["Localidade"].isin(localidades)]

    query = query.strip().lower()
    if query:
        columns = [
            "Linha",
            "Codigo",
            "Nome",
            "Equipe",
            "GrupoEquipe",
            "Segmento",
            "EquipePadrao",
            "TipoEquipe",
            "Localidade",
            "Gestor",
            "Supervisor",
            "Aparelho",
            "Modelo",
            "Aba",
        ]
        mask = pd.Series(False, index=result.index)
        for col in columns:
            mask = mask | result[col].astype(str).str.lower().str.contains(query, na=False)
        result = result[mask]

    return result.sort_values(by=["Segmento", "GrupoEquipe", "EquipePadrao", "Nome", "Linha"], kind="stable")


def non_empty_or_default(value: Any, default: str) -> str:
    text = str(value).strip() if value is not None else ""
    return text if text else default


def normalize_name(value: Any) -> str:
    return normalize_text(value).replace(" ", "")


def classify_papel(row: pd.Series) -> str:
    nome = normalize_name(row.get("Nome", ""))
    gestor = normalize_name(row.get("Gestor", ""))
    supervisor = normalize_name(row.get("Supervisor", ""))
    if nome != "" and gestor != "" and nome == gestor:
        return "Gerente"
    if nome != "" and supervisor != "" and nome == supervisor:
        return "Supervisor"
    if "vago" in normalize_text(row.get("Nome", "")):
        return "Vago"
    return "Vendedor"


def build_full_table(df: pd.DataFrame) -> pd.DataFrame:
    from src.core.config import NEW_COLUMNS_ORDER, HIDDEN_COLUMNS
    table = df.copy()
    table["Papel"] = table.apply(classify_papel, axis=1)
    _hidden = set(HIDDEN_COLUMNS)
    existing_front = [c for c in NEW_COLUMNS_ORDER if c in table.columns]
    tail = [c for c in table.columns if c not in existing_front and not c.startswith("__") and c not in _hidden]
    return table[existing_front + tail]


def _get_gerentes_medicamento(rules_path: Path) -> set[str]:
    """Retorna o conjunto de nomes normalizados dos gerentes do Medicamento."""
    out: set[str] = set()
    for nome in GESTORES_MEDICAMENTO.values():
        if nome:
            out.add(normalize_text(nome))
    med_path = rules_path.parent / "equipes_medicamento.csv"
    if med_path.exists():
        try:
            ref = pd.read_csv(med_path, dtype=str).fillna("")
            for _, r in ref.iterrows():
                g = str(r.get("gerente", "")).strip()
                if g:
                    out.add(normalize_text(g))
        except Exception:
            pass
    return out


def _get_gerentes_alimento(rules_path: Path) -> set[str]:
    """Retorna o conjunto de nomes normalizados dos gerentes do Alimento (equipes_alimento.csv)."""
    ali_path = rules_path.parent / "equipes_alimento.csv"
    out: set[str] = set()
    if not ali_path.exists():
        return out
    try:
        ref = pd.read_csv(ali_path, dtype=str).fillna("")
        for _, r in ref.iterrows():
            g = str(r.get("gerente", "")).strip()
            if g:
                out.add(normalize_text(g))
        if "marco antonio neves suzart" not in out:
            out.add("marco antonio neves suzart")
        if "marcelo neves" not in out:
            out.add("marcelo neves")
    except Exception:
        pass
    return out


def _get_supervisor_from_rules(equipe: str, rules_path: Path) -> str:
    """Obtém supervisor da equipe. Medicamento usa equipes_medicamento (geralmente vago)."""
    eq_key = normalize_team_key(equipe)
    if eq_key in ("prosper norte", "prosper sul"):
        med_path = rules_path.parent / "equipes_medicamento.csv"
        if med_path.exists():
            try:
                ref = pd.read_csv(med_path, dtype=str).fillna("")
                for _, r in ref.iterrows():
                    if normalize_team_key(str(r.get("equipe_real", ""))) == eq_key:
                        sup = str(r.get("supervisor", "")).strip()
                        return sup if sup else ""
            except Exception:
                pass
        return ""
    ali_path = rules_path.parent / "equipes_alimento.csv"
    if ali_path.exists():
        try:
            ref = pd.read_csv(ali_path, dtype=str).fillna("")
            fallback_sup = ""
            for _, r in ref.iterrows():
                er = str(r.get("equipe_real", "")).strip()
                loc = str(r.get("localidade", "")).strip()
                if normalize_team_key(er) != eq_key:
                    continue
                sup = str(r.get("supervisor", "")).strip()
                if not sup:
                    continue
                if not loc:
                    return sup
                fallback_sup = sup
            if fallback_sup:
                return fallback_sup
        except Exception:
            pass
    if rules_path.exists():
        try:
            rules = pd.read_csv(rules_path, dtype=str).fillna("")
            for _, r in rules.iterrows():
                if normalize_team_key(str(r.get("equipe_padrao", ""))) == eq_key:
                    sup = str(r.get("supervisor", "")).strip()
                    if sup:
                        return sup
        except Exception:
            pass
    return ""


def _supervisor_display(dfe: pd.DataFrame, eq: str, rules_path: Path) -> str:
    """Nome do supervisor para exibição no cabeçalho. Vago se vazio, senão nome. Fallback em regras."""
    sup_col = dfe["Supervisor"].fillna("").astype(str).str.strip()
    sup = ""
    for v in sup_col:
        if v and v not in ("—", "-", "vago", "Sem Supervisor"):
            sup = v
            break
    if sup:
        return sup
    if eq == "Promotores":
        for _, row in dfe.iterrows():
            loc = normalize_text(str(row.get("Localidade", "") or row.get("Equipe", "")).strip())
            if "supervisor" in loc:
                return non_empty_or_default(row.get("Nome", ""), "—")
    sup_from_rules = _get_supervisor_from_rules(eq, rules_path)
    if sup_from_rules:
        return sup_from_rules
    return "Vago"


def _codigo_supervisor(dfe: pd.DataFrame) -> str:
    """Obtem o codigo do supervisor (linha onde Nome == Supervisor)."""
    sup = dfe["Supervisor"].fillna("").astype(str).str.strip()
    nomes = dfe["Nome"].fillna("").astype(str).str.strip()
    cods = dfe["Codigo"].fillna("").astype(str).str.strip()
    for i, s in enumerate(sup):
        if s and s != "Sem Supervisor":
            nm = nomes.iloc[i]
            if nm and normalize_text(nm) == normalize_text(s):
                return cods.iloc[i] or "—"
    return "—"


def _celula_vago(v: Any) -> str:
    txt = str(v).strip() if v is not None else ""
    if normalize_text(txt) == "vago":
        return '<span style="color:#c00;font-weight:600">VAGO</span>'
    return _escape_html(txt) if txt else ""


def _escape_html(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _render_equipe_tabela(dfe: pd.DataFrame, eq: str) -> str:
    """Tabela com supervisor como 1ª linha e demais vendedores. Cabeçalho com rótulos claros."""
    cod_sup = _codigo_supervisor(dfe)
    nome_sup = non_empty_or_default(dfe["Supervisor"].iloc[0], "—")
    linha_contexto = str(((st.session_state.get("chamado_context") or {}).get("linha") or "")).strip()

    cols_full = [
        "Codigo", "Nome", "Equipe", "Linha", "E-mail", "Gerenciamento",
        "Aparelho", "Modelo", "CHIP", "IMEI A", "IMEI B",
        "Setor", "Cargo", "Desconto", "Perfil", "Empresa", "Ativo",
        "Numero de Serie", "Operadora",
    ]
    labels_full = [
        "Código", "Nome", "Localidade", "Linha", "E-mail", "Gerenciamento",
        "Aparelho", "Modelo", "CHIP", "IMEI A", "IMEI B",
        "Setor", "Cargo", "Desconto", "Perfil", "Empresa", "Ativo",
        "Nº Série", "Operadora",
    ]
    cols = [c for c in cols_full if c in dfe.columns]
    labels = [labels_full[cols_full.index(c)] for c in cols]

    sup_norm = normalize_text(nome_sup)
    row_supervisor = None
    rows_data = []

    def _is_supervisor_row(row: pd.Series) -> bool:
        if sup_norm and normalize_text(str(row.get("Nome", "")).strip()) == sup_norm:
            return True
        if eq == "Promotores":
            loc = normalize_text(str(row.get("Localidade", "") or row.get("Equipe", "")).strip())
            if "supervisor" in loc:
                return True
        return False

    for _, row in dfe.iterrows():
        if _is_supervisor_row(row):
            row_supervisor = row
            continue
        rows_data.append(row)

    html = ['<table class="tbl-equipe">']
    html.append("<thead><tr>")
    for L in labels:
        html.append(f"<th>{L}</th>")
    html.append("</tr></thead><tbody>")

    if row_supervisor is not None:
        display_nome_sup = nome_sup if (nome_sup and nome_sup != "—") else non_empty_or_default(row_supervisor.get("Nome", ""), "—")
        row_classes = ["row-supervisor"]
        linha_sup = str(row_supervisor.get("Linha", "") or "").strip()
        if linha_contexto and linha_sup == linha_contexto:
            row_classes.append("row-target")
        html.append(f'<tr class="{" ".join(row_classes)}">')
        for c in cols:
            v = row_supervisor.get(c, "")
            s = str(v).strip() if v is not None else ""
            if not s and c == "Codigo":
                s = cod_sup or non_empty_or_default(row_supervisor.get("Codigo", ""), "—")
            elif not s and c == "Nome":
                s = display_nome_sup
            elif not s and c == "Equipe":
                s = eq
            html.append(f'<td><strong>{_escape_html(s)}</strong></td>')
        html.append("</tr>")

    for row in rows_data:
        linha_row = str(row.get("Linha", "") or "").strip()
        row_class = ' class="row-target"' if linha_contexto and linha_row == linha_contexto else ""
        html.append(f"<tr{row_class}>")
        for c in cols:
            v = row.get(c, "")
            s = str(v).strip() if v is not None else ""
            if c in ("Codigo", "Nome") and normalize_text(s) == "vago":
                html.append(f'<td>{_celula_vago(s)}</td>')
            else:
                html.append(f'<td>{_escape_html(s)}</td>')
        html.append("</tr>")
    html.append("</tbody></table>")
    return '<div class="tbl-wrapper">' + "\n".join(html) + '</div>'


def _garantir_usuario_local(username: str, is_admin: bool) -> None:
    """Upsert do usuário no PostgreSQL para manter FKs de auditoria."""
    if not HAS_DB:
        return
    try:
        from src.db.repository import get_connection
        conn = get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO usuarios_app (username, is_admin, auth_provider, ativo)
                    VALUES (%s, %s, 'keycloak', TRUE)
                    ON CONFLICT (username) DO UPDATE
                      SET is_admin = EXCLUDED.is_admin,
                          auth_provider = 'keycloak',
                          atualizado_em = NOW()
                    """,
                    (username, is_admin),
                )
            conn.commit()
        finally:
            conn.close()
    except Exception:
        pass


def _autenticar_keycloak() -> bool:
    """
    Login via Keycloak Direct Grant — sem popup nem segunda aba.
    Exibe a página de login customizada com form user/senha.
    """
    if st.session_state.get("authenticated"):
        return True

    _render_login_page()

    with st.form("login_form", clear_on_submit=False):
        username = st.text_input("Usuário", placeholder="seu.usuario")
        password = st.text_input("Senha", type="password", placeholder="••••••••")
        submitted = st.form_submit_button("Entrar", use_container_width=True)

    if not submitted:
        return False

    if not username or not password:
        st.error("Preencha usuário e senha.")
        return False

    token_url = f"{_KEYCLOAK_URL}/realms/{_KEYCLOAK_REALM}/protocol/openid-connect/token"
    client_secret = os.environ.get("KEYCLOAK_CLIENT_SECRET_STREAMLIT", "3PYcJYM5wXzmoNhhkaLgTjcJ5pJqwwxi")
    try:
        import httpx
        resp = httpx.post(
            token_url,
            data={
                "grant_type": "password",
                "client_id": _KEYCLOAK_CLIENT_ID,
                "client_secret": client_secret,
                "username": username,
                "password": password,
                "scope": "openid profile",
            },
            timeout=10,
        )
    except Exception as _exc:
        st.error(f"Erro de conexão com Keycloak: {_exc}")
        return False

    if resp.status_code != 200:
        err = resp.json().get("error_description", resp.text)
        st.error(f"Falha no login ({resp.status_code}): {err}")
        return False

    token_data = resp.json()
    access_token = token_data.get("access_token", "")

    # Decodifica claims sem verificar assinatura (já validado pelo Keycloak)
    import base64 as _b64
    import json as _json
    try:
        payload_b64 = access_token.split(".")[1]
        payload_b64 += "=" * (-len(payload_b64) % 4)
        claims = _json.loads(_b64.b64decode(payload_b64))
    except Exception:
        claims = {}

    kc_username = claims.get("preferred_username", username)
    roles = claims.get("realm_access", {}).get("roles", [])
    is_admin = "admin" in roles

    st.session_state.authenticated = True
    st.session_state.user = {"username": kc_username, "is_admin": is_admin}
    st.session_state.access_token = access_token

    _garantir_usuario_local(kc_username, is_admin)
    st.rerun()
    return True


_LOGIN_PAGE_HTML = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');

/* ── Ocultar chrome do Streamlit ── */
header[data-testid="stHeader"],
[data-testid="stToolbar"],
[data-testid="stDecoration"],
[data-testid="stMainMenu"],
[data-testid="stSidebar"] { display: none !important; }

/* ── Fundo geral (lado direito) ── */
.stApp,
[data-testid="stAppViewContainer"] { background: #111827 !important; }

/* ── Desloca conteúdo para a direita do painel de marca ── */
[data-testid="stMain"] {
    margin-left: 44% !important;
    width: 56% !important;
    background: #111827 !important;
    overflow-x: hidden !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    min-height: 100vh !important;
}

[data-testid="stMainBlockContainer"] {
    max-width: 360px !important;
    width: 100% !important;
    padding: 0 24px !important;
    margin: 0 !important;
    box-sizing: border-box !important;
}

/* ── Inputs ── */
[data-baseweb="input"],
[data-baseweb="input"] > div,
div[data-baseweb="input"],
[data-testid="stTextInput"] > div > div,
[data-testid="stTextInput"] > div > div > div {
    background-color: #ffffff !important;
    background: #ffffff !important;
    border: 1px solid rgba(0,0,0,0.1) !important;
    border-radius: 10px !important;
    transition: border-color 0.2s, box-shadow 0.2s !important;
}
[data-baseweb="base-input"],
div[data-baseweb="base-input"],
[data-baseweb="base-input"] > div {
    background-color: #ffffff !important;
    background: #ffffff !important;
    border-radius: 10px !important;
}
[data-baseweb="base-input"] input,
[data-baseweb="base-input"] input[type="text"],
[data-baseweb="base-input"] input[type="password"] {
    background-color: #ffffff !important;
    background: #ffffff !important;
    color: #111827 !important;
    -webkit-text-fill-color: #111827 !important;
    -webkit-box-shadow: 0 0 0 1000px #ffffff inset !important;
    box-shadow: 0 0 0 1000px #ffffff inset !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-size: 0.9375rem !important;
    font-weight: 500 !important;
    caret-color: #f2c230 !important;
}
[data-baseweb="base-input"] input:-webkit-autofill,
[data-baseweb="base-input"] input:-webkit-autofill:hover,
[data-baseweb="base-input"] input:-webkit-autofill:focus {
    -webkit-box-shadow: 0 0 0 1000px #ffffff inset !important;
    -webkit-text-fill-color: #111827 !important;
    caret-color: #f2c230 !important;
}
[data-baseweb="base-input"] input::placeholder,
[data-baseweb="base-input"] input::-webkit-input-placeholder {
    color: rgba(0,0,0,0.35) !important;
    -webkit-text-fill-color: rgba(0,0,0,0.35) !important;
    opacity: 1 !important;
}
[data-baseweb="input"]:focus-within,
div[data-baseweb="input"]:focus-within {
    border-color: #f2c230 !important;
    box-shadow: 0 0 0 3px rgba(242,194,48,0.18) !important;
    background-color: #ffffff !important;
}

/* ── Botão olho (mostrar/ocultar senha) ── */
[data-baseweb="base-input"] button,
[data-baseweb="base-input"] [role="button"] {
    background: transparent !important;
    border: none !important;
    color: rgba(0,0,0,0.35) !important;
    cursor: pointer !important;
    padding: 0 8px !important;
    display: flex !important;
    align-items: center !important;
    transition: color 0.15s !important;
}
[data-baseweb="base-input"] button:hover,
[data-baseweb="base-input"] [role="button"]:hover {
    color: #d4a820 !important;
    background: transparent !important;
}
[data-baseweb="base-input"] button svg,
[data-baseweb="base-input"] [role="button"] svg {
    width: 18px !important;
    height: 18px !important;
    fill: currentColor !important;
    stroke: none !important;
}

/* Labels dos inputs */
[data-testid="stTextInput"] label p {
    color: rgba(255,255,255,0.5) !important;
    font-size: 0.72rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.07em !important;
    text-transform: uppercase !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
}

/* ── Botão Entrar ── */
[data-testid="stFormSubmitButton"] button,
[data-testid^="stBaseButton-secondary"] {
    background: #f2c230 !important;
    border: none !important;
    border-radius: 10px !important;
    color: #0d1117 !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-size: 0.9375rem !important;
    font-weight: 700 !important;
    padding: 14px !important;
    width: 100% !important;
    transition: background 0.2s, transform 0.15s, box-shadow 0.2s !important;
    margin-top: 4px !important;
}
[data-testid="stFormSubmitButton"] button:hover {
    background: #d4a820 !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 8px 24px rgba(242,194,48,0.28) !important;
}
[data-testid="stFormSubmitButton"] button p { color: #0d1117 !important; font-weight: 700 !important; }

/* ── Card do formulário ── */
[data-testid="stForm"] {
    background: rgba(255,255,255,0.03) !important;
    border: 1px solid rgba(255,255,255,0.07) !important;
    border-radius: 16px !important;
    padding: 28px !important;
}

/* ── Alerta de erro ── */
[data-testid="stAlert"] {
    background: rgba(239,68,68,0.1) !important;
    border: 1px solid rgba(239,68,68,0.25) !important;
    border-radius: 8px !important;
    color: #fca5a5 !important;
    margin-bottom: 8px !important;
}
</style>

<!-- ── Painel de marca fixo à esquerda ── -->
<div id="login-brand-panel" style="
  position:fixed; top:0; left:0;
  width:44%; height:100vh;
  background:#0d1117;
  border-right:1px solid rgba(242,194,48,0.12);
  display:flex; flex-direction:column;
  align-items:center; justify-content:center;
  padding:48px; box-sizing:border-box;
  z-index:9999; overflow:hidden;
  font-family:'Plus Jakarta Sans',sans-serif;
">
  <!-- grade sutil -->
  <div style="position:absolute;inset:0;
    background-image:
      linear-gradient(rgba(242,194,48,0.04) 1px,transparent 1px),
      linear-gradient(90deg,rgba(242,194,48,0.04) 1px,transparent 1px);
    background-size:40px 40px;
    -webkit-mask-image:radial-gradient(ellipse 90% 80% at 50% 50%,black 30%,transparent 100%);
    mask-image:radial-gradient(ellipse 90% 80% at 50% 50%,black 30%,transparent 100%);
  "></div>
  <!-- brilho superior -->
  <div style="position:absolute;inset:0;
    background:radial-gradient(ellipse 80% 60% at 50% 0%,rgba(242,194,48,0.09) 0%,transparent 70%);
  "></div>
  <!-- conteúdo -->
  <div style="position:relative;z-index:2;display:flex;flex-direction:column;align-items:center;text-align:center;">
    <!-- logo P -->
    <div style="
      width:80px;height:80px;border-radius:20px;
      background:linear-gradient(135deg,#f2c230,#c9971a);
      display:flex;align-items:center;justify-content:center;
      box-shadow:0 0 52px rgba(242,194,48,0.35),0 8px 28px rgba(0,0,0,0.6);
      margin-bottom:28px;
      font-size:38px;font-weight:800;color:#0d1117;line-height:1;
    "><img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAHAAAABfCAYAAAAwGkOoAAAQAElEQVR4Aex8CWBdRdX/78x9a7Y2afZ0oxQKZSlQAduigoLAJwKCoqgIRUC2ioCoIKBYCn7IIoL6IYrgAlgRFzYB2Yq0yGZb1pa2adqkSbPvL3nv3Zn/79yXgKVpaZbW9k9vZu5sZ86cOb85s90kBjufHVoDOwHcoeEDdgK4E8AdXAM7uPg7LXAngDu4BnZw8Xda4E4Ad3AN7ODi77TADyaAO3iv/z8Sf6cF7uBg7gRwJ4A7uAY2I/703csL95s6/nOHf7jijtOOLl0657NFDbM/Xdz9lU8V9556THnytGNLkqd8qqT3nBOKui74bHHtmUcXPXXEQWPPmD590qjNsB3Rop0W2KfO6dOnh/fda+Kpn/5YxRtHf7Q8edjBFdZGIw3JtDd/xfrI7EeWZO9z78K8wkdeyYv/Y0l25PHF8fCjL2eFn1icG7n/X7lZ9zyXV/rIGzmHrm0N3x52va1HzBiXPvnI8sWfOHj81L4mtkrwgQZwypRx5R87aOwDHz1wXDLR2dTb3WPuXLomsuerq2PhVbURae4AupMuuDDOjlnkRn1kx9KIR4TeRywmiMZ8ZEctchiGjaA3BdQ2R/FmjectWh6ZVt2C14+YMbb3wweM+/zWQNBsDabbM8/JkyvGHnZwxePT95ngdyVCNZUNkWOqGsLh7pSIdUKwBB6ANH0q5REQg86EoLNLwzDaOkJo7wY6ujwkEkBvj0EyLfAdIOIC74lFJCwwEkJvOoTltZFIXWvo3uMPK+84eO9dSjCCzw4N4NEHT877ylHl8y/4XHHn+Z8rSp57QnHi+EPLHqZ+hP4dN316edaHpo297chDypO96ejat6ojhzd1hkwkBHhUOiBI0dK6eyxSPlBalMKMaUkc+4lOnPw/nbjo1BZceW4TLj+nEVee14grzm5lvBmnHduG44/owP57pJCf40ALRpKgW3AIOOWaJmcLj1o2cHh5ZTQnYV3tPntMOB8j9JD1CHHahmxmTSnM/eJRZfVdfm/rgmWxz92zMDf7voU54fv/lR17bln86FkHjOuYPHly3pQpY/c9+iMltX5PpHNdY+isN6piYY/WEQux287Bt7ScpGB8WRInHdWFGy5pxPwb1+GXV9fh5u+sx48uWo8ffqMel5zTgvPObMWcM1voNWxm2IK532jC/17UgJ9eUY9f/3AdfnbFepxweCdysn2k0gY+4RPHtpCZhsMcMM1dRnrS5pZpUydeNRIqU+4jwWfLeIwA1ZGzSu/t8LLaFrwVL1pVHxYhT11/4lyPsqKC3LhFdbOXHQmlWp0LLV66Jru0udNImKA5WluI5pCXm8Z+e3bj219twWO/qsbfflaNay5swBeP68CMg3qw954plJenkZMHhGOASbORrv/w3YzTh8IgWA7jxiWx3z5JfOaoTlx3yXr89Sc1OOPEVhTkWKR9B0sQ05yehT5sLHo45SbT7srdd9/lI+Q0LGeGVXvbVpbjD6voXFYT/3xbIiQRSh4NOUDoqRjjLADLsS4wxiDRK5JMGjE+c1lUVJDCzP17cNmZjXj49mr8+dY6nHNyC6bslsToPIdYFCQUuF6GKfo0PeuSJciUiQEc+QblKU6QvQJdOKMRwaQJPq6c04T5N9fg8JndnELTcM6RjeO6KAhR5q5eDyWj/KfJlRX5HqKjGoZYcxtW+xzgHTmrIvmvFV62eAKPChCqAwI6vgDq0RFL9SCADr4vHPkO06b24KwvtOLmbzfjdz+qxewvtKO8yAfnN4DAO06jTkPHJBz5kdmgnQUbB/prWwvne9h1oo+fXtmAr57YgSitlTCyESGVg8fBV7neM8d8pOyfGMaz3QOo4LXPquh9vSoeygp7IFLsrgkC6wiEKt0xSzJdEREkua5NGp/CDd9uxK1cly4/vxkzZ3QhTMu0SdIqYMKQjmMB3H6CbOgFLGLuIF3AS1+0MkfP6uIskHaIcUq/+KxWfPnYdnjIyM1icCZFhINxZX145tSpUyOaNxRvhlJpW9ZpOrCi5/W1US8SoWKoHBhFS+c3B+kDzdeQmncEz5GGxzF8/OAEPntCByaMJR2nOJuigh1AEgLFEEyj72EmXZCjgPblbnlArMCKZE8eHqMmMxA0PwUCZXHR6a2YNa0bussFG7OqefYlwal0VKzjd1ve2IaUymbDnO0o9clDKmreXh8NRTjdBMqhzoUAgWMZAWiGyhKO6xQESgEOesGYvDSO+0Q7kAACU6U1GMsBECQElsoG6TVHNzbqQQ7g44SvwToDykEf1NVZgcixDecBEIHjjjQ7y+LbZzeiKD8Ny2nbaD/YmE4I8Zg9AUN82PQQa27lauMmTjxtZW24PEQlCJVBNVDlKq56BAoDH+qAb+aRhhFuFoBZ0xPYd7cUdFNBbWk2QOXSsZ4j4Aof+EhAz8i7zmqZUmrY75kmEO8SvSeWGTsZYi1ScuapSDpI2ApBFOw9JY3PfKIblgPKp0BCEEPsVVuX583i0UirDtaz54Otsm3oJxb6v+rl8FTwaDybbVSoDNBb+qJRPi7i8SCDkmBTjypXFchZDLBKRVVToSKswxDkJRwdQiBAJW+MNLbsIT+d1skV2s6XePjPirMtMrY8k+qGrDth0O5lfXHLGG5ItV0COG7CLt98uy5swmHtNn2/EjeU/Z2Uo5I4rGldwJxTWjFxbDpQFlTx71BtGAlYaj1dMAlWUEptWIInwjWMIdUMkEYgQOAx+IfTJVkgmKbZaGlhGgfu2QvLGxvhUqCc27o9jM1PzcYQHoo8hFpbucru5em5elbSzjlYmMCcNt2oo4nq5uDkT3Xg80e3QUe6UrsAHI1t7EWUO6B11cqJFweAgNixug1CkUw96h2ZzVMmPZi36HZTwQ+YCNT6PvHhBFLcoeoMADaYtkDRaH/XwfDtp90eATSVdaFoOCTsNsUTKplTDYdwv8wbhKr4RNLD/xzSg4tPb0I8m8Uc9XyzfqA1jW7CO4jHInrRgzx3uprmaYOZtEK+lVWwhvlMDMURnP4plOMM4Hlwym69IG5wRvvmYNh+Iikq+aBbIItB19mqFT4+o/xVWpOIAsfF3rE1dpHvjZ2Cl0oLTj6qFZec1YSCAlLzAA8GjvDRlDaupD3WU1fUAQy7OgyamkN4860o/r0kisqqKFpbQrC0HOH1nKEVO5WFfmNm75/jtCP0jvV1IKhM2TGH4tE87FN2tUKhKCEjKtn7M3wPxZAqvYfHiCUPO6ii+u2a+NQQjw3QQzo7bdhzBgAVCrFsi71VR9MQpr92Uguu+nojbz2466RT8KAaUXSpOFZgXb49AQjGmuoQ/vxwDubdko+vXlKC064oxhmXl+DrcwsxZ14xzvlBIU5n3uxvlWLujQVY+FKY95kCo5oS8NFXxlOMvjSDTTil1LFEURkEKURp7bk5Pi+DbCaP2ezOJjhsPlvF2jzFNig99NBDQ4cdNL77zXXRCs9z/Eqg05eHdzGgqugswfR9gBaK8pIkrr+kGd85qxm5eYAjeCymtCTkm5/iAPJK0yLX1YXx27/m4IsXl+H4r5fj8p8U4dd/HIWHn83Gc//OwkuvR/FWVQRraiN4/e0wXlwaxz8WxfHLv4zG2VeV48us9+xLMTZCxtSY4ygJpkU2GLTGkCUDOx1EWi6g8ZGaLo+3M2VFFtayjzCZ8eYLezYwi83lms0VbouyiRMnxhJtqxIrak08GnYIeT7KC304C34o9dCTtEimBIkeD9YHxpWmcOrxbbj7R+vwmaPbSS/BvSdUSazDjR0c15m2VoNnXsjCt24oxGFfGY/v3VyIf74SRWN7BAl+SfBpjfE4kBX2EQnRRwXhsA3uLCP8wh6LWIRpNh1dBi+8FsPsS0vxy/tz4evmg3WJB3QNAwEKdpibUpYlYqKQAxIIx7dn2WaaAFroY63lZbqlVJoanP+vAjh27Nh4WYHfsa7JC+mB3aM0Z3++Hf93ZT0vgFtx0N7dPJCnMY3b7k99rBPfOKUFt/1gPeZe0oRxFT6gVkcdGJqqKlRoJF38ev7AYzmYc00RzrqyGPMfptKpQ1DpYW6Mwkgi2KhQT4SeSgV9GGpRSqY+WKskeLOa5fRJa+EAue4Xhbj7wVxWcABlVYpgF0sO2OQjCAYXaYgjqRz49R8tHSF4niYtRAxa27wapgbtKMag64xIBQWvYozXXtMcCYVD2kcJALqQX7+n7ZfApec249Yr6nDL5Q249bJ6/OS79bjw7FZMnZKCI0jiG+qFyqFW2H+A1vskp8QL5hXhouvHcArMRop3oBFuVhzXT4EPIS2EWqP+WZP9MPBZ2dHSDPPUg2bl1FJYGsxt4iD6I4LuJPDze0djKTc8NE84LlwcF4BT4oG9tg1OnmQDp7OkAbq6QmhsNpCgQUE0kgbPvQ8OzGHzuWbzxVunlF/LoyWF0r6u2QuJcCphBy86tQ3nndwK3VI7fltTwcpLLXadlMTECSnE9fai2/AADATKCEBxVIJBTV0El19fiDlXF+HhZ7LgE7g4p0QFTMDHpxfWpRUhMB2d0kQzQWyQeRx0tgPBDBTNTNE1ivUcHFOOiraoXBfB489x7tWpVAUhGxcAQZIBnLBN0TYDHgge/fWN9i6PucI2BTkxg46Ed1dQOMiXGST9SJB7+TnJzvrmcEh3dgYhXPTldpz7pWYYWiIsMjplSPMA+GkIvAzO9N8FndbNiRpSkmvjPxZGcdKFJfjtg9no4uV1dtgQVGpVLEQEqkBGwFkWagmgFVCn0JDqY3kGnICzaDxTxykRacH1CUHjBvA9cLuPZ1/JQs16CktjVsZBNWzqUfgtlJ2DQF0719XmNgNDltZZHu7Trrp61QoM4SGLIdQaRpVPzizvXNcSCnkctaqMOQTunC+1wGPfHKck7aSuR6I9Faou0A6V4AARIb6MUHfNLQY3/LqA61wpqrnLJEMoTz2/gaZkfccpNI00dz4+t67Wt9wEuWATYmk9luXkBLZADCxD9aoONgqBiMDpcBHN81hOalq9brKWrYqiuZV5dCx4H0depFA2wXTLQff22jBlZSa0DbVACqmCaNYgvUo3yCpDJz/ykLFNr6+JxaLchQl3LBec0oSvn9IC3cA4dkaxErKn7uDYW90ggPmEj2+wixY6xa6iAi+9vgg/n59HC/IQCrHsPS4vLw8VFRUoKysb0IdCYfT08E7SAiKGIIIPW3KO4cBOiwzlam0LoaVVwNFEkAmCCjxwFVDADI2yJQO9oH9tWYwAsh7THoBk2rzMYEhumwE440PjF76xJlIQ4Va9h5uBzx7RgrO55nlUPvtB4dlDWqVaj4NAOHUFehEtYmc1EQbeWhHBuXOL8Ag3LCGuV0JPisApH7VeVfLxxx+He+65B3feeeeAfv78P+Lqq6/BhAkTkEymIAQRqmptL+C28aufxvB8ubae1zgKviMdLZ7vgR3nzoAlQxCtHq7vL7waQVgvK1gQCgtWrwufMnDl98/dJgBOnDTxrhAVeAAAEABJREFUwprG8Aw9pKe4wThyVgLf4iefcJQCctoUVQI76KgQ7aQqEgoYi4XK0d2aUGlvvR3BOVcVB4dtQzBBpQdTJukywBlakkNWVjZmzJgZgDNlyhQM5Pfff3/Mnn0afv/7u0k7A+l0OqhLVu/jHISml+aUTLzfhzZTTLsmqQO4mVpbF0J1fZSis7McceGQb7n+vZ2hHPx7qwPIEV6WG7U3cPXh7hCYtkcKl36tBYVF7BbBc1SH65eb0jimhd4RVfYPXIbAfQ5e59b9PFreSt6YRIIdpkBIIDYDmohAD8QKZDQaQWlpaT/XzYY6zc6dO5f3qAXvCyCbAEVDmuvnpLFEIwRQCAQyYlOPA4KKgA7ORxbk9LVDefnJPieGRSwZsjNDrrmFFSeV+lXNnVHh7IExBRbf+EoLdts1Cb36UgAAy/4JuTkIAeVQpSO4DhCuN9rpt5ZH8M3rirCsMoIQz3UkgFFzFWG0vz7pyUYBjMXi2HXXLf86oxY6a9asPsVik49jm44y5vBIk5ftA2wv8woiGPDRIvbFkDjFHfVjz2aBQpNUkEwLqtrMiUwM2Zkh19yCigfvP2Hh27WRcDScpvgGX+PF8+GzOqE3KNovZaEh4WI5U7Q6qGfUaA7XxLr6EK64tRBLlkd5DhMWC0upEb5BGhGC6CwyWtE4UFxcjNzc3A0AWbZsGafM2bjqqqvQ1tZG+g3dfvvtF1jwhrnvTQlsGpjAD8ajciiDNkt5MwPxvbSZtHPsCWnANe95fu2oYX8Md7M6A+9a2pOsr6xcn6Ec2tsMrdr716IFzKpv9WZEaHq97PRHD+rB6Se2U88EQDv+HyyY805KuBYSF6YJKy+ib7ojH88vDiMe5YhnbsYJA/UMSEwMNRJ4zwuBtzxBXKSfBlAA77vvPtx2221YtGjjWUt3rUGlzbwc20rSAqfwcmFMfgqcPDhINlOBRYQ8QxMC/vZkLhK8jBDxkGT12hb5AUmG5bYagAU5yWd0yrFcL8qKfFxxdgMiUe0zgXlXrwMILxAFmJuU+x/PwvxHcxHmpaF2egDijbJ0B3rAAQdskO+4VtbU1ARWqQUe+Wmo+RqqX7VqFTJ4a2pgL9yFKIh7Te5B7ijSOIDSZgDCwI8jU/1YvGxZGC8sicNyVvG5wy7L993KlWuvHbjWluduFQD322fiX9e3hXg2N9AOzPliS+Z7nQUk6DQ285CIO/S3V0Vw4535/LREUrFUEvMZfT8nYlBeXr4BWW9vL1566SV0d3cHO1KdLpVARDQI/PLlyxm+m2ZiI5fmTczEshQO3LsXunFxfHE4QkQ2ou3PCKZXCzy6IBer1wlMyHITBJ5d3eOA2jDfw3BmGHUHrDp16tgC3vV9WoFKJB0O+VAnjjmM6x77SEMAEVU3YN0gkxLZXsNL43x2OIwwOyyQoOj9Xo5roepyr7323oA0lUpBRHDmmWfihhtuQGFh4QblVVVVWLx4MYwxHCgcYRuUvptIpT3szy8jAYC8URFL2vcTjTPJSl48PMTNi7UhOD+E0lHWNXdGjn2X89BjVNfQKw9Uc/wYt6SxS7ePBkX5Dl/6VDcKih3AHRcYqAnSoAaqmsnjWrFoaQRPLorxhkZYhRpy9JnSTb4dRwcdb1/GYsyYgg3owuEwTj/9dMybNw86vYpsyO9OHvabmpoIoARAb1C5L8FZD/l5KRx1SBf0sxUUPJWOsmnbfWQbBsIkl+7Hn4/jVZ5hY2EflvVokE+sWLGCZszyYboRBXD33Sfu8WZVvCKLU6B1aehfAx39EVpfjxqeg/AnkFeC98YvStPLRf6BJ+OoawwTQJKosgP64MWMgZ0IwSaCu+wyEf1rXD/l2rVroQf3SISCMdORjkHg/vCHP0C9iNYPsgZ8WW5e9t4tCf0u6QLVC3gEDSxWZBOysT+6i77ngRyEOCVxL4fRuWm3+LWqowZsZAiZbGIItTZRpaLALnTCntLlZQOz+eVcz3GqL83m4Y0QsrMctRyFcEHHNd0nBq1v+aoQHnw6j7vOJFFXKqtGu4kW/zObCrXAHnvsuRGAbTw2xPXzex+5iATr4T28arvuuuv4fa4rqCPCLbOjPNDBZqExy4GoXwyiUYtzvtAKL0ImpHEMhBRkRTkDBxcktC8SlHKJxN0P5WLVmgg4CXDtE/Aa8WYW+vQj4rS1EWE0efLksStrTX6IF9WWa/OsA3twMD/MBvfsYIfYOReEDmBcRBCsIRB2HGAAx7Xvyeez0MRPLR6VKQH6Ws46UI8BH8cRQnacniwmTZoEEdmArqenhxfXPdCd6MKFC3HHHXdgzpw5wZmwoaGBGwquTeSRqWQZsC3HPMaMGH7VEJx4eCc+dnA3wMEJCH/Q96h8gDgBzRFOHElYn1d/S1+L4ld/ykOYlw86dY4t9NNvr1hzIUbwGTEAi/MSzwo3AewJFSk47Rie+chddSmBcpwWaR8pvuvz7DMcleGYBrp6JDgrxSMC36hiLAxUoVpMZhoM4EWoPOYrkBxIjG3o5s+fj5NPPhlnnHEGLrzwQvzgBz/Aww8/DN2d6vooIhCRTCUCZqBtp/lGcFuy75Q0zv9SK1iUkd8hCIN+BdUcFDhoDQJp6FMcjDf/bjTaOw2JAZ/XZmvXmwMxwg+5D5/jvvuWZDe2hydwAuJo9fCRD3XzzpMLH3UvvImgA/vEhthzCXpMMDMhI8ynoyRra8NYVROCEDxD0OngCKFqwL0DJGkHcL7vB5+NRo8eHZQqmEGELz24q3/ttdewfv166G5Tp1QNWUwwKJdG6IWCWgWC3ueGIzsKnPP5FlSMS0E3YkKJVB4WQ0Vy7JyweuD1pZm0uN88mIdnXszmOi7o5RXahOLUazU1qxeziRF1VNvw+dlU/De9aY+IcJfF2f3YQxPI0qsmKsOJZQNCnLQpdp+osL/MoyNQgT4Y1XDJ8jD0oBuARWWIiJbQS6AvRjbpfN5N6fkvKysroBGRIKysrEQymYQCptbmeQbSx/s/QVZiTVNkRimBpaegZ/L675hDO4E0s+m0XAwj6jQUdpgZju2xCoRr5KucOu+6Pxe9aQuKhfIi3/7zxTXTtMpIexVh2DzzstLH6ULvWw9jedCdtifXCi+jQGiPFTTng/ogkAJhixllMUbHTOYI3uS3PsOtnfA855TKaTYJgvqMb8apBer6l5OTswGV7kC1TIR8ghIbvPUl0p/HVF9cB5xj22mudccd2o2vfrYdEmI5r/VoqoD2h0mQxgHQao7ygo8hXXuHh5/dm4fKqhB3nh5081Lb4H2Sxe82zMRIuWEDOHnyhANaOj1Pf51B/0vRzP27MaGMwzWt3XMQuEBW7ajGNaQWoDOjbmKECiE+UBBr1/PUG9B7QT3h6HY6JXGEq9VgM4+1NjgD9ltgP6ke0HUT49iICCXglAdyR/CoTlU+xxwfIgJxAD8NYtb+PfjuuY0YPcrC8QzrWOC0nGEGSEC0EwRaQwF75QR3/SkXDy3IDjYuvEXE6Gz/qZqaVU+weKu4YQNYmIsbWztDPCEY6C/mTt8zhbguQ6qbPpGlL8wEjoHmCPXg6JFRBLMkzDKOZuqJ+R6hpHg6L5ECqixs+rGkGz9+HERkA6LVq1cHmxXDDZaCqCFAGrYDKhzBY9ge22I8mXKYMa0bN3yrHiUlPnQXrYNHcaPEpONbBxS/KCiohA0B6iGHRxbEcevv82l5BmrBYwuTna8srfo4tuKTkXoYDRTkpWZ6ns+53qGk2MfkiVzsA/CopE3xpTXQQajUoPPCCsTOcjQ75jHKmikqCzA8Trjg+xNpmDuQ8znUdf0bN27cRsV1dXXcFb9bV0FUIisCBUbjoJVbCtTTCxxyQBI3fLsBFeUcXOyKKiiQR0B5+AJTpNVAGAacI8Abb8cw7+cFSCYMdSEYnQPb0BYtzPDfem+Vb8jcuWWPVjeGQ54xwYjjYo2JFTyA++zsZrkKgVEaS/w4ogNLMBiV5cNYgQAQAmc47C0tRYLzIDb5WGsxatQo6E1LiveeXV1dgdXxugqNjY3IWB2CsB9Aj21aHThEwrHNNAfBsR/vwNUXNmDsOB8uRdkIMiAqKEgOEcbBh1rTNGMwYWBtdQjfu3lM8LcVNuQjFEq7ljbsz/Z7lWZreooydPY96fSH61tDIkY46oDCUWmU8NOR/h5nZnrZBG/jwOFM8FjOOCFk2mLqbgn4KlGmGA4CtRIHzRQSD+w8z0NraytuvPFGnHfeebjgggtw/vnn48orr0RtbS0Vyt0Fq/aDxyhUPuHZjLgx6XD+yS249pIm7DqBls9tv2LlLMXSZumFaydXQ4j+ZCoBnqCl1eAaWp5+rNUvDfqrkr1J7+Lq6sqlZLzVnWpmyI2MyfFPUaV4tIAQL2rHcgeqf3PHyQfEZZN8OfOwTKgddQwVIAYz99cB66CWoRbCLNJ5LKUmqTgmBnRqYXpdtmDBAjzwwAN48MEH6f+Gp556Gp2dndDyoCIZGjFBVOVOcnNSynvvmy5rxCVfbeO0R4hShi1RBqXiDKADTaMacqWE1VLyEQ7aHk6X1/58DP7+XJyDxEE3Oyz6W83aVTcFdbbBywynjYlF/ic4f7JTQCxsUJLPudNCuwjwjc08rr8sQJOVOI3tUuHjyBkJ9FKJTMKRh+M059QU+ukHDC0MexKJhJGVFUN2dpznvizG48xnAfoetqVTcpKfgkATO+SABH49rw7HHdkB45GG4sNQFkaDWkRDkPmxzHAapyyc3ZFIAJf/OB/z/57DmUTAPRRK8/1XV66sPI7Vt5mjWMNqq1SM9hpUGDcwnH4CNJ1a1qb5CouE8FAf6piio3KjcYvZJ7QhL8tCv+QHdLrYiI590pB5UE+jWqg8NN7niQkNRQKeJkgISxQQldEiRYtT8PbaNYErzm7GXf9biz137wUIKJsnrQRgQOuGAKHXP5oBpxOjBPTCNa+1y8P3f1KI+Y/nQAeOWvKuJenqF5es2ReDfIZLPiwA+dU94mlnqWRHk/F1mArBM0IdyOBlo54P3KcHXzupFULJ0qp75cIBIXoRQAVaIYSc2oRxYbvCMqd5AZ1ju4xoms0LrVcx9nkI7+n1MGlsCt88tRW/+H49Zn+hDfoP88gWjrwcq4F1EOVUyPjq1SG88u84XlkSQ1MLkYwAQt/a7uHaW8bg3kf4iQheMCjGliYbn3ph7cZbYGz9h2oaeiOKE+GC046TjQgRYJhx/drPpLboTS2Gwg5f+0IrzjqpHfoP4fxgw8AGaA6injQKmvLTdh0Tah1CxB09dMV0asHgzpie4I0vt7js7CbccXUt5pzRjPET03C8qiUrCByIN9SSGptCuOPuPJx2WSnO/F4J5lxbRF+M2ZeW4K4/5nE9Nbj2ZwW49+/ZYN+5cXMoK/Ab/vVSdZHK89/wwwKQuqPMVIEABgLhTg0ODEFcDV+Dd44Kj8cNLp7dhOsubsC4ModQ2PKrgEMy5UMB9blVTQfeQbf/aQKH070AAApRSURBVCu8RKenlfncUIVpMEWjLT45sxu3fb8Bf76lGuec3MpPTWl4XF8dz3ciAooMGhE/NRncdm8uPn1+Ka69vQBPPBfDm5VhVNeFsLbWwytvRnDtL/Jx0gWlnDazArAdF8KxRamVLy6pKsZ/8RmalvsEdlzwiRfBAnxOQ138hKJK0UnIcr3qIxtUIBwBjjfACsKJR3Ti8V9VYe4FjThyZgJTd02hgrcjxYVJlKofY1FCX1FosfukHkzfJ4GTj27D9Rc34++3V+MXV6/HUR/txJh8C7UY/VM1x/UsGHjac25clrwWwymXlmPebUVY1xBiP4Bo1CDKT1phDwjRRyM6eAzeWBkFuEykOcjKRve8tvDltZMH1bmtQGyGw1PU4shA/3YhkRBUruEKT45CFEWEJUNxrEcH8tAxEAsLTvqfLvzyh7W45/r1uPXyetz0nUbcdGkTbvqOxhtw03cb8JtranHfrTW49rJmfPqTHSjQHbEPQOdZDal0ppjmm6DQCPGnR3Jx7tVFWPTvMCJeGmHjsVVtXCuQjkgLvRMDDQEDw34VF6Sfen5x9T6k+K87MxwJfG4VDRw7h2DrX13vQZWuPMXqeyie/Bzr0aKh0lGBUG2ngYKCNPbftwezDkpgxoHdmHlwIogfxAv08lILzpzB2hZ8+gl4kA/lgzAkH8swsHCC8Mt7RuHKW/Kxel04sDbHxix8kpJI61hA11ftD6vymCDIivqIRf3rX1myZqveb1LaLXaqoi0mfi8hp660TieGmlF9r66OorGBLHmx66AafG+NQab7WagGNe6zvoKZZNjvNa1ey0ij6mdp4IIJgi9HWNSr9SjKv7g7Fz++czQ6Ex6yOD0qqCCAlDzYkcJxIIr2gCgyU2Pav93Kk41vvFV1ScB8O3lRvKFLws7X6tpH/BAmaKvXeXiVl7qqJBgZOuORqknBdAbVI0ews4pYPLUwjp/eU4CutEGEINk+cGlvHHIGwVSp6yRlEIIvXPMEwu4YvFAZORDb2WOGI8/q+tAdvnaQw9MjpzpuAha/GYNOYQI3HNYjUjcAg0cKlUS4PK9YHsG8nxWiqc1DTADh2gZYRiylVU9KnfsZOG5PBewFQfZ5zTJmVG+6Tr9NYbt43hGCan8nPujI2obQL/T/c1odqeysF7J44MksrKyitjgLDZrhCFcgDhAhDOxlT6fB7X8cjdd5PIhFLfRXN2x/e4wIpw3hGuigq7oFEVVH70AY9Z8aHNNPvj2F7NrQxeGArJtcnErzeAbLQ3wkJFjOG4y/PpED/TMs9nzozEeipigTB3A8LaH1/e2ZLMS45oHzqtAytdjR0hBYoi6iHtdAj2KzhNMoYebZUzCpPN1YW7v6UeW2vXkzXIFaEviV/i6ocCp1VIyCeMf9o/DKG1FwUA+X/bDqC9dAosGPrIL5D41CW4eB3msSJfKVwMpELEkIHqdKQk3787lcqs0BftqD/hWRjVSXscJ26dij4cm14IWa8yYUpx0vQDiwLcA7qQ5e9l5+0xjUrSd7jn6nmqGaQO/owUdDJ8I6wpS6/rim6ek0F6TX+koPpXeqdxZqqGkwHnjwIQ++WaRE5M2YB3Rw+nyMtytR3uhkrt0QgASCR6Lg8K6hQG1OwFqwXPciIaCmIXTQ00/rqo7t8qGGhy2X79Lup071SE3riPaMj7cqY7jsxkJ+rQ5DNxCBRtmUo9IDDTkJ9KfW4PryhPWhlqCeoGiSAURIq4k+i1IFQ0BSF3i9jAZpHMA02BQLPYHEga4ug4cWxNHG0BPtLqfJPhI4BcvBsC50KtXjA8ss13TdlKV9/4za2lUvMWu7ddqjYQv37CvVc/ad3NOTZscdlaxMjRE88Xwu5swrxvOvZAFBpgYWJMt45yD6wzpO9LQl4MCHMC2MiAjxZT7LECifojoEsyAYgnUhQkc6YYpep23hhXhjo4f/u2sUzri8BDf8Oh/RUIYGyouVg/YokwErKY8gj7JRpng0jWSvfLmqqupX2M4fM1Ly1Vdll4wvSNs0lU/dk60P4/lY/HoUZ11RhBvuzA/+wZuzgLBVIxa8D6baAGsYpyKpSuqXhapk0RSAvtAqJafDIKlFnN70X271e5tyaG0zeHRBNi6YOwZHn1NG4MZgwUsxNLeHgktvcIqE1iVIhtYn9I5eWacptOVFd9kY33b3YFJ19arfYwd4qK2RkfJfK1a0d3TJ5AlFKdfLD6cQmzG6kI+OboOf3JWPY8+twK135+ONt2LoaGfTCkjcwSgYtBrQasWz0NmsHxgRhwBwLecaBk6L4P4okTBY3+AF//jnr4/n4Ds3FuGIMyrw1StL8Jen8tBAC+zhJyPDaXF0luMHZwvLLxgKlvDOk6ZN5xNYoIeX8LlZaUwen3h90UtrQ7S8ypHRytbnQi2OXCNLl1VVNneb8gMmJ9PJZITToccBb2BE4BGYOir12ttH40vfKcW3birCLb8dhb88mIuXl8SxbEUYLW0OHR2CNn40bWfYyQN3R7sEn3WWLY9i4fNx/PWhXNzOq7Dv/XgM5swtxYnfKMXZ3y/hN7rczAHdOFgOoLT1sEtZr18wOvmbzgTymzq9EyNhP63XZ109Dp389JRIeSganXLTJvWu7EqESp94bt3egELM9w7iBgfgFnTqjTdW1z34VE3k4N27nwvT+vSuVH9JySiUtKY4v3i30Pr+9mQ2rv5pIb5xTREumFeIC64uxfn8MnDe1cUMS3D+3CJo/Lyry5guxtevKaQvwdfnFeGymwrx24fysGhxFIluQVbc5wABvxkahMNwe43vXecZd/izL9aElr6+5lSeV1vrq1fdbxHOgW/PHZVtf5YTt9ekfPnQy0vXmEeerZ5cOcx/94H/0mO2Urvu/ifXHRKPp8Z9aEpPXdRzwT6cCqP+BDotxiJATraFYdmaOg+vr4zgmX9l4YlFWfjH8zGG2dA/TX7i+ShefjUefItraKG4HATZMYcIp1MvDIgxiIXgdilN18Yj9nNvvV0ZevS5moo1azb+dXb9Pc36+lU/X7a86ryVKyu/21i78mXs4A81svV6sHRpTfX9j60rKzSxvIox/u+mjkum9eqNCqe1OHT3CnpSgmD3KkIrMvTCj6gGnngwXKs4E3KdctBvjtEIMDrboWxMyk2u6E2UjUr+OeHbqW8sX22eXrSmfMWKyvsA3anw/QFxZlv087llyzoWvrTmlL//syac02bi6TQ+v0tx7xMf3zfROnPPRPqQvTotpz03sbzHTSpNut0YTt+j282Y2uEfMqW79yN7J+p2KU792Vr7hdZOW/D8K9XmyYXrsvhR9YTaqqo3t0Ufttc2tgmA/9n5p1ev7llRWTn/2ZfWHX73I7X5f3q8NvzHv9d7j/1zrVnw/DrzzAvV5qlF1eYv/1hn7nu0LnTf4+tjf3ystmzBi9UnrFy5+g9r1qxp+U9+H/T4Ngfwg67wke7/TgBHWqPbmN9OALexwke6uZ0AjrRGtzG/bQLgNu7TB6q5nQDu4HDvBHAngDu4BnZw8f8fAAAA//+F7GgyAAAABklEQVQDAPPhUr4uxcntAAAAAElFTkSuQmCC" style="width:52px;height:52px;object-fit:contain;" /></div>
    <!-- nome -->
    <div style="font-size:2rem;font-weight:800;color:#f9fafb;letter-spacing:-0.03em;line-height:1.15;margin-bottom:8px;">
      Prosper<br><span style="color:#f2c230;">Distribuidora</span>
    </div>
    <div style="font-size:0.8rem;color:rgba(255,255,255,0.35);font-weight:600;
      letter-spacing:0.08em;text-transform:uppercase;margin-bottom:44px;">
      Gerenciamento de Telefones
    </div>
    <!-- features -->
    <div style="display:flex;flex-direction:column;gap:13px;width:100%;max-width:270px;text-align:left;">
      <div style="display:flex;align-items:center;gap:11px;font-size:0.8125rem;color:rgba(255,255,255,0.5);font-weight:500;">
        <span style="width:6px;height:6px;border-radius:50%;background:#f2c230;opacity:.7;flex-shrink:0;display:inline-block;"></span>
        Controle de linhas telefônicas
      </div>
      <div style="display:flex;align-items:center;gap:11px;font-size:0.8125rem;color:rgba(255,255,255,0.5);font-weight:500;">
        <span style="width:6px;height:6px;border-radius:50%;background:#f2c230;opacity:.7;flex-shrink:0;display:inline-block;"></span>
        Gestão por equipe e segmento
      </div>
      <div style="display:flex;align-items:center;gap:11px;font-size:0.8125rem;color:rgba(255,255,255,0.5);font-weight:500;">
        <span style="width:6px;height:6px;border-radius:50%;background:#f2c230;opacity:.7;flex-shrink:0;display:inline-block;"></span>
        Auditoria de movimentações
      </div>
      <div style="display:flex;align-items:center;gap:11px;font-size:0.8125rem;color:rgba(255,255,255,0.5);font-weight:500;">
        <span style="width:6px;height:6px;border-radius:50%;background:#f2c230;opacity:.7;flex-shrink:0;display:inline-block;"></span>
        Integração com Chamados TI
      </div>
    </div>
  </div>
</div>

<!-- ── Cabeçalho do formulário (lado direito) ── -->
<div style="font-family:'Plus Jakarta Sans',sans-serif;margin-bottom:28px;padding-top:16px;">
  <div style="font-size:1.625rem;font-weight:800;color:#f9fafb;letter-spacing:-0.02em;margin-bottom:6px;">
    Bem-vindo de volta
  </div>
  <div style="font-size:0.875rem;color:rgba(255,255,255,0.38);font-weight:400;">
    Acesse com suas credenciais de TI.
  </div>
</div>
"""


def _render_login_page() -> None:
    """Injeta painel de marca + CSS completo na tela de login."""
    st.markdown(_LOGIN_PAGE_HTML, unsafe_allow_html=True)


def _render_login_or_first_user() -> bool:
    """Wrapper para compatibilidade — delega ao Keycloak OIDC."""
    return _autenticar_keycloak()


def _audit(
    acao: str,
    entidade: str,
    chave: str = "",
    chamado_id: str = "",
    antes: dict | None = None,
    depois: dict | None = None,
    detalhes: str = "",
) -> None:
    """Wrapper seguro de auditoria (não quebra fluxo principal)."""
    if not (HAS_DB and registrar_auditoria):
        return
    try:
        user = st.session_state.get("user", {}) or {}
        chamado_ref = str(chamado_id or st.session_state.get("chamado_id") or "").strip()
        registrar_auditoria(
            acao=acao,
            entidade=entidade,
            chave_registro=chave,
            chamado_id=chamado_ref,
            antes=antes,
            depois=depois,
            detalhes=detalhes,
            user_id=str(user.get("id", "")),
            username=str(user.get("username", "")),
            origem="app",
        )
        # Unificacao progressiva: também registra no `chamado_eventos` quando houver contexto de chamado.
        if registrar_chamado_evento and chamado_ref and str(chamado_ref).isdigit():
            try:
                registrar_chamado_evento(
                    chamado_id=chamado_ref,
                    tipo_evento=str(acao or "").strip(),
                    descricao=str(detalhes or "").strip(),
                    antes=antes,
                    depois=depois,
                    user_id=str(user.get("id", "")),
                )
            except Exception:
                pass
    except Exception:
        pass


def _set_post_chamado_banner(default_text: str = "Alteração concluída com sucesso.") -> None:
    """Prepara um banner de retorno ao chamado após ações relevantes."""
    chamado_context = st.session_state.get("chamado_context") or {}
    chamado_id = str(chamado_context.get("chamado_id") or st.session_state.get("chamado_id") or "").strip()
    return_url = str(chamado_context.get("return_url") or "").strip()
    if not chamado_id and not return_url:
        return
    st.session_state["_post_chamado_return_banner"] = {
        "text": default_text,
        "chamado_id": chamado_id,
        "return_url": return_url,
    }


def _auto_mark_context_line_vago(df_current: pd.DataFrame, modo_db: str, dados_do_banco: bool) -> pd.DataFrame:
    """Marca Nome/Codigo como VAGO ao abrir um chamado com linha em contexto."""
    chamado_context = st.session_state.get("chamado_context") or {}
    chamado_id = str(chamado_context.get("chamado_id") or st.session_state.get("chamado_id") or "").strip()
    linha_ctx = str(chamado_context.get("linha") or "").strip()
    auto_signature = f"{chamado_id}|{linha_ctx}|{modo_db}"
    if not chamado_id or not linha_ctx:
        return df_current
    if st.session_state.get("_auto_vago_signature") == auto_signature:
        return df_current
    if not (HAS_DB and dados_do_banco and save_linhas and load_linhas):
        st.session_state["_auto_vago_signature"] = auto_signature
        st.session_state["_post_conflict_feedback"] = {
            "type": "warning",
            "text": "Nao foi possivel marcar a linha como VAGO automaticamente porque o banco nao esta ativo.",
        }
        return df_current

    modos_tentativa = [modo_db] + [m for m in ["ativas", "desativadas"] if m != modo_db]
    for modo_target in modos_tentativa:
        if not has_data(modo_target):
            continue
        df_mode = load_linhas(modo=modo_target)
        if "Linha" not in df_mode.columns:
            continue
        mask_linha = df_mode["Linha"].fillna("").astype(str).str.strip() == linha_ctx
        if not mask_linha.any():
            continue
        if int(mask_linha.sum()) > 1:
            st.session_state["_auto_vago_signature"] = auto_signature
            st.session_state["_post_conflict_feedback"] = {
                "type": "warning",
                "text": f"A linha `{linha_ctx}` apareceu mais de uma vez e nao foi marcada como VAGO automaticamente.",
            }
            return df_current

        row_idx = df_mode[mask_linha].index[0]
        before_row = df_mode.loc[row_idx].to_dict()
        nome_norm = normalize_text(before_row.get("Nome", ""))
        codigo_norm = normalize_text(before_row.get("Codigo", ""))
        if nome_norm == "vago" and codigo_norm == "vago":
            st.session_state["_auto_vago_signature"] = auto_signature
            return df_mode if modo_target == modo_db else df_current

        df_mode.loc[row_idx, "Nome"] = "VAGO"
        if "Codigo" in df_mode.columns:
            df_mode.loc[row_idx, "Codigo"] = "VAGO"
        save_linhas(df_mode, modo=modo_target)
        _audit(
            acao="marcar_linha_vaga_chamado",
            entidade="linhas",
            chave=linha_ctx,
            chamado_id=chamado_id,
            antes=before_row,
            depois={"Nome": "VAGO", "Codigo": "VAGO", "modo": modo_target},
            detalhes="Nome e codigo marcados como VAGO automaticamente ao abrir chamado.",
        )
        st.session_state["_auto_vago_signature"] = auto_signature
        st.session_state["_post_conflict_feedback"] = {
            "type": "success",
            "text": f"Linha `{linha_ctx}` marcada como VAGO automaticamente para o chamado `{chamado_id}`.",
        }
        return df_mode if modo_target == modo_db else df_current

    st.session_state["_auto_vago_signature"] = auto_signature
    st.session_state["_post_conflict_feedback"] = {
        "type": "warning",
        "text": f"A linha `{linha_ctx}` nao foi encontrada para marcar como VAGO automaticamente.",
    }
    return df_current


def main() -> None:
    st.set_page_config(
        page_title="Gerenciamento de Telefones",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    # Restaura preferência de modo claro/escuro.
    if "color_mode" not in st.session_state:
        theme_from_url = st.query_params.get("theme", "")
        if theme_from_url in ("light", "dark"):
            st.session_state["color_mode"] = theme_from_url
        else:
            st.session_state["color_mode"] = "dark"

    if not _render_login_or_first_user():
        return

    cwd = Path.cwd()
    rules_path = DOC_DIR / RULES_FILE if (DOC_DIR / RULES_FILE).exists() else cwd / RULES_FILE

    query_params = st.query_params

    def _qp_value(*keys: str) -> str:
        for qp_key in keys:
            raw_qp = query_params.get(qp_key)
            if isinstance(raw_qp, list):
                raw_qp = raw_qp[0] if raw_qp else ""
            raw_qp = str(raw_qp or "").strip()
            if raw_qp:
                return raw_qp
        return ""

    chamado_qp = _qp_value("chamado_id", "id_chamado", "ticket_id", "chamado")
    linha_ctx_qp = _qp_value("linha", "linha_telefone", "numero_linha", "phone_line")
    segmento_ctx_explicit = _qp_value("segmento_chamado", "ctx_segmento")
    equipe_ctx_explicit = _qp_value("equipe_chamado", "ctx_equipe")
    busca_ctx_explicit = _qp_value("busca_chamado", "ctx_busca")
    return_url_qp = _qp_value("return_url", "return_to", "back_url", "chamados_url")
    contexto_externo_base = bool(chamado_qp or linha_ctx_qp or segmento_ctx_explicit or equipe_ctx_explicit or busca_ctx_explicit or return_url_qp)
    segmento_ctx_qp = segmento_ctx_explicit or (_qp_value("segmento") if contexto_externo_base else "")
    equipe_ctx_qp = equipe_ctx_explicit or (_qp_value("equipe") if contexto_externo_base else "")
    busca_ctx_qp = busca_ctx_explicit or (_qp_value("busca") if contexto_externo_base else "")

    if chamado_qp:
        st.session_state["chamado_id"] = chamado_qp
    elif "chamado_id" not in st.session_state:
        st.session_state["chamado_id"] = ""

    chamado_context = {
        "chamado_id": chamado_qp,
        "linha": linha_ctx_qp,
        "segmento": segmento_ctx_qp,
        "equipe": equipe_ctx_qp,
        "busca": busca_ctx_qp,
        "return_url": return_url_qp,
    }
    has_chamado_context = any(bool(v) for v in chamado_context.values())
    ctx_signature = "|".join(chamado_context.get(k, "") for k in ["chamado_id", "linha", "segmento", "equipe", "busca", "return_url"])
    segmentos_validos_ctx = ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"]
    if has_chamado_context and st.session_state.get("_chamado_context_signature") != ctx_signature:
        st.session_state["chamado_context"] = chamado_context
        st.session_state["_chamado_context_signature"] = ctx_signature
        if chamado_context["segmento"] in segmentos_validos_ctx:
            st.session_state["nav_segmento"] = chamado_context["segmento"]
        if chamado_context["equipe"]:
            st.session_state["filtro_equipe"] = chamado_context["equipe"]
        if chamado_context["linha"]:
            st.session_state["filtro_busca"] = chamado_context["linha"]
        elif chamado_context["busca"]:
            st.session_state["filtro_busca"] = chamado_context["busca"]
    elif not has_chamado_context:
        st.session_state["chamado_context"] = {}
        st.session_state["_chamado_context_signature"] = ""
    if "_qp_hydrated" not in st.session_state:
        modo_qp = query_params.get("modo")
        seg_qp = query_params.get("segmento")
        equipe_qp = query_params.get("equipe")
        busca_qp = query_params.get("busca")
        busca_tipo_qp = query_params.get("busca_tipo")
        somente_vagas_qp = query_params.get("somente_vagas")
        if modo_qp in ["Linhas ativas", "Linhas desativadas"]:
            st.session_state["nav_modo"] = modo_qp
        if seg_qp in ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"]:
            st.session_state["nav_segmento"] = seg_qp
        if equipe_qp:
            st.session_state["filtro_equipe"] = equipe_qp
        if busca_qp is not None:
            st.session_state["filtro_busca"] = busca_qp
        if busca_tipo_qp in ["Geral", "Linha", "IMEI", "Aparelho", "Motivo"]:
            st.session_state["filtro_busca_tipo"] = busca_tipo_qp
        if somente_vagas_qp in ["0", "1"]:
            st.session_state["filtro_somente_vagas"] = somente_vagas_qp == "1"
        st.session_state["_qp_hydrated"] = True
    chamados_app_url = get_chamados_app_url()

    # ── Sidebar Prosper ───────────────────────────────────────────────────────
    _seg_opts = ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"]
    _seg_pref = st.session_state.get("pending_segmento") or st.session_state.get("nav_segmento")
    if _seg_pref in _seg_opts:
        st.session_state["nav_segmento"] = _seg_pref
    if st.session_state.get("nav_segmento") not in _seg_opts:
        st.session_state["nav_segmento"] = _seg_opts[0]
    if st.session_state.get("nav_modo") not in ["Linhas ativas", "Linhas desativadas"]:
        st.session_state["nav_modo"] = "Linhas ativas"
    # Honra pagina via query_param na primeira carga
    _pagina_qp = str(query_params.get("pagina") or "painel").strip()
    if _pagina_qp in ["painel", "config"] and "current_page" not in st.session_state:
        st.session_state["current_page"] = _pagina_qp

    nav = render_sidebar(
        user=st.session_state.get("user", {}),
        current_page=st.session_state.get("current_page", "painel"),
    )

    if nav["logout_clicked"]:
        st.session_state.authenticated = False
        st.session_state.user = None
        st.session_state.pop("access_token", None)
        st.rerun()

    modo_sel = nav["modo"]
    segmento_sel = nav["segmento"]
    current_page = nav["pagina"]
    agrupar_por = nav.get("agrupar_por", "Equipe")
    st.session_state["nav_modo"] = modo_sel
    st.session_state["nav_segmento"] = segmento_sel
    st.session_state["current_page"] = current_page
    modo_db = "ativas" if modo_sel == "Linhas ativas" else "desativadas"
    segmento_sem_filtro_modo = segmento_sel in ("Manutenção", "Roubo e Perda")


    def _render_config_content() -> None:
        if st.session_state.get("user", {}).get("is_admin"):
            with st.expander("Gerenciar usuários", expanded=False):
                keycloak_admin_url = f"{_KEYCLOAK_URL}/admin/master/console/#/{_KEYCLOAK_REALM}/users"
                st.info(
                    "Usuários são gerenciados no **Keycloak Admin Console**. "
                    "Crie, edite, desative ou redefina senhas diretamente lá."
                )
                st.link_button("Abrir Keycloak Admin Console", url=keycloak_admin_url)
            st.divider()
        st.markdown("**Configurações gerais**")
        if HAS_DB:
            if is_postgres_configured():
                st.success("PostgreSQL: conexão ok.")
            else:
                st.warning("PostgreSQL: não configurado.")
        if st.session_state.get("user", {}).get("is_admin"):
            st.divider()
            with st.expander("Histórico de alterações (auditoria)", expanded=False):
                chamado_context_audit = str(((st.session_state.get("chamado_context") or {}).get("chamado_id") or "")).strip()
                c_audit_1, c_audit_2 = st.columns([2.2, 1])
                with c_audit_1:
                    audit_chamado = st.text_input(
                        "Filtrar por chamado",
                        value=st.session_state.get("audit_chamado_filter", ""),
                        placeholder="Ex.: 123",
                        key="audit_chamado_filter",
                    ).strip()
                with c_audit_2:
                    st.caption(" ")
                    if chamado_context_audit and st.button("Usar chamado atual", key="btn_use_current_chamado_filter"):
                        st.session_state["audit_chamado_filter"] = chamado_context_audit
                        st.rerun()
                audit_limit = st.number_input(
                    "Quantidade de registros",
                    min_value=20,
                    max_value=2000,
                    value=200,
                    step=20,
                    key="audit_limit",
                )
                logs = listar_auditoria(
                    limit=int(audit_limit),
                    chamado_id=audit_chamado,
                ) if listar_auditoria else []
                if not logs:
                    st.caption("Sem registros de auditoria.")
                else:
                    def _parse_json_safe(value: Any) -> Any:
                        if value is None:
                            return None
                        if isinstance(value, float) and pd.isna(value):
                            return None
                        if isinstance(value, (dict, list, int, float, bool)):
                            return value
                        txt = str(value).strip()
                        if not txt or txt.lower() in {"none", "nan"}:
                            return None
                        try:
                            return json.loads(txt)
                        except Exception:
                            return txt

                    def _friendly_action_name(action: str) -> str:
                        action_map = {
                            "salvar_edicoes": "Salvar edições",
                            "criar_linha": "Criar linha",
                            "excluir_linha": "Excluir linha",
                            "mover_equipe": "Mover equipe",
                            "mover_setor": "Mover setor",
                            "mudar_modo_linha": "Ativar/Desativar linha",
                            "enviar_manutencao": "Enviar para manutenção",
                            "sincronizar_banco": "Sincronizar banco",
                            "login": "Login",
                            "logout": "Logout",
                        }
                        action_norm = str(action or "").strip()
                        return action_map.get(action_norm, action_norm.replace("_", " ").title())

                    def _build_what_edited(row: pd.Series) -> str:
                        acao_ev = str(row.get("acao", "") or "").strip()
                        antes_ev = _parse_json_safe(row.get("antes_json"))
                        depois_ev = _parse_json_safe(row.get("depois_json"))
                        detalhes_ev = str(row.get("detalhes", "") or "").strip()
                        chave_ev = str(row.get("chave_registro", "") or "").strip()

                        if acao_ev == "salvar_edicoes" and isinstance(depois_ev, dict):
                            alteracoes: list[dict[str, Any]] = []
                            total_alteracoes = 0
                            if isinstance(antes_ev, dict):
                                raw_alter = antes_ev.get("alteracoes") or []
                                if isinstance(raw_alter, list):
                                    alteracoes = raw_alter
                                total_alteracoes = int(antes_ev.get("alteracoes_total") or 0)
                            if not alteracoes and isinstance(depois_ev, dict):
                                raw_alter = depois_ev.get("alteracoes") or []
                                if isinstance(raw_alter, list):
                                    alteracoes = raw_alter
                                total_alteracoes = int(depois_ev.get("alteracoes_total") or total_alteracoes or 0)

                            def _fmt_val(v: Any) -> str:
                                txt = "—" if v is None else str(v).strip()
                                if not txt:
                                    txt = "—"
                                return txt if len(txt) <= 36 else (txt[:33] + "...")

                            if alteracoes:
                                partes: list[str] = []
                                for ch in alteracoes[:2]:
                                    if not isinstance(ch, dict):
                                        continue
                                    ln = str(ch.get("linha") or "—").strip() or "—"
                                    campo = str(ch.get("campo") or "—").strip() or "—"
                                    antes_v = _fmt_val(ch.get("antes"))
                                    depois_v = _fmt_val(ch.get("depois"))
                                    partes.append(f"Linha {ln} | {campo}: {antes_v} -> {depois_v}")
                                if partes:
                                    total_base = total_alteracoes if total_alteracoes > 0 else len(alteracoes)
                                    extra = max(0, total_base - len(partes))
                                    sufixo = f" (+{extra} alteração(ões))" if extra > 0 else ""
                                    return " ; ".join(partes) + sufixo

                            segmento = str(depois_ev.get("segmento", "") or "").strip() or "—"
                            modo_ev = str(depois_ev.get("modo", "") or "").strip() or "—"
                            linhas_editadas: list[str] = []
                            campos_alterados: list[str] = []
                            if isinstance(antes_ev, dict):
                                linhas_editadas = antes_ev.get("linhas_editadas") or []
                                campos_alterados = antes_ev.get("campos_alterados") or []
                            if not campos_alterados and isinstance(depois_ev, dict):
                                campos_alterados = depois_ev.get("campos_alterados") or []
                            qtd_linhas = len(linhas_editadas) if isinstance(linhas_editadas, list) else 0
                            campos_validos = [str(c).strip() for c in campos_alterados if str(c).strip()]
                            campos_preview = ", ".join(campos_validos[:3])
                            if len(campos_validos) > 3:
                                campos_preview += f" +{len(campos_validos) - 3}"
                            if qtd_linhas == 1 and isinstance(linhas_editadas, list) and linhas_editadas:
                                alvo = f"linha {str(linhas_editadas[0]).strip()}"
                            else:
                                alvo = f"{qtd_linhas} linha(s)"
                            if campos_preview:
                                return f"Alterou {alvo} em {segmento} ({modo_ev}) | Campos: {campos_preview}"
                            return f"Alterou {alvo} em {segmento} ({modo_ev})"
                        if acao_ev == "criar_linha":
                            return f"Criou a linha {chave_ev or '—'}"
                        if acao_ev == "excluir_linha":
                            return f"Excluiu a linha {chave_ev or '—'}"
                        if acao_ev == "mover_equipe":
                            eq_dest = ""
                            if isinstance(depois_ev, dict):
                                eq_dest = str(depois_ev.get('EquipePadrao') or depois_ev.get('Equipe') or "").strip()
                            return f"Moveu a linha {chave_ev or '—'} para equipe {eq_dest or '—'}"
                        if acao_ev == "mover_setor":
                            seg_dest = ""
                            if isinstance(depois_ev, dict):
                                seg_dest = str(depois_ev.get('Segmento') or "").strip()
                            return f"Moveu a linha {chave_ev or '—'} para o setor {seg_dest or '—'}"
                        if acao_ev == "mudar_modo_linha":
                            modo_dest = ""
                            if isinstance(depois_ev, dict):
                                modo_dest = str(depois_ev.get('modo') or "").strip()
                            return f"Alterou a linha {chave_ev or '—'} para {modo_dest or '—'}"
                        if acao_ev == "enviar_manutencao":
                            return f"Enviou para manutenção: {chave_ev or '—'}"
                        if acao_ev in ("sincronizar_banco", "login", "logout"):
                            return f"{_friendly_action_name(acao_ev)}"

                        if detalhes_ev:
                            return detalhes_ev
                        return _friendly_action_name(acao_ev) or "Edição"

                    def _extract_chamado_id(row: pd.Series) -> str:
                        raw = row.get("chamado_id", "")
                        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                            return "—"
                        txt = str(raw).strip()
                        if not txt or txt.lower() == "nan":
                            return "—"
                        return txt

                    def _to_local_datetime(series: pd.Series) -> pd.Series:
                        # `criado_em` no SQLite vem em UTC; converter para horário local melhora a leitura.
                        dt_utc = pd.to_datetime(series, errors="coerce", utc=True)
                        tz_name = os.environ.get("APP_TIMEZONE", "America/Sao_Paulo")
                        try:
                            return dt_utc.dt.tz_convert(tz_name)
                        except Exception:
                            return dt_utc

                    df_logs = pd.DataFrame(logs)
                    acoes_edicao = {
                        "salvar_edicoes",
                        "criar_linha",
                        "excluir_linha",
                        "mover_equipe",
                        "mover_setor",
                        "mudar_modo_linha",
                        "enviar_manutencao",
                    }
                    if "acao" in df_logs.columns:
                        df_logs = df_logs[df_logs["acao"].isin(acoes_edicao)]
                    if "criado_em" in df_logs.columns:
                        df_logs["__ord_dt"] = _to_local_datetime(df_logs["criado_em"])
                        sort_cols = ["__ord_dt"] + (["id"] if "id" in df_logs.columns else [])
                        sort_asc = [False] * len(sort_cols)
                        df_logs = df_logs.sort_values(sort_cols, ascending=sort_asc, kind="stable")
                    elif "id" in df_logs.columns:
                        df_logs = df_logs.sort_values(["id"], ascending=[False], kind="stable")
                    df_logs = df_logs.reset_index(drop=True)
                    if df_logs.empty:
                        st.caption("Sem alterações de edição registradas.")
                    else:
                        df_view = pd.DataFrame()
                        if "username" in df_logs.columns:
                            df_view["Quem editou"] = df_logs["username"].fillna("").astype(str).str.strip().replace("", "—")
                        else:
                            df_view["Quem editou"] = "—"
                        df_view["Chamado"] = df_logs.apply(_extract_chamado_id, axis=1)
                        df_view["O que editou"] = df_logs.apply(_build_what_edited, axis=1)
                        dt_fmt = df_logs["__ord_dt"] if "__ord_dt" in df_logs.columns else _to_local_datetime(df_logs.get("criado_em"))
                        df_view["Quando editou"] = dt_fmt.dt.strftime("%d/%m/%Y %H:%M:%S").fillna("—")
                        st.dataframe(df_view[["Quem editou", "Chamado", "O que editou", "Quando editou"]], width="stretch", hide_index=True)

    def _render_chamados_content() -> None:
        st.markdown("### 📌 Sistema de Chamados")
        if not HAS_DB or not criar_chamado or not listar_chamados:
            st.error("Módulo de chamados indisponível no momento.")
            return

        st.caption("Abra, acompanhe e atualize chamados no mesmo sistema e no mesmo banco.")
        known_linhas = []
        try:
            df_linhas_ref = load_linhas(modo=modo_db)
            if "Linha" in df_linhas_ref.columns:
                known_linhas = sorted(
                    {
                        str(v).strip()
                        for v in df_linhas_ref["Linha"].fillna("").astype(str).tolist()
                        if str(v).strip()
                    }
                )
        except Exception:
            known_linhas = []

        with st.expander("Abrir chamado", expanded=True):
            c_new_1, c_new_2 = st.columns([1.2, 1.2])
            with c_new_1:
                ch_titulo = st.text_input("Título", key="ch_novo_titulo", placeholder="Ex.: Troca de aparelho")
                ch_tipo = st.selectbox(
                    "Tipo",
                    options=["gerenciamento", "incidente", "solicitacao", "manutencao", "roubo_perda"],
                    key="ch_novo_tipo",
                )
                ch_prioridade = st.selectbox(
                    "Prioridade",
                    options=["baixa", "normal", "alta", "critica"],
                    index=1,
                    key="ch_novo_prioridade",
                )
            with c_new_2:
                ch_status = st.selectbox(
                    "Status inicial",
                    options=["aberto", "em_andamento", "aguardando", "resolvido", "fechado"],
                    key="ch_novo_status",
                )
                ch_linha = st.selectbox(
                    "Linha (opcional)",
                    options=[""] + known_linhas,
                    key="ch_novo_linha",
                )
                ch_origem = st.selectbox(
                    "Origem",
                    options=["app", "gerenciamento", "manual", "externo"],
                    index=0,
                    key="ch_novo_origem",
                )
            ch_desc = st.text_area(
                "Descrição",
                key="ch_novo_descricao",
                placeholder="Descreva o chamado com detalhes do que precisa ser feito.",
                height=110,
            )
            if st.button("Abrir chamado", key="btn_abrir_chamado"):
                titulo = str(ch_titulo or "").strip()
                descricao = str(ch_desc or "").strip()
                if not titulo or not descricao:
                    st.error("Título e descrição são obrigatórios.")
                else:
                    user = st.session_state.get("user", {}) or {}
                    novo = criar_chamado(
                        titulo=titulo,
                        descricao=descricao,
                        tipo=str(ch_tipo or "").strip(),
                        prioridade=str(ch_prioridade or "").strip(),
                        status=str(ch_status or "").strip(),
                        linha_num=str(ch_linha or "").strip(),
                        solicitante_id=str(user.get("id", "")),
                        origem=str(ch_origem or "").strip(),
                    )
                    if not novo:
                        st.error("Não foi possível abrir o chamado.")
                    else:
                        chamado_id_new = str(novo.get("id") or "")
                        if chamado_id_new and registrar_chamado_evento:
                            try:
                                registrar_chamado_evento(
                                    chamado_id=chamado_id_new,
                                    tipo_evento="chamado_aberto",
                                    descricao=f"Abertura do chamado {novo.get('numero_chamado')}",
                                    depois={
                                        "status": novo.get("status"),
                                        "prioridade": novo.get("prioridade"),
                                        "linha": novo.get("linha_numero"),
                                        "tipo": novo.get("tipo"),
                                    },
                                    user_id=str(user.get("id", "")),
                                )
                            except Exception:
                                pass
                        _audit(
                            acao="abrir_chamado",
                            entidade="chamados",
                            chave=str(novo.get("numero_chamado") or novo.get("id") or ""),
                            chamado_id=chamado_id_new,
                            depois=novo,
                            detalhes="Chamado aberto pela página unificada.",
                        )
                        st.success(f"Chamado {novo.get('numero_chamado')} aberto com sucesso.")
                        st.session_state["chamado_detalhe_id"] = int(novo.get("id") or 0)
                        st.rerun()

        st.divider()
        c_f1, c_f2, c_f3 = st.columns([1.2, 2.0, 0.8])
        with c_f1:
            status_filter = st.selectbox(
                "Filtrar status",
                options=["", "aberto", "em_andamento", "aguardando", "resolvido", "fechado"],
                format_func=lambda x: "Todos" if x == "" else x,
                key="ch_filtro_status",
            )
        with c_f2:
            busca_filter = st.text_input(
                "Buscar chamado",
                key="ch_filtro_busca",
                placeholder="Número, título, descrição ou linha",
            ).strip()
        with c_f3:
            limit_filter = st.number_input("Limite", min_value=20, max_value=500, value=120, step=20, key="ch_filtro_limite")

        chamados_rows = listar_chamados(limit=int(limit_filter), status=status_filter, busca=busca_filter)
        if not chamados_rows:
            st.info("Nenhum chamado encontrado com os filtros atuais.")
            return

        df_ch = pd.DataFrame(chamados_rows)
        df_ch["aberto_em_fmt"] = pd.to_datetime(df_ch["aberto_em"], errors="coerce").dt.strftime("%d/%m/%Y %H:%M:%S")
        df_ch["atualizado_em_fmt"] = pd.to_datetime(df_ch["atualizado_em"], errors="coerce").dt.strftime("%d/%m/%Y %H:%M:%S")
        st.dataframe(
            df_ch[["id", "numero_chamado", "status", "prioridade", "tipo", "linha_numero", "titulo", "aberto_em_fmt", "atualizado_em_fmt"]],
            width="stretch",
            hide_index=True,
        )

        detail_options = [f"{int(r.get('id'))} • #{str(r.get('numero_chamado') or '')} • {str(r.get('titulo') or '')}" for r in chamados_rows if r.get("id")]
        selected_opt = st.selectbox("Detalhar chamado", options=detail_options, key="ch_detail_select")
        selected_id = int(str(selected_opt).split("•")[0].strip()) if selected_opt else 0
        if selected_id and obter_chamado:
            detalhe = obter_chamado(str(selected_id))
            if detalhe:
                st.markdown("#### Detalhe do chamado")
                d1, d2, d3 = st.columns([1, 1, 2])
                with d1:
                    st.caption(f"ID: {detalhe.get('id')}")
                    st.caption(f"Número: {detalhe.get('numero_chamado')}")
                with d2:
                    st.caption(f"Status atual: {detalhe.get('status')}")
                    st.caption(f"Prioridade: {detalhe.get('prioridade')}")
                with d3:
                    st.caption(f"Linha: {detalhe.get('linha_numero') or '—'}")
                    st.caption(f"Tipo: {detalhe.get('tipo')}")
                st.text_area("Descrição atual", value=str(detalhe.get("descricao") or ""), height=100, disabled=True, key=f"desc_current_{selected_id}")

                new_status = st.selectbox(
                    "Novo status",
                    options=["aberto", "em_andamento", "aguardando", "resolvido", "fechado"],
                    index=["aberto", "em_andamento", "aguardando", "resolvido", "fechado"].index(str(detalhe.get("status") or "aberto")) if str(detalhe.get("status") or "aberto") in ["aberto", "em_andamento", "aguardando", "resolvido", "fechado"] else 0,
                    key=f"status_new_{selected_id}",
                )
                if st.button("Atualizar status do chamado", key=f"btn_update_status_{selected_id}"):
                    status_old = str(detalhe.get("status") or "").strip()
                    status_new = str(new_status or "").strip()
                    if status_new == status_old:
                        st.info("O chamado já está com esse status.")
                    else:
                        ok_status = atualizar_status_chamado(str(selected_id), status_new) if atualizar_status_chamado else False
                        if not ok_status:
                            st.error("Não foi possível atualizar o status.")
                        else:
                            user = st.session_state.get("user", {}) or {}
                            if registrar_chamado_evento:
                                try:
                                    registrar_chamado_evento(
                                        chamado_id=str(selected_id),
                                        tipo_evento="status_alterado",
                                        descricao=f"Status alterado de {status_old} para {status_new}",
                                        antes={"status": status_old},
                                        depois={"status": status_new},
                                        user_id=str(user.get("id", "")),
                                    )
                                except Exception:
                                    pass
                            _audit(
                                acao="atualizar_status_chamado",
                                entidade="chamados",
                                chave=str(detalhe.get("numero_chamado") or selected_id),
                                chamado_id=str(selected_id),
                                antes={"status": status_old},
                                depois={"status": status_new},
                                detalhes="Atualização de status pela página de chamados.",
                            )
                            st.success("Status atualizado com sucesso.")
                            st.rerun()
    # ── Barra de ações (topo do conteúdo principal) ──────────────────────────
    if current_page != "config":
        _bar_add, _bar_chamados, _ = st.columns([1, 1.2, 3])
        with _bar_add:
            with st.popover("➕ Adicionar"):
                if HAS_DB:
                    st.caption("Passo 1: escolha Segmento e Equipe. Passo 2: complete os dados na tabela.")
                    seg_add = st.selectbox("Segmento", ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"], key="add_seg")
                    eq_ref = (
                        EQUIPES_ALIMENTO
                        if seg_add == "Alimento"
                        else (
                            EQUIPES_MEDICAMENTO
                            if seg_add == "Medicamento"
                            else (
                                EQUIPES_PROMOTORES
                                if seg_add == "Promotores"
                                else (
                                    EQUIPES_INTERNOS
                                    if seg_add == "Internos"
                                    else (EQUIPES_MANUTENCAO if seg_add == "Manutenção" else EQUIPES_ROUBO_PERDA)
                                )
                            )
                        )
                    )
                    eq_add = st.selectbox("Equipe", eq_ref, key="add_equipe")
                    if st.button("Criar linha na equipe", key="add_criar_linha"):
                        try:
                            df_atual = load_linhas(modo=modo_db)
                            cols = df_atual.columns.tolist() or [
                                "Codigo", "Nome", "Equipe", "EquipePadrao", "GrupoEquipe", "TipoEquipe",
                                "Localidade", "Gestor", "Supervisor", "Segmento", "Papel", "Linha",
                                "E-mail", "Gerenciamento", "IMEI A", "IMEI B", "CHIP", "Aparelho", "Modelo",
                                "Setor", "Cargo", "Desconto", "Perfil", "Empresa", "Ativo",
                                "Numero de Serie", "Operadora", "Aba",
                            ]
                            linha_temp = f"NOVA-{secrets.token_hex(3).upper()}"
                            new_row = {c: "" for c in cols}
                            new_row.update({
                                "Segmento": seg_add,
                                "EquipePadrao": eq_add,
                                "Equipe": eq_add,
                                "GrupoEquipe": seg_add,
                                "TipoEquipe": "",
                                "Linha": linha_temp,
                            })
                            df_novo = pd.concat([df_atual, pd.DataFrame([new_row])], ignore_index=True)
                            df_novo = df_novo.sort_values(
                                ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
                                kind="stable",
                            )
                            save_linhas(df_novo, modo=modo_db)
                            _audit(
                                acao="criar_linha",
                                entidade="linhas",
                                chave=linha_temp,
                                depois={
                                    "segmento": seg_add,
                                    "equipe": eq_add,
                                    "modo": modo_db,
                                    "linha": linha_temp,
                                },
                                detalhes="Linha criada via botão Adicionar",
                            )
                            _set_post_chamado_banner("Linha criada com sucesso.")
                            st.session_state.pending_segmento = seg_add
                            st.session_state.pending_equipe = eq_add
                            st.session_state.pending_linha = linha_temp
                            st.session_state.filtro_busca = ""
                            st.success("Linha criada! Complete os dados na tabela abaixo.")
                            st.rerun()
                        except Exception as exc:
                            import logging as _log
                            _log.getLogger(__name__).error("Erro ao criar linha: %s", exc, exc_info=True)
                            st.error("Não foi possível criar a linha. Verifique os dados e tente novamente.")
                else:
                    st.caption("Banco indisponivel no momento para adicionar linhas.")
        with _bar_chamados:
            if chamados_app_url and st.session_state.get("authenticated"):
                _token = st.session_state.get("access_token", "")
                _sep = "&amp;" if "?" in chamados_app_url else "?"
                _dest = f"{chamados_app_url}{_sep}sso_token={_token}" if _token else chamados_app_url
                _safe_url = str(_dest).replace('"', "&quot;").replace("<", "&lt;").replace(">", "&gt;")
                st.markdown(
                    f'<a href="{_safe_url}" target="_blank" style="display:inline-flex;align-items:center;gap:6px;'
                    f'height:2.5rem;padding:0 1rem;border-radius:6px;border:none;background:#f2c230;'
                    f'color:#111827;font-weight:600;text-decoration:none;white-space:nowrap;font-size:0.875rem;'
                    f'box-shadow:0 1px 3px rgba(0,0,0,.2);">📌 Chamados</a>',
                    unsafe_allow_html=True,
                )

    if current_page == "config":
        try:
            from src.pages.config_admin import render_config_admin
            render_config_admin()
        except Exception as _cfg_exc:
            import logging as _log
            _log.getLogger(__name__).error("Erro ao carregar config_admin: %s", _cfg_exc, exc_info=True)
            st.error("Não foi possível carregar a página de configuração. Recarregue e tente novamente.")
        return

    active_chamado_context = st.session_state.get("chamado_context", {}) or {}
    if any(bool(v) for v in active_chamado_context.values()):
        contexto_partes: list[str] = []
        if active_chamado_context.get("chamado_id"):
            contexto_partes.append(f"**Chamado:** {active_chamado_context['chamado_id']}")
        if active_chamado_context.get("linha"):
            contexto_partes.append(f"**Linha alvo:** {active_chamado_context['linha']}")
        if active_chamado_context.get("segmento"):
            contexto_partes.append(f"**Segmento:** {active_chamado_context['segmento']}")
        if active_chamado_context.get("equipe"):
            contexto_partes.append(f"**Equipe:** {active_chamado_context['equipe']}")
        contexto_partes.append(f"**Usuário:** {st.session_state.get('user', {}).get('username', '—')}")
        st.info("Contexto de chamado ativo. Filtros aplicados automaticamente.  \n" + " | ".join(contexto_partes))
        if active_chamado_context.get("return_url"):
            st.link_button("Voltar para o chamado", url=str(active_chamado_context["return_url"]), use_container_width=False)

    post_conflict_feedback = st.session_state.get("_post_conflict_feedback") or {}
    if post_conflict_feedback:
        feedback_type = str(post_conflict_feedback.get("type") or "").strip().lower()
        feedback_text = str(post_conflict_feedback.get("text") or "").strip()
        if feedback_text:
            if feedback_type == "success":
                st.success(feedback_text)
            else:
                st.warning(feedback_text)
        st.session_state["_post_conflict_feedback"] = {}

    post_chamado_banner = st.session_state.get("_post_chamado_return_banner") or {}
    if post_chamado_banner:
        banner_text = str(post_chamado_banner.get("text") or "").strip() or "Alteração concluída com sucesso."
        chamado_banner = str(post_chamado_banner.get("chamado_id") or "").strip()
        return_url_banner = str(post_chamado_banner.get("return_url") or "").strip()
        if chamado_banner:
            banner_text = f"{banner_text} Chamado: {chamado_banner}."
        st.success(banner_text)
        if return_url_banner:
            st.link_button("Retornar ao chamado", url=return_url_banner, use_container_width=False)
        st.session_state["_post_chamado_return_banner"] = {}

    # Carregar dados: somente banco de dados
    if not HAS_DB:
        st.error("O backend de banco de dados nao esta disponivel.")
        st.stop()
    dados_do_banco = False
    if has_data(modo_db) or (segmento_sem_filtro_modo and (has_data("ativas") or has_data("desativadas"))):
        try:
            if segmento_sem_filtro_modo:
                df_ativas = load_linhas(modo="ativas") if has_data("ativas") else pd.DataFrame()
                df_des = load_linhas(modo="desativadas") if has_data("desativadas") else pd.DataFrame()
                df = pd.concat([df_ativas, df_des], ignore_index=True)
            else:
                df = load_linhas(modo=modo_db)
            dados_do_banco = True
        except Exception as exc:
            import logging as _log
            _log.getLogger(__name__).error("Erro ao ler o banco: %s", exc, exc_info=True)
            st.error("Não foi possível carregar os dados. Verifique a conexão com o banco e recarregue a página.")
            st.stop()
    if not dados_do_banco:
        st.error("O banco de dados esta vazio ou nao possui dados para este modo. O sistema nao usa mais planilhas como fonte.")
        st.stop()

    # Unificacao progressiva com chamados:
    # quando `chamado_id` vem via URL, garantimos que existe um registro mínimo no backend.
    # Isso evita problemas futuros ao evoluir para `chamado_eventos/movimentacoes`.
    if dados_do_banco and active_chamado_context.get("chamado_id"):
        try:
            garantir_chamado_stub(active_chamado_context.get("chamado_id"))
            if active_chamado_context.get("linha"):
                vincular_chamado_linha(active_chamado_context.get("chamado_id"), active_chamado_context.get("linha"))
        except Exception:
            pass

    if dados_do_banco and not segmento_sem_filtro_modo:
        df = _auto_mark_context_line_vago(df, modo_db, dados_do_banco)

    prefixo = "Linhas" if segmento_sem_filtro_modo else ("Linhas Ativas" if modo_sel == "Linhas ativas" else "Linhas Desativadas")
    df_full_mode = df.copy()
    # ID técnico por linha para merge seguro no data_editor (evita depender da coluna Linha).
    df_full_mode = df_full_mode.reset_index(drop=True)
    df_full_mode["__row_key"] = [f"{modo_db}:{i}" for i in range(len(df_full_mode))]
    df_full_mode["__base_linha"] = df_full_mode["Linha"].fillna("").astype(str).str.strip() if "Linha" in df_full_mode.columns else ""

    df = df.reset_index(drop=True).copy()
    df["__row_key"] = df_full_mode["__row_key"].values
    df["__base_linha"] = df_full_mode["__base_linha"].values
    if segmento_sel == "Alimento":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "alimento"]
        if (not dados_do_banco) and "Aba" in df.columns:
            aba_norm = df["Aba"].fillna("").astype(str).str.strip()
            df = df[aba_norm.str.lower().isin([a.lower() for a in ABAS_ALIMENTO])]
        equipes_ref = EQUIPES_ALIMENTO
        titulo = f"{prefixo} — Alimento"
    elif segmento_sel == "Medicamento":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "medicamento"]
        equipes_ref = EQUIPES_MEDICAMENTO
        titulo = f"{prefixo} — Medicamento"
    elif segmento_sel == "Promotores":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "promotores"]
        equipes_ref = EQUIPES_PROMOTORES
        titulo = f"{prefixo} — Promotores"
    elif segmento_sel == "Internos":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "internos"]
        equipes_ref = EQUIPES_INTERNOS
        titulo = f"{prefixo} — Internos"
    elif segmento_sel == "Manutenção":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "manutenção"]
        equipes_ref = EQUIPES_MANUTENCAO
        titulo = f"{prefixo} — Manutenção"
    else:
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "roubo e perda"]
        equipes_ref = EQUIPES_ROUBO_PERDA
        titulo = f"{prefixo} — Roubo e Perda"

    if df.empty:
        st.warning(f"Nenhum dado de {segmento_sel} encontrado.")
        st.stop()

    linhas_preenchidas_segmento = (
        df["Linha"].fillna("").astype(str).str.strip() != ""
        if "Linha" in df.columns
        else pd.Series(False, index=df.index)
    )
    linhas_vagas_segmento = build_vaga_mask(df)
    linhas_em_uso_segmento = linhas_preenchidas_segmento & ~linhas_vagas_segmento
    escopo_modo = "ativas" if modo_sel == "Linhas ativas" else "desativadas"
    total_segmento = int(len(df))
    painel_contexto = segmento_sel if segmento_sem_filtro_modo else f"{segmento_sel} | {modo_sel}"
    total_label = f"Total {segmento_sel}" if segmento_sem_filtro_modo else f"Total {escopo_modo}"
    st.caption(f"Painel operacional: {painel_contexto}")
    d1, d2, d3 = st.columns(3)
    d1.metric(f"Em uso ({segmento_sel})", int(linhas_em_uso_segmento.sum()))
    d2.metric(f"Vagas ({segmento_sel})", int(linhas_vagas_segmento.sum()))
    d3.metric(total_label, total_segmento)

    if dados_do_banco:
        tbl = df.copy()
        tbl = tbl.sort_values(
            ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
            kind="stable",
        )
    else:
        tbl = build_full_table(df)
    # Normaliza valores nulos para não exibir "None" no editor/tabelas.
    tbl = tbl.replace({None: ""}).fillna("")
    from src.core.config import NEW_COLUMNS_ORDER, HIDDEN_COLUMNS
    cols_front = NEW_COLUMNS_ORDER
    _hidden_set = set(HIDDEN_COLUMNS)
    existing_front = [c for c in cols_front if c in tbl.columns]
    tail = [c for c in tbl.columns if c not in existing_front and not c.startswith("__") and c not in _hidden_set]
    tbl = tbl[existing_front + tail]
    if not dados_do_banco and segmento_sel == "Alimento":
        gerentes_set = _get_gerentes_alimento(rules_path)
        mask_papel = tbl["Papel"].fillna("").astype(str).str.strip().str.lower() == "gerente"
        mask_nome = tbl["Nome"].fillna("").astype(str).apply(
            lambda x: normalize_text(x.strip()) in gerentes_set
        )
        mask_gerente = mask_papel | mask_nome
        if mask_gerente.any():
            tbl.loc[mask_gerente, "EquipePadrao"] = "Gerentes do Alimento"
            tbl = tbl.sort_values(
                ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
                kind="stable",
            )
    if not dados_do_banco and segmento_sel == "Medicamento":
        gerentes_set = _get_gerentes_medicamento(rules_path)
        mask_papel = tbl["Papel"].fillna("").astype(str).str.strip().str.lower() == "gerente"
        mask_nome = tbl["Nome"].fillna("").astype(str).apply(
            lambda x: normalize_text(x.strip()) in gerentes_set
        )
        mask_gerente = mask_papel | mask_nome
        if mask_gerente.any():
            tbl.loc[mask_gerente, "EquipePadrao"] = "Gerentes do Medicamento"
            tbl = tbl.sort_values(
                ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
                kind="stable",
            )
    equipes_validas = sorted(set(tbl["EquipePadrao"].dropna().astype(str).str.strip().unique()))
    equipes = [e for e in equipes_ref if e in equipes_validas]
    equipes += [e for e in equipes_validas if e not in equipes]
    # No segmento Manutenção, "Mover equipe" = mover entre as duas seções (não entre equipes externas)
    equipes_mover_manut = ["Manutenção", "Manutenções antigas"]
    tbl_segment = tbl.copy()

    st.markdown(f'<p class="main-title">{titulo}</p>', unsafe_allow_html=True)
    st.markdown('<p class="main-subtitle">Controle de linhas telefônicas por equipe. Selecione uma equipe para ver os detalhes.</p>', unsafe_allow_html=True)

    col_filtro, col_busca_tipo, col_busca, col_vagas = st.columns([1, 1.1, 2.0, 0.9])
    with col_filtro:
        equipe_options = ["Todas as equipes"] + equipes
        pending_equipe = st.session_state.get("pending_equipe")
        if pending_equipe in equipe_options:
            st.session_state.filtro_equipe = pending_equipe
        current_equipe_filter = st.session_state.get("filtro_equipe")
        if current_equipe_filter not in equipe_options:
            st.session_state["filtro_equipe"] = "Todas as equipes"
        equipe_sel = st.selectbox(
            "Equipe",
            options=equipe_options,
            format_func=lambda x: "Todas as equipes" if x == "Todas as equipes" else x,
            label_visibility="collapsed",
            key="filtro_equipe",
        )
    with col_busca_tipo:
        busca_tipo_options = ["Geral", "Linha", "IMEI", "Aparelho", "Motivo"]
        if st.session_state.get("filtro_busca_tipo") not in busca_tipo_options:
            st.session_state["filtro_busca_tipo"] = "Geral"
        busca_tipo = st.selectbox(
            "Tipo de busca",
            options=busca_tipo_options,
            label_visibility="collapsed",
            key="filtro_busca_tipo",
        )
    with col_busca:
        placeholder_map = {
            "Geral": "Buscar por nome, codigo, linha, aparelho, IMEI...",
            "Linha": "Buscar numero da linha...",
            "IMEI": "Buscar IMEI A ou IMEI B...",
            "Aparelho": "Buscar aparelho ou modelo...",
            "Motivo": "Buscar motivo ou observacao...",
        }
        busca = st.text_input(
            "Buscar",
            placeholder=placeholder_map.get(busca_tipo, "Buscar..."),
            label_visibility="collapsed",
            key="filtro_busca",
        )
    with col_vagas:
        if "filtro_somente_vagas" not in st.session_state:
            st.session_state["filtro_somente_vagas"] = False
        somente_vagas = st.checkbox("Somente vagas", key="filtro_somente_vagas")
    editor_context_key = f"{modo_db}|{segmento_sel}|{equipe_sel}|{busca_tipo}|{busca}|{int(somente_vagas)}"

    query = busca.strip().lower()
    if query:
        mask = pd.Series(False, index=tbl.index)
        if busca_tipo == "Linha":
            columns_to_search = ["Linha"]
        elif busca_tipo == "IMEI":
            columns_to_search = ["IMEI A", "IMEI B"]
        elif busca_tipo == "Aparelho":
            columns_to_search = ["Aparelho", "Modelo", "Marca"]
        elif busca_tipo == "Motivo":
            columns_to_search = ["Motivo", "Observação", "Gerenciamento"]
        else:
            columns_to_search = [
                "Nome",
                "Codigo",
                "Linha",
                "Aparelho",
                "Modelo",
                "Marca",
                "EquipePadrao",
                "Supervisor",
                "Gestor",
                "Motivo",
                "Observação",
                "IMEI A",
                "IMEI B",
                "E-mail",
            ]
        for c in columns_to_search:
            if c in tbl.columns:
                mask = mask | tbl[c].astype(str).str.lower().str.contains(query, na=False)
        tbl = tbl[mask]

    if equipe_sel != "Todas as equipes":
        tbl = tbl[tbl["EquipePadrao"] == equipe_sel]

    if somente_vagas:
        tbl = tbl[build_vaga_mask(tbl)]

    for qp_key in ["com_motivo", "sem_aparelho", "imei_pendente", "sem_email"]:
        if qp_key in st.query_params:
            del st.query_params[qp_key]

    pending_linha = st.session_state.get("pending_linha")
    if pending_linha and not tbl.empty and "Linha" in tbl.columns:
        mask_pending = tbl["Linha"].fillna("").astype(str) == pending_linha
        if mask_pending.any():
            st.info(f"Nova linha criada na equipe **{equipe_sel}**. Procure por `{pending_linha}` e edite os dados.")
            st.session_state.pending_segmento = None
            st.session_state.pending_equipe = None
            st.session_state.pending_linha = None

    linha_contexto_ativa = str(((st.session_state.get("chamado_context") or {}).get("linha") or "")).strip()
    if linha_contexto_ativa and not tbl.empty and "Linha" in tbl.columns:
        mask_contexto = tbl["Linha"].fillna("").astype(str).str.strip() == linha_contexto_ativa
        if mask_contexto.any():
            st.success(f"Linha do chamado localizada: `{linha_contexto_ativa}`")

    conflict_payload = st.session_state.get("_editor_conflict_payload") or {}

    def _build_conflict_payload_from_pending(
        pending_changes: list[dict[str, str]],
        context_key: str,
    ) -> dict[str, Any]:
        if not pending_changes:
            return {}
        resumo: list[dict[str, str]] = []
        detalhes: list[dict[str, str]] = []
        grouped: dict[str, list[dict[str, str]]] = {}
        for ch in pending_changes:
            base_linha = str(ch.get("base_linha") or "").strip()
            if not base_linha:
                continue
            grouped.setdefault(base_linha, []).append(ch)
        for base_linha, items in grouped.items():
            campos = [str(it.get("campo") or "").strip() for it in items if str(it.get("campo") or "").strip()]
            resumo.append(
                {
                    "Linha": base_linha or "—",
                    "Motivo": "Alguns campos continuam em conflito.",
                    "Campos com diferenca": ", ".join(campos[:6]) + (f" (+{len(campos) - 6})" if len(campos) > 6 else ""),
                }
            )
            for it in items:
                detalhes.append(
                    {
                        "Linha": base_linha or "—",
                        "Campo": str(it.get("campo") or "—"),
                        "Valor original": str(it.get("original") or "—"),
                        "Minha edição": str(it.get("mine") or "—"),
                        "Valor atual no banco": str(it.get("current") or "—"),
                    }
                )
        return {
            "context": context_key,
            "resumo": resumo,
            "detalhes": detalhes,
            "pending_changes": pending_changes,
        }

    if conflict_payload and conflict_payload.get("context") == editor_context_key:
        st.warning("Existe um conflito de edição pendente. Revise as diferenças antes de tentar salvar novamente.")
        resumo_conf = conflict_payload.get("resumo") or []
        detalhes_conf = conflict_payload.get("detalhes") or []
        if resumo_conf:
            st.dataframe(pd.DataFrame(resumo_conf), width="stretch", hide_index=True)
        if detalhes_conf:
            with st.expander("Ver diferenças do conflito", expanded=False):
                st.dataframe(pd.DataFrame(detalhes_conf), width="stretch", hide_index=True)
            linhas_conflito_opts = sorted(
                {
                    str(item.get("Linha") or "").strip()
                    for item in detalhes_conf
                    if str(item.get("Linha") or "").strip()
                }
            )
            if linhas_conflito_opts:
                linha_revisao = st.selectbox(
                    "Revisar conflito por linha",
                    options=linhas_conflito_opts,
                    key="conflict_line_review",
                )
                detalhes_linha = [
                    item for item in detalhes_conf
                    if str(item.get("Linha") or "").strip() == linha_revisao
                ]
                if detalhes_linha:
                    st.caption(f"Comparativo da linha `{linha_revisao}`")
                    st.dataframe(pd.DataFrame(detalhes_linha), width="stretch", hide_index=True)
                    campos_linha_opts = [
                        str(item.get("Campo") or "").strip()
                        for item in detalhes_linha
                        if str(item.get("Campo") or "").strip()
                    ]
                    if campos_linha_opts:
                        campo_revisao = st.selectbox(
                            "Resolver campo",
                            options=campos_linha_opts,
                            key="conflict_field_review",
                        )
                        detalhe_campo = next(
                            (
                                item for item in detalhes_linha
                                if str(item.get("Campo") or "").strip() == campo_revisao
                            ),
                            None,
                        )
                        if detalhe_campo:
                            a_res1, a_res2 = st.columns([1, 1.2])
                            with a_res1:
                                if st.button("Aceitar valor atual do banco", key="btn_accept_db_value_conflict"):
                                    pending_changes = conflict_payload.get("pending_changes") or []
                                    pending_changes = [
                                        ch
                                        for ch in pending_changes
                                        if not (
                                            str(ch.get("base_linha") or "").strip() == linha_revisao
                                            and str(ch.get("campo") or "").strip() == campo_revisao
                                        )
                                    ]
                                    st.session_state["_editor_conflict_payload"] = _build_conflict_payload_from_pending(
                                        pending_changes,
                                        editor_context_key,
                                    )
                                    st.session_state["_post_conflict_feedback"] = {
                                        "type": "success",
                                        "text": f"Campo `{campo_revisao}` da linha `{linha_revisao}` resolvido com o valor atual do banco.",
                                    }
                                    st.rerun()
                            with a_res2:
                                if st.button("Sobrescrever com minha edição", key="btn_force_my_value_conflict"):
                                    pending_changes = conflict_payload.get("pending_changes") or []
                                    pending_change = next(
                                        (
                                            ch for ch in pending_changes
                                            if str(ch.get("base_linha") or "").strip() == linha_revisao
                                            and str(ch.get("campo") or "").strip() == campo_revisao
                                        ),
                                        None,
                                    )
                                    if pending_change is None:
                                        st.warning("Nao foi possivel localizar o conflito selecionado.")
                                    else:
                                        df_current_mode = load_linhas(modo=modo_db)
                                        seg_lower_force = segmento_sel.strip().lower()
                                        mask_seg_force = df_current_mode["Segmento"].fillna("").astype(str).str.strip().str.lower() == seg_lower_force
                                        base_linha_force = str(pending_change.get("base_linha") or "").strip()
                                        campo_force = str(pending_change.get("campo") or "").strip()
                                        novo_valor_raw = str(pending_change.get("mine_raw") or "")
                                        current_matches = df_current_mode[
                                            mask_seg_force & (df_current_mode["Linha"].fillna("").astype(str).str.strip() == base_linha_force)
                                        ]
                                        if current_matches.empty:
                                            st.warning("A linha nao foi encontrada no banco para sobrescrever.")
                                        elif len(current_matches) > 1:
                                            st.warning("Existe mais de uma linha com esse identificador. Resolva manualmente para evitar sobrescrita indevida.")
                                        elif not campo_force or campo_force not in df_current_mode.columns:
                                            st.warning("Campo invalido para sobrescrita.")
                                        else:
                                            row_idx_force = current_matches.index[0]
                                            before_force = df_current_mode.loc[row_idx_force].to_dict()
                                            df_current_mode.loc[row_idx_force, campo_force] = novo_valor_raw
                                            save_linhas(df_current_mode, modo=modo_db)
                                            _audit(
                                                acao="resolver_conflito_edicao",
                                                entidade="linhas",
                                                chave=base_linha_force,
                                                antes=before_force,
                                                depois={campo_force: novo_valor_raw, "modo": modo_db},
                                                detalhes=f"Conflito resolvido manualmente mantendo a edicao do usuario no campo {campo_force}.",
                                            )
                                            _set_post_chamado_banner("Conflito resolvido e alteração aplicada com sucesso.")
                                            pending_changes = [
                                                ch
                                                for ch in pending_changes
                                                if not (
                                                    str(ch.get("base_linha") or "").strip() == linha_revisao
                                                    and str(ch.get("campo") or "").strip() == campo_revisao
                                                )
                                            ]
                                            st.session_state["_editor_conflict_payload"] = _build_conflict_payload_from_pending(
                                                pending_changes,
                                                editor_context_key,
                                            )
                                            st.session_state["_editor_base_context"] = ""
                                            st.session_state["_editor_base_snapshot"] = []
                                            st.session_state["_post_conflict_feedback"] = {
                                                "type": "success",
                                                "text": f"Campo `{campo_revisao}` da linha `{linha_revisao}` sobrescrito com a sua edicao.",
                                            }
                                            st.rerun()
        c_conf1, c_conf2, c_conf3 = st.columns([1, 1, 1.2])
        with c_conf1:
            if st.button("Atualizar dados", key="btn_refresh_after_conflict"):
                st.session_state["_editor_base_context"] = ""
                st.session_state["_editor_base_snapshot"] = []
                st.session_state["_editor_conflict_payload"] = {}
                st.rerun()
        with c_conf2:
            if st.button("Fechar aviso", key="btn_close_conflict_warning"):
                st.session_state["_editor_conflict_payload"] = {}
                st.rerun()
        with c_conf3:
            if st.button("Reaplicar alterações seguras", key="btn_reapply_safe_changes"):
                st.session_state["_reapply_conflict_request"] = editor_context_key

    conflict_map_by_line: dict[str, list[str]] = {}
    if conflict_payload and conflict_payload.get("context") == editor_context_key:
        pending_conflict_changes = conflict_payload.get("pending_changes") or []
        for ch in pending_conflict_changes:
            base_linha = str(ch.get("base_linha") or "").strip()
            campo = str(ch.get("campo") or "").strip()
            if not base_linha or not campo:
                continue
            conflict_map_by_line.setdefault(base_linha, [])
            if campo not in conflict_map_by_line[base_linha]:
                conflict_map_by_line[base_linha].append(campo)

    def _decorate_editor_df(df_in: pd.DataFrame, manut_antigos: bool = False) -> pd.DataFrame:
        df_out = df_in.copy()
        df_out = df_out.drop(columns=["__manut_antigo", "Histórico"], errors="ignore")
        if segmento_sel == "Manutenção" and manut_antigos:
            df_out = df_out.drop(columns=["↕ Ordem"], errors="ignore")
        elif segmento_sel == "Manutenção" and not manut_antigos:
            if "↕ Ordem" not in df_out.columns:
                df_out["↕ Ordem"] = list(range(1, len(df_out) + 1))
            else:
                df_out["↕ Ordem"] = pd.to_numeric(df_out["↕ Ordem"], errors="coerce")
        else:
            if "↕ Ordem" not in df_out.columns:
                df_out["↕ Ordem"] = list(range(1, len(df_out) + 1))
            else:
                ordem_num = pd.to_numeric(df_out["↕ Ordem"], errors="coerce")
                ordem_seq = pd.Series(range(1, len(df_out) + 1), index=df_out.index, dtype="int64")
                df_out["↕ Ordem"] = ordem_num.where(ordem_num.notna(), ordem_seq).astype(int)
        if "Mover equipe" not in df_out.columns:
            df_out["Mover equipe"] = ""
        if conflict_map_by_line:
            base_linhas = df_out["__base_linha"].fillna("").astype(str).str.strip() if "__base_linha" in df_out.columns else pd.Series([""] * len(df_out), index=df_out.index)
            conflito_textos = []
            conflito_ord = []
            for ln in base_linhas:
                campos = conflict_map_by_line.get(ln, [])
                if campos:
                    txt = ", ".join(campos[:3])
                    if len(campos) > 3:
                        txt += f" +{len(campos) - 3}"
                    conflito_textos.append(f"Conflito: {txt}")
                    conflito_ord.append(0)
                else:
                    conflito_textos.append("")
                    conflito_ord.append(1)
            df_out["Conflito"] = conflito_textos
            df_out["__conflict_ord"] = conflito_ord
            sort_cols_conf = ["__conflict_ord"] + [c for c in ["Papel", "Nome", "Linha"] if c in df_out.columns]
            ascending_conf = [True] + [True] * (len(sort_cols_conf) - 1)
            df_out = df_out.sort_values(sort_cols_conf, ascending=ascending_conf, kind="stable")
            df_out = df_out.drop(columns=["__conflict_ord"], errors="ignore")
        cols = df_out.columns.tolist()
        controls_front = [c for c in ["↕ Ordem", "Mover equipe"] if c in cols]
        cols = controls_front + [c for c in cols if c not in controls_front]
        if "Conflito" in cols:
            cols = controls_front + ["Conflito"] + [c for c in cols if c not in (set(controls_front) | {"Conflito"})]
        df_out = df_out[cols]
        return df_out

    n_linhas = tbl["Linha"].nunique()
    n_pessoas = len(tbl)
    lbl_linhas = "linhas" if segmento_sem_filtro_modo else ("linhas desativadas" if modo_sel == "Linhas desativadas" else "linhas ativas")
    st.caption(f"**{n_pessoas}** registros | **{n_linhas}** {lbl_linhas}")

    if tbl.empty:
        st.info("Nenhum resultado. Tente outro filtro ou termo de busca.")
        st.stop()

    if segmento_sel == "Medicamento":
        sort_cols = ["Localidade", "Papel", "Nome", "Linha"]
        sort_ascending = [True, True, True, True]
    elif segmento_sel == "Manutenção":
        sort_cols = ["Data da Troca", "Data Retorno", "Nome", "Linha"]
        sort_ascending = [False, False, True, True]
    elif segmento_sel == "Roubo e Perda":
        sort_cols = ["Data Ocorrência", "Data da Troca", "Nome", "Linha"]
        sort_ascending = [False, False, True, True]
    else:
        sort_cols = ["Papel", "Nome", "Linha"]
        sort_ascending = [True, True, True]
    header_cls = "team-header-inactive" if modo_sel == "Linhas desativadas" else "team-header"
    usar_editor = HAS_DB and dados_do_banco
    save_clicked_top = False
    if usar_editor:
        st.caption("✏️ Edite direto na linha. Em '↕ Ordem' informe a posição no bloco e em 'Mover equipe' troque a equipe da linha.")
        col_save_btn, _ = st.columns([0.38, 0.62])
        with col_save_btn:
            save_clicked_top = st.button(
                "Salvar alterações",
                key="btn_save_changes_top",
                type="primary",
                use_container_width=False,
            )

    def _sort_for_display(df_in: pd.DataFrame) -> pd.DataFrame:
        """Ordena visualização, priorizando datas em ordem decrescente quando aplicável."""
        df_sort = df_in.copy()
        if "↕ Ordem" in df_sort.columns:
            df_sort["↕ Ordem"] = pd.to_numeric(df_sort["↕ Ordem"], errors="coerce")
        # Regra específica de Manutenção:
        # o mais recente sempre deve ser Ordem 1, depois 2, 3...
        # (independente de ordem manual anterior).
        # Manutenção: lógica movida para _sort_manut_block (duas seções separadas)
        sort_eff: list[str] = []
        asc_eff: list[bool] = []
        date_cols = {"Data da Troca", "Data Retorno", "Data Ocorrência", "Data Solicitação TBS"}
        if "↕ Ordem" in df_sort.columns:
            sort_eff.append("↕ Ordem")
            asc_eff.append(True)
        for col, asc in zip(sort_cols, sort_ascending):
            if col not in df_sort.columns:
                continue
            if col in date_cols:
                aux = "__ord_" + normalize_text(col).replace(" ", "_")
                df_sort[aux] = pd.to_datetime(df_sort[col], errors="coerce")
                sort_eff.append(aux)
            else:
                sort_eff.append(col)
            asc_eff.append(asc)
        if sort_eff:
            df_sort = df_sort.sort_values(sort_eff, kind="stable", ascending=asc_eff)
        aux_cols = [c for c in df_sort.columns if c.startswith("__ord_")]
        if aux_cols:
            df_sort = df_sort.drop(columns=aux_cols, errors="ignore")
        return df_sort

    def _apply_manual_team_order(df_in: pd.DataFrame, equipe_nome: str) -> pd.DataFrame:
        """Aplica ordem manual persistida em sessão para uma equipe."""
        if segmento_sel == "Manutenção":
            # Em manutenção a ordem é automática por recência (mais novo = 1).
            return df_in
        if df_in.empty or "Linha" not in df_in.columns:
            return df_in
        order_map = st.session_state.get("_manual_order_map") or {}
        order_key = f"{modo_db}|{segmento_sel}|{str(equipe_nome).strip()}"
        ordered_lines = order_map.get(order_key) or []
        if not ordered_lines:
            return df_in
        rank = {str(ln).strip(): i for i, ln in enumerate(ordered_lines)}
        df_ord = df_in.copy()
        df_ord["__manual_rank"] = df_ord["Linha"].fillna("").astype(str).str.strip().map(lambda v: rank.get(v, 10**9))
        df_ord["__manual_seq"] = range(len(df_ord))
        df_ord = df_ord.sort_values(["__manual_rank", "__manual_seq"], kind="stable")
        return df_ord.drop(columns=["__manual_rank", "__manual_seq"], errors="ignore")

    def _split_manut_novos_antigos(tbl_manut: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
        """Separa linhas de manutenção em novas (hoje+) e antigas."""
        from datetime import date
        today_ref = date.today()
        if "Data da Troca" not in tbl_manut.columns or tbl_manut.empty:
            return pd.DataFrame(), tbl_manut.copy()
        df = tbl_manut.copy()
        # Aceita YYYY-MM-DD, DD/MM/YYYY e outros formatos
        df["__ord_dt"] = pd.to_datetime(df["Data da Troca"], errors="coerce", dayfirst=True)
        df["__ord_date"] = df["__ord_dt"].dt.date
        # data_troca >= hoje = novos; NaT/data inválida vai para antigos
        mask_novos = (df["__ord_date"].notna()) & (df["__ord_date"] >= today_ref)
        df = df.drop(columns=["__ord_dt", "__ord_date"], errors="ignore")
        return df[mask_novos].copy(), df[~mask_novos].copy()

    def _sort_manut_block(df_in: pd.DataFrame, is_novos: bool) -> pd.DataFrame:
        """Ordena bloco de manutenção. Novos: ordem 1,2,3. Antigos: só por data."""
        df_sort = df_in.copy()
        if df_sort.empty:
            return df_sort
        sort_cols: list[str] = []
        asc_list: list[bool] = []
        for col, asc in [("Data da Troca", False), ("Data Retorno", False), ("Nome", True), ("Linha", True)]:
            if col not in df_sort.columns:
                continue
            if col in {"Data da Troca", "Data Retorno"}:
                aux = "__ord_" + normalize_text(col).replace(" ", "_")
                df_sort[aux] = pd.to_datetime(df_sort[col], errors="coerce")
                sort_cols.append(aux)
            else:
                sort_cols.append(col)
            asc_list.append(asc)
        if sort_cols:
            df_sort = df_sort.sort_values(sort_cols, kind="stable", ascending=asc_list)
        aux_cols = [c for c in df_sort.columns if c.startswith("__ord_")]
        if aux_cols:
            df_sort = df_sort.drop(columns=aux_cols, errors="ignore")
        if is_novos and len(df_sort) > 0:
            df_sort["↕ Ordem"] = list(range(1, len(df_sort) + 1))
        return df_sort

    def _render_team_actions(eq_nome: str, df_equipe: pd.DataFrame, df_editor_view: pd.DataFrame | None = None) -> None:
        """Ações por bloco de equipe: excluir, mover e ativar/desativar."""
        linhas_team = sorted([str(x).strip() for x in df_equipe["Linha"].dropna().astype(str).tolist() if str(x).strip()])
        if not linhas_team:
            return
        key_base = normalize_text(eq_nome).replace(" ", "_")
        linhas_selecionadas_editor: list[str] = []
        if isinstance(df_editor_view, pd.DataFrame) and not df_editor_view.empty:
            if "Selecionar" in df_editor_view.columns and "Linha" in df_editor_view.columns:
                mask_sel = df_editor_view["Selecionar"].fillna(False).astype(bool)
                if mask_sel.any():
                    linhas_selecionadas_editor = sorted(
                        {
                            str(v).strip()
                            for v in df_editor_view.loc[mask_sel, "Linha"].tolist()
                            if str(v).strip()
                        }
                    )
        with st.expander(f"Ações da equipe: {eq_nome}", expanded=False):
            action_options = [
                "Excluir linha",
                "Mover para outra equipe",
                "Mover para outro setor",
            ]
            if segmento_sel != "Manutenção":
                action_options.append("Enviar aparelho para manutenção")
            action_options.append("Desativar linha" if modo_db == "ativas" else "Ativar linha")

            c1, c2, c3 = st.columns([1.3, 1.4, 1.3])
            with c1:
                linha_default = ""
                if len(linhas_selecionadas_editor) == 1:
                    linha_default = linhas_selecionadas_editor[0]
                elif len(linhas_selecionadas_editor) > 1:
                    st.warning("Mais de uma linha foi selecionada no editor. Para esta ação, mantenha somente uma selecionada.")
                linha_options = [""] + linhas_team
                idx_default = linha_options.index(linha_default) if linha_default in linha_options else 0
                linha_alvo = st.selectbox("Linha", options=linha_options, index=idx_default, key=f"acao_linha_{key_base}")
                if linha_default:
                    st.caption("Linha carregada automaticamente da seleção do editor.")
            with c2:
                acao_linha = st.selectbox(
                    "Ação",
                    options=action_options,
                    key=f"acao_tipo_{key_base}",
                )
            linha_selecionada = {}
            if linha_alvo:
                matches = df_equipe[df_equipe["Linha"].fillna("").astype(str).str.strip() == linha_alvo]
                if not matches.empty:
                    linha_selecionada = matches.iloc[0].to_dict()
            destino_setor = ""
            destino_equipe = ""
            with c3:
                if acao_linha == "Mover para outra equipe":
                    destino_equipe = st.selectbox("Equipe destino", options=equipes, key=f"acao_eq_dest_{key_base}")
                elif acao_linha == "Mover para outro setor":
                    destino_setor = st.selectbox("Setor destino", options=["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"], key=f"acao_seg_dest_{key_base}")
                    eq_dest_ref = (
                        EQUIPES_ALIMENTO
                        if destino_setor == "Alimento"
                        else (
                            EQUIPES_MEDICAMENTO
                            if destino_setor == "Medicamento"
                            else (
                                EQUIPES_PROMOTORES
                                if destino_setor == "Promotores"
                                else (
                                    EQUIPES_INTERNOS
                                    if destino_setor == "Internos"
                                    else (EQUIPES_MANUTENCAO if destino_setor == "Manutenção" else EQUIPES_ROUBO_PERDA)
                                )
                            )
                        )
                    )
                    destino_equipe = st.selectbox("Equipe destino", options=eq_dest_ref, key=f"acao_eq_dest_setor_{key_base}")

            if linha_selecionada:
                resumo_linha = []
                for campo in ["Nome", "Aparelho", "Modelo", "Segmento", "EquipePadrao"]:
                    valor = str(linha_selecionada.get(campo, "") or "").strip()
                    if valor:
                        resumo_linha.append(f"{campo}: {valor}")
                if resumo_linha:
                    st.caption("Linha selecionada: " + " | ".join(resumo_linha))

            action_guidance = {
                "Excluir linha": "Remove a linha deste modo. Use apenas quando o registro nao deve mais existir aqui.",
                "Mover para outra equipe": "Mantem o mesmo setor e apenas troca a equipe responsavel.",
                "Mover para outro setor": "Move a linha para outro setor e atualiza tambem a equipe destino.",
                "Enviar aparelho para manutenção": "Cria uma copia em Manutenção e mantem a linha original para voce atualizar o aparelho depois.",
                "Desativar linha": "Move a linha de ativas para desativadas sem perder o historico do registro.",
                "Ativar linha": "Move a linha de desativadas para ativas para voltar ao uso operacional.",
            }
            if acao_linha in action_guidance:
                guidance_text = action_guidance[acao_linha]
                if acao_linha == "Mover para outro setor" and destino_setor == "Roubo e Perda":
                    guidance_text += " Depois da movimentacao, revise Data Ocorrencia, Motivo e observacoes."
                elif acao_linha == "Mover para outro setor" and destino_setor == "Manutenção":
                    guidance_text += " Depois da movimentacao, revise datas e dados do aparelho em manutencao."
                elif acao_linha == "Mover para outra equipe" and destino_equipe:
                    guidance_text += f" Destino selecionado: {destino_equipe}."
                st.info(guidance_text)

            def _format_action_date(value: Any) -> str:
                if value in (None, ""):
                    return ""
                try:
                    return pd.Timestamp(value).strftime("%d/%m/%Y")
                except Exception:
                    return str(value).strip()

            motivo_operacao = ""
            observacao_operacao = ""
            data_ocorrencia_operacao = ""
            data_troca_operacao = ""
            data_retorno_operacao = ""

            if acao_linha == "Mover para outro setor" and destino_setor == "Roubo e Perda":
                f_roubo1, f_roubo2 = st.columns([1, 1.2])
                with f_roubo1:
                    data_ocorrencia_operacao = _format_action_date(
                        st.date_input(
                            "Data ocorrencia",
                            value=pd.Timestamp.today().date(),
                            key=f"acao_data_ocorrencia_{key_base}",
                        )
                    )
                    motivo_operacao = st.text_input(
                        "Motivo",
                        key=f"acao_motivo_roubo_{key_base}",
                        placeholder="Ex.: Roubo, perda ou furto",
                    ).strip()
                with f_roubo2:
                    observacao_operacao = st.text_area(
                        "Observacao",
                        key=f"acao_obs_roubo_{key_base}",
                        placeholder="Detalhes importantes para acompanhamento",
                        height=80,
                    ).strip()
            elif acao_linha in {"Enviar aparelho para manutenção"} or (
                acao_linha == "Mover para outro setor" and destino_setor == "Manutenção"
            ):
                f_manut1, f_manut2 = st.columns([1, 1.2])
                with f_manut1:
                    data_troca_operacao = _format_action_date(
                        st.date_input(
                            "Data da troca",
                            value=pd.Timestamp.today().date(),
                            key=f"acao_data_troca_{key_base}",
                        )
                    )
                    data_retorno_operacao = _format_action_date(
                        st.date_input(
                            "Data retorno prevista",
                            value=None,
                            key=f"acao_data_retorno_{key_base}",
                        )
                    )
                with f_manut2:
                    motivo_operacao = st.text_input(
                        "Motivo da manutencao",
                        key=f"acao_motivo_manut_{key_base}",
                        placeholder="Ex.: Defeito no aparelho, troca preventiva",
                    ).strip()
                    observacao_operacao = st.text_area(
                        "Observacao",
                        key=f"acao_obs_manut_{key_base}",
                        placeholder="Informacoes para a equipe de manutencao",
                        height=80,
                    ).strip()
            elif acao_linha in {"Desativar linha", "Ativar linha"}:
                f_status1, f_status2 = st.columns([1, 1.2])
                with f_status1:
                    motivo_operacao = st.text_input(
                        "Motivo operacional",
                        key=f"acao_motivo_status_{key_base}",
                        placeholder="Ex.: Desligamento, substituicao, reativacao",
                    ).strip()
                with f_status2:
                    observacao_operacao = st.text_area(
                        "Observacao",
                        key=f"acao_obs_status_{key_base}",
                        placeholder="Contexto opcional da alteracao de status",
                        height=80,
                    ).strip()

            requires_confirmation = (
                acao_linha in {"Excluir linha", "Enviar aparelho para manutenção", "Desativar linha", "Ativar linha"}
                or (acao_linha == "Mover para outro setor" and destino_setor in {"Manutenção", "Roubo e Perda"})
            )
            action_confirmed = True
            if requires_confirmation:
                confirmation_labels = {
                    "Excluir linha": "Confirmo a exclusao desta linha.",
                    "Enviar aparelho para manutenção": "Confirmo o envio para manutencao e vou revisar a linha original depois.",
                    "Desativar linha": "Confirmo que esta linha deve sair das ativas.",
                    "Ativar linha": "Confirmo que esta linha deve voltar para as ativas.",
                }
                confirm_label = confirmation_labels.get(
                    acao_linha,
                    "Confirmo a movimentacao desta linha para um fluxo operacional sensivel.",
                )
                action_confirmed = st.checkbox(
                    confirm_label,
                    key=f"acao_confirm_{key_base}_{normalize_text(acao_linha).replace(' ', '_')}",
                )

            if linhas_selecionadas_editor:
                st.caption(f"{len(linhas_selecionadas_editor)} linha(s) selecionada(s) no editor.")

            mover_lote_disponivel = acao_linha in {"Mover para outra equipe", "Mover para outro setor"}
            if mover_lote_disponivel and st.button("Mover linhas selecionadas", key=f"acao_move_selected_{key_base}"):
                if not linhas_selecionadas_editor:
                    st.warning("Selecione ao menos uma linha no checkbox 'Selecionar'.")
                    return
                if acao_linha == "Mover para outra equipe" and (not destino_equipe or destino_equipe == eq_nome):
                    st.warning("Selecione uma equipe de destino diferente da atual.")
                    return
                if acao_linha == "Mover para outro setor":
                    if not destino_setor or not destino_equipe:
                        st.warning("Selecione o setor e a equipe de destino.")
                        return
                    if destino_setor == "Roubo e Perda" and not motivo_operacao:
                        st.warning("Informe o motivo para mover a linha para Roubo e Perda.")
                        return
                if requires_confirmation and not action_confirmed:
                    st.warning("Confirme a operacao para continuar.")
                    return
                try:
                    modo_origem = modo_db
                    df_origem = load_linhas(modo=modo_origem)
                    mask_lote = df_origem["Linha"].fillna("").astype(str).isin(linhas_selecionadas_editor)
                    if not mask_lote.any():
                        st.error("Nenhuma linha selecionada foi encontrada no modo atual.")
                        return
                    before_rows = df_origem[mask_lote].copy()

                    if acao_linha == "Mover para outra equipe":
                        df_origem.loc[mask_lote, "EquipePadrao"] = destino_equipe
                        df_origem.loc[mask_lote, "Equipe"] = destino_equipe
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="mover_equipe_lote",
                            entidade="linhas",
                            chave=f"{eq_nome}:{len(before_rows)}",
                            antes={"linhas": before_rows["Linha"].fillna("").astype(str).tolist()},
                            depois={"EquipePadrao": destino_equipe, "Equipe": destino_equipe, "quantidade": int(len(before_rows)), "modo": modo_origem},
                            detalhes=f"Movimentacao em lote para equipe {destino_equipe}",
                        )
                        try:
                            if active_chamado_context.get("chamado_id"):
                                for _, row_mv in before_rows.iterrows():
                                    linha_mv = str(row_mv.get("Linha", "")).strip()
                                    if not linha_mv:
                                        continue
                                    registrar_movimentacao_linha(
                                        chamado_id=str(active_chamado_context.get("chamado_id")),
                                        linha_num=linha_mv,
                                        tipo_movimentacao="mover_equipe",
                                        antes=row_mv.to_dict(),
                                        depois={"EquipePadrao": destino_equipe, "Equipe": destino_equipe},
                                        motivo="",
                                        observacao=f"Origem equipe: {eq_nome} (lote)",
                                        user_id=str(st.session_state.get("user", {}).get("id", "")),
                                    )
                        except Exception:
                            pass
                        _set_post_chamado_banner("Linhas movidas com sucesso.")
                        st.session_state.pending_equipe = destino_equipe
                        st.success(f"{len(before_rows)} linha(s) movida(s) para equipe {destino_equipe}.")
                        st.rerun()

                    elif acao_linha == "Mover para outro setor":
                        df_origem.loc[mask_lote, "Segmento"] = destino_setor
                        df_origem.loc[mask_lote, "GrupoEquipe"] = destino_setor
                        df_origem.loc[mask_lote, "EquipePadrao"] = destino_equipe
                        df_origem.loc[mask_lote, "Equipe"] = destino_equipe
                        if destino_setor == "Roubo e Perda":
                            df_origem.loc[mask_lote, "Data Ocorrência"] = data_ocorrencia_operacao
                            df_origem.loc[mask_lote, "Motivo"] = motivo_operacao
                            if observacao_operacao:
                                df_origem.loc[mask_lote, "Observação"] = observacao_operacao
                        elif destino_setor == "Manutenção":
                            df_origem.loc[mask_lote, "Data da Troca"] = data_troca_operacao
                            if data_retorno_operacao:
                                df_origem.loc[mask_lote, "Data Retorno"] = data_retorno_operacao
                            if motivo_operacao:
                                df_origem.loc[mask_lote, "Motivo"] = motivo_operacao
                            if observacao_operacao:
                                df_origem.loc[mask_lote, "Observação"] = observacao_operacao
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="mover_setor_lote",
                            entidade="linhas",
                            chave=f"{eq_nome}:{len(before_rows)}",
                            antes={"linhas": before_rows["Linha"].fillna("").astype(str).tolist()},
                            depois={"Segmento": destino_setor, "EquipePadrao": destino_equipe, "quantidade": int(len(before_rows)), "modo": modo_origem},
                            detalhes=f"Movimentacao em lote para setor {destino_setor} / equipe {destino_equipe}",
                        )
                        try:
                            if active_chamado_context.get("chamado_id"):
                                for _, row_mv in before_rows.iterrows():
                                    linha_mv = str(row_mv.get("Linha", "")).strip()
                                    if not linha_mv:
                                        continue
                                    registrar_movimentacao_linha(
                                        chamado_id=str(active_chamado_context.get("chamado_id")),
                                        linha_num=linha_mv,
                                        tipo_movimentacao="mover_setor",
                                        antes=row_mv.to_dict(),
                                        depois={"Segmento": destino_setor, "EquipePadrao": destino_equipe, "Motivo": motivo_operacao, "Observação": observacao_operacao},
                                        motivo=motivo_operacao,
                                        observacao=observacao_operacao,
                                        user_id=str(st.session_state.get("user", {}).get("id", "")),
                                    )
                        except Exception:
                            pass
                        _set_post_chamado_banner("Linhas movidas de setor com sucesso.")
                        st.session_state.pending_segmento = destino_setor
                        st.session_state.pending_equipe = destino_equipe
                        st.success(f"{len(before_rows)} linha(s) movida(s) para setor {destino_setor} / equipe {destino_equipe}.")
                        st.rerun()
                except Exception as exc:
                    import logging as _log
                    _log.getLogger(__name__).error("Erro ao mover linhas: %s", exc, exc_info=True)
                    st.error("Não foi possível mover as linhas selecionadas. Tente novamente.")

            if st.button("Executar ação", key=f"acao_exec_{key_base}"):
                if not linha_alvo:
                    st.error("Selecione uma linha.")
                    return
                if acao_linha == "Mover para outra equipe" and (not destino_equipe or destino_equipe == eq_nome):
                    st.warning("Selecione uma equipe de destino diferente da atual.")
                    return
                if acao_linha == "Mover para outro setor":
                    segmento_atual = str(linha_selecionada.get("Segmento", "") or "").strip()
                    equipe_atual = str(linha_selecionada.get("EquipePadrao", "") or "").strip()
                    if not destino_setor or not destino_equipe:
                        st.warning("Selecione o setor e a equipe de destino.")
                        return
                    if destino_setor == segmento_atual and destino_equipe == equipe_atual:
                        st.warning("Selecione um destino diferente do setor e equipe atuais.")
                        return
                    if destino_setor == "Roubo e Perda" and not motivo_operacao:
                        st.warning("Informe o motivo para mover a linha para Roubo e Perda.")
                        return
                if acao_linha == "Enviar aparelho para manutenção" and not motivo_operacao:
                    st.warning("Informe o motivo da manutencao antes de continuar.")
                    return
                if acao_linha == "Desativar linha" and not motivo_operacao:
                    st.warning("Informe o motivo operacional da desativacao.")
                    return
                if requires_confirmation and not action_confirmed:
                    st.warning("Confirme a operacao para continuar.")
                    return
                try:
                    modo_origem = modo_db
                    df_origem = load_linhas(modo=modo_origem)
                    mask_origem = df_origem["Linha"].fillna("").astype(str) == linha_alvo
                    if not mask_origem.any():
                        st.error("Linha não encontrada no modo atual.")
                        return
                    before_row = df_origem[mask_origem].iloc[0].to_dict()

                    if acao_linha == "Excluir linha":
                        try:
                            if active_chamado_context.get("chamado_id"):
                                # Registra movimentação antes de remover a linha do banco,
                                # para respeitar a FK de `movimentacoes_linha.linha_id`.
                                registrar_movimentacao_linha(
                                    chamado_id=str(active_chamado_context.get("chamado_id")),
                                    linha_num=linha_alvo,
                                    tipo_movimentacao="excluir_linha",
                                    antes=before_row,
                                    depois=None,
                                    motivo="",
                                    observacao=f"Exclusão na equipe {eq_nome}",
                                    user_id=str(st.session_state.get("user", {}).get("id", "")),
                                )
                        except Exception:
                            pass
                        df_origem = df_origem[~mask_origem].copy()
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="excluir_linha",
                            entidade="linhas",
                            chave=linha_alvo,
                            antes=before_row,
                            detalhes=f"Exclusão na equipe {eq_nome}",
                        )
                        _set_post_chamado_banner("Linha excluída com sucesso.")
                        st.success("Linha excluída com sucesso.")
                        st.rerun()

                    elif acao_linha == "Mover para outra equipe":
                        df_origem.loc[mask_origem, "EquipePadrao"] = destino_equipe
                        df_origem.loc[mask_origem, "Equipe"] = destino_equipe
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="mover_equipe",
                            entidade="linhas",
                            chave=linha_alvo,
                            antes=before_row,
                            depois={
                                "EquipePadrao": destino_equipe,
                                "Equipe": destino_equipe,
                                "modo": modo_origem,
                            },
                            detalhes=f"Origem equipe: {eq_nome}",
                        )
                        try:
                            if active_chamado_context.get("chamado_id"):
                                registrar_movimentacao_linha(
                                    chamado_id=str(active_chamado_context.get("chamado_id")),
                                    linha_num=linha_alvo,
                                    tipo_movimentacao="mover_equipe",
                                    antes=before_row,
                                    depois={
                                        "EquipePadrao": destino_equipe,
                                        "Equipe": destino_equipe,
                                    },
                                    motivo="",
                                    observacao=f"Origem equipe: {eq_nome}",
                                    user_id=str(st.session_state.get("user", {}).get("id", "")),
                                )
                        except Exception:
                            pass
                        _set_post_chamado_banner("Linha movida com sucesso.")
                        st.session_state.pending_equipe = destino_equipe
                        st.success(f"Linha movida para equipe {destino_equipe}.")
                        st.rerun()

                    elif acao_linha == "Mover para outro setor":
                        df_origem.loc[mask_origem, "Segmento"] = destino_setor
                        df_origem.loc[mask_origem, "GrupoEquipe"] = destino_setor
                        df_origem.loc[mask_origem, "EquipePadrao"] = destino_equipe
                        df_origem.loc[mask_origem, "Equipe"] = destino_equipe
                        if destino_setor == "Roubo e Perda":
                            df_origem.loc[mask_origem, "Data Ocorrência"] = data_ocorrencia_operacao
                            df_origem.loc[mask_origem, "Motivo"] = motivo_operacao
                            if observacao_operacao:
                                df_origem.loc[mask_origem, "Observação"] = observacao_operacao
                        elif destino_setor == "Manutenção":
                            df_origem.loc[mask_origem, "Data da Troca"] = data_troca_operacao
                            if data_retorno_operacao:
                                df_origem.loc[mask_origem, "Data Retorno"] = data_retorno_operacao
                            if motivo_operacao:
                                df_origem.loc[mask_origem, "Motivo"] = motivo_operacao
                            if observacao_operacao:
                                df_origem.loc[mask_origem, "Observação"] = observacao_operacao
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="mover_setor",
                            entidade="linhas",
                            chave=linha_alvo,
                            antes=before_row,
                            depois={
                                "Segmento": destino_setor,
                                "GrupoEquipe": destino_setor,
                                "EquipePadrao": destino_equipe,
                                "Equipe": destino_equipe,
                                "Data Ocorrência": data_ocorrencia_operacao if destino_setor == "Roubo e Perda" else "",
                                "Data da Troca": data_troca_operacao if destino_setor == "Manutenção" else "",
                                "Data Retorno": data_retorno_operacao if destino_setor == "Manutenção" else "",
                                "Motivo": motivo_operacao,
                                "Observação": observacao_operacao,
                                "modo": modo_origem,
                            },
                        )
                        try:
                            if active_chamado_context.get("chamado_id"):
                                registrar_movimentacao_linha(
                                    chamado_id=str(active_chamado_context.get("chamado_id")),
                                    linha_num=linha_alvo,
                                    tipo_movimentacao="mover_setor",
                                    antes=before_row,
                                    depois={
                                        "Segmento": destino_setor,
                                        "EquipePadrao": destino_equipe,
                                        "Motivo": motivo_operacao,
                                        "Observação": observacao_operacao,
                                    },
                                    motivo=motivo_operacao,
                                    observacao=observacao_operacao,
                                    user_id=str(st.session_state.get("user", {}).get("id", "")),
                                )
                        except Exception:
                            pass
                        _set_post_chamado_banner("Linha movida de setor com sucesso.")
                        st.session_state.pending_segmento = destino_setor
                        st.session_state.pending_equipe = destino_equipe
                        st.success(f"Linha movida para setor {destino_setor} / equipe {destino_equipe}.")
                        st.rerun()

                    elif acao_linha == "Enviar aparelho para manutenção":
                        linha_origem = df_origem[mask_origem].iloc[0].copy()
                        linha_manut = linha_origem.copy()
                        linha_manut["Segmento"] = "Manutenção"
                        linha_manut["GrupoEquipe"] = "Manutenção"
                        linha_manut["EquipePadrao"] = "Manutenção"
                        linha_manut["Equipe"] = "Manutenção"
                        linha_manut["Linha"] = f"MANUT-{secrets.token_hex(3).upper()}"
                        linha_manut["Nome"] = f"{str(linha_origem.get('Nome', '')).strip()} (Manutenção)".strip()
                        base_ger = str(linha_manut.get("Gerenciamento", "") or "").strip()
                        linha_manut["Gerenciamento"] = f"{base_ger} | Origem: {str(linha_origem.get('Linha', '')).strip()}".strip(" |")
                        linha_manut["Data da Troca"] = data_troca_operacao
                        linha_manut["Data Retorno"] = data_retorno_operacao
                        linha_manut["Motivo"] = motivo_operacao
                        if observacao_operacao:
                            linha_manut["Observação"] = observacao_operacao
                        df_origem = pd.concat([df_origem, pd.DataFrame([linha_manut])], ignore_index=True)
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="enviar_manutencao",
                            entidade="linhas",
                            chave=str(linha_manut["Linha"]),
                            antes=before_row,
                            depois=linha_manut.to_dict(),
                            detalhes="Criada cópia para manutenção",
                        )
                        try:
                            if active_chamado_context.get("chamado_id"):
                                registrar_movimentacao_linha(
                                    chamado_id=str(active_chamado_context.get("chamado_id")),
                                    linha_num=str(linha_manut["Linha"]),
                                    tipo_movimentacao="enviar_manutencao",
                                    antes=before_row,
                                    depois=linha_manut.to_dict(),
                                    motivo=motivo_operacao,
                                    observacao=observacao_operacao,
                                    user_id=str(st.session_state.get("user", {}).get("id", "")),
                                )
                        except Exception:
                            pass
                        _set_post_chamado_banner("Envio para manutenção registrado com sucesso.")
                        st.session_state.pending_segmento = "Manutenção"
                        st.session_state.pending_equipe = "Manutenção"
                        st.session_state.pending_linha = linha_manut["Linha"]
                        st.success("Aparelho enviado para Manutenção. Atualize a linha original com o novo aparelho.")
                        st.rerun()

                    elif acao_linha in ("Desativar linha", "Ativar linha"):
                        destino_modo = "desativadas" if acao_linha == "Desativar linha" else "ativas"
                        if destino_modo == modo_origem:
                            st.info("A linha já está neste modo.")
                            return
                        df_dest = load_linhas(modo=destino_modo)
                        movidas = df_origem[mask_origem].copy()
                        df_origem = df_origem[~mask_origem].copy()
                        linhas_movidas = set(movidas["Linha"].fillna("").astype(str))
                        if not df_dest.empty:
                            mask_rm = ~df_dest["Linha"].fillna("").astype(str).isin(linhas_movidas)
                            df_dest = df_dest[mask_rm].copy()
                        df_dest = pd.concat([df_dest, movidas], ignore_index=True)
                        if motivo_operacao:
                            df_dest.loc[df_dest["Linha"].fillna("").astype(str) == linha_alvo, "Motivo"] = motivo_operacao
                        if observacao_operacao:
                            df_dest.loc[df_dest["Linha"].fillna("").astype(str) == linha_alvo, "Observação"] = observacao_operacao
                        save_linhas(df_origem, modo=modo_origem)
                        save_linhas(df_dest, modo=destino_modo)
                        _audit(
                            acao="mudar_modo_linha",
                            entidade="linhas",
                            chave=linha_alvo,
                            antes=before_row,
                            depois={**before_row, "modo": destino_modo, "Motivo": motivo_operacao, "Observação": observacao_operacao},
                        )
                        try:
                            if active_chamado_context.get("chamado_id"):
                                registrar_movimentacao_linha(
                                    chamado_id=str(active_chamado_context.get("chamado_id")),
                                    linha_num=linha_alvo,
                                    tipo_movimentacao="mudar_modo_linha",
                                    antes=before_row,
                                    depois={"modo": destino_modo, "Motivo": motivo_operacao, "Observação": observacao_operacao},
                                    motivo=motivo_operacao,
                                    observacao=observacao_operacao,
                                    user_id=str(st.session_state.get("user", {}).get("id", "")),
                                )
                        except Exception:
                            pass
                        _set_post_chamado_banner("Alteração de status da linha concluída com sucesso.")
                        st.success(f"Linha movida para {destino_modo}.")
                        st.rerun()
                except Exception as exc:
                    import logging as _log
                    _log.getLogger(__name__).error("Erro ao executar ação na linha: %s", exc, exc_info=True)
                    st.error("Não foi possível executar a ação. Tente novamente.")

    all_edited: list[pd.DataFrame] = []
    all_base_views: list[pd.DataFrame] = []

    if segmento_sel == "Manutenção":
        # Remove coluna Histórico – usamos seções separadas em vez disso
        tbl = tbl.drop(columns=["Histórico", "__manut_antigo"], errors="ignore")
        tbl_novos, tbl_antigos = _split_manut_novos_antigos(tbl)
        for block_name, block_df, is_novos in [
            ("Manutenção", tbl_novos, True),
            ("Manutenções antigas", tbl_antigos, False),
        ]:
            dfe = _sort_manut_block(block_df, is_novos)
            header_text = f'<div class="{header_cls}">📋 {block_name}</div>'
            st.markdown(header_text, unsafe_allow_html=True)
            if usar_editor:
                all_base_views.append(dfe.copy())
                dfe_editor = _decorate_editor_df(dfe, manut_antigos=not is_novos)
                ed_key = "ed_manut_novos" if is_novos else "ed_manut_antigos"
                col_config = {
                    "__row_key": None,
                    "__base_linha": None,
                    "↕ Ordem": st.column_config.NumberColumn("↕ Ordem", min_value=1, step=1, format="%d"),
                    "Mover equipe": st.column_config.SelectboxColumn("Mover equipe", options=[""] + equipes_mover_manut),
                }
                if not is_novos and "↕ Ordem" not in dfe_editor.columns:
                    col_config.pop("↕ Ordem", None)
                ed = st.data_editor(
                    dfe_editor,
                    width="stretch",
                    num_rows="fixed",
                    hide_index=True,
                    key=ed_key,
                    column_config=col_config,
                    disabled=["__row_key", "__base_linha", "Conflito"],
                )
                all_edited.append(ed)
            else:
                st.markdown(_render_equipe_tabela(dfe, block_name), unsafe_allow_html=True)
    elif equipe_sel == "Todas as equipes":
        _col_map = {"Equipe": "EquipePadrao", "Setor": "Setor", "Gestor": "Gestor", "Cargo": "Cargo"}
        _group_col = _col_map.get(agrupar_por, "EquipePadrao")
        if _group_col not in tbl.columns:
            _group_col = "EquipePadrao"
        eq_order = [e for e in equipes if e in tbl[_group_col].unique()]
        eq_order += [e for e in sorted(tbl[_group_col].dropna().astype(str).unique()) if e not in eq_order]
        for eq in eq_order:
            dfe = _sort_for_display(tbl[tbl[_group_col].fillna("").astype(str) == str(eq)])
            dfe = _apply_manual_team_order(dfe, eq)
            if segmento_sel in ("Internos", "Roubo e Perda"):
                header_text = f'<div class="{header_cls}">📋 {eq}</div>'
            elif eq in ("Gerentes do Alimento", "Gerentes do Medicamento"):
                gestor = "—"
                supervisor = "—"
                header_text = f'<div class="{header_cls}">📋 {eq}</div>'
            else:
                gestor = non_empty_or_default(dfe["Gestor"].iloc[0], "—")
                supervisor = _supervisor_display(dfe, eq, rules_path)
                header_text = f'<div class="{header_cls}">📋 {eq}</div>'
            st.markdown(header_text, unsafe_allow_html=True)
            if usar_editor:
                all_base_views.append(dfe.copy())
                dfe_editor = _decorate_editor_df(dfe)
                ed = st.data_editor(
                    dfe_editor,
                    width="stretch",
                    num_rows="fixed",
                    hide_index=True,
                    key=f"ed_{eq}",
                    column_config={
                        "__row_key": None,
                        "__base_linha": None,
                        "↕ Ordem": st.column_config.NumberColumn("↕ Ordem", min_value=1, step=1, format="%d"),
                        "Mover equipe": st.column_config.SelectboxColumn("Mover equipe", options=[""] + equipes),
                    },
                    disabled=["__row_key", "__base_linha", "Conflito"],
                )
                all_edited.append(ed)
            else:
                st.markdown(_render_equipe_tabela(dfe, eq), unsafe_allow_html=True)
    else:
        dfe = _sort_for_display(tbl)
        dfe = _apply_manual_team_order(dfe, equipe_sel)
        if segmento_sel in ("Internos", "Roubo e Perda"):
            header_text = f'<div class="{header_cls}">📋 {equipe_sel}</div>'
        elif equipe_sel in ("Gerentes do Alimento", "Gerentes do Medicamento"):
            gestor = "—"
            supervisor = "—"
            header_text = f'<div class="{header_cls}">📋 {equipe_sel}</div>'
        else:
            gestor = non_empty_or_default(dfe["Gestor"].iloc[0], "—")
            supervisor = _supervisor_display(dfe, equipe_sel, rules_path)
            header_text = f'<div class="{header_cls}">📋 {equipe_sel}</div>'
        st.markdown(header_text, unsafe_allow_html=True)
        if usar_editor:
            all_base_views.append(dfe.copy())
            dfe_editor = _decorate_editor_df(dfe)
            ed = st.data_editor(
                dfe_editor,
                width="stretch",
                num_rows="fixed",
                hide_index=True,
                key=f"ed_{equipe_sel}",
                column_config={
                    "__row_key": None,
                    "__base_linha": None,
                    "↕ Ordem": st.column_config.NumberColumn("↕ Ordem", min_value=1, step=1, format="%d"),
                    "Mover equipe": st.column_config.SelectboxColumn("Mover equipe", options=[""] + equipes),
                },
                disabled=["__row_key", "__base_linha", "Conflito"],
            )
            all_edited.append(ed)
        else:
            st.markdown(_render_equipe_tabela(dfe, equipe_sel), unsafe_allow_html=True)

    if usar_editor and all_base_views:
        stored_editor_context = st.session_state.get("_editor_base_context")
        if stored_editor_context != editor_context_key:
            base_snapshot_df = pd.concat(all_base_views, ignore_index=True)
            st.session_state["_editor_base_context"] = editor_context_key
            st.session_state["_editor_base_snapshot"] = base_snapshot_df.to_dict("records")
            st.session_state["_editor_conflict_payload"] = {}

        conflict_payload = st.session_state.get("_editor_conflict_payload") or {}
        reapply_request = st.session_state.get("_reapply_conflict_request")
        if conflict_payload and conflict_payload.get("context") == editor_context_key and reapply_request == editor_context_key:
            pending_changes = conflict_payload.get("pending_changes") or []
            st.session_state["_reapply_conflict_request"] = ""
            if pending_changes:
                seg_lower_reapply = segmento_sel.strip().lower()
                full_current = df_full_mode.copy()
                mask_seg_reapply = full_current["Segmento"].fillna("").astype(str).str.strip().str.lower() == seg_lower_reapply
                segment_current = full_current[mask_seg_reapply].copy()
                others_current = full_current[~mask_seg_reapply].copy()
                grouped_changes: dict[str, list[dict[str, str]]] = {}
                for ch in pending_changes:
                    base_linha = str(ch.get("base_linha") or "").strip()
                    if not base_linha:
                        continue
                    grouped_changes.setdefault(base_linha, []).append(ch)

                def _norm_cmp(v: Any) -> str:
                    if v is None:
                        return ""
                    if isinstance(v, float) and pd.isna(v):
                        return ""
                    return str(v).strip()

                applied_fields = 0
                applied_rows: set[str] = set()
                remaining_pending: list[dict[str, str]] = []
                remaining_resumo: list[dict[str, str]] = []
                remaining_detalhes: list[dict[str, str]] = []

                if "Linha" in segment_current.columns:
                    segment_current["__base_linha_norm"] = segment_current["Linha"].fillna("").astype(str).str.strip()

                for base_linha, changes in grouped_changes.items():
                    if "__base_linha_norm" not in segment_current.columns:
                        remaining_resumo.append({"Linha": base_linha or "—", "Motivo": "Nao foi possivel comparar a linha no banco."})
                        remaining_pending.extend(changes)
                        continue
                    current_matches = segment_current[segment_current["__base_linha_norm"] == base_linha]
                    if current_matches.empty:
                        remaining_resumo.append({"Linha": base_linha or "—", "Motivo": "A linha nao foi encontrada no banco na versao atual."})
                        remaining_pending.extend(changes)
                        continue
                    row_idx = current_matches.index[0]
                    row_remaining_fields: list[str] = []
                    for ch in changes:
                        campo = str(ch.get("campo") or "").strip()
                        valor_original = _norm_cmp(ch.get("original"))
                        minha_edicao = _norm_cmp(ch.get("mine"))
                        if not campo or campo not in segment_current.columns:
                            remaining_pending.append(ch)
                            row_remaining_fields.append(campo or "—")
                            continue
                        valor_atual = _norm_cmp(segment_current.at[row_idx, campo])
                        if valor_atual == valor_original or valor_atual == minha_edicao:
                            if valor_atual != minha_edicao:
                                segment_current.at[row_idx, campo] = minha_edicao
                            applied_fields += 1
                            applied_rows.add(base_linha)
                        else:
                            remaining_pending.append(ch)
                            row_remaining_fields.append(campo)
                            remaining_detalhes.append(
                                {
                                    "Linha": base_linha or "—",
                                    "Campo": campo,
                                    "Valor original": valor_original or "—",
                                    "Minha edição": minha_edicao or "—",
                                    "Valor atual no banco": valor_atual or "—",
                                }
                            )
                    if row_remaining_fields:
                        remaining_resumo.append(
                            {
                                "Linha": base_linha or "—",
                                "Motivo": "Alguns campos continuam em conflito.",
                                "Campos com diferenca": ", ".join(row_remaining_fields[:6]) + (f" (+{len(row_remaining_fields) - 6})" if len(row_remaining_fields) > 6 else ""),
                            }
                        )

                if applied_fields > 0:
                    segment_current = segment_current.drop(columns=["__base_linha_norm"], errors="ignore")
                    updated_full = pd.concat([others_current, segment_current], ignore_index=True)
                    updated_full = updated_full.drop(columns=["__row_key", "__base_linha"], errors="ignore")
                    sort_cols_save = [c for c in ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"] if c in updated_full.columns]
                    if sort_cols_save:
                        updated_full = updated_full.sort_values(sort_cols_save, kind="stable")
                    n_reapply = save_linhas(updated_full, modo=modo_db)
                    _audit(
                        acao="reaplicar_edicoes_seguras",
                        entidade="linhas",
                        chave=modo_db,
                        antes={
                            "linhas_reaplicadas": sorted(applied_rows),
                            "campos_reaplicados": int(applied_fields),
                        },
                        depois={
                            "registros_gravados_modo": int(n_reapply),
                            "segmento": segmento_sel,
                            "modo": modo_db,
                        },
                        detalhes=f"Reaplicacao segura executada em {len(applied_rows)} linha(s).",
                    )

                if remaining_pending:
                    st.session_state["_editor_conflict_payload"] = {
                        "context": editor_context_key,
                        "resumo": remaining_resumo,
                        "detalhes": remaining_detalhes,
                        "pending_changes": remaining_pending,
                    }
                else:
                    st.session_state["_editor_conflict_payload"] = {}

                st.session_state["_editor_base_context"] = ""
                st.session_state["_editor_base_snapshot"] = []

                if applied_fields > 0 and remaining_pending:
                    st.session_state["_post_conflict_feedback"] = {
                        "type": "warning",
                        "text": f"Algumas alteracoes seguras foram reaplicadas ({applied_fields} campo(s)), mas ainda existem conflitos pendentes.",
                    }
                elif applied_fields > 0:
                    st.session_state["_post_conflict_feedback"] = {
                        "type": "success",
                        "text": f"Reaplicacao segura concluida com sucesso. {applied_fields} campo(s) reaplicado(s).",
                    }
                else:
                    st.session_state["_post_conflict_feedback"] = {
                        "type": "warning",
                        "text": "Nenhuma alteracao segura pode ser reaplicada automaticamente.",
                    }
                st.rerun()

    live_changed_rows_df = pd.DataFrame()
    live_validation_errors: list[dict[str, str]] = []
    if usar_editor and all_base_views and all_edited:
        live_changed_rows_df = collect_changed_editor_rows(all_base_views, all_edited)
        if not live_changed_rows_df.empty:
            live_candidate_df = build_candidate_full_df(
                df_full_mode,
                live_changed_rows_df.drop(columns=["Conflito", "Validacao", "↕ Ordem", "Mover equipe"], errors="ignore"),
            )
            live_validation_errors = collect_editor_validation_errors(
                live_changed_rows_df.drop(columns=["Conflito", "Validacao", "__row_key", "__base_linha", "↕ Ordem", "Mover equipe"], errors="ignore"),
                live_candidate_df.drop(columns=["Conflito", "Validacao", "__base_linha", "↕ Ordem", "Mover equipe"], errors="ignore"),
            )
            if live_validation_errors:
                linhas_com_pendencia = sorted(
                    {
                        str(item.get("Linha") or "").strip()
                        for item in live_validation_errors
                        if str(item.get("Linha") or "").strip()
                    }
                )
                st.warning(
                    f"Existem pendencias de validacao em {len(linhas_com_pendencia)} linha(s) alterada(s). Corrija antes de salvar."
                )
                with st.expander("Ver pendencias de validacao", expanded=False):
                    st.dataframe(pd.DataFrame(live_validation_errors), width="stretch", hide_index=True)

    if usar_editor and all_edited:
        if save_clicked_top:
            seg_lower = segmento_sel.strip().lower()
            mask_seg = df_full_mode["Segmento"].fillna("").astype(str).str.strip().str.lower() == seg_lower
            outros = df_full_mode[~mask_seg].copy()
            segment_atual = df_full_mode[mask_seg].copy()
            edited_df = pd.concat(all_edited, ignore_index=True)
            edited_df = edited_df.drop(columns=["Conflito"], errors="ignore")

            has_reorder_action = "↕ Ordem" in edited_df.columns
            has_team_move_action = False
            if "Mover equipe" in edited_df.columns:
                has_team_move_action = bool(edited_df["Mover equipe"].fillna("").astype(str).str.strip().ne("").any())
            if "Mover equipe" in edited_df.columns:
                dest_series = edited_df["Mover equipe"].fillna("").astype(str).str.strip()
                mask_move_team = dest_series != ""
                if segmento_sel == "Manutenção":
                    # No segmento Manutenção, "Mover equipe" = mover entre seções (Data da Troca)
                    today_str = datetime.now().strftime("%Y-%m-%d")
                    yesterday_str = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
                    mask_para_antigos = mask_move_team & (dest_series == "Manutenções antigas")
                    mask_para_novos = mask_move_team & (dest_series == "Manutenção")
                    if "Data da Troca" in edited_df.columns:
                        edited_df.loc[mask_para_antigos, "Data da Troca"] = yesterday_str
                        edited_df.loc[mask_para_novos, "Data da Troca"] = today_str
                    # EquipePadrao permanece "Manutenção" (não alterar para equipes externas)
                else:
                    if "EquipePadrao" in edited_df.columns:
                        edited_df.loc[mask_move_team, "EquipePadrao"] = dest_series[mask_move_team]
                    if "Equipe" in edited_df.columns:
                        edited_df.loc[mask_move_team, "Equipe"] = dest_series[mask_move_team]
            if has_reorder_action and "EquipePadrao" in edited_df.columns:
                ordem_num = pd.to_numeric(edited_df["↕ Ordem"], errors="coerce")
                edited_df["__ordem_manual"] = ordem_num.fillna(10**9)
                edited_df["__ordem_idx"] = range(len(edited_df))
                edited_df = edited_df.sort_values(
                    ["EquipePadrao", "__ordem_manual", "__ordem_idx"],
                    kind="stable",
                ).drop(columns=["__ordem_manual", "__ordem_idx"], errors="ignore")
                # Renumeração automática da coluna de ordem por equipe (1..N).
                if "↕ Ordem" in edited_df.columns:
                    edited_df["↕ Ordem"] = (
                        edited_df.groupby("EquipePadrao", sort=False).cumcount() + 1
                    )
            elif "↕ Ordem" in edited_df.columns:
                # Mesmo sem mudança manual de ordem, garante sequência válida na equipe.
                edited_df["↕ Ordem"] = (
                    edited_df.groupby("EquipePadrao", sort=False).cumcount() + 1
                )
                # Persiste ordem manual em sessão para exibição consistente.
                if "Linha" in edited_df.columns:
                    manual_map = st.session_state.get("_manual_order_map") or {}
                    for eq_name, grp in edited_df.groupby("EquipePadrao", sort=False):
                        key = f"{modo_db}|{segmento_sel}|{str(eq_name).strip()}"
                        linhas_ord = [
                            str(v).strip()
                            for v in grp["Linha"].fillna("").astype(str).tolist()
                            if str(v).strip()
                        ]
                        manual_map[key] = linhas_ord
                    st.session_state["_manual_order_map"] = manual_map
            row_keys_editadas = set(edited_df["__row_key"].dropna().astype(str).unique()) if "__row_key" in edited_df.columns else set()
            base_snapshot_records = st.session_state.get("_editor_base_snapshot") or []
            base_snapshot_df = pd.DataFrame(base_snapshot_records) if base_snapshot_records else pd.DataFrame()

            def _norm_cmp(v: Any) -> str:
                if v is None:
                    return ""
                if isinstance(v, float) and pd.isna(v):
                    return ""
                return str(v).strip()

            row_keys_alteradas: set[str] = set()
            campos_alterados_set: set[str] = set()
            linhas_editadas_lista: list[str] = []
            celulas_alteradas = 0
            alteracoes_detalhadas: list[dict[str, str]] = []
            max_alteracoes_detalhadas = 200
            conflitos_detectados: list[dict[str, str]] = []
            conflitos_detalhados: list[dict[str, str]] = []
            pending_conflict_changes: list[dict[str, str]] = []

            if row_keys_editadas and "__row_key" in edited_df.columns and "__row_key" in base_snapshot_df.columns:
                base_idx = base_snapshot_df.set_index("__row_key", drop=False)
                edit_idx = edited_df.set_index("__row_key", drop=False)
                cols_compare = [
                    c for c in edited_df.columns
                    if c in base_snapshot_df.columns and c not in ("__row_key", "↕ Ordem", "Mover equipe")
                ]
                for rk in row_keys_editadas:
                    if rk not in base_idx.index or rk not in edit_idx.index:
                        continue
                    base_row = base_idx.loc[rk]
                    edit_row = edit_idx.loc[rk]
                    linha_ref = _norm_cmp(edit_row.get("Linha")) or _norm_cmp(base_row.get("Linha")) or str(rk)
                    houve_mudanca = False
                    for col in cols_compare:
                        base_norm = _norm_cmp(base_row.get(col))
                        edit_norm = _norm_cmp(edit_row.get(col))
                        if base_norm != edit_norm:
                            houve_mudanca = True
                            celulas_alteradas += 1
                            campos_alterados_set.add(str(col))
                            if len(alteracoes_detalhadas) < max_alteracoes_detalhadas:
                                alteracoes_detalhadas.append(
                                    {
                                        "linha": linha_ref,
                                        "campo": str(col),
                                        "antes": base_norm or "—",
                                        "depois": edit_norm or "—",
                                    }
                                )
                    if houve_mudanca:
                        row_keys_alteradas.add(str(rk))
                        linhas_editadas_lista.append(linha_ref)

            linhas_editadas_lista = sorted(set(linhas_editadas_lista))
            campos_alterados_lista = sorted(campos_alterados_set)

            if has_reorder_action and not row_keys_alteradas and "__row_key" in edited_df.columns:
                # Reordenação sem alteração de campo: força salvamento da nova ordem.
                row_keys_alteradas = set(edited_df["__row_key"].dropna().astype(str).unique())
                linhas_editadas_lista = sorted(
                    {
                        str(v).strip()
                        for v in edited_df["Linha"].fillna("").astype(str).tolist()
                        if str(v).strip()
                    }
                )
                if not campos_alterados_lista:
                    campos_alterados_lista = ["ordem_linha"]
            if has_team_move_action and "__row_key" in edited_df.columns:
                row_keys_alteradas.update(set(edited_df["__row_key"].dropna().astype(str).unique()))
                if "mover_equipe_inline" not in campos_alterados_lista:
                    campos_alterados_lista.append("mover_equipe_inline")

            if not row_keys_alteradas:
                st.info("Nenhuma alteração detectada para salvar.")
            else:
                latest_segment_idx = segment_atual.copy()
                latest_segment_idx["__base_linha_norm"] = latest_segment_idx["Linha"].fillna("").astype(str).str.strip() if "Linha" in latest_segment_idx.columns else ""
                base_snapshot_idx = base_snapshot_df.set_index("__row_key", drop=False) if "__row_key" in base_snapshot_df.columns else pd.DataFrame()
                edit_idx = edited_df.set_index("__row_key", drop=False) if "__row_key" in edited_df.columns else pd.DataFrame()
                cols_conf_compare = [c for c in base_snapshot_df.columns if c in segment_atual.columns and c not in ("__row_key", "__base_linha")]
                if segmento_sel == "Manutenção" and has_team_move_action:
                    cols_conf_compare = [c for c in cols_conf_compare if c != "Data da Troca"]
                for rk in sorted(row_keys_alteradas):
                    if base_snapshot_idx.empty or rk not in base_snapshot_idx.index:
                        continue
                    if not edit_idx.empty and rk in edit_idx.index:
                        edit_row = edit_idx.loc[rk]
                    else:
                        edit_row = base_snapshot_idx.loc[rk]
                    base_row = base_snapshot_idx.loc[rk]
                    base_linha = _norm_cmp(base_row.get("__base_linha") or base_row.get("Linha"))
                    current_matches = latest_segment_idx[latest_segment_idx["__base_linha_norm"] == base_linha] if base_linha else pd.DataFrame()
                    if current_matches.empty:
                        conflitos_detectados.append(
                            {
                                "Linha": base_linha or "—",
                                "Motivo": "A linha nao foi encontrada no banco na versao atual.",
                            }
                        )
                        continue
                    current_row = current_matches.iloc[0]
                    campos_conflito: list[str] = []
                    for col in cols_conf_compare:
                        if _norm_cmp(base_row.get(col)) != _norm_cmp(current_row.get(col)):
                            campos_conflito.append(str(col))
                            minha_edicao = _norm_cmp(edit_row.get(col))
                            valor_original = _norm_cmp(base_row.get(col)) or "—"
                            valor_banco = _norm_cmp(current_row.get(col)) or "—"
                            pending_conflict_changes.append(
                                {
                                    "base_linha": base_linha or "—",
                                    "campo": str(col),
                                    "original": valor_original,
                                    "mine": minha_edicao or "—",
                                    "current": valor_banco,
                                    "original_raw": _norm_cmp(base_row.get(col)),
                                    "mine_raw": _norm_cmp(edit_row.get(col)),
                                    "current_raw": _norm_cmp(current_row.get(col)),
                                }
                            )
                            conflitos_detalhados.append(
                                {
                                    "Linha": base_linha or "—",
                                    "Campo": str(col),
                                    "Valor original": valor_original,
                                    "Minha edição": minha_edicao or "—",
                                    "Valor atual no banco": valor_banco,
                                }
                            )
                    if campos_conflito:
                        conflitos_detectados.append(
                            {
                                "Linha": base_linha or "—",
                                "Motivo": "Outro usuario alterou esta linha apos a abertura da tela.",
                                "Campos com diferenca": ", ".join(campos_conflito[:6]) + (f" (+{len(campos_conflito) - 6})" if len(campos_conflito) > 6 else ""),
                            }
                        )

                if conflitos_detectados:
                    st.session_state["_editor_conflict_payload"] = {
                        "context": editor_context_key,
                        "resumo": conflitos_detectados,
                        "detalhes": conflitos_detalhados,
                        "pending_changes": pending_conflict_changes,
                    }
                    st.error("Conflito de edicao detectado. Atualize a tela antes de salvar para evitar sobrescrever alteracoes de outra pessoa.")
                    st.dataframe(pd.DataFrame(conflitos_detectados), width="stretch", hide_index=True)
                    if conflitos_detalhados:
                        with st.expander("Ver diferenças do conflito", expanded=False):
                            st.dataframe(pd.DataFrame(conflitos_detalhados), width="stretch", hide_index=True)
                    return

                mask_rows_alteradas = edited_df["__row_key"].fillna("").astype(str).isin(row_keys_alteradas)
                edited_changed_df = edited_df[mask_rows_alteradas].copy()
                if has_reorder_action:
                    novos_seg = edited_df.copy()
                else:
                    mask_manter = ~segment_atual["__row_key"].fillna("").astype(str).isin(row_keys_alteradas)
                    segment_nao_editado = segment_atual[mask_manter].copy()
                    novos_seg = pd.concat([segment_nao_editado, edited_changed_df], ignore_index=True)
                novos = pd.concat([outros, novos_seg], ignore_index=True)
                novos = novos.drop(columns=["__row_key"], errors="ignore")
                control_only_change = celulas_alteradas == 0 and (has_reorder_action or has_team_move_action)
                validation_errors: list[dict[str, str]] = []
                if not control_only_change:
                    validation_errors = collect_editor_validation_errors(
                        edited_changed_df.drop(columns=["__row_key", "__base_linha", "↕ Ordem", "Mover equipe"], errors="ignore"),
                        novos.drop(columns=["__base_linha", "↕ Ordem", "Mover equipe"], errors="ignore"),
                    )
                if validation_errors:
                    st.error("Nao foi possivel salvar porque existem campos invalidos nas linhas alteradas.")
                    st.dataframe(pd.DataFrame(validation_errors), width="stretch", hide_index=True)
                    return
                novos = novos.drop(columns=["Mover equipe"], errors="ignore")
                n = save_linhas(novos, modo=modo_db)
                campos_txt = ", ".join(campos_alterados_lista[:4]) if campos_alterados_lista else "campos não identificados"
                if campos_alterados_lista and len(campos_alterados_lista) > 4:
                    campos_txt += f" (+{len(campos_alterados_lista) - 4})"
                _audit(
                    acao="salvar_edicoes",
                    entidade="linhas",
                    chave=modo_db,
                    antes={
                        "registros_segmento_antes": int(len(segment_atual)),
                        "linhas_editadas": linhas_editadas_lista,
                        "campos_alterados": campos_alterados_lista,
                        "celulas_alteradas": int(celulas_alteradas),
                        "alteracoes_total": int(celulas_alteradas),
                        "alteracoes": alteracoes_detalhadas,
                    },
                    depois={
                        "registros_segmento_depois": int(len(novos_seg)),
                        "registros_gravados_modo": int(n),
                        "segmento": segmento_sel,
                        "modo": modo_db,
                        "campos_alterados": campos_alterados_lista,
                        "alteracoes_total": int(celulas_alteradas),
                        "alteracoes": alteracoes_detalhadas,
                    },
                    detalhes=f"Edição de {len(linhas_editadas_lista)} linha(s). Campos: {campos_txt}",
                )
                _set_post_chamado_banner("Alterações salvas com sucesso.")
                st.session_state["_editor_base_context"] = ""
                st.session_state["_editor_base_snapshot"] = []
                st.session_state["_editor_conflict_payload"] = {}
                st.success(f"Alterações salvas! {len(linhas_editadas_lista)} linha(s) alterada(s).")
                st.rerun()

    st.markdown("---")
    def _strip_tz(df: pd.DataFrame) -> pd.DataFrame:
        for col in df.select_dtypes(include=["datetimetz"]).columns:
            df = df.copy()
            df[col] = df[col].dt.tz_convert("America/Sao_Paulo").dt.tz_localize(None)
        return df

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for segmento in ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"]:
            seg_lower = segmento.strip().lower()
            mask = df_full_mode["Segmento"].fillna("").astype(str).str.strip().str.lower() == seg_lower
            df_seg = df_full_mode[mask].drop(columns=["__row_key", "__base_linha"], errors="ignore")
            if not df_seg.empty:
                _strip_tz(df_seg).to_excel(writer, sheet_name=segmento, index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    header_fill = PatternFill(start_color="2D5A24", end_color="2D5A24", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    gerente_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    supervisor_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for ws in wb.worksheets:
        header_map = {str(cell.value or "").strip(): col_idx for col_idx, cell in enumerate(ws[1], 1)}
        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center")
            row_num = row[0].row
            papel_col = header_map.get("Papel")
            equipe_col = header_map.get("EquipePadrao")
            is_gerente = False
            is_supervisor = False
            if papel_col:
                papel_val = str(ws.cell(row=row_num, column=papel_col).value or "").strip().lower()
                if papel_val == "gerente":
                    is_gerente = True
                elif papel_val == "supervisor":
                    is_supervisor = True
            if equipe_col:
                eq_val = str(ws.cell(row=row_num, column=equipe_col).value or "").strip()
                if eq_val in ("Gerentes do Alimento", "Gerentes do Medicamento"):
                    is_gerente = True
            row_fill = gerente_fill if is_gerente else (supervisor_fill if is_supervisor else None)
            if row_fill:
                for c in row:
                    c.fill = row_fill
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 14
            for cell in ws[col_letter]:
                try:
                    max_len = max(max_len, min(len(str(cell.value or "")), 50))
                except TypeError:
                    pass
            ws.column_dimensions[col_letter].width = max_len
        ws.freeze_panes = "A2"
    buf_out = io.BytesIO()
    wb.save(buf_out)
    buf_out.seek(0)
    st.download_button(
        "Exportar tudo para Excel",
        data=buf_out,
        file_name="linhas_telefones.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
