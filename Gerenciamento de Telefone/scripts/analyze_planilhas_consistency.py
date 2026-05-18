"""Analisa consistência entre planilhas legadas e banco atual."""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.db.repository import load_linhas


ROOT = Path(__file__).resolve().parent.parent
PLANILHAS_DIR = ROOT / "data" / "planilhas"
TELEFONES_PATH = PLANILHAS_DIR / "Telefones11.25.xlsx"
RELACAO_PATH = PLANILHAS_DIR / "Relação Linhas Prosper270226.xlsx"
ATIVAS_PATH = PLANILHAS_DIR / "Telefones11.25_SomenteAtivas.xlsx"
ABAS_FOCO = [
    "Prosper Norte",
    "Prosper Sul",
    "Nova Prosper",
    "Promotores",
    "Internos",
    "Troca de Aparelho",
    "Devolução Manutenção",
    "Devolucao Manutencao",
    "Roubo-Perda",
]


def digits_only(value: object) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def _find_header(ws) -> tuple[int | None, int | None, list[str]]:
    for row_idx in range(1, min(25, ws.max_row) + 1):
        values = [str(ws.cell(row_idx, col_idx).value or "").strip() for col_idx in range(1, min(35, ws.max_column) + 1)]
        lowered = [value.lower() for value in values]
        if "linha" in lowered:
            return row_idx, lowered.index("linha") + 1, values
    return None, None, []


def load_telefones_df() -> pd.DataFrame:
    wb = load_workbook(TELEFONES_PATH, data_only=True)
    rows: list[dict[str, object]] = []
    for aba in wb.sheetnames:
        if aba not in ABAS_FOCO:
            continue
        ws = wb[aba]
        header_row, line_col, headers = _find_header(ws)
        if header_row is None or line_col is None:
            continue
        max_cols = min(ws.max_column, len(headers))
        for row_idx in range(header_row + 1, ws.max_row + 1):
            linha = digits_only(ws.cell(row_idx, line_col).value)
            if not (10 <= len(linha) <= 13):
                continue
            row: dict[str, object] = {}
            for col_idx in range(1, max_cols + 1):
                key = headers[col_idx - 1] if headers[col_idx - 1] else f"col_{col_idx}"
                row[key] = ws.cell(row_idx, col_idx).value
            row["Aba"] = aba
            row["LinhaNorm"] = linha
            rows.append(row)
    return pd.DataFrame(rows)


def load_lines_from_workbook(path: Path, abas_foco: list[str] | None = None) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    rows: list[dict[str, object]] = []
    for aba in wb.sheetnames:
        if abas_foco and aba not in abas_foco:
            continue
        ws = wb[aba]
        header_row, line_col, headers = _find_header(ws)
        if header_row is None or line_col is None:
            continue
        nome_idx = None
        lowered = [value.lower() for value in headers]
        if "nome" in lowered:
            nome_idx = lowered.index("nome") + 1
        max_cols = min(ws.max_column, len(headers))
        for row_idx in range(header_row + 1, ws.max_row + 1):
            linha = digits_only(ws.cell(row_idx, line_col).value)
            if not (10 <= len(linha) <= 13):
                continue
            nome = str(ws.cell(row_idx, nome_idx).value or "").strip() if nome_idx else ""
            rows.append({"LinhaNorm": linha, "Aba": aba, "Nome": nome})
    return pd.DataFrame(rows)


def load_relacao_lines() -> list[str]:
    wb = load_workbook(RELACAO_PATH, data_only=True)
    ws = wb["ListaAtual"] if "ListaAtual" in wb.sheetnames else wb[wb.sheetnames[0]]
    linhas: list[str] = []
    for row_idx in range(2, ws.max_row + 1):
        linha = digits_only(ws.cell(row_idx, 4).value)
        if 10 <= len(linha) <= 13:
            linhas.append(linha)
    return linhas


def first_nonempty(df: pd.DataFrame, cols: list[str]) -> pd.Series:
    out = pd.Series([""] * len(df), index=df.index, dtype=object)
    for col in cols:
        if col in df.columns:
            values = df[col].fillna("").astype(str).str.strip()
            out = out.where(out != "", values)
    return out


def main() -> int:
    telefones_df = load_telefones_df()
    relacao_linhas = load_relacao_lines()
    relacao_set = set(relacao_linhas)
    ativas_file_df = load_lines_from_workbook(ATIVAS_PATH)
    ativas_file_set = set(ativas_file_df["LinhaNorm"].astype(str).unique()) if not ativas_file_df.empty else set()

    ativas_db = load_linhas("ativas")
    desativadas_db = load_linhas("desativadas")
    db_df = pd.concat(
        [ativas_db.assign(__modo="ativas"), desativadas_db.assign(__modo="desativadas")],
        ignore_index=True,
    )

    telefones_set = set(telefones_df["LinhaNorm"].astype(str).unique()) if not telefones_df.empty else set()
    db_set = set(db_df["Linha"].fillna("").astype(str).str.strip()) if "Linha" in db_df.columns else set()
    db_set = {value for value in db_set if value}

    telefones_dup = (
        telefones_df.groupby("LinhaNorm").size().reset_index(name="qtd").query("qtd > 1").sort_values(["qtd", "LinhaNorm"], ascending=[False, True])
        if not telefones_df.empty
        else pd.DataFrame(columns=["LinhaNorm", "qtd"])
    )
    relacao_dup = pd.Series(relacao_linhas).value_counts().reset_index(name="qtd") if relacao_linhas else pd.DataFrame(columns=["LinhaNorm", "qtd"])
    if not relacao_dup.empty:
        relacao_dup.columns = ["LinhaNorm", "qtd"]
        relacao_dup = relacao_dup[relacao_dup["qtd"] > 1]

    codigo_series = first_nonempty(telefones_df, ["Código", "Códigos", "Fernando Gonçalves De Mello", "ASSET"])
    nome_series = first_nonempty(telefones_df, ["Nome", "NOME COLABORADOR", "Nomes"])
    email_series = first_nonempty(telefones_df, ["E-mail", "Email"])
    imei_a_series = first_nonempty(telefones_df, ["IMEI A", "IMEI 1", "IMEI"])
    aparelho_series = first_nonempty(telefones_df, ["Aparelho"])

    resumo = {
        "telefones_rows": int(len(telefones_df)),
        "telefones_unique_linhas": int(len(telefones_set)),
        "telefones_duplicate_linhas": int(len(telefones_dup)),
        "relacao_rows": int(len(relacao_linhas)),
        "relacao_unique_linhas": int(len(relacao_set)),
        "relacao_duplicate_linhas": int(len(relacao_dup)),
        "ativas_file_rows": int(len(ativas_file_df)),
        "ativas_file_unique_linhas": int(len(ativas_file_set)),
        "db_rows": int(len(db_df)),
        "db_unique_linhas": int(len(db_set)),
        "ativas_db_rows": int(len(ativas_db)),
        "ativas_db_unique_linhas": int(ativas_db["Linha"].fillna("").astype(str).str.strip().replace("", pd.NA).dropna().nunique()) if "Linha" in ativas_db.columns else 0,
        "linhas_telefones_not_in_db": int(len(telefones_set - db_set)),
        "linhas_relacao_not_in_db": int(len(relacao_set - db_set)),
        "linhas_telefones_not_in_relacao": int(len(telefones_set - relacao_set)),
        "linhas_relacao_not_in_telefones": int(len(relacao_set - telefones_set)),
        "linhas_relacao_not_in_ativas_file": int(len(relacao_set - ativas_file_set)),
        "linhas_ativas_file_not_in_ativas_db": int(len(ativas_file_set - set(ativas_db["Linha"].fillna("").astype(str).str.strip()))) if "Linha" in ativas_db.columns else int(len(ativas_file_set)),
        "linhas_ativas_db_not_in_ativas_file": int(len(set(ativas_db["Linha"].fillna("").astype(str).str.strip()) - ativas_file_set)) if "Linha" in ativas_db.columns else 0,
        "amostra_telefones_not_in_db": sorted(telefones_set - db_set)[:20],
        "amostra_relacao_not_in_db": sorted(relacao_set - db_set)[:20],
        "amostra_telefones_not_in_relacao": sorted(telefones_set - relacao_set)[:20],
        "amostra_relacao_not_in_telefones": sorted(relacao_set - telefones_set)[:20],
        "amostra_relacao_not_in_ativas_file": sorted(relacao_set - ativas_file_set)[:20],
        "amostra_ativas_file_not_in_ativas_db": sorted(ativas_file_set - set(ativas_db["Linha"].fillna("").astype(str).str.strip()))[:20] if "Linha" in ativas_db.columns else sorted(ativas_file_set)[:20],
        "amostra_ativas_db_not_in_ativas_file": sorted(set(ativas_db["Linha"].fillna("").astype(str).str.strip()) - ativas_file_set)[:20] if "Linha" in ativas_db.columns else [],
        "telefones_missing_nome": int((nome_series == "").sum()),
        "telefones_missing_codigo": int((codigo_series == "").sum()),
        "telefones_missing_email": int((email_series == "").sum()),
        "telefones_missing_imei_a": int((imei_a_series == "").sum()),
        "telefones_missing_aparelho": int((aparelho_series == "").sum()),
    }

    print(json.dumps(resumo, ensure_ascii=False, indent=2))
    print("\nTOP_DUP_TELEFONES")
    print(telefones_dup.head(20).to_string(index=False) if not telefones_dup.empty else "nenhum")
    print("\nTOP_DUP_RELACAO")
    print(relacao_dup.head(20).to_string(index=False) if not relacao_dup.empty else "nenhum")

    relacao_not_in_telefones = sorted(relacao_set - telefones_set)
    if relacao_not_in_telefones:
        print("\nRELACAO_SEM_TELEFONES_DETALHE")
        rel_det = ativas_file_df[ativas_file_df["LinhaNorm"].isin(relacao_not_in_telefones)].copy()
        if rel_det.empty:
            print("Nao encontradas em Telefones11.25_SomenteAtivas.xlsx")
        else:
            print(rel_det.sort_values(["Aba", "LinhaNorm"]).head(30).to_string(index=False))

    ativas_db_set = set(ativas_db["Linha"].fillna("").astype(str).str.strip()) if "Linha" in ativas_db.columns else set()
    ativas_missing_in_db = sorted(ativas_file_set - ativas_db_set)
    if ativas_missing_in_db:
        print("\nATIVAS_FILE_SEM_BANCO_DETALHE")
        atv_det = ativas_file_df[ativas_file_df["LinhaNorm"].isin(ativas_missing_in_db)].copy()
        print(atv_det.sort_values(["Aba", "LinhaNorm"]).head(50).to_string(index=False))

    ativas_extra_in_db = sorted(ativas_db_set - ativas_file_set)
    if ativas_extra_in_db:
        print("\nATIVAS_BANCO_SEM_ATIVAS_FILE_DETALHE")
        db_det = ativas_db[ativas_db["Linha"].fillna("").astype(str).str.strip().isin(ativas_extra_in_db)].copy()
        cols = [c for c in ["Linha", "Nome", "EquipePadrao", "Segmento", "Aba", "Cargo"] if c in db_det.columns]
        print(db_det[cols].sort_values(["Aba", "Linha"]).head(50).to_string(index=False))

    print("\nAMOSTRAS_FALTAS_TELEFONES")
    detail_df = telefones_df.copy()
    detail_df["CodigoNorm"] = codigo_series
    detail_df["NomeNorm"] = nome_series
    detail_df["EmailNorm"] = email_series
    detail_df["ImeiANorm"] = imei_a_series
    detail_df["AparelhoNorm"] = aparelho_series

    def _sample_missing(col_norm: str, label: str) -> None:
        missing = detail_df[detail_df[col_norm] == ""]
        if missing.empty:
            print(f"{label}: nenhum")
            return
        cols = [c for c in ["LinhaNorm", "Aba", "NomeNorm", "CodigoNorm", "EmailNorm", "AparelhoNorm", "ImeiANorm"] if c in missing.columns]
        print(f"{label}:")
        print(missing[cols].head(15).to_string(index=False))

    _sample_missing("CodigoNorm", "SEM_CODIGO")
    _sample_missing("EmailNorm", "SEM_EMAIL")
    _sample_missing("ImeiANorm", "SEM_IMEI_A")
    _sample_missing("AparelhoNorm", "SEM_APARELHO")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
