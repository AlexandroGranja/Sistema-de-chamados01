"""Reconstrói as linhas ativas a partir das abas operacionais principais da planilha."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from src.db.repository import save_linhas
from src.utils.text import digits_only, normalize_team_key, normalize_text, title_case_safe
from src.utils.validators import is_valid_phone


ROOT = Path(__file__).resolve().parent.parent
PLANILHAS_DIR = ROOT / "data" / "planilhas"
ATIVAS_PATH = PLANILHAS_DIR / "Telefones11.25_SomenteAtivas.xlsx"
ALIMENTO_RULES_PATH = ROOT / "doc" / "equipes_alimento.csv"
MEDICAMENTO_RULES_PATH = ROOT / "doc" / "equipes_medicamento.csv"

PRIMARY_SHEETS = [
    "Nova Prosper",
    "Prosper Norte",
    "Prosper Sul",
    "Promotores",
    "Internos",
    "Diretoria",
]

GESTORES_MEDICAMENTO = {
    "Prosper Norte": "Priscila Rangel Manhães",
    "Prosper Sul": "Gustavo Luis Dias De Armada",
}

OUTPUT_COLUMNS = [
    "Codigo",
    "Nome",
    "Equipe",
    "EquipePadrao",
    "GrupoEquipe",
    "TipoEquipe",
    "Localidade",
    "Gestor",
    "Supervisor",
    "Segmento",
    "Papel",
    "Linha",
    "E-mail",
    "Gerenciamento",
    "Data da Troca",
    "Data Retorno",
    "Data Ocorrência",
    "Data Solicitação TBS",
    "Motivo",
    "Observação",
    "IMEI A",
    "IMEI B",
    "Marca",
    "CHIP",
    "Aparelho",
    "Modelo",
    "Setor",
    "Cargo",
    "Desconto",
    "Perfil",
    "Empresa",
    "Ativo",
    "Numero de Serie",
    "Patrimonio",
    "Operadora",
    "Aba",
]


def _map_known_column(column_name: str) -> str:
    c = normalize_text(column_name)
    mapping = {
        "codigo": "Codigo",
        "codigos": "Codigo",
        "asset": "Codigo",
        "fernando goncalves de mello": "Codigo",
        "nome": "Nome",
        "nomes": "Nome",
        "nome colaborador": "Nome",
        "nome de guerra": "Nome de Guerra",
        "equipe": "Equipe",
        "local": "Equipe",
        "localidade": "Equipe",
        "linha": "Linha",
        "email": "E-mail",
        "e-mail": "E-mail",
        "gerenciamento": "Gerenciamento",
        "bloqueio": "Gerenciamento",
        "imei": "IMEI A",
        "imei a": "IMEI A",
        "imei 1": "IMEI A",
        "imei b": "IMEI B",
        "imei 2": "IMEI B",
        "imei2": "IMEI B",
        "chip": "CHIP",
        "marca": "Marca",
        "aparelho": "Aparelho",
        "modelo": "Modelo",
        "setor": "Setor",
        "cargo": "Cargo",
        "desconto": "Desconto",
        "perfil": "Perfil",
        "empresa": "Empresa",
        "ativo": "Ativo",
        "ativos": "Ativo",
        "numero de serie": "Numero de Serie",
        "número de série": "Numero de Serie",
        "s/n": "Numero de Serie",
        "ns": "Numero de Serie",
        "operadora": "Operadora",
        "patrimonio": "Patrimonio",
        "patrimônio": "Patrimonio",
        "data da troca": "Data da Troca",
        "data retorno": "Data Retorno",
        "data ocorrencia": "Data Ocorrência",
        "data ocorrência": "Data Ocorrência",
        "data solicitacao tbs": "Data Solicitação TBS",
        "data solicitação tbs": "Data Solicitação TBS",
        "motivo troca": "Motivo",
        "motivo": "Motivo",
        "obs": "Observação",
    }
    return mapping.get(c, column_name)


def _find_header(ws) -> tuple[int | None, int | None, list[str]]:
    for row_idx in range(1, min(25, ws.max_row) + 1):
        values = [str(ws.cell(row_idx, col_idx).value or "").strip() for col_idx in range(1, min(35, ws.max_column) + 1)]
        lowered = [normalize_text(value) for value in values]
        if "linha" in lowered:
            return row_idx, lowered.index("linha") + 1, values
    return None, None, []


def _build_alimento_map() -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    team_defaults: dict[str, dict[str, str]] = {}
    ref = pd.read_csv(ALIMENTO_RULES_PATH, dtype=str).fillna("")
    for _, row in ref.iterrows():
        equipe_real = str(row.get("equipe_real") or "").strip()
        localidade = str(row.get("localidade") or "").strip()
        gerente = str(row.get("gerente") or "").strip()
        supervisor = str(row.get("supervisor") or "").strip()
        if localidade:
            out[normalize_team_key(localidade)] = {
                "equipe": equipe_real,
                "gestor": gerente,
                "supervisor": supervisor,
            }
        elif equipe_real:
            team_defaults[normalize_team_key(equipe_real)] = {
                "equipe": equipe_real,
                "gestor": gerente,
                "supervisor": supervisor,
            }

    for equipe in ["consumo baixada", "consumo oeste", "consumo zona norte", "consumo niteroi"]:
        out[equipe] = team_defaults.get(
            equipe,
            {"equipe": title_case_safe(equipe), "gestor": "Marcelo Neves", "supervisor": ""},
        )

    out["especial consumo"] = {
        "equipe": "Equipe Especial",
        "gestor": "Marco Antonio Neves Suzart",
        "supervisor": "Ricardo Cascao",
    }
    for localidade, equipe, supervisor in [
        ("alcantara", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
        ("cabucu", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
        ("piabeta", "Consumo Baixada", ""),
        ("piabetá", "Consumo Baixada", ""),
        ("realengo", "Consumo Oeste", "Marcelo Martins Da Costa"),
        ("barra da tijuca", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
        ("recreio", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
        ("botafogo", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
        ("taquara", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
        ("anchieta", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
        ("lote xv", "Consumo Baixada", ""),
        ("niteroi i", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
        ("niterói i", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
        ("conveniencia farma 12", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
        ("conveniencia farma 20", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
        ("key account", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
        ("supervisora senior", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
        ("gerencia varejo ii", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
        ("consumo zona sul", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
        ("itaborai", "Consumo Baixada", ""),
    ]:
        out[normalize_team_key(localidade)] = {
            "equipe": equipe,
            "gestor": "Marcelo Neves",
            "supervisor": supervisor,
        }

    for key in [
        "rede especial 1",
        "rede especial 2",
        "rede especial 3",
        "rede especial 4",
        "rota especial 1",
        "rota especial 2",
        "rota especial 3",
        "rota especial 4",
        "impulso 01",
        "impulso 02",
        "impulso 03",
        "impulso 04",
        "impulso 05",
        "auto servico 01",
        "auto servico 02",
        "auto servico 03",
        "auto servico 04",
        "auto servico 05",
        "auto servico 06",
        "a s impulso 2",
        "a s impulso 3",
        "a s impulso 4",
        "a s impulso 5",
    ]:
        out[key] = {
            "equipe": "Equipe Especial",
            "gestor": "Marco Antonio Neves Suzart",
            "supervisor": "Ricardo Cascao",
        }
    return out


def _detect_tipo_equipe(equipe: str, aba: str) -> str:
    eq = normalize_text(equipe)
    ab = normalize_text(aba)
    if "intern" in eq or "intern" in ab or normalize_text(aba) == "diretoria":
        return "Interna"
    return "Externa"


def _detect_grupo_equipe(equipe_padrao: str, tipo_equipe: str) -> str:
    equipe_norm = normalize_text(equipe_padrao)
    if any(token in equipe_norm for token in ["especial", "impulso", "auto servico", "rota especial"]):
        return "Equipe Especial"
    if normalize_text(tipo_equipe) == "interna":
        return "Equipe Interna"
    return "Outras Equipes Externas"


def _classify_papel(nome: str, gestor: str, supervisor: str) -> str:
    nome_norm = normalize_text(nome).replace(" ", "")
    gestor_norm = normalize_text(gestor).replace(" ", "")
    supervisor_norm = normalize_text(supervisor).replace(" ", "")
    if nome_norm and gestor_norm and nome_norm == gestor_norm:
        return "Gerente"
    if nome_norm and supervisor_norm and nome_norm == supervisor_norm:
        return "Supervisor"
    if "vago" in normalize_text(nome):
        return "Vago"
    return "Vendedor"


def _parse_workbook() -> pd.DataFrame:
    wb = load_workbook(ATIVAS_PATH, data_only=True)
    rows: list[dict[str, Any]] = []
    for aba in PRIMARY_SHEETS:
        if aba not in wb.sheetnames:
            continue
        ws = wb[aba]
        header_row, line_col, headers = _find_header(ws)
        if header_row is None or line_col is None:
            continue

        mapped_headers = [_map_known_column(header) for header in headers]
        for row_idx in range(header_row + 1, ws.max_row + 1):
            line_value = ws.cell(row_idx, line_col).value
            if not is_valid_phone(line_value):
                continue
            row_data: dict[str, Any] = {}
            seen_keys: dict[str, int] = {}
            for col_idx in range(1, min(ws.max_column, len(mapped_headers)) + 1):
                key = mapped_headers[col_idx - 1] if mapped_headers[col_idx - 1] else f"col_{col_idx}"
                seen_keys[key] = seen_keys.get(key, 0) + 1
                if key == "IMEI A" and seen_keys[key] == 2:
                    key = "IMEI B"
                row_data[key] = ws.cell(row_idx, col_idx).value
            row_data["Linha"] = digits_only(line_value)
            row_data["Aba"] = aba
            rows.append(row_data)
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    for col in OUTPUT_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df


def _normalize_base_fields(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in OUTPUT_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    for col in ["Codigo", "Nome", "Equipe", "Linha", "E-mail", "Gerenciamento", "Aparelho", "Modelo", "Setor", "Cargo", "Desconto", "Perfil", "Empresa", "Ativo", "Numero de Serie", "Patrimonio", "Operadora", "Aba", "Motivo", "Observação"]:
        out[col] = out[col].fillna("").astype(str).str.strip()
    out["IMEI A"] = out["IMEI A"].map(digits_only)
    out["IMEI B"] = out["IMEI B"].map(digits_only)
    out["CHIP"] = out["CHIP"].map(digits_only)
    out["Linha"] = out["Linha"].map(digits_only)
    out["Marca"] = out["Aparelho"]
    return out


def _apply_primary_mapping(df: pd.DataFrame) -> pd.DataFrame:
    out = _normalize_base_fields(df)
    alimento_map = _build_alimento_map()

    for idx in out.index:
        aba = str(out.at[idx, "Aba"]).strip()
        equipe_raw = str(out.at[idx, "Equipe"]).strip()
        tipo = _detect_tipo_equipe(equipe_raw, aba)

        out.at[idx, "TipoEquipe"] = tipo
        out.at[idx, "GrupoEquipe"] = _detect_grupo_equipe(equipe_raw, tipo)
        out.at[idx, "Localidade"] = equipe_raw

        if aba == "Nova Prosper":
            out.at[idx, "Segmento"] = "Alimento"
            key = normalize_team_key(equipe_raw)
            if key in alimento_map:
                out.at[idx, "EquipePadrao"] = alimento_map[key]["equipe"]
                out.at[idx, "Gestor"] = alimento_map[key]["gestor"]
                out.at[idx, "Supervisor"] = alimento_map[key]["supervisor"]
            elif any(token in key for token in ["especial", "impulso", "rota especial", "auto servico", "a s impulso"]):
                out.at[idx, "EquipePadrao"] = "Equipe Especial"
                out.at[idx, "Gestor"] = "Marco Antonio Neves Suzart"
                out.at[idx, "Supervisor"] = "Ricardo Cascao"
            else:
                out.at[idx, "EquipePadrao"] = "Consumo Zona Norte"
                out.at[idx, "Gestor"] = "Marcelo Neves"
                out.at[idx, "Supervisor"] = "Paulo Roberto Ferreira Chaves"

        elif aba in {"Prosper Norte", "Prosper Sul"}:
            out.at[idx, "Segmento"] = "Medicamento"
            out.at[idx, "EquipePadrao"] = aba
            out.at[idx, "Gestor"] = GESTORES_MEDICAMENTO.get(aba, "")
            out.at[idx, "Supervisor"] = ""
            out.at[idx, "GrupoEquipe"] = "Outras Equipes Externas"

        elif aba == "Promotores":
            out.at[idx, "Segmento"] = "Promotores"
            out.at[idx, "EquipePadrao"] = "Promotores"
            out.at[idx, "Gestor"] = ""
            out.at[idx, "Supervisor"] = ""
            out.at[idx, "GrupoEquipe"] = "Outras Equipes Externas"

        elif aba == "Internos":
            out.at[idx, "Segmento"] = "Internos"
            out.at[idx, "EquipePadrao"] = "Internos"
            out.at[idx, "Gestor"] = ""
            out.at[idx, "Supervisor"] = ""
            out.at[idx, "GrupoEquipe"] = "Equipe Interna"

        elif aba == "Diretoria":
            out.at[idx, "Segmento"] = "Internos"
            out.at[idx, "Equipe"] = "Diretoria"
            out.at[idx, "EquipePadrao"] = "Diretoria"
            out.at[idx, "Localidade"] = "Diretoria"
            out.at[idx, "Gestor"] = ""
            out.at[idx, "Supervisor"] = ""
            out.at[idx, "GrupoEquipe"] = "Equipe Interna"
            out.at[idx, "TipoEquipe"] = "Interna"

        out.at[idx, "Papel"] = _classify_papel(
            str(out.at[idx, "Nome"]),
            str(out.at[idx, "Gestor"]),
            str(out.at[idx, "Supervisor"]),
        )

    ordered = [col for col in OUTPUT_COLUMNS if col in out.columns]
    return out[ordered].copy()


def main() -> int:
    df = _parse_workbook()
    if df.empty:
        print("Nenhuma linha ativa encontrada nas abas principais.")
        return 1

    final_df = _apply_primary_mapping(df)
    final_df = final_df.drop_duplicates(subset=["Linha"], keep="first").copy()
    total = save_linhas(final_df, modo="ativas")

    print(f"Linhas ativas reconstruidas: {total}")
    print("Totais por aba:")
    print(final_df.groupby("Aba")["Linha"].nunique().sort_index().to_string())
    print("\nTotais por segmento:")
    print(final_df.groupby("Segmento")["Linha"].nunique().sort_index().to_string())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
